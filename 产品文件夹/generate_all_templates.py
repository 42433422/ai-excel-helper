#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sqlite3
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def create_template_directories():
    """创建模板目录"""
    dirs = ['出货单模板', '发货单模板']
    for dir_name in dirs:
        if not os.path.exists(dir_name):
            os.makedirs(dir_name)
            print(f"✓ 创建目录: {dir_name}")
        else:
            print(f"目录已存在: {dir_name}")

def get_data_from_db():
    """从数据库获取数据"""
    conn = sqlite3.connect('customer_products_final_corrected.db')
    
    # 获取客户信息
    customers = pd.read_sql_query("SELECT * FROM customers", conn)
    
    # 获取产品信息
    products = pd.read_sql_query("SELECT * FROM products", conn)
    
    # 获取订单信息
    orders = pd.read_sql_query("SELECT * FROM orders", conn)
    
    conn.close()
    
    return customers, products, orders

def create_delivery_order_template(customer, products, file_path):
    """创建发货单模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "发货单"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    # 标题样式
    title_font = Font(size=20, bold=True)
    header_font = Font(size=12, bold=True)
    cell_font = Font(size=11)
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题
    ws.merge_cells('A1:H1')
    ws['A1'] = "成都国圣工业有限公司（五星花）送货单"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 客户信息
    ws['A3'] = "购货单位："
    ws['B3'] = customer['客户名称']
    ws['D3'] = "联系人："
    ws['E3'] = customer['联系人'] if customer['联系人'] and customer['联系人'].strip() else " "
    
    ws['A4'] = "日期："
    ws['B4'] = datetime.now().strftime("%Y年%m月%d日")
    ws['D4'] = "订单编号："
    ws['E4'] = customer.get('订单编号', ' ')
    
    ws['A5'] = "电话："
    ws['B5'] = customer.get('电话', ' ')
    ws['D5'] = "地址："
    ws['E5'] = customer.get('地址', ' ')
    
    # 产品表头
    headers = ['产品型号', '产品名称', '数量/件', '规格/KG', '数量/KG', '单价/元', '金额/元', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(color="FFFFFF")
    
    # 产品数据
    row_num = 8
    total_amount = 0
    
    for _, product in products.iterrows():
        for col, field in enumerate(['产品型号', '产品名称', '数量_件', '规格_KG', '数量_KG', '单价', '金额'], 1):
            value = product.get(field, '')
            if field == '金额' and value:
                total_amount += float(value)
            if value:
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.font = cell_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
        
        # 备注列
        ws.cell(row=row_num, column=8, value='').border = thin_border
        row_num += 1
    
    # 合计行
    ws.merge_cells(f'A{row_num}:F{row_num}')
    ws.cell(row=row_num, column=1, value="合计金额").alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=1).font = header_font
    ws.cell(row=row_num, column=1).border = thin_border
    
    ws.cell(row=row_num, column=7, value=total_amount).font = header_font
    ws.cell(row=row_num, column=7).border = thin_border
    ws.cell(row=row_num, column=7).number_format = '¥#,##0.00'
    
    # 备注条款
    row_num += 2
    ws.merge_cells(f'A{row_num}:H{row_num}')
    ws.cell(row=row_num, column=1, value="备注：1、产品质量检验期为七天。")
    ws.cell(row=row_num, column=1).font = cell_font
    
    row_num += 1
    ws.merge_cells(f'A{row_num}:H{row_num}')
    ws.cell(row=row_num, column=1, value="   2、此送货单为一式四联（白色存根、粉红色客户、蓝色结款、黄色会计）。")
    ws.cell(row=row_num, column=1).font = cell_font
    
    row_num += 1
    ws.merge_cells(f'A{row_num}:H{row_num}')
    ws.cell(row=row_num, column=1, value="   3、货款月结30天，逾期未付按银行贷款利息计算。")
    ws.cell(row=row_num, column=1).font = cell_font
    
    row_num += 2
    ws[f'A{row_num}'] = "购货单位签收："
    ws[f'D{row_num}'] = "送货人："
    ws[f'F{row_num}'] = "复核："
    
    wb.save(file_path)
    print(f"✓ 创建发货单模板: {file_path}")

def create_shipping_record_template(customer, products, file_path):
    """创建出货记录模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "出货记录"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 20
    
    # 标题
    ws.merge_cells('A1:K1')
    ws['A1'] = f"{customer['客户名称']} 出货记录"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 客户信息
    ws['A3'] = "客户名称："
    ws['B3'] = customer['客户名称']
    ws['D3'] = "联系人："
    ws['E3'] = customer['联系人'] if customer['联系人'] and customer['联系人'].strip() else " "
    
    ws['A4'] = "供应商："
    ws['B4'] = customer.get('供应商名称', ' ')
    ws['D4'] = "供应商电话："
    ws['E4'] = customer.get('供应商电话', ' ')
    
    ws['A5'] = "创建时间："
    ws['B5'] = customer.get('创建时间', ' ')
    
    # 表头
    headers = ['购货单位', '联系人', '产品型号', '产品名称', '规格(KG/桶)', '数量(件)', '重量(KG)', '单价(元)', '金额(元)', '订单日期', '文件来源']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 产品数据
    row_num = 8
    total_amount = 0
    total_weight = 0
    
    for _, product in products.iterrows():
        # 购货单位
        ws.cell(row=row_num, column=1, value=customer['客户名称'])
        # 联系人
        ws.cell(row=row_num, column=2, value=customer['联系人'] if customer['联系人'] and customer['联系人'].strip() else " ")
        # 产品型号
        ws.cell(row=row_num, column=3, value=product.get('产品型号', ''))
        # 产品名称
        ws.cell(row=row_num, column=4, value=product.get('产品名称', ''))
        # 规格
        spec = product.get('规格_KG', 0)
        ws.cell(row=row_num, column=5, value=float(spec) if spec else 0)
        # 数量(件)
        qty = product.get('数量_件', 0)
        ws.cell(row=row_num, column=6, value=float(qty) if qty else 0)
        # 重量(KG)
        weight = product.get('数量_KG', 0)
        ws.cell(row=row_num, column=7, value=float(weight) if weight else 0)
        total_weight += float(weight) if weight else 0
        # 单价
        price = product.get('单价', 0)
        ws.cell(row=row_num, column=8, value=float(price) if price else 0)
        # 金额
        amount = product.get('金额', 0)
        ws.cell(row=row_num, column=9, value=float(amount) if amount else 0)
        total_amount += float(amount) if amount else 0
        # 订单日期
        ws.cell(row=row_num, column=10, value=product.get('日期', ''))
        # 文件来源
        ws.cell(row=row_num, column=11, value=product.get('来源', ''))
        
        row_num += 1
    
    # 合计行
    ws.merge_cells(f'A{row_num}:F{row_num}')
    ws.cell(row=row_num, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=row_num, column=7, value=total_weight).font = Font(bold=True)
    ws.cell(row=row_num, column=9, value=total_amount).font = Font(bold=True)
    
    wb.save(file_path)
    print(f"✓ 创建出货记录模板: {file_path}")

def generate_all_templates():
    """生成所有32个模板文件"""
    print("=" * 80)
    print("=== 生成32个Excel模板文件 ===")
    print("=" * 80)
    
    # 创建目录
    create_template_directories()
    
    # 获取数据
    customers, products, orders = get_data_from_db()
    
    print(f"\n客户总数: {len(customers)}")
    print(f"产品总数: {len(products)}")
    print()
    
    # 为每个客户创建模板
    template_count = 0
    for _, customer in customers.iterrows():
        customer_name = customer['客户名称']
        print(f"\n处理客户: {customer_name}")
        
        # 获取该客户的产品
        customer_products = products[products['客户ID'] == customer['customer_id']]
        print(f"  产品数量: {len(customer_products)}")
        
        # 创建文件名（避免非法字符）
        safe_name = customer_name.replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_').replace('[', '_').replace(']', '_')
        
        # 发货单模板
        delivery_file = f"发货单模板/{safe_name}-发货单模板.xlsx"
        create_delivery_order_template(customer, customer_products, delivery_file)
        template_count += 1
        
        # 出货记录模板
        shipping_file = f"出货单模板/{safe_name}-出货记录模板.xlsx"
        create_shipping_record_template(customer, customer_products, shipping_file)
        template_count += 1
    
    print("\n" + "=" * 80)
    print(f"✓ 成功生成 {template_count} 个模板文件!")
    print(f"  - 发货单模板: 16个")
    print(f"  - 出货记录模板: 16个")
    print("=" * 80)

if __name__ == "__main__":
    generate_all_templates()