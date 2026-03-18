import pandas as pd
import openpyxl

# 定义文件路径
file_path = 'e:\\女娲1号\\发货单\\七彩乐园.xlsx'

print(f"分析文件：{file_path}")
print("=" * 60)

# 使用openpyxl加载工作簿
wb = openpyxl.load_workbook(file_path)

# 查看所有工作表
print(f"工作表列表：{wb.sheetnames}")

# 分析25出货工作表（之前处理过的）
ws = wb['25出货']
print(f"\n分析工作表：{ws.title}")
print(f"行数：{ws.max_row}")
print(f"列数：{ws.max_column}")

# 查看前20行数据，包括标题
print("\n前20行数据：")
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
    # 过滤掉全空的行
    if any(cell is not None for cell in row):
        print(f"行 {row_idx}: {row}")

# 关闭工作簿
wb.close()