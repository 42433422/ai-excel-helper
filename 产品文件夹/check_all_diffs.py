#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

def check_all_differences():
    """检查所有可能的差异点"""
    wb1 = load_workbook('发货单/尹玉华1.xlsx')
    ws1 = wb1['尹']
    
    wb2 = load_workbook('尹.xlsx')
    ws2 = wb2.active
    
    print('=== 完整差异检查 ===')
    
    # 1. 工作表标签颜色
    print('\n1. 工作表标签颜色:')
    print(f'   原始: {ws1.sheet_properties.tabColor}')
    print(f'   复制: {ws2.sheet_properties.tabColor}')
    
    # 2. 冻结窗格
    print('\n2. 冻结窗格:')
    print(f'   原始: 水平={ws1.freeze_panes.row if ws1.freeze_panes else None}, 垂直={ws1.freeze_panes.column if ws1.freeze_panes else None}')
    print(f'   复制: 水平={ws2.freeze_panes.row if ws2.freeze_panes else None}, 垂直={ws2.freeze_panes.column if ws2.freeze_panes else None}')
    
    # 3. 拆分窗格
    print('\n3. 拆分窗格:')
    print(f'   原始: 水平拆分={ws1.sheet_view.splitHorizontal}, 垂直拆分={ws1.sheet_view.splitVertical}')
    print(f'   复制: 水平拆分={ws2.sheet_view.splitHorizontal}, 垂直拆分={ws2.sheet_view.splitVertical}')
    
    # 4. 滚动位置
    print('\n4. 滚动位置:')
    print(f'   原始: 上={ws1.sheet_view.topLeftCell.row}, 左={ws1.sheet_view.topLeftCell.column}')
    print(f'   复制: 上={ws2.sheet_view.topLeftCell.row}, 左={ws2.sheet_view.topLeftCell.column}')
    
    # 5. 公式
    print('\n5. 检查公式（A4单元格）:')
    cell1 = ws1.cell(row=4, column=7)  # 数量/KG列
    cell2 = ws2.cell(row=4, column=7)
    print(f'   原始: 值={cell1.value}, 公式={cell1.data_type}')
    print(f'   复制: 值={cell2.value}, 公式={cell2.data_type}')
    
    # 6. 条件格式
    print('\n6. 条件格式:')
    # openpyxl 3.1.2中条件格式的正确访问方式
    cf1 = getattr(ws1, '_cf_rules', [])
    cf2 = getattr(ws2, '_cf_rules', [])
    print(f'   原始: {len(cf1)}个条件格式')
    print(f'   复制: {len(cf2)}个条件格式')
    
    # 7. 自定义名称
    print('\n7. 工作簿自定义名称:')
    print(f'   原始: {[name for name in wb1.defined_names] if hasattr(wb1, "defined_names") else "无"}')
    print(f'   复制: {[name for name in wb2.defined_names] if hasattr(wb2, "defined_names") else "无"}')
    
    # 8. 页眉页脚
    print('\n8. 页眉页脚:')
    print(f'   原始页眉: {ws1.header_footer.oddHeader if ws1.header_footer else "无"}')
    print(f'   原始页脚: {ws1.header_footer.oddFooter if ws1.header_footer else "无"}')
    print(f'   复制页眉: {ws2.header_footer.oddHeader if ws2.header_footer else "无"}')
    print(f'   复制页脚: {ws2.header_footer.oddFooter if ws2.header_footer else "无"}')
    
    # 9. 工作表保护
    print('\n9. 工作表保护:')
    print(f'   原始: 保护状态={ws1.protection.sheet}, 密码={ws1.protection.password}')
    print(f'   复制: 保护状态={ws2.protection.sheet}, 密码={ws2.protection.password}')
    
    wb1.close()
    wb2.close()

if __name__ == "__main__":
    check_all_differences()