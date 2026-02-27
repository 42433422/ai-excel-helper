#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import load_workbook

def inspect_template(file_path):
    """检查模板文件的格式"""
    print(f"\n=== 检查模板文件: {file_path} ===")
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"工作表名称: {ws.title}")
        print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")
        
        print("\n前5行内容:")
        for row in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(15, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                value = cell.value
                if value is not None:
                    value_str = str(value)
                    if len(value_str) > 25:
                        value_str = value_str[:25] + "..."
                    row_data.append(f"{col}:{value_str}")
            if row_data:
                print(f"  行{row}: {' | '.join(row_data)}")
        
        # 检查合并单元格
        if ws.merged_cells.ranges:
            print(f"\n合并单元格 ({len(list(ws.merged_cells.ranges))}个):")
            for merged_range in list(ws.merged_cells.ranges)[:5]:
                print(f"  {merged_range}")
        
        return True
        
    except Exception as e:
        print(f"检查失败: {e}")
        return False

if __name__ == "__main__":
    # 检查几个发货单模板
    templates = [
        "发货单/尹玉华1.xlsx",
        "发货单/七彩乐园.xlsx",
        "发货单/志泓.xlsx"
    ]
    
    for template in templates:
        inspect_template(template)