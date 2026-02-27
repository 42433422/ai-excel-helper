import pandas as pd
import openpyxl

# 定义要分析的文件和工作表
files_to_analyze = [
    {'path': 'e:\\女娲1号\\发货单\\尹玉华1.xlsx', 'sheet': '出货'},
    {'path': 'e:\\女娲1号\\发货单\\温总.xlsx', 'sheet': None},  # 自动选择活动工作表
    {'path': 'e:\\女娲1号\\发货单\\七彩乐园.xlsx', 'sheet': None}  # 自动选择活动工作表
]

for file_info in files_to_analyze:
    file_path = file_info['path']
    sheet_name = file_info['sheet']
    
    print(f"\n{'='*70}")
    print(f"分析文件：{file_path}")
    
    # 使用openpyxl加载工作簿
    wb = openpyxl.load_workbook(file_path)
    
    # 查看所有工作表
    print(f"工作表列表：{wb.sheetnames}")
    
    # 选择要分析的工作表
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        print(f"自动选择工作表：{ws.title}")
    
    print(f"分析工作表：{ws.title}")
    print(f"行数：{ws.max_row}")
    print(f"列数：{ws.max_column}")
    
    # 查看前10行数据，包括标题
    print(f"前10行数据：")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        # 过滤掉全空的行
        if any(cell is not None for cell in row):
            print(f"行 {row_idx}: {row}")
    
    # 使用pandas尝试读取，找到正确的标题行
    print(f"\n使用pandas尝试读取，自动识别标题行：")
    for header_row in range(0, 8):
        try:
            df = pd.read_excel(file_path, sheet_name=ws.title, header=header_row)
            if len(df.columns) > 3 and len(df) > 5:
                print(f"标题行 {header_row + 1} 的列名：")
                print(df.columns.tolist())
                print(f"数据行数：{len(df)}")
                print(f"前3行数据：")
                print(df.head(3))
                print()
                # 只显示第一个合适的标题行
                break
        except Exception as e:
            pass
    
    wb.close()
    print(f"{'='*70}")