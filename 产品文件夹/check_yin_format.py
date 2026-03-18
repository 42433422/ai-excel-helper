#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

wb = load_workbook('发货单/尹玉华1.xlsx')
ws = wb['尹']

print('=== 尹工作表原始格式 ===')
print(f'合并单元格: {list(ws.merged_cells.ranges)}')
print(f'列数: {ws.max_column}, 行数: {ws.max_row}')

print('\n列宽:')
for col in range(1, ws.max_column + 1):
    col_letter = chr(64 + col) if col <= 26 else chr(64 + (col-1)//26) + chr(64 + (col-1)%26)
    width = ws.column_dimensions[col_letter].width if col_letter in ws.column_dimensions else 8
    print(f'  列{col}: {col_letter}, 宽={width}')

print('\n行高:')
for row in range(1, min(5, ws.max_row + 1)):
    height = ws.row_dimensions[row].height if row in ws.row_dimensions else None
    print(f'  行{row}: 高={height}')

print('\n单元格值:')
for row in range(1, min(5, ws.max_row + 1)):
    row_data = []
    for col in range(1, min(10, ws.max_column + 1)):
        cell = ws.cell(row=row, column=col)
        value = cell.value
        if value:
            row_data.append(f'{cell.coordinate}={value}')
    if row_data:
        print(f'  行{row}: {" | ".join(row_data)}')