import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# 定义要分析的文件路径
file_path = 'e:\\女娲1号\\发货单\\尹玉华1.xlsx'

print(f"分析文件：{file_path}")
print("=" * 60)

# 使用openpyxl加载工作簿
wb = openpyxl.load_workbook(file_path)

# 查看所有工作表
print(f"工作表列表：{wb.sheetnames}")

# 分析第一个工作表
ws = wb.active
print(f"分析工作表：{ws.title}")
print("=" * 60)

# 查看工作表的基本信息
print(f"行数：{ws.max_row}")
print(f"列数：{ws.max_column}")
print("=" * 60)

# 查看前几行数据，包括标题
print("前5行数据（包括标题）：")
for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
    # 过滤掉全空的行
    if any(cell is not None for cell in row):
        print(row)

print("=" * 60)

# 使用pandas读取文件，尝试自动识别标题行
print("使用pandas读取，自动识别标题：")
df = pd.read_excel(file_path)
print("列名：")
print(df.columns.tolist())
print("\n数据前5行：")
print(df.head())

print("=" * 60)

# 尝试不同的标题行
print("尝试不同的标题行：")
for header_row in range(0, 5):
    try:
        df_test = pd.read_excel(file_path, header=header_row)
        if len(df_test.columns) > 2:  # 至少有3列数据
            print(f"标题行 {header_row + 1} 的列名：")
            print(df_test.columns.tolist())
            print(f"数据行数：{len(df_test)}")
            print()
    except Exception as e:
        pass

# 关闭工作簿
wb.close()