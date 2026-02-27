#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook
from openpyxl import Workbook

def copy_with_formulas():
    """复制工作表，保留所有属性，包括公式"""
    print("=== 复制工作表（保留公式） ===\n")
    
    mapping = {
        '发货单/尹玉华1.xlsx': '尹',
        '发货单/七彩乐园.xlsx': '七彩乐园',
        '发货单/侯雪梅.xlsx': '侯雪梅',
        '发货单/刘英.xlsx': '刘英',
        '发货单/国圣化工.xlsx': '国圣',
        '发货单/宗南.xlsx': '宗南',
        '发货单/宜榢.xlsx': '宜榢',
        '发货单/小洋杨总、.xlsx': '24徐',
        '发货单/志泓.xlsx': '志泓',
        '发货单/新旺博旺.xlsx': '新旺博旺',
        '发货单/温总.xlsx': '温总',
        '发货单/澜宇电视柜.xlsx': '澜宇电视柜',
        '发货单/现金.xlsx': '现金',
        '发货单/迎扬李总(1).xlsx': '迎扬电视墙',
        '发货单/迎扬李总.xlsx': '迎扬电视墙',
        '发货单/邻居杨总.xlsx': '邻居',
        '发货单/邻居贾总.xlsx': '贾总',
    }
    
    count = 0
    for file_path, sheet_name in mapping.items():
        output_file = f"{sheet_name}.xlsx"
        print(f"复制: {os.path.basename(file_path)} -> {output_file}")
        
        try:
            # 方法1：使用完整复制（包括公式）
            wb = load_workbook(file_path, data_only=False)  # 关键：data_only=False 保留公式
            ws = wb[sheet_name]
            
            # 创建新工作簿
            new_wb = Workbook()
            
            # 移除默认工作表
            new_ws = new_wb.active
            new_wb.remove(new_ws)
            
            # 复制整个工作表
            new_ws = new_wb.copy_worksheet(ws)  # 关键：使用copy_worksheet方法
            new_ws.title = sheet_name
            
            # 保存工作簿
            new_wb.save(output_file)
            new_wb.close()
            wb.close()
            
            print(f"  ✓ 完成 (公式已保留)")
            count += 1
        except Exception as e:
            print(f"  ✗ 错误: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"\n完成！共 {count} 个文件")

if __name__ == "__main__":
    copy_with_formulas()