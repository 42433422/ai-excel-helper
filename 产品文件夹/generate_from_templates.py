#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import sqlite3
import os
from openpyxl import load_workbook
from datetime import datetime

def copy_template_format(source_file, target_file, customer_name, products_data):
    """复制模板格式并填充数据"""
    try:
        # 加载模板文件
        wb = load_workbook(source_file)
        
        # 获取第一个工作表
        ws = wb.active
        
        # 找到数据开始行
        data_start_row = 2  # 默认从第2行开始
        
        # 查找产品数据区域
        product_cols = {}
        for col in range(1, 15):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                cell_str = str(cell_value)
                if '产品型号' in cell_str:
                    product_cols['产品型号'] = col
                elif '产品名称' in cell_str:
                    product_cols['产品名称'] = col
                elif '规格' in cell_str and 'KG' in cell_str:
                    product_cols['规格_KG'] = col
                elif '数量' in cell_str and '件' in cell_str:
                    product_cols['数量_件'] = col
                elif '数量' in cell_str and 'KG' in cell_str:
                    product_cols['数量_KG'] = col
                elif '单价' in cell_str and '元' in cell_str:
                    product_cols['单价'] = col
                elif '金额' in cell_str:
                    product_cols['金额'] = col
        
        # 找到数据区域开始行
        for row in range(2, 10):
            has_data = False
            for col in product_cols.values():
                if ws.cell(row=row, column=col).value:
                    has_data = True
                    break
            if has_data:
                data_start_row = row
                break
        
        # 清除现有数据（保留格式）
        for row in range(data_start_row, 50):
            for col in product_cols.values():
                ws.cell(row=row, column=col).value = None
        
        # 填充新产品数据
        row = data_start_row
        for _, product in products_data.iterrows():
            for field, col in product_cols.items():
                value = product.get(field, '')
                if field in ['规格_KG', '数量_件', '数量_KG', '单价', '金额'] and value:
                    try:
                        ws.cell(row=row, column=col, value=float(value))
                    except:
                        ws.cell(row=row, column=col, value=value)
                else:
                    ws.cell(row=row, column=col, value=value)
            row += 1
        
        # 保存文件
        wb.save(target_file)
        return True
        
    except Exception as e:
        print(f"  ✗ 复制模板失败: {e}")
        return False

def generate_templates_from_existing():
    """使用现有发货单模板生成所有客户的模板"""
    print("=" * 80)
    print("=== 使用现有发货单模板生成所有客户的模板 ===")
    print("=" * 80)
    
    # 获取数据
    conn = sqlite3.connect('customer_products_final_corrected.db')
    customers = pd.read_sql_query("SELECT * FROM customers", conn)
    products = pd.read_sql_query("SELECT * FROM products", conn)
    conn.close()
    
    # 模板目录
    template_dir = "发货单模板"
    shipping_dir = "出货单模板"
    
    if not os.path.exists(template_dir):
        os.makedirs(template_dir)
    if not os.path.exists(shipping_dir):
        os.makedirs(shipping_dir)
    
    print(f"\n客户总数: {len(customers)}")
    print()
    
    # 获取一个发货单模板作为参考
    source_templates = [
        "发货单/尹玉华1.xlsx",
        "发货单/七彩乐园.xlsx",
        "发货单/志泓.xlsx",
        "发货单/现金.xlsx"
    ]
    
    template_used = None
    for source in source_templates:
        if os.path.exists(source):
            template_used = source
            print(f"使用模板: {source}")
            break
    
    if not template_used:
        print("未找到可用的发货单模板文件！")
        return
    
    # 为每个客户生成模板
    template_count = 0
    for _, customer in customers.iterrows():
        customer_name = customer['客户名称']
        print(f"\n处理客户: {customer_name}")
        
        # 获取该客户的产品
        customer_products = products[products['客户ID'] == customer['customer_id']]
        print(f"  产品数量: {len(customer_products)}")
        
        # 安全文件名
        safe_name = customer_name.replace('/', '_').replace('\\', '_')
        
        # 发货单模板
        delivery_file = f"{template_dir}/{safe_name}-发货单模板.xlsx"
        if os.path.exists(template_used):
            if copy_template_format(template_used, delivery_file, customer_name, customer_products):
                print(f"  ✓ 发货单模板: {delivery_file}")
                template_count += 1
        else:
            print(f"  ✗ 模板文件不存在: {template_used}")
    
    print("\n" + "=" * 80)
    print(f"✓ 成功生成 {template_count} 个发货单模板文件!")
    print("=" * 80)

if __name__ == "__main__":
    generate_templates_from_existing()