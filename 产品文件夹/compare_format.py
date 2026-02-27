#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

# 检查原始文件
print('=== 原始尹工作表 ===')
wb1 = load_workbook('发货单/尹玉华1.xlsx')
ws1 = wb1['尹']
print(f'合并单元格: {list(ws1.merged_cells.ranges)}')

print('\n=== 复制的尹.xlsx ===')
wb2 = load_workbook('尹.xlsx')
ws2 = wb2.active
print(f'合并单元格: {list(ws2.merged_cells.ranges)}')

print('\n列宽对比:')
print('原始:')
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    width = ws1.column_dimensions[col_letter].width if col_letter in ws1.column_dimensions else '未设置'
    print(f'  {col_letter}: {width}')

print('\n复制后:')
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    width = ws2.column_dimensions[col_letter].width if col_letter in ws2.column_dimensions else '未设置'
    print(f'  {col_letter}: {width}')