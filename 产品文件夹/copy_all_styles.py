#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook
from openpyxl import Workbook
from copy import copy

def copy_with_all_styles():
    """复制工作表，保留所有样式（字体、颜色、边框、对齐等）"""
    print("=== 复制工作表（保留所有样式） ===\n")
    
    mapping = {
        '发货单/尹玉华1.xlsx': '尹',
        '发货单/七彩乐园.xlsx': '七彩乐园',
        '发货单/侯雪梅.xlsx': '侯雪梅',
        '发货单/刘英.xlsx': '刘英',
        '发货单/国圣化工.xlsx': '国圣',
        '发货单/宗南.xlsx': '宗南',
        '发货单/宜榢.xlsx': '宜榢',
        '发货单/小洋杨总、.xlsx': '24徐',
        '发货单/志泓.xlsx': '志泓',
        '发货单/新旺博旺.xlsx': '新旺博旺',
        '发货单/温总.xlsx': '温总',
        '发货单/澜宇电视柜.xlsx': '澜宇电视柜',
        '发货单/现金.xlsx': '现金',
        '发货单/迎扬李总(1).xlsx': '迎扬电视墙',
        '发货单/迎扬李总.xlsx': '迎扬电视墙',
        '发货单/邻居杨总.xlsx': '邻居',
        '发货单/邻居贾总.xlsx': '贾总',
    }
    
    count = 0
    for file_path, sheet_name in mapping.items():
        output_file = f"{sheet_name}.xlsx"
        print(f"复制: {os.path.basename(file_path)} -> {output_file}")
        
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
            
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name
            
            # 复制列宽和行高
            for col_letter, dim in ws.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = dim.width
            
            for row_num, dim in ws.row_dimensions.items():
                new_ws.row_dimensions[row_num].height = dim.height
            
            # 复制单元格（值和样式）
            for row in ws.iter_rows():
                for cell in row:
                    new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    
                    # 复制字体
                    if cell.font:
                        from openpyxl.styles import Font
                        new_cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            vertAlign=cell.font.vertAlign,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color
                        )
                    
                    # 复制对齐
                    if cell.alignment:
                        from openpyxl.styles import Alignment
                        new_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    
                    # 复制边框
                    if cell.border:
                        from openpyxl.styles import Border, Side
                        new_cell.border = Border(
                            left=cell.border.left,
                            right=cell.border.right,
                            top=cell.border.top,
                            bottom=cell.border.bottom,
                            diagonal=cell.border.diagonal,
                            diagonalDown=cell.border.diagonalDown,
                            diagonalUp=cell.border.diagonalUp
                        )
                    
                    # 复制填充
                    if cell.fill:
                        from openpyxl.styles import PatternFill, GradientFill
                        if isinstance(cell.fill, PatternFill):
                            new_cell.fill = PatternFill(
                                fill_type=cell.fill.fill_type,
                                start_color=cell.fill.start_color,
                                end_color=cell.fill.end_color
                            )
                        elif isinstance(cell.fill, GradientFill):
                            new_cell.fill = GradientFill(
                                type=cell.fill.type,
                                degree=cell.fill.degree,
                                left=cell.fill.left,
                                right=cell.fill.right,
                                top=cell.fill.top,
                                bottom=cell.fill.bottom,
                                stop=cell.fill.stop
                            )
                    
                    # 复制数字格式
                    new_cell.number_format = cell.number_format
            
            # 复制合并单元格
            for merged_range in ws.merged_cells.ranges:
                new_ws.merge_cells(str(merged_range))
            
            new_wb.save(output_file)
            new_wb.close()
            wb.close()
            
            print(f"  ✓ 完成 (所有样式已保留)")
            count += 1
        except Exception as e:
            print(f"  ✗ 错误: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"\n完成！共 {count} 个文件")

if __name__ == "__main__":
    copy_with_all_styles()