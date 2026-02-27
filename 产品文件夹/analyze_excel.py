#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import openpyxl
from openpyxl import load_workbook

def analyze_excel_layout(file_path):
    print('=== Excel合并单元格布局分析 ===')
    
    # 使用openpyxl分析合并单元格
    wb = load_workbook(file_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        if sheet_name in ['出货', '尹']:
            print(f'\n【工作表: {sheet_name}】')
            ws = wb[sheet_name]
            
            # 获取合并单元格信息
            merged_ranges = list(ws.merged_cells.ranges)
            print(f'合并单元格数量: {len(merged_ranges)}')
            
            if merged_ranges:
                print('合并单元格详细信息:')
                for i, merged_range in enumerate(merged_ranges):
                    print(f'{i+1}. {merged_range}')
                    
                    # 获取合并区域的内容
                    top_left = ws[merged_range.coord.split(':')[0]]
                    print(f'   内容: {top_left.value}')
                    
            # 获取最大行列数
            max_row = ws.max_row
            max_col = ws.max_column
            print(f'数据范围: {max_row}行 × {max_col}列')
            
            # 分析数据结构
            print('\n数据布局分析:')
            
            # 读取前几行来分析布局
            for row in range(1, min(11, max_row + 1)):
                row_data = []
                for col in range(1, min(11, max_col + 1)):
                    cell = ws.cell(row=row, column=col)
                    value = cell.value
                    if value is not None:
                        row_data.append(f'[{col}:{value}]')
                if row_data:
                    print(f'第{row}行: {" ".join(row_data)}')
            
            print('-' * 50)

if __name__ == "__main__":
    # 分析Excel文件
    analyze_excel_layout('e:\\女娲1号\\发货单\\尹玉华1.xlsx')