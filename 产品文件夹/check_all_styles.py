#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

def check_all_styles():
    """检查所有样式"""
    wb1 = load_workbook('发货单/尹玉华1.xlsx')
    ws1 = wb1['尹']
    
    wb2 = load_workbook('尹.xlsx')
    ws2 = wb2.active
    
    print('=== 完整样式对比 ===')
    
    # 检查几个关键单元格
    for row in [1, 2, 3, 4]:
        cell1 = ws1.cell(row=row, column=1)
        cell2 = ws2.cell(row=row, column=1)
        
        print(f'\n行{row} A列:')
        print(f'  原始: 字体={cell1.font.name}/{cell1.font.size}, 颜色RGB={cell1.font.color}, 行高={ws1.row_dimensions[row].height if row in ws1.row_dimensions else "默认"}')
        print(f'  复制: 字体={cell2.font.name}/{cell2.font.size}, 颜色RGB={cell2.font.color}, 行高={ws2.row_dimensions[row].height if row in ws2.row_dimensions else "默认"}')
        
        # 检查对齐
        print(f'  原始对齐: 水平={cell1.alignment.horizontal}, 垂直={cell1.alignment.vertical}')
        print(f'  复制对齐: 水平={cell2.alignment.horizontal}, 垂直={cell2.alignment.vertical}')
        
        # 检查边框
        print(f'  原始边框: left={cell1.border.left.style if cell1.border.left else None}')
        print(f'  复制边框: left={cell2.border.left.style if cell2.border.left else None}')
    
    wb1.close()
    wb2.close()

check_all_styles()