#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

wb1 = load_workbook('发货单/尹玉华1.xlsx')
ws1 = wb1['尹']

wb2 = load_workbook('尹.xlsx')
ws2 = wb2.active

print('=== 行高（单元格大小）对比 ===')
print('原始文件行高:')
for row in range(1, min(6, ws1.max_row + 1)):
    height = ws1.row_dimensions[row].height if row in ws1.row_dimensions else '默认'
    print(f'  行{row}: {height}')

print('\n复制后行高:')
for row in range(1, min(6, ws2.max_row + 1)):
    height = ws2.row_dimensions[row].height if row in ws2.row_dimensions else '默认'
    print(f'  行{row}: {height}')

wb1.close()
wb2.close()