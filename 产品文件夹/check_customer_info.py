import openpyxl

# 打开Excel文件
file_path = 'e:\\女娲1号\\发货单\\七彩乐园.xlsx'
wb = openpyxl.load_workbook(file_path)

print(f"检查文件: {file_path}")
print(f"工作表列表: {wb.sheetnames}")

# 遍历所有工作表，查找购买单位信息
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n检查工作表: {sheet_name}")
    
    # 遍历前20行，查找包含购买单位的行
    for row_idx in range(1, 21):
        row = ws[row_idx]
        row_values = []
        for cell in row:
            row_values.append(cell.value)
        
        # 过滤掉全空的行
        if any(cell is not None for cell in row_values):
            # 检查是否包含购买单位相关信息
            row_str = str(row_values)
            if '购货单位' in row_str or '购买单位' in row_str or '客户' in row_str:
                print(f"行 {row_idx}: {row_values[:10]}")

# 关闭工作簿
wb.close()