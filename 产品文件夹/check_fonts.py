#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

# 比较字体样式
wb1 = load_workbook('发货单/尹玉华1.xlsx')
ws1 = wb1['尹']

wb2 = load_workbook('尹.xlsx')
ws2 = wb2.active

print('=== 字体样式对比 ===')
print('原始文件:')
cell = ws1.cell(row=1, column=1)
print(f'  A1 字体: {cell.font.name}, 大小: {cell.font.size}, 加粗: {cell.font.bold}')

print('\n复制后:')
cell2 = ws2.cell(row=1, column=1)
print(f'  A1 字体: {cell2.font.name}, 大小: {cell2.font.size}, 加粗: {cell2.font.bold}')

# 检查几个关键单元格
print('\n关键单元格对比:')
for row in [1, 2, 3]:
    cell1 = ws1.cell(row=row, column=1)
    cell2 = ws2.cell(row=row, column=1)
    print(f'  行{row}: 原始={cell1.font.name}/{cell1.font.size} -> 复制={cell2.font.name}/{cell2.font.size}')

wb1.close()
wb2.close()