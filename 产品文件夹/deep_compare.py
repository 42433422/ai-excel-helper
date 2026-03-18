#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

def deep_copy_workbook():
    """深度复制工作簿（包含所有属性）"""
    print("=== 深度复制工作簿对比 ===\n")
    
    # 加载原始文件
    wb1 = load_workbook('发货单/尹玉华1.xlsx')
    ws1 = wb1['尹']
    
    # 加载复制后的文件
    wb2 = load_workbook('尹.xlsx')
    ws2 = wb2.active
    
    print('1. 页面设置对比:')
    print(f'   原始: 页面方向={ws1.page_setup.paperSize}, 缩放={ws1.page_setup.scale}')
    print(f'   复制: 页面方向={ws2.page_setup.paperSize}, 缩放={ws2.page_setup.scale}')
    
    print('\n2. 打印标题:')
    print(f'   原始: 顶端标题行={ws1.print_titles}, 打印区域={ws1.print_area}')
    print(f'   复制: 顶端标题行={ws2.print_titles}, 打印区域={ws2.print_area}')
    
    print('\n3. 视图设置:')
    print(f'   原始: 显示网格线={ws1.sheet_view.showGridLines}, 显示行列表头={ws1.sheet_view.showRowColHeaders}')
    print(f'   复制: 显示网格线={ws2.sheet_view.showGridLines}, 显示行列表头={ws2.sheet_view.showRowColHeaders}')
    
    print('\n4. 边距设置:')
    print(f'   原始: 上={ws1.page_margins.top}, 下={ws1.page_margins.bottom}, 左={ws1.page_margins.left}, 右={ws1.page_margins.right}')
    print(f'   复制: 上={ws2.page_margins.top}, 下={ws2.page_margins.bottom}, 左={ws2.page_margins.left}, 右={ws2.page_margins.right}')
    
    print('\n5. 所有单元格值对比（前10行）:')
    all_match = True
    for row in range(1, min(11, ws1.max_row + 1)):
        for col in range(1, min(6, ws1.max_column + 1)):
            v1 = ws1.cell(row=row, column=col).value
            v2 = ws2.cell(row=row, column=col).value
            if v1 != v2:
                print(f'   ✗ 行{row}列{col}: 原始={v1} 复制={v2}')
                all_match = False
    
    if all_match:
        print('   ✓ 所有单元格值完全一致')
    
    wb1.close()
    wb2.close()

deep_copy_workbook()