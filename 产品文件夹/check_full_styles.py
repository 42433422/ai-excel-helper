#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

def check_full_styles():
    """检查完整复制的文件是否包含所有样式"""
    wb1 = load_workbook('发货单/尹玉华1.xlsx', data_only=False)
    ws1 = wb1['尹']
    
    wb2 = load_workbook('尹_full.xlsx', data_only=False)
    ws2 = wb2.active
    
    print('=== 完整样式检查 ===')
    
    # 检查几个关键单元格的样式
    for row in [1, 2, 3, 4]:
        for col in [1, 4, 5, 7, 9]:
            cell1 = ws1.cell(row=row, column=col)
            cell2 = ws2.cell(row=row, column=col)
            
            print(f'\n单元格{row},{col}:')
            print(f'  原始: 字体={cell1.font.name}/{cell1.font.size}, 对齐={cell1.alignment.horizontal}/{cell1.alignment.vertical}')
            print(f'  复制: 字体={cell2.font.name}/{cell2.font.size}, 对齐={cell2.alignment.horizontal}/{cell2.alignment.vertical}')
            
            # 检查边框
            print(f'  原始边框: 左={cell1.border.left.style}, 右={cell1.border.right.style}')
            print(f'  复制边框: 左={cell2.border.left.style}, 右={cell2.border.right.style}')
    
    wb1.close()
    wb2.close()

check_full_styles()