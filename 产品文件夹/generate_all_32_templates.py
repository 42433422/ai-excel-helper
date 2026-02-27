#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import sqlite3
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime

def create_shipping_record(source_file, target_file, customer_name, products_data):
    """创建出货记录模板（使用发货单格式）"""
    try:
        if not os.path.exists(source_file):
            print(f"  ✗ 模板不存在: {source_file}")
            return False
        
        # 加载模板
        wb = load_workbook(source_file)
        ws = wb.active
        
        # 获取合并单元格列表
        merged_ranges = list(ws.merged_cells.ranges)
        
        # 取消所有合并单元格
        for merged_range in merged_ranges:
            ws.unmerge_cells(str(merged_range))
        
        # 查找数据开始行
        data_start_row = 2
        for row in range(2, 10):
            cell_value = ws.cell(row=row, column=3).value
            if cell_value and isinstance(cell_value, str) and len(cell_value) < 20:
                data_start_row = row
                break
        
        # 清除现有数据行
        for row in range(data_start_row, 100):
            for col in range(1, 15):
                ws.cell(row=row, column=col).value = None
        
        # 填充新产品数据
        row = data_start_row
        for _, product in products_data.iterrows():
            ws.cell(row=row, column=1, value=product.get('日期', datetime.now().strftime('%Y-%m-%d')))
            ws.cell(row=row, column=2, value=product.get('单号', ''))
            ws.cell(row=row, column=3, value=product.get('产品型号', ''))
            ws.cell(row=row, column=6, value=product.get('产品名称', ''))
            
            qty = product.get('数量_件', 0)
            ws.cell(row=row, column=7, value=int(qty) if qty else 0)
            
            spec = product.get('规格_KG', 0)
            ws.cell(row=row, column=8, value=float(spec) if spec else 0)
            
            weight = product.get('数量_KG', 0)
            ws.cell(row=row, column=9, value=float(weight) if weight else 0)
            
            price = product.get('单价', 0)
            ws.cell(row=row, column=10, value=float(price) if price else 0)
            
            amount = product.get('金额', 0)
            ws.cell(row=row, column=11, value=float(amount) if amount else 0)
            
            ws.cell(row=row, column=12, value=product.get('来源', ''))
            
            row += 1
        
        # 恢复合并单元格
        for merged_range in merged_ranges:
            ws.merge_cells(str(merged_range))
        
        # 保存文件
        wb.save(target_file)
        return True
        
    except Exception as e:
        print(f"  ✗ 生成失败: {e}")
        return False

def generate_all_templates():
    """生成所有模板（发货单 + 出货记录）"""
    print("=" * 80)
    print("=== 生成所有模板文件 ===")
    print("=" * 80)
    
    # 获取数据
    conn = sqlite3.connect('customer_products_final_corrected.db')
    customers = pd.read_sql_query("SELECT * FROM customers", conn)
    products = pd.read_sql_query("SELECT * FROM products", conn)
    conn.close()
    
    # 模板目录
    delivery_dir = "发货单模板"
    shipping_dir = "出货单模板"
    
    # 确保目录存在
    os.makedirs(delivery_dir, exist_ok=True)
    os.makedirs(shipping_dir, exist_ok=True)
    
    # 发货单模板文件映射
    template_mapping = {
        '尹玉华1': '发货单/尹玉华1.xlsx',
        '七彩乐园家私': '发货单/七彩乐园.xlsx',
        '志泓家私': '发货单/志泓.xlsx',
        '陈洪强': '发货单/现金.xlsx',
        '中江博郡家私': '发货单/迎扬李总(1).xlsx',
        '温总': '发货单/温总.xlsx',
        '宗南家私': '发货单/宗南.xlsx',
        '宜榢家私': '发货单/宜榢.xlsx',
        '新旺博旺': '发货单/新旺博旺.xlsx',
        '澜宇家私': '发货单/澜宇电视柜.xlsx',
        '侯雪梅': '发货单/侯雪梅.xlsx',
        '国圣化工': '发货单/国圣化工.xlsx',
        '杨总': '发货单/邻居杨总.xlsx',
        '邻居贾总': '发货单/邻居贾总.xlsx',
        '刘英': '发货单/刘英.xlsx',
        '小火洋': '发货单/小洋杨总、.xlsx',
    }
    
    print(f"\n客户总数: {len(customers)}")
    print()
    
    delivery_count = 0
    shipping_count = 0
    
    for _, customer in customers.iterrows():
        customer_name = customer['客户名称']
        print(f"\n处理客户: {customer_name}")
        
        customer_products = products[products['客户ID'] == customer['customer_id']].copy()
        customer_products = customer_products.sort_values('记录顺序', ascending=True)
        print(f"  产品数量: {len(customer_products)}")
        
        safe_name = customer_name.replace('/', '_').replace('\\', '_')
        source_template = template_mapping.get(customer_name, '发货单/尹玉华1.xlsx')
        
        # 发货单模板
        delivery_file = f"{delivery_dir}/{safe_name}-发货单模板.xlsx"
        if create_shipping_record(source_template, delivery_file, customer_name, customer_products):
            print(f"  ✓ 发货单模板: {delivery_file}")
            delivery_count += 1
        
        # 出货记录模板
        shipping_file = f"{shipping_dir}/{safe_name}-出货记录模板.xlsx"
        if create_shipping_record(source_template, shipping_file, customer_name, customer_products):
            print(f"  ✓ 出货记录模板: {shipping_file}")
            shipping_count += 1
    
    print("\n" + "=" * 80)
    print(f"✓ 成功生成 {delivery_count} 个发货单模板!")
    print(f"✓ 成功生成 {shipping_count} 个出货记录模板!")
    print(f"✓ 总计: {delivery_count + shipping_count} 个模板文件!")
    print("=" * 80)

if __name__ == "__main__":
    generate_all_templates()