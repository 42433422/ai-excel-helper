#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

def check_formulas():
    """检查复制的文件是否包含公式"""
    # 检查尹.xlsx中的公式
    wb = load_workbook('尹.xlsx', data_only=False)
    ws = wb.active
    
    print('=== 检查公式 ===')
    print('尹.xlsx 文件中的公式:')
    
    # 检查几个关键单元格
    cells_to_check = [
        (4, 7),   # G4 数量/KG
        (4, 9),   # I4 金额/元
        (5, 7),   # G5 数量/KG
        (5, 9),   # I5 金额/元
    ]
    
    for row, col in cells_to_check:
        cell = ws.cell(row=row, column=col)
        print(f'  单元格{row},{col}值: {cell.value} (类型: {type(cell.value)})')
    
    wb.close()
    
    # 与原始文件对比
    print('\n=== 与原始文件对比 ===')
    wb_original = load_workbook('发货单/尹玉华1.xlsx', data_only=False)
    ws_original = wb_original['尹']
    
    for row, col in cells_to_check:
        cell_original = ws_original.cell(row=row, column=col)
        cell_copy = ws.cell(row=row, column=col)
        
        match = '✓' if cell_original.value == cell_copy.value else '✗'
        print(f'  单元格{row},{col}: 原始="{cell_original.value}" 复制="{cell_copy.value}" {match}')
    
    wb_original.close()

check_formulas()