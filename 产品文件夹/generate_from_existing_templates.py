#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import sqlite3
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime

def use_existing_template(source_file, target_file, customer_name, products_data):
    """使用现有模板格式生成客户模板"""
    try:
        # 尝试加载现有模板
        if os.path.exists(source_file):
            wb = load_workbook(source_file)
            ws = wb.active
            
            # 查找工作表名称
            sheet_name = ws.title
            print(f"  使用模板工作表: {sheet_name}")
            
            # 查找数据开始行
            data_start_row = 2
            for row in range(2, 10):
                cell_value = ws.cell(row=row, column=3).value  # 产品型号列
                if cell_value and isinstance(cell_value, str) and len(cell_value) < 20:
                    data_start_row = row
                    break
            
            print(f"  数据开始行: {data_start_row}")
            
            # 清除现有数据行
            for row in range(data_start_row, 100):
                for col in [1, 2, 3, 6, 7, 8, 9, 10, 11, 12, 13]:
                    ws.cell(row=row, column=col).value = None
            
            # 填充新产品数据
            row = data_start_row
            for _, product in products_data.iterrows():
                # 日期
                ws.cell(row=row, column=1, value=product.get('日期', datetime.now().strftime('%Y-%m-%d')))
                # 单号
                ws.cell(row=row, column=2, value=product.get('单号', ''))
                # 产品型号
                ws.cell(row=row, column=3, value=product.get('产品型号', ''))
                # 产品名称
                ws.cell(row=row, column=6, value=product.get('产品名称', ''))
                # 数量/件
                qty = product.get('数量_件', 0)
                ws.cell(row=row, column=7, value=int(qty) if qty else 0)
                # 规格/KG
                spec = product.get('规格_KG', 0)
                ws.cell(row=row, column=8, value=float(spec) if spec else 0)
                # 数量/KG
                weight = product.get('数量_KG', 0)
                ws.cell(row=row, column=9, value=float(weight) if weight else 0)
                # 单价/元
                price = product.get('单价', 0)
                ws.cell(row=row, column=10, value=float(price) if price else 0)
                # 金额/元
                amount = product.get('金额', 0)
                ws.cell(row=row, column=11, value=float(amount) if amount else 0)
                # 备注
                ws.cell(row=row, column=12, value=product.get('来源', ''))
                
                row += 1
            
            wb.save(target_file)
            return True
        else:
            print(f"  ✗ 模板文件不存在: {source_file}")
            return False
            
    except Exception as e:
        print(f"  ✗ 生成失败: {e}")
        return False

def generate_from_existing_templates():
    """使用现有发货单模板生成所有客户的模板"""
    print("=" * 80)
    print("=== 使用现有发货单模板生成所有客户的模板 ===")
    print("=" * 80)
    
    # 获取数据
    conn = sqlite3.connect('customer_products_final_corrected.db')
    customers = pd.read_sql_query("SELECT * FROM customers", conn)
    products = pd.read_sql_query("SELECT * FROM products", conn)
    conn.close()
    
    # 模板目录
    template_dir = "发货单模板"
    shipping_dir = "出货单模板"
    
    # 清空现有模板目录
    if os.path.exists(template_dir):
        for f in os.listdir(template_dir):
            os.remove(os.path.join(template_dir, f))
    else:
        os.makedirs(template_dir)
        
    if os.path.exists(shipping_dir):
        for f in os.listdir(shipping_dir):
            os.remove(os.path.join(shipping_dir, f))
    else:
        os.makedirs(shipping_dir)
    
    # 发货单模板文件映射
    template_mapping = {
        '尹玉华1': '发货单/尹玉华1.xlsx',
        '七彩乐园家私': '发货单/七彩乐园.xlsx',
        '志泓家私': '发货单/志泓.xlsx',
        '陈洪强': '发货单/现金.xlsx',
        '中江博郡家私': '发货单/迎扬李总(1).xlsx',
        '温总': '发货单/温总.xlsx',
        '宗南家私': '发货单/宗南.xlsx',
        '宜榢家私': '发货单/宜榢.xlsx',
        '新旺博旺': '发货单/新旺博旺.xlsx',
        '澜宇家私': '发货单/澜宇电视柜.xlsx',
        '侯雪梅': '发货单/侯雪梅.xlsx',
        '国圣化工': '发货单/国圣化工.xlsx',
        '杨总': '发货单/邻居杨总.xlsx',
        '邻居贾总': '发货单/邻居贾总.xlsx',
        '刘英': '发货单/刘英.xlsx',
        '小火洋': '发货单/小洋杨总、.xlsx',
    }
    
    print(f"\n客户总数: {len(customers)}")
    print()
    
    template_count = 0
    for _, customer in customers.iterrows():
        customer_name = customer['客户名称']
        print(f"\n处理客户: {customer_name}")
        
        # 获取该客户的产品
        customer_products = products[products['客户ID'] == customer['customer_id']].copy()
        customer_products = customer_products.sort_values('记录顺序', ascending=True)  # 按最新记录排序
        print(f"  产品数量: {len(customer_products)}")
        
        # 安全文件名
        safe_name = customer_name.replace('/', '_').replace('\\', '_')
        
        # 获取对应的发货单模板
        source_template = template_mapping.get(customer_name, '发货单/尹玉华1.xlsx')
        
        # 发货单模板
        delivery_file = f"{template_dir}/{safe_name}-发货单模板.xlsx"
        if os.path.exists(source_template):
            if use_existing_template(source_template, delivery_file, customer_name, customer_products):
                print(f"  ✓ 发货单模板: {delivery_file}")
                template_count += 1
        else:
            print(f"  ✗ 模板不存在: {source_template}")
    
    print("\n" + "=" * 80)
    print(f"✓ 成功生成 {template_count} 个发货单模板文件!")
    print("=" * 80)

if __name__ == "__main__":
    generate_from_existing_templates()