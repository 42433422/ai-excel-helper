#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook
from openpyxl import Workbook

def copy_with_full_styles():
    """完整复制工作表，包括所有样式"""
    print("=== 完整复制（包括字体和对齐） ===\n")
    
    mapping = {
        '发货单/尹玉华1.xlsx': '尹',
    }
    
    count = 0
    for file_path, sheet_name in mapping.items():
        output_file = f"{sheet_name}_full.xlsx"
        print(f"复制: {os.path.basename(file_path)} -> {output_file}")
        
        try:
            # 加载原始文件，保留公式
            wb = load_workbook(file_path, data_only=False)
            ws = wb[sheet_name]
            
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name
            
            # 复制列宽
            for col_letter, dim in ws.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = dim.width
            
            # 复制行高
            for row_num, dim in ws.row_dimensions.items():
                new_ws.row_dimensions[row_num].height = dim.height
            
            # 复制所有单元格（包括公式和样式）
            for row in ws.iter_rows():
                for cell in row:
                    new_cell = new_ws.cell(row=cell.row, column=cell.column)
                    
                    # 复制值（包括公式）
                    new_cell.value = cell.value
                    
                    # 复制字体
                    if cell.font:
                        from openpyxl.styles import Font
                        new_cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            color=cell.font.color
                        )
                    
                    # 复制对齐方式
                    if cell.alignment:
                        from openpyxl.styles import Alignment
                        new_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit
                        )
                    
                    # 复制边框
                    if cell.border:
                        from openpyxl.styles import Border
                        new_cell.border = Border(
                            left=cell.border.left,
                            right=cell.border.right,
                            top=cell.border.top,
                            bottom=cell.border.bottom
                        )
                    
                    # 复制填充
                    if cell.fill:
                        from openpyxl.styles import PatternFill
                        new_cell.fill = PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color
                        )
            
            # 复制合并单元格
            for merged_range in ws.merged_cells.ranges:
                new_ws.merge_cells(str(merged_range))
            
            # 复制页面设置
            if ws.page_setup:
                new_ws.page_setup.paperSize = ws.page_setup.paperSize
                new_ws.page_setup.scale = ws.page_setup.scale
                new_ws.page_setup.orientation = ws.page_setup.orientation
            
            # 复制打印区域
            if ws.print_area:
                new_ws.print_area = ws.print_area
            
            new_wb.save(output_file)
            new_wb.close()
            wb.close()
            
            print(f"  ✓ 完成（包括字体和对齐）")
            count += 1
        except Exception as e:
            print(f"  ✗ 错误: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"\n完成！共 {count} 个文件")

if __name__ == "__main__":
    copy_with_full_styles()