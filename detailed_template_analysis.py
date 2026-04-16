import openpyxl
from openpyxl import load_workbook
import os

folder = r'e:\FHD\424'
files = [f for f in os.listdir(folder) if f.endswith('.xlsx') and '考勤统计表' in f and not f.startswith('~$')]

if files:
    template_file = os.path.join(folder, files[0])
    wb = load_workbook(template_file)
    ws = wb["明细"]
    
    print('=== 模板结构详细分析 ===')
    print(f'最大行：{ws.max_row}, 最大列：{ws.max_column}')
    
    # 分析表头
    print('\n=== 表头分析 (前 3 行) ===')
    for row_idx in range(1, 4):
        print(f'\nRow {row_idx}:')
        for col_idx in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_type = '公式' if isinstance(cell.value, str) and cell.value.startswith('=') else '文本/数字'
                print(f'  {cell.coordinate}: {cell.value} [{cell_type}]')
    
    # 分析数据区的前几行
    print('\n=== 数据区结构 (第 4-15 行) ===')
    for row_idx in range(4, min(16, ws.max_row + 1)):
        print(f'\nRow {row_idx}:')
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_type = '公式' if isinstance(cell.value, str) and cell.value.startswith('=') else '文本/数字'
                print(f'  {cell.coordinate}: {cell.value} [{cell_type}]')
    
    # 找出所有需要手动填充的列（非公式列）
    print('\n=== 需要填充的数据列 (非公式列) ===')
    data_rows_with_names = []
    for row_idx in range(4, min(50, ws.max_row + 1)):
        name_cell = ws.cell(row=row_idx, column=3)  # C 列是姓名
        if name_cell.value and not (isinstance(name_cell.value, str) and name_cell.value.startswith('=')):
            data_rows_with_names.append(row_idx)
    
    print(f'包含姓名的行：{data_rows_with_names[:10]}...')
    
    # 查看第 4 行的所有列
    print(f'\n=== 第 4 行 (第一个数据行) 的所有列 ===')
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=4, column=col_idx)
        if cell.value:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            is_formula = isinstance(cell.value, str) and cell.value.startswith('=')
            value_str = str(cell.value)[:50] if len(str(cell.value)) > 50 else cell.value
            cell_type = "公式" if is_formula else "数据"
            print(f'  {col_letter}4: {value_str} [{cell_type}]')
