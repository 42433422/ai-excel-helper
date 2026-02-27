import sqlite3
import openpyxl

print("=== 检查出货记录Excel中的产品数据 ===\n")

# 检查七彩乐园的出货记录
excel_path = r"C:\Users\Administrator\Desktop\新建文件夹 (4)\出货记录\七彩乐园\七彩乐园.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
    print(f"工作表列表: {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"=== {sheet_name} ===")
        print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
        
        # 读取前20行看看结构
        for i, row in enumerate(ws.iter_rows(max_row=min(20, ws.max_row), values_only=True), 1):
            if any(row):
                print(f"  Row {i}: {list(row)[:8]}")
        print()
except Exception as e:
    print(f"错误: {e}")
