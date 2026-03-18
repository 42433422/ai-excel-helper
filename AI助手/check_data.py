import openpyxl

# 检查25出货工作表
wb = openpyxl.load_workbook(r"c:\Users\97088\Desktop\新建文件夹 (4)\AI助手\templates\七彩乐园.xlsx")
ws = wb["25出货"]

print("=== 检查25出货工作表 ===\n")

# 从后往前找有数据的行
last_data_row = 0
for row in range(ws.max_row, 0, -1):
    has_data = False
    for col in range(1, 12):  # 检查前11列
        val = ws.cell(row=row, column=col).value
        if val is not None and str(val).strip():
            has_data = True
            break
    if has_data:
        last_data_row = row
        break

print(f"最后有数据的行: {last_data_row}\n")

# 显示最后有数据的几行
print("最后5行数据:")
for row in range(max(1, last_data_row - 4), last_data_row + 1):
    row_data = []
    for col in range(1, 12):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        if val is not None:
            if isinstance(val, str) and val.startswith('='):
                row_data.append(f"{chr(64+col)}{row}:公式")
            else:
                row_data.append(f"{chr(64+col)}{row}:{val}")
    if row_data:
        print(f"行 {row}: {' | '.join(row_data)}")

wb.close()
