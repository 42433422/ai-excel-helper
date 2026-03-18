import openpyxl
wb = openpyxl.load_workbook(r'templates\七彩乐园.xlsx')
ws = wb['25出货']

print('=== 新写入的数据 (行420) ===')
print('A日期:', ws.cell(row=420, column=1).value)
print('B单号:', ws.cell(row=420, column=2).value)
print('C型号:', ws.cell(row=420, column=3).value)
print('F产品:', ws.cell(row=420, column=6).value)
print('G数量:', ws.cell(row=420, column=7).value)
print('H规格:', ws.cell(row=420, column=8).value)
print('I公式:', ws.cell(row=420, column=9).value)
print('J单价:', ws.cell(row=420, column=10).value)
print('K公式:', ws.cell(row=420, column=11).value)

wb.close()
