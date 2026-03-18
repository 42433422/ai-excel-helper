#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查Excel同步数据的脚本
"""

import openpyxl

def check_excel_sync():
    """检查Excel同步数据"""
    excel_path = '尹玉华132.xlsx'
    
    # 打开Excel文件
    wb = openpyxl.load_workbook(excel_path)
    
    print('Excel文件中的所有工作表:')
    print('=' * 40)
    for sheet_name in wb.sheetnames:
        print(f'  - {sheet_name}')
    
    # 检查25年出货工作表
    if '25年出货' in wb.sheetnames:
        ws = wb['25年出货']
        print('\n25年出货工作表数据检查:')
        print('=' * 80)
        
        max_row = ws.max_row
        print(f'总行数: {max_row}')
        
        # 检查最后几行的数据
        print('\n最后5行的数据:')
        for row in range(max_row-4, max_row+1):
            row_data = []
            for col in range(1, 15):  # 检查前15列
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    row_data.append(f'列{col}: {cell_value}')
            
            if row_data:
                print(f'第{row}行: ' + ' | '.join(row_data))
        
        # 查找包含我们同步数据的行
        print('\n查找包含"蕊芯"的行:')
        print('=' * 60)
        found_ruixin = False
        for row in range(1, min(max_row+1, 20)):  # 检查前20行
            for col in range(1, 15):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and '蕊芯' in str(cell_value):
                    print(f'第{row}行列{col}: {cell_value}')
                    found_ruixin = True
        
        if not found_ruixin:
            print('未找到包含"蕊芯"的行')
        
        # 查找包含"测试客户"的行
        print('\n查找包含"测试客户"的行:')
        print('=' * 60)
        found_test = False
        for row in range(1, min(max_row+1, 20)):  # 检查前20行
            for col in range(1, 15):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and '测试客户' in str(cell_value):
                    print(f'第{row}行列{col}: {cell_value}')
                    found_test = True
        
        if not found_test:
            print('未找到包含"测试客户"的行')
        
        # 查找包含"26-010001"的行
        print('\n查找包含"26-010001"的行:')
        print('=' * 60)
        found_order = False
        for row in range(1, min(max_row+1, 20)):  # 检查前20行
            for col in range(1, 15):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and '26-010001' in str(cell_value):
                    print(f'第{row}行列{col}: {cell_value}')
                    found_order = True
        
        if not found_order:
            print('未找到包含订单编号的行')
    
    else:
        print('\n未找到"25年出货"工作表!')
    
    wb.close()

if __name__ == '__main__':
    check_excel_sync()
