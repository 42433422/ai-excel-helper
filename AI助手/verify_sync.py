import openpyxl
from openpyxl.utils import get_column_letter

# 检查备份文件
wb = openpyxl.load_workbook(r"c:\Users\97088\Desktop\新建文件夹 (4)\AI助手\templates\七彩乐园_备份.xlsx")
ws = wb["25出货"]

print(f"备份文件行数: {ws.max_row}\n")

# 显示最后5行
print("最后5行数据:")
for row in range(max(1, ws.max_row - 4), ws.max_row + 1):
    row_data = []
    for col in range(1, 12):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            row_data.append(f"{get_column_letter(col)}{row}:{cell.value}")
    if row_data:
        print(f"行 {row}: {' | '.join(row_data)}")

wb.close()

print("\n" + "="*50 + "\n")

# 检查原始文件
wb2 = openpyxl.load_workbook(r"c:\Users\97088\Desktop\新建文件夹 (4)\AI助手\templates\七彩乐园.xlsx")
ws2 = wb2["25出货"]

print(f"原始文件行数: {ws2.max_row}\n")

# 显示最后5行
print("最后5行数据:")
for row in range(max(1, ws2.max_row - 4), ws2.max_row + 1):
    row_data = []
    for col in range(1, 12):
        cell = ws2.cell(row=row, column=col)
        if cell.value is not None:
            row_data.append(f"{get_column_letter(col)}{row}:{cell.value}")
    if row_data:
        print(f"行 {row}: {' | '.join(row_data)}")

wb2.close()
