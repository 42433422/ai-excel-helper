#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
解析尹玉华Excel文件中的蕊芯产品数据
"""

import openpyxl
import sqlite3

def parse_ruixin_excel():
    file_path = 'templates/尹玉华1 - 副本.xlsx'
    
    print('=== 解析Excel文件中的蕊芯产品数据 ===')
    print(f'文件路径: {file_path}')
    
    try:
        # 使用openpyxl读取Excel文件
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # 解析Sheet2 - 蕊芯家私产品价格表
        print('\n--- 解析Sheet2: 蕊芯家私产品价格表 ---')
        worksheet = workbook['Sheet2']
        
        # 获取所有产品数据
        products_data = []
        
        # 找到数据开始的行（跳过标题行）
        start_row = 2  # 从第3行开始（索引为2）
        
        # 遍历每一行
        for row_num in range(start_row, worksheet.max_row + 1):
            # 获取产品编号
            product_code = worksheet.cell(row=row_num, column=1).value
            if not product_code or str(product_code).strip() == '':
                continue  # 跳过空行
                
            # 获取产品名称
            product_name = worksheet.cell(row=row_num, column=2).value
            if not product_name:
                continue
                
            # 获取内单价和外单价
            inner_price = worksheet.cell(row=row_num, column=4).value  # 内单价
            outer_price = worksheet.cell(row=row_num, column=6).value  # 外单价
            spec = worksheet.cell(row=row_num, column=3).value  # 规格
            
            # 处理价格数据
            try:
                inner_price = float(inner_price) if inner_price else 0.0
                outer_price = float(outer_price) if outer_price else 0.0
            except (ValueError, TypeError):
                inner_price = 0.0
                outer_price = 0.0
                
            # 处理规格数据
            spec = str(spec) if spec else ''
            
            product_data = {
                'code': str(product_code).strip(),
                'name': str(product_name).strip(),
                'inner_price': inner_price,
                'outer_price': outer_price,
                'specification': spec
            }
            
            products_data.append(product_data)
            print(f"产品: {product_data['code']} - {product_data['name']}")
            print(f"  规格: {product_data['specification']} | 内单价: ¥{product_data['inner_price']} | 外单价: ¥{product_data['outer_price']}")
        
        print(f'\n✅ 解析完成，共提取 {len(products_data)} 个产品')
        
        # 解析净味系列报价表
        print('\n--- 解析净味系列报价表 ---')
        worksheet = workbook['净味系列报价']
        
        jingwei_products = []
        
        # 遍历净味系列报价表
        for row_num in range(3, worksheet.max_row + 1):  # 从第3行开始
            # 获取产品编号
            product_code = worksheet.cell(row=row_num, column=1).value
            if not product_code or str(product_code).strip() == '':
                continue
                
            # 获取产品名称
            product_name = worksheet.cell(row=row_num, column=2).value
            if not product_name:
                continue
                
            # 获取内单价和外单价
            inner_price = worksheet.cell(row=row_num, column=4).value
            outer_price = worksheet.cell(row=row_num, column=5).value
            spec = worksheet.cell(row=row_num, column=3).value
            
            try:
                inner_price = float(inner_price) if inner_price else 0.0
                outer_price = float(outer_price) if outer_price else 0.0
            except (ValueError, TypeError):
                inner_price = 0.0
                outer_price = 0.0
                
            spec = str(spec) if spec else ''
            
            product_data = {
                'code': str(product_code).strip(),
                'name': str(product_name).strip(),
                'inner_price': inner_price,
                'outer_price': outer_price,
                'specification': spec
            }
            
            jingwei_products.append(product_data)
            print(f"净味产品: {product_data['code']} - {product_data['name']}")
            print(f"  规格: {product_data['specification']} | 内单价: ¥{product_data['inner_price']} | 外单价: ¥{product_data['outer_price']}")
        
        print(f'\n✅ 净味系列解析完成，共提取 {len(jingwei_products)} 个产品')
        
        # 合并所有产品数据
        all_products = products_data + jingwei_products
        
        # 去重（基于产品编号）
        unique_products = {}
        for product in all_products:
            code = product['code']
            if code not in unique_products:
                unique_products[code] = product
            else:
                # 如果有重复，保留价格更完整的那个
                existing = unique_products[code]
                if product['inner_price'] > existing['inner_price'] or product['outer_price'] > existing['outer_price']:
                    unique_products[code] = product
        
        final_products = list(unique_products.values())
        print(f'\n📊 总计提取产品: {len(final_products)} 个（去重后）')
        
        workbook.close()
        
        return final_products
        
    except Exception as e:
        print(f'❌ 解析Excel文件失败: {e}')
        return []

def save_products_to_db(products_data):
    """将产品数据保存到数据库"""
    conn = sqlite3.connect('products.db')
    cursor = conn.cursor()
    
    # 获取新创建的客户ID
    cursor.execute("SELECT id FROM purchase_units WHERE unit_name = '蕊芯家私1'")
    inner_customer = cursor.fetchone()
    cursor.execute("SELECT id FROM purchase_units WHERE unit_name = '蕊芯家私'")
    outer_customer = cursor.fetchone()
    
    if not inner_customer or not outer_customer:
        print('❌ 找不到新的客户单位')
        conn.close()
        return
    
    inner_customer_id = inner_customer[0]
    outer_customer_id = outer_customer[0]
    
    print(f'\n=== 保存产品到数据库 ===')
    print(f'蕊芯家私1（内单价）ID: {inner_customer_id}')
    print(f'蕊芯家私（外单价）ID: {outer_customer_id}')
    
    # 为每个产品创建记录
    for product in products_data:
        try:
            # 首先尝试查找是否已有相同型号的产品
            cursor.execute('SELECT id FROM products WHERE model_number = ?', [product['code']])
            existing_product = cursor.fetchone()
            
            if existing_product:
                product_id = existing_product[0]
                print(f'找到现有产品: {product_id} - {product["name"]}')
            else:
                # 创建新产品
                cursor.execute('''
                    INSERT INTO products (model_number, name, specification, price, quantity, 
                                       description, category, brand, unit, is_active, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, '个', 1, datetime('now'), datetime('now'))
                ''', (
                    product['code'],
                    product['name'],
                    f'规格: {product["specification"]}',
                    max(product['inner_price'], product['outer_price']),
                    1,
                    f'来源: 尹玉华Excel-蕊芯价格表',
                    '蕊芯系列',
                    ''
                ))
                product_id = cursor.lastrowid
                print(f'创建新产品: {product_id} - {product["name"]}')
            
            # 创建客户产品关联（内单价）
            if product['inner_price'] > 0:
                cursor.execute('''
                    INSERT OR REPLACE INTO customer_products (unit_id, product_id, custom_price, is_active, created_at, updated_at)
                    VALUES (?, ?, ?, 1, datetime('now'), datetime('now'))
                ''', (inner_customer_id, product_id, product['inner_price']))
            
            # 创建客户产品关联（外单价）
            if product['outer_price'] > 0:
                cursor.execute('''
                    INSERT OR REPLACE INTO customer_products (unit_id, product_id, custom_price, is_active, created_at, updated_at)
                    VALUES (?, ?, ?, 1, datetime('now'), datetime('now'))
                ''', (outer_customer_id, product_id, product['outer_price']))
            
        except Exception as e:
            print(f'保存产品 {product["code"]} 失败: {e}')
    
    conn.commit()
    conn.close()
    print(f'\n🎉 产品数据保存完成！')

if __name__ == "__main__":
    # 解析Excel文件
    products = parse_ruixin_excel()
    
    if products:
        # 保存到数据库
        save_products_to_db(products)
    else:
        print('❌ 没有提取到产品数据')