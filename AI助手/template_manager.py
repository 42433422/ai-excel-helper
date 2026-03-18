# 模板管理模块
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from config import EXCEL_CONFIG
from utils import setup_logger
from template_model import TemplateModel

# 设置日志
logger = setup_logger()

# 初始化数据库
TemplateModel.init_db()

class TemplateManager:
    """模板管理类"""
    
    def __init__(self):
        """初始化模板管理器"""
        self.template_dir = os.path.join(os.getcwd(), "templates")
        self.save_path = EXCEL_CONFIG["save_path"]
        
        # 确保模板目录存在
        if not os.path.exists(self.template_dir):
            os.makedirs(self.template_dir)
    
    def read_template(self, file_path, sheet_name=None):
        """读取XLSX模板文件，可以指定工作表名称"""
        try:
            logger.info(f"读取模板文件: {file_path}")
            
            # 检查文件是否存在
            if not os.path.exists(file_path):
                logger.error(f"模板文件不存在: {file_path}")
                return None
            
            # 打开工作簿
            workbook = openpyxl.load_workbook(file_path)
            
            # 获取所有工作表名称
            all_sheets = workbook.sheetnames
            
            # 如果指定了工作表名称，使用指定的工作表，否则使用活动工作表
            if sheet_name and sheet_name in all_sheets:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
                sheet_name = worksheet.title
            
            # 读取数据
            data = []
            for row in worksheet.iter_rows(values_only=True):
                data.append(list(row))
            
            # 关闭工作簿
            workbook.close()
            
            # 获取模板信息
            template_info = {
                "file_path": file_path,
                "sheet_name": sheet_name,
                "all_sheets": all_sheets,
                "data": data,
                "name": os.path.basename(file_path)
            }
            
            logger.info(f"模板文件 {file_path} 读取成功，工作表: {sheet_name}")
            return template_info
        except Exception as e:
            logger.error(f"读取模板文件失败: {e}")
            return None
    
    def save_template(self, template_info, template_name, company_id=None, description=''):
        """保存模板，支持关联公司"""
        try:
            logger.info(f"保存模板: {template_name}, 公司ID: {company_id}")
            
            # 构建模板文件路径
            template_path = os.path.join(self.template_dir, f"{template_name}.xlsx")
            
            # 如果原始模板有多个工作表，我们需要完整保存
            original_path = template_info.get("file_path")
            if original_path and os.path.exists(original_path):
                # 直接复制整个工作簿
                import shutil
                shutil.copy2(original_path, template_path)
                logger.info(f"复制整个工作簿作为模板: {template_path}")
            else:
                # 否则创建新工作簿
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = template_info.get("sheet_name", "Sheet1")
                
                # 写入数据
                for i, row in enumerate(template_info["data"]):
                    for j, cell_value in enumerate(row):
                        cell = worksheet.cell(row=i+1, column=j+1, value=cell_value)
                
                # 保存工作簿
                workbook.save(template_path)
                workbook.close()
                
            # 保存到数据库
            TemplateModel.add_template(template_name, template_path, company_id, description)
            
            logger.info(f"模板 {template_name} 保存成功")
            return template_path
        except Exception as e:
            logger.error(f"保存模板失败: {e}")
            return None
    
    def get_companies(self):
        """获取所有公司"""
        try:
            logger.info("获取所有公司")
            companies = TemplateModel.get_companies()
            logger.info(f"找到 {len(companies)} 个公司")
            return companies
        except Exception as e:
            logger.error(f"获取公司列表失败: {e}")
            return []
    
    def add_company(self, name, description=''):
        """添加公司"""
        try:
            logger.info(f"添加公司: {name}")
            company_id = TemplateModel.add_company(name, description)
            logger.info(f"公司 {name} 添加成功，ID: {company_id}")
            return company_id
        except Exception as e:
            logger.error(f"添加公司失败: {e}")
            return None
    
    def list_templates(self, company_id=None):
        """列出可用模板，可以按公司筛选"""
        try:
            logger.info(f"列出可用模板，公司ID: {company_id}")
            
            # 从数据库获取模板
            db_templates = TemplateModel.get_templates(company_id)
            
            # 同时从文件系统获取模板，确保所有模板都被记录
            file_templates = []
            for file in os.listdir(self.template_dir):
                if file.endswith(".xlsx"):
                    file_path = os.path.join(self.template_dir, file)
                    template_name = os.path.splitext(file)[0]
                    
                    # 检查是否已存在于数据库中
                    existing = next((t for t in db_templates if t['name'] == template_name), None)
                    if not existing:
                        # 添加到数据库
                        TemplateModel.add_template(template_name, file_path)
                        file_templates.append({
                            "id": None,
                            "name": template_name,
                            "file_name": file,
                            "path": file_path,
                            "company_id": None,
                            "company_name": None
                        })
            
            # 重新从数据库获取所有模板
            db_templates = TemplateModel.get_templates(company_id)
            
            # 转换为前端需要的格式
            templates = []
            for template in db_templates:
                file_name = os.path.basename(template['file_path'])
                templates.append({
                    "id": template['id'],
                    "name": template['name'],
                    "file_name": file_name,
                    "path": template['file_path'],
                    "company_id": template['company_id'],
                    "company_name": template['company_name']
                })
            
            # 添加文件系统中新增的模板
            templates.extend(file_templates)
            
            logger.info(f"找到 {len(templates)} 个模板")
            return templates
        except Exception as e:
            logger.error(f"列出模板失败: {e}")
            return []
    
    def use_template(self, template_name, custom_data=None, sheet_name=None, filename=None):
        """使用模板生成新文件，可以指定工作表名称和自定义文件名"""
        try:
            logger.info(f"使用模板 {template_name} 生成新文件")
            
            # 构建模板文件路径
            template_path = os.path.join(self.template_dir, f"{template_name}.xlsx")
            
            # 打开工作簿，保留所有样式和公式
            workbook = openpyxl.load_workbook(template_path, read_only=False, keep_vba=False)
            
            # 如果指定了工作表名称，使用指定的工作表，否则使用活动工作表
            if sheet_name and sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # 应用自定义数据，保留原有样式
            if custom_data:
                # 获取所有合并单元格
                merged_cells = list(worksheet.merged_cells.ranges)
                
                for cell_address, value in custom_data.items():
                    # 检查单元格是否是合并单元格的一部分
                    is_merged = False
                    for merged_cell in merged_cells:
                        if cell_address in merged_cell:
                            is_merged = True
                            # 只修改合并区域的第一个单元格
                            cell_address = merged_cell.start_cell.coordinate
                            break
                    
                    try:
                        # 获取原单元格
                        original_cell = worksheet[cell_address]
                        
                        # 设置新值
                        worksheet[cell_address] = value
                        
                        # 直接复制所有样式（不分别保存和恢复）
                        new_cell = worksheet[cell_address]
                        new_cell.font = original_cell.font.copy()
                        new_cell.alignment = original_cell.alignment.copy()
                        new_cell.border = original_cell.border.copy()
                        new_cell.fill = original_cell.fill.copy()
                        new_cell.number_format = original_cell.number_format
                    except Exception as e:
                        logger.error(f"处理单元格 {cell_address} 时出错: {e}")
                        continue
            
            # 生成文件名
            from utils import generate_unique_filename
            if filename:
                # 如果提供了自定义文件名，使用它（确保包含扩展名）
                if not filename.endswith('.xlsx'):
                    filename += '.xlsx'
            else:
                # 否则生成唯一文件名
                filename = generate_unique_filename(prefix=f"{template_name}_", ext=".xlsx")
            
            file_path = os.path.join(self.save_path, filename)
            
            # 保存新文件，确保保留所有样式和格式
            workbook.save(file_path)
            workbook.close()
            
            logger.info(f"使用模板 {template_name} 生成文件 {file_path} 成功")
            return file_path
        except Exception as e:
            logger.error(f"使用模板生成文件失败: {e}")
            return None
    
    def create_template_from_file(self, original_file, template_name, company_id=None, description=''):
        """从现有文件创建模板，支持关联公司"""
        try:
            logger.info(f"从文件 {original_file} 创建模板 {template_name}, 公司ID: {company_id}")
            
            # 读取原始文件
            template_info = self.read_template(original_file)
            if template_info is None:
                logger.error(f"读取原始文件失败")
                return None
            
            # 保存为模板，关联公司
            template_path = self.save_template(template_info, template_name, company_id, description)
            
            logger.info(f"从文件 {original_file} 创建模板 {template_name} 成功")
            return template_path
        except Exception as e:
            logger.error(f"从文件创建模板失败: {e}")
            return None
    
    def get_template_preview(self, template_name, sheet_name=None, limit_rows=None):
        """获取模板预览数据，可以指定工作表名称和行数限制"""
        try:
            logger.info(f"获取模板 {template_name} 的预览数据")
            
            # 构建模板文件路径
            template_path = os.path.join(self.template_dir, f"{template_name}.xlsx")
            
            # 读取模板
            template_info = self.read_template(template_path, sheet_name)
            if isinstance(template_info, str):
                logger.error(f"获取模板预览失败: {template_info}")
                return None
            if template_info is None:
                logger.error(f"模板 {template_name} 不存在")
                return None
            
            # 如果指定了行数限制，只返回指定行数，否则返回全部数据
            data = template_info["data"]
            if limit_rows:
                data = data[:limit_rows]
            
            preview_data = {
                "name": template_name,
                "sheet_name": template_info["sheet_name"],
                "all_sheets": template_info["all_sheets"],
                "data": data
            }
            
            logger.info(f"获取模板 {template_name} 预览数据成功")
            return preview_data
        except Exception as e:
            logger.error(f"获取模板预览数据失败: {e}")
            return None
    
    def save_customized_template(self, template_name, custom_data, new_template_name, sheet_name=None):
        """保存自定义修改后的模板为新模板"""
        try:
            logger.info(f"保存自定义模板: {new_template_name}")
            
            # 构建模板文件路径
            template_path = os.path.join(self.template_dir, f"{template_name}.xlsx")
            
            # 打开工作簿
            workbook = openpyxl.load_workbook(template_path)
            
            # 如果指定了工作表名称，使用指定的工作表
            if sheet_name and sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # 应用自定义数据
            if custom_data:
                for cell_address, value in custom_data.items():
                    worksheet[cell_address] = value
            
            # 构建新模板文件路径
            new_template_path = os.path.join(self.template_dir, f"{new_template_name}.xlsx")
            
            # 保存新模板
            workbook.save(new_template_path)
            workbook.close()
            
            # 保存到数据库
            TemplateModel.add_template(new_template_name, new_template_path)
            
            logger.info(f"自定义模板 {new_template_name} 保存成功")
            return new_template_path
        except Exception as e:
            logger.error(f"保存自定义模板失败: {e}")
            return None
    
    def delete_template(self, template_name):
        """删除模板，同时处理文件系统和数据库"""
        try:
            logger.info(f"删除模板: {template_name}")
            
            # 构建模板文件路径
            template_path = os.path.join(self.template_dir, f"{template_name}.xlsx")
            
            # 删除文件系统中的模板
            if os.path.exists(template_path):
                os.remove(template_path)
                logger.info(f"删除文件系统中的模板: {template_path}")
            
            # 从数据库中删除模板记录
            # 先获取模板ID
            templates = TemplateModel.get_templates()
            deleted = False
            for template in templates:
                if template['name'] == template_name:
                    TemplateModel.delete_template(template['id'])
                    deleted = True
                    logger.info(f"从数据库中删除模板记录，ID: {template['id']}")
                    break
            
            return True
        except Exception as e:
            logger.error(f"删除模板失败: {e}")
            return False
    
    def delete_all_templates(self):
        """删除所有模板，同时处理文件系统和数据库"""
        try:
            logger.info("删除所有模板")
            
            # 删除数据库中的所有模板记录
            TemplateModel.delete_all_templates()
            logger.info("从数据库中删除所有模板记录")
            
            # 删除文件系统中的所有模板文件
            for file in os.listdir(self.template_dir):
                if file.endswith(".xlsx"):
                    file_path = os.path.join(self.template_dir, file)
                    os.remove(file_path)
                    logger.info(f"删除文件系统中的模板: {file_path}")
            
            return True
        except Exception as e:
            logger.error(f"删除所有模板失败: {e}")
            return False
    
    def delete_all_companies(self):
        """删除所有公司"""
        try:
            logger.info("删除所有公司")
            
            # 先删除所有模板（外键约束）
            self.delete_all_templates()
            
            # 删除所有公司记录
            TemplateModel.delete_all_companies()
            logger.info("从数据库中删除所有公司记录")
            
            return True
        except Exception as e:
            logger.error(f"删除所有公司失败: {e}")
            return False
    
    def delete_all_data(self):
        """删除所有数据（公司、模板）"""
        try:
            logger.info("删除所有数据")
            
            # 删除所有模板和公司
            self.delete_all_companies()
            
            return True
        except Exception as e:
            logger.error(f"删除所有数据失败: {e}")
            return False
    
    def check_format_consistency(self, original_template, generated_file, sheet_name=None):
        """检查生成文件与原模板的格式一致性"""
        try:
            logger.info(f"检查生成文件 {generated_file} 与原模板 {original_template} 的格式一致性")
            
            # 打开两个工作簿
            original_workbook = openpyxl.load_workbook(original_template)
            generated_workbook = openpyxl.load_workbook(generated_file)
            
            # 选择工作表
            if sheet_name and sheet_name in original_workbook.sheetnames and sheet_name in generated_workbook.sheetnames:
                original_sheet = original_workbook[sheet_name]
                generated_sheet = generated_workbook[sheet_name]
            else:
                original_sheet = original_workbook.active
                generated_sheet = generated_workbook.active
            
            # 检查项目
            consistency_report = {
                "success": True,
                "differences": [],
                "summary": "格式一致性检查通过"
            }
            
            # 1. 检查单元格值（非自定义单元格）
            logger.info("检查单元格值...")
            for row in original_sheet.iter_rows():
                for cell in row:
                    cell_address = cell.coordinate
                    original_value = cell.value
                    generated_value = generated_sheet[cell_address].value
                    
                    # 只检查非空单元格
                    if original_value is not None and original_value != "" and generated_value != original_value:
                        consistency_report["differences"].append({
                            "type": "value",
                            "cell": cell_address,
                            "original": original_value,
                            "generated": generated_value,
                            "message": f"单元格 {cell_address} 值不一致"
                        })
            
            # 2. 检查单元格样式
            logger.info("检查单元格样式...")
            for row in original_sheet.iter_rows():
                for cell in row:
                    cell_address = cell.coordinate
                    original_cell = cell
                    generated_cell = generated_sheet[cell_address]
                    
                    # 检查字体 - 使用更详细的比较
                    font_different = False
                    try:
                        # 比较字体的各个属性，而不是直接比较StyleProxy对象
                        if original_cell.font.name != generated_cell.font.name or \
                           original_cell.font.size != generated_cell.font.size or \
                           original_cell.font.bold != generated_cell.font.bold or \
                           original_cell.font.italic != generated_cell.font.italic or \
                           original_cell.font.underline != generated_cell.font.underline or \
                           original_cell.font.color != generated_cell.font.color:
                            font_different = True
                    except Exception as e:
                        font_different = True
                        logger.debug(f"比较单元格 {cell_address} 字体时出错: {e}")
                    
                    if font_different:
                        consistency_report["differences"].append({
                            "type": "font",
                            "cell": cell_address,
                            "message": f"单元格 {cell_address} 字体样式不一致"
                        })
                    
                    # 检查对齐方式 - 使用更详细的比较
                    alignment_different = False
                    try:
                        # 比较对齐方式的各个属性
                        if original_cell.alignment.horizontal != generated_cell.alignment.horizontal or \
                           original_cell.alignment.vertical != generated_cell.alignment.vertical or \
                           original_cell.alignment.text_rotation != generated_cell.alignment.text_rotation or \
                           original_cell.alignment.wrap_text != generated_cell.alignment.wrap_text:
                            alignment_different = True
                    except Exception as e:
                        alignment_different = True
                        logger.debug(f"比较单元格 {cell_address} 对齐方式时出错: {e}")
                    
                    if alignment_different:
                        consistency_report["differences"].append({
                            "type": "alignment",
                            "cell": cell_address,
                            "message": f"单元格 {cell_address} 对齐方式不一致"
                        })
                    
                    # 检查数字格式
                    if original_cell.number_format != generated_cell.number_format:
                        consistency_report["differences"].append({
                            "type": "number_format",
                            "cell": cell_address,
                            "message": f"单元格 {cell_address} 数字格式不一致"
                        })
            
            # 3. 检查合并单元格
            logger.info("检查合并单元格...")
            original_merged_cells = list(original_sheet.merged_cells.ranges)
            generated_merged_cells = list(generated_sheet.merged_cells.ranges)
            
            if len(original_merged_cells) != len(generated_merged_cells):
                consistency_report["differences"].append({
                    "type": "merged_cells_count",
                    "message": f"合并单元格数量不一致: 原模板 {len(original_merged_cells)} 个, 生成文件 {len(generated_merged_cells)} 个"
                })
            
            # 4. 检查列宽
            logger.info("检查列宽...")
            for column in original_sheet.column_dimensions:
                if column in generated_sheet.column_dimensions:
                    original_width = original_sheet.column_dimensions[column].width
                    generated_width = generated_sheet.column_dimensions[column].width
                    if original_width != generated_width:
                        consistency_report["differences"].append({
                            "type": "column_width",
                            "column": column,
                            "original": original_width,
                            "generated": generated_width,
                            "message": f"列 {column} 宽度不一致"
                        })
                else:
                    consistency_report["differences"].append({
                        "type": "column_missing",
                        "column": column,
                        "message": f"生成文件缺少列 {column} 的维度设置"
                    })
            
            # 5. 检查行高
            logger.info("检查行高...")
            for row in original_sheet.row_dimensions:
                if row in generated_sheet.row_dimensions:
                    original_height = original_sheet.row_dimensions[row].height
                    generated_height = generated_sheet.row_dimensions[row].height
                    if original_height != generated_height:
                        consistency_report["differences"].append({
                            "type": "row_height",
                            "row": row,
                            "original": original_height,
                            "generated": generated_height,
                            "message": f"行 {row} 高度不一致"
                        })
                else:
                    consistency_report["differences"].append({
                        "type": "row_missing",
                        "row": row,
                        "message": f"生成文件缺少行 {row} 的维度设置"
                    })
            
            # 6. 检查工作表数量
            logger.info("检查工作表数量...")
            if len(original_workbook.sheetnames) != len(generated_workbook.sheetnames):
                consistency_report["differences"].append({
                    "type": "sheet_count",
                    "original": len(original_workbook.sheetnames),
                    "generated": len(generated_workbook.sheetnames),
                    "message": f"工作表数量不一致"
                })
            
            # 关闭工作簿
            original_workbook.close()
            generated_workbook.close()
            
            # 生成检查报告
            if len(consistency_report["differences"]) > 0:
                consistency_report["success"] = False
                consistency_report["summary"] = f"发现 {len(consistency_report['differences'])} 处格式差异"
                logger.warning(f"格式一致性检查失败，发现 {len(consistency_report['differences'])} 处差异")
            else:
                logger.info("格式一致性检查通过")
            
            return consistency_report
        except Exception as e:
            logger.error(f"格式一致性检查失败: {e}")
            return {
                "success": False,
                "differences": [],
                "summary": f"格式一致性检查失败: {e}"
            }
    
    def extract_template_info(self, template_path, sheet_name=None):
        """从模板中提取关键信息"""
        try:
            logger.info(f"从模板 {template_path} 提取关键信息")
            
            # 读取模板
            template_info = self.read_template(template_path, sheet_name)
            if not template_info:
                return None
            
            # 提取信息
            extracted_info = {
                "companies": self.extract_company_info(template_info),
                "participants": self.extract_participant_info(template_info),
                "metadata": self.extract_metadata(template_info)
            }
            
            logger.info(f"从模板 {template_path} 提取信息成功")
            return extracted_info
        except Exception as e:
            logger.error(f"提取模板信息失败: {e}")
            return None
    
    def extract_company_info(self, template_info):
        """从模板中提取公司信息"""
        try:
            logger.info("从模板中提取公司信息")
            
            companies = []
            data = template_info.get("data", [])
            
            # 定义公司名称关键词
            company_keywords = ["公司", "企业", "集团", "有限公司", "股份有限公司", "责任有限公司"]
            
            # 遍历模板数据，寻找包含公司关键词的单元格
            for i, row in enumerate(data):
                for j, cell_value in enumerate(row):
                    if cell_value and isinstance(cell_value, str):
                        # 检查是否包含公司关键词
                        for keyword in company_keywords:
                            if keyword in cell_value:
                                # 提取公司名称（简单处理，实际应用中可能需要更复杂的正则表达式）
                                company_name = cell_value.strip()
                                companies.append({
                                    "name": company_name,
                                    "position": f"行{i+1},列{j+1}"
                                })
                                break
            
            # 去重
            unique_companies = []
            seen_names = set()
            for company in companies:
                if company["name"] not in seen_names:
                    seen_names.add(company["name"])
                    unique_companies.append(company)
            
            logger.info(f"提取到 {len(unique_companies)} 个公司信息")
            return unique_companies
        except Exception as e:
            logger.error(f"提取公司信息失败: {e}")
            return []
    
    def extract_participant_info(self, template_info):
        """从模板中提取项目参与方信息"""
        try:
            logger.info("从模板中提取项目参与方信息")
            
            participants = []
            data = template_info.get("data", [])
            
            # 定义参与方关键词
            participant_keywords = ["甲方", "乙方", "丙方", "丁方", "供应商", "客户", "合作伙伴", "参与方"]
            
            # 遍历模板数据，寻找包含参与方关键词的单元格
            for i, row in enumerate(data):
                for j, cell_value in enumerate(row):
                    if cell_value and isinstance(cell_value, str):
                        # 检查是否包含参与方关键词
                        for keyword in participant_keywords:
                            if keyword in cell_value:
                                # 提取参与方信息
                                # 简单处理：假设参与方名称在关键词后面的单元格或同一单元格
                                participant_info = {
                                    "role": keyword,
                                    "name": "",
                                    "position": f"行{i+1},列{j+1}"
                                }
                                
                                # 检查下一个单元格是否包含名称
                                if j + 1 < len(row) and row[j+1]:
                                    participant_info["name"] = str(row[j+1]).strip()
                                else:
                                    # 尝试从当前单元格提取名称
                                    # 简单处理：假设格式为 "甲方：XXX公司"
                                    if "：" in cell_value or ":" in cell_value:
                                        separator = "：" if "：" in cell_value else ":"
                                        name_part = cell_value.split(separator)[1].strip() if len(cell_value.split(separator)) > 1 else ""
                                        participant_info["name"] = name_part
                                
                                participants.append(participant_info)
                                break
            
            logger.info(f"提取到 {len(participants)} 个项目参与方信息")
            return participants
        except Exception as e:
            logger.error(f"提取项目参与方信息失败: {e}")
            return []
    
    def extract_metadata(self, template_info):
        """从模板中提取元数据"""
        try:
            logger.info("从模板中提取元数据")
            
            metadata = []
            data = template_info.get("data", [])
            
            # 定义元数据关键词
            metadata_keywords = ["模板类型", "版本", "创建日期", "创建人", "描述", "用途", "状态"]
            
            # 遍历模板数据，寻找包含元数据关键词的单元格
            for i, row in enumerate(data):
                for j, cell_value in enumerate(row):
                    if cell_value and isinstance(cell_value, str):
                        # 检查是否包含元数据关键词
                        for keyword in metadata_keywords:
                            if keyword in cell_value:
                                # 提取元数据
                                metadata_value = ""
                                if j + 1 < len(row) and row[j+1]:
                                    metadata_value = str(row[j+1]).strip()
                                elif "：" in cell_value or ":" in cell_value:
                                    separator = "：" if "：" in cell_value else ":"
                                    metadata_value = cell_value.split(separator)[1].strip() if len(cell_value.split(separator)) > 1 else ""
                                
                                metadata.append({
                                    "key": keyword,
                                    "value": metadata_value,
                                    "position": f"行{i+1},列{j+1}"
                                })
                                break
            
            logger.info(f"提取到 {len(metadata)} 条元数据")
            return metadata
        except Exception as e:
            logger.error(f"提取元数据失败: {e}")
            return []
    
    def save_extracted_info(self, template_id, extracted_info):
        """保存提取的信息到数据库"""
        try:
            logger.info(f"保存提取的信息到数据库，模板ID: {template_id}")
            
            # 保存公司信息
            for company_info in extracted_info.get("companies", []):
                company_name = company_info.get("name", "")
                if company_name:
                    # 检查公司是否已存在
                    existing_companies = TemplateModel.get_companies()
                    company_id = next((c["id"] for c in existing_companies if c["name"] == company_name), None)
                    
                    if not company_id:
                        # 添加新公司
                        company_id = TemplateModel.add_company(company_name)
                    
                    # 更新模板的公司关联
                    TemplateModel.update_template(template_id, company_id=company_id)
            
            # 保存项目参与方信息
            for participant_info in extracted_info.get("participants", []):
                participant_name = participant_info.get("name", "")
                participant_role = participant_info.get("role", "")
                
                if participant_name:
                    # 检查参与方是否已存在
                    existing_participants = TemplateModel.get_project_participants()
                    participant_id = next((p["id"] for p in existing_participants if p["name"] == participant_name), None)
                    
                    if not participant_id:
                        # 添加新参与方
                        participant_id = TemplateModel.add_project_participant(
                            name=participant_name,
                            participant_type=participant_role,
                            description=f"从模板中提取的参与方: {participant_role}"
                        )
                    
                    # 关联模板和参与方
                    TemplateModel.add_template_participant(
                        template_id=template_id,
                        participant_id=participant_id,
                        role=participant_role
                    )
            
            # 保存元数据
            for metadata in extracted_info.get("metadata", []):
                key = metadata.get("key", "")
                value = metadata.get("value", "")
                
                if key and value:
                    TemplateModel.add_template_metadata(
                        template_id=template_id,
                        key=key,
                        value=value
                    )
            
            logger.info(f"保存提取的信息到数据库成功，模板ID: {template_id}")
            return True
        except Exception as e:
            logger.error(f"保存提取的信息到数据库失败: {e}")
            return False
    
    def auto_extract_and_save(self, template_path, template_id, sheet_name=None):
        """自动提取模板信息并保存到数据库"""
        try:
            logger.info(f"自动提取模板信息并保存，模板ID: {template_id}")
            
            # 提取信息
            extracted_info = self.extract_template_info(template_path, sheet_name)
            if not extracted_info:
                return False
            
            # 保存信息到数据库
            return self.save_extracted_info(template_id, extracted_info)
        except Exception as e:
            logger.error(f"自动提取并保存信息失败: {e}")
            return False
    
    # 数据关联功能
    def get_template_with_relations(self, template_id):
        """获取包含关联数据的模板信息"""
        try:
            logger.info(f"获取包含关联数据的模板信息，模板ID: {template_id}")
            
            # 获取模板基本信息
            template = TemplateModel.get_template(template_id)
            if not template:
                return None
            
            # 获取关联的参与方
            participants = TemplateModel.get_template_participants(template_id)
            
            # 获取关联的标签
            tags = TemplateModel.get_template_tags_by_template(template_id)
            
            # 获取模板元数据
            metadata = TemplateModel.get_template_metadata(template_id)
            
            # 获取模板版本
            versions = TemplateModel.get_template_versions(template_id)
            
            # 构建完整的模板信息
            template_with_relations = {
                **template,
                "participants": participants,
                "tags": tags,
                "metadata": metadata,
                "versions": versions
            }
            
            logger.info(f"获取包含关联数据的模板信息成功，模板ID: {template_id}")
            return template_with_relations
        except Exception as e:
            logger.error(f"获取包含关联数据的模板信息失败: {e}")
            return None
    
    def get_templates_by_company(self, company_id):
        """根据公司ID获取模板列表"""
        try:
            logger.info(f"根据公司ID获取模板列表，公司ID: {company_id}")
            
            # 获取公司信息
            company = TemplateModel.get_company(company_id)
            if not company:
                return []
            
            # 获取公司关联的模板
            templates = TemplateModel.get_templates(company_id)
            
            logger.info(f"获取到 {len(templates)} 个模板，公司ID: {company_id}")
            return templates
        except Exception as e:
            logger.error(f"根据公司ID获取模板列表失败: {e}")
            return []
    
    def get_templates_by_participant(self, participant_id):
        """根据参与方ID获取相关模板"""
        try:
            logger.info(f"根据参与方ID获取相关模板，参与方ID: {participant_id}")
            
            # 获取参与方信息
            participant = TemplateModel.get_project_participant(participant_id)
            if not participant:
                return []
            
            # 获取参与方关联的模板
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
            SELECT t.*, c.name as company_name
            FROM templates t
            JOIN template_participants tp ON t.id = tp.template_id
            LEFT JOIN companies c ON t.company_id = c.id
            WHERE tp.participant_id = ?
            ORDER BY t.name
            ''', (participant_id,))
            
            templates = cursor.fetchall()
            conn.close()
            
            # 转换为字典列表
            template_list = []
            for template in templates:
                template_list.append({
                    "id": template[0],
                    "name": template[1],
                    "file_path": template[2],
                    "company_id": template[3],
                    "description": template[4],
                    "created_at": template[5],
                    "company_name": template[6]
                })
            
            logger.info(f"获取到 {len(template_list)} 个模板，参与方ID: {participant_id}")
            return template_list
        except Exception as e:
            logger.error(f"根据参与方ID获取相关模板失败: {e}")
            return []
    
    def get_templates_by_tag(self, tag_id):
        """根据标签ID获取相关模板"""
        try:
            logger.info(f"根据标签ID获取相关模板，标签ID: {tag_id}")
            
            # 获取标签信息
            tag = TemplateModel.get_template_tag(tag_id)
            if not tag:
                return []
            
            # 获取标签关联的模板
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
            SELECT t.*, c.name as company_name
            FROM templates t
            JOIN template_tag_relations ttr ON t.id = ttr.template_id
            LEFT JOIN companies c ON t.company_id = c.id
            WHERE ttr.tag_id = ?
            ORDER BY t.name
            ''', (tag_id,))
            
            templates = cursor.fetchall()
            conn.close()
            
            # 转换为字典列表
            template_list = []
            for template in templates:
                template_list.append({
                    "id": template[0],
                    "name": template[1],
                    "file_path": template[2],
                    "company_id": template[3],
                    "description": template[4],
                    "created_at": template[5],
                    "company_name": template[6]
                })
            
            logger.info(f"获取到 {len(template_list)} 个模板，标签ID: {tag_id}")
            return template_list
        except Exception as e:
            logger.error(f"根据标签ID获取相关模板失败: {e}")
            return []
    
    def get_related_templates(self, template_id):
        """获取相关模板（基于相同公司、参与方或标签）"""
        try:
            logger.info(f"获取相关模板，模板ID: {template_id}")
            
            # 获取当前模板信息
            current_template = self.get_template_with_relations(template_id)
            if not current_template:
                return []
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # 基于公司、参与方或标签查找相关模板
            cursor.execute('''
            SELECT DISTINCT t.id, t.name, t.file_path, t.company_id, t.description, t.created_at, c.name as company_name
            FROM templates t
            LEFT JOIN companies c ON t.company_id = c.id
            WHERE t.id != ? 
            AND (
                t.company_id = ? OR 
                t.id IN (SELECT template_id FROM template_participants WHERE participant_id IN 
                    (SELECT participant_id FROM template_participants WHERE template_id = ?)) OR
                t.id IN (SELECT template_id FROM template_tag_relations WHERE tag_id IN 
                    (SELECT tag_id FROM template_tag_relations WHERE template_id = ?))
            )
            ORDER BY t.created_at DESC
            ''', (template_id, current_template.get("company_id"), template_id, template_id))
            
            related_templates = cursor.fetchall()
            conn.close()
            
            # 转换为字典列表
            template_list = []
            for template in related_templates:
                template_list.append({
                    "id": template[0],
                    "name": template[1],
                    "file_path": template[2],
                    "company_id": template[3],
                    "description": template[4],
                    "created_at": template[5],
                    "company_name": template[6]
                })
            
            logger.info(f"获取到 {len(template_list)} 个相关模板，模板ID: {template_id}")
            return template_list
        except Exception as e:
            logger.error(f"获取相关模板失败: {e}")
            return []
    
    def add_template_tag(self, template_id, tag_name):
        """为模板添加标签"""
        try:
            logger.info(f"为模板添加标签，模板ID: {template_id}, 标签名称: {tag_name}")
            
            # 检查标签是否已存在
            existing_tags = TemplateModel.get_template_tags()
            tag_id = next((t["id"] for t in existing_tags if t["name"] == tag_name), None)
            
            if not tag_id:
                # 添加新标签
                tag_id = TemplateModel.add_template_tag(tag_name, description=f"自动创建的标签: {tag_name}")
            
            # 关联模板和标签
            TemplateModel.add_template_tag_relation(template_id, tag_id)
            
            logger.info(f"为模板添加标签成功，模板ID: {template_id}, 标签名称: {tag_name}")
            return True
        except Exception as e:
            logger.error(f"为模板添加标签失败: {e}")
            return False
    
    def remove_template_tag(self, template_id, tag_id):
        """移除模板的标签"""
        try:
            logger.info(f"移除模板的标签，模板ID: {template_id}, 标签ID: {tag_id}")
            
            # 删除标签关联
            result = TemplateModel.delete_template_tag_relation(template_id, tag_id)
            
            logger.info(f"移除模板的标签成功，模板ID: {template_id}, 标签ID: {tag_id}")
            return result
        except Exception as e:
            logger.error(f"移除模板的标签失败: {e}")
            return False
    
    def update_template_metadata(self, metadata_id, value):
        """更新模板元数据"""
        try:
            logger.info(f"更新模板元数据，元数据ID: {metadata_id}")
            
            result = TemplateModel.update_template_metadata(metadata_id, value)
            
            logger.info(f"更新模板元数据成功，元数据ID: {metadata_id}")
            return result
        except Exception as e:
            logger.error(f"更新模板元数据失败: {e}")
            return False
    
    def delete_template_metadata(self, metadata_id):
        """删除模板元数据"""
        try:
            logger.info(f"删除模板元数据，元数据ID: {metadata_id}")
            
            result = TemplateModel.delete_template_metadata(metadata_id)
            
            logger.info(f"删除模板元数据成功，元数据ID: {metadata_id}")
            return result
        except Exception as e:
            logger.error(f"删除模板元数据失败: {e}")
            return False
    
    def get_participant_templates_map(self):
        """获取参与方与模板的映射关系"""
        try:
            logger.info("获取参与方与模板的映射关系")
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
            SELECT pp.id, pp.name, pp.type, GROUP_CONCAT(t.name, ', ') as template_names
            FROM project_participants pp
            JOIN template_participants tp ON pp.id = tp.participant_id
            JOIN templates t ON tp.template_id = t.id
            GROUP BY pp.id, pp.name, pp.type
            ORDER BY pp.name
            ''')
            
            mapping = cursor.fetchall()
            conn.close()
            
            participant_map = []
            for item in mapping:
                participant_map.append({
                    "participant_id": item[0],
                    "participant_name": item[1],
                    "participant_type": item[2],
                    "template_names": item[3].split(', ') if item[3] else []
                })
            
            logger.info(f"获取到 {len(participant_map)} 个参与方模板映射关系")
            return participant_map
        except Exception as e:
            logger.error(f"获取参与方与模板的映射关系失败: {e}")
            return []
    
    def get_company_templates_map(self):
        """获取公司与模板的映射关系"""
        try:
            logger.info("获取公司与模板的映射关系")
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
            SELECT c.id, c.name, GROUP_CONCAT(t.name, ', ') as template_names
            FROM companies c
            JOIN templates t ON c.id = t.company_id
            GROUP BY c.id, c.name
            ORDER BY c.name
            ''')
            
            mapping = cursor.fetchall()
            conn.close()
            
            company_map = []
            for item in mapping:
                company_map.append({
                    "company_id": item[0],
                    "company_name": item[1],
                    "template_names": item[2].split(', ') if item[2] else []
                })
            
            logger.info(f"获取到 {len(company_map)} 个公司模板映射关系")
            return company_map
        except Exception as e:
            logger.error(f"获取公司与模板的映射关系失败: {e}")
            return []
    
    # 版本控制功能
    def add_template_version(self, template_id, version_number, description='', created_by=''):
        """为模板添加新版本"""
        try:
            logger.info(f"为模板添加新版本，模板ID: {template_id}, 版本号: {version_number}")
            
            # 获取模板基本信息
            template = TemplateModel.get_template(template_id)
            if not template:
                return None
            
            # 构建新版本文件路径
            original_path = template.get("file_path")
            if not original_path or not os.path.exists(original_path):
                logger.error(f"原模板文件不存在: {original_path}")
                return None
            
            # 创建版本目录
            version_dir = os.path.join(self.template_dir, "versions", str(template_id))
            if not os.path.exists(version_dir):
                os.makedirs(version_dir, exist_ok=True)
            
            # 复制原模板文件到版本目录
            import shutil
            version_filename = f"{template['name']}_v{version_number}.xlsx"
            version_path = os.path.join(version_dir, version_filename)
            shutil.copy2(original_path, version_path)
            
            # 添加版本记录到数据库
            version_id = TemplateModel.add_template_version(
                template_id=template_id,
                version_number=version_number,
                file_path=version_path,
                description=description,
                created_by=created_by
            )
            
            logger.info(f"为模板添加新版本成功，模板ID: {template_id}, 版本号: {version_number}, 版本ID: {version_id}")
            return version_id
        except Exception as e:
            logger.error(f"为模板添加新版本失败: {e}")
            return None
    
    def get_template_versions(self, template_id):
        """获取模板的版本历史"""
        try:
            logger.info(f"获取模板的版本历史，模板ID: {template_id}")
            
            # 获取版本列表
            versions = TemplateModel.get_template_versions(template_id)
            
            logger.info(f"获取到 {len(versions)} 个版本，模板ID: {template_id}")
            return versions
        except Exception as e:
            logger.error(f"获取模板的版本历史失败: {e}")
            return []
    
    def get_template_version(self, version_id):
        """获取特定版本的模板信息"""
        try:
            logger.info(f"获取特定版本的模板信息，版本ID: {version_id}")
            
            # 获取版本信息
            version = TemplateModel.get_template_version(version_id)
            if not version:
                return None
            
            # 读取版本文件
            if os.path.exists(version['file_path']):
                template_info = self.read_template(version['file_path'])
                if template_info:
                    version['template_info'] = template_info
            
            logger.info(f"获取特定版本的模板信息成功，版本ID: {version_id}")
            return version
        except Exception as e:
            logger.error(f"获取特定版本的模板信息失败: {e}")
            return None
    
    def rollback_to_version(self, template_id, version_id):
        """回滚模板到特定版本"""
        try:
            logger.info(f"回滚模板到特定版本，模板ID: {template_id}, 版本ID: {version_id}")
            
            # 获取版本信息
            version = TemplateModel.get_template_version(version_id)
            if not version:
                return False
            
            # 验证版本是否属于该模板
            if version['template_id'] != template_id:
                logger.error(f"版本 {version_id} 不属于模板 {template_id}")
                return False
            
            # 获取模板信息
            template = TemplateModel.get_template(template_id)
            if not template:
                return False
            
            # 复制版本文件到原模板位置
            import shutil
            shutil.copy2(version['file_path'], template['file_path'])
            
            logger.info(f"回滚模板到特定版本成功，模板ID: {template_id}, 版本ID: {version_id}")
            return True
        except Exception as e:
            logger.error(f"回滚模板到特定版本失败: {e}")
            return False
    
    def compare_versions(self, version_id1, version_id2):
        """比较两个版本的差异"""
        try:
            logger.info(f"比较两个版本的差异，版本ID1: {version_id1}, 版本ID2: {version_id2}")
            
            # 获取两个版本的信息
            version1 = TemplateModel.get_template_version(version_id1)
            version2 = TemplateModel.get_template_version(version_id2)
            
            if not version1 or not version2:
                return None
            
            # 验证两个版本是否属于同一模板
            if version1['template_id'] != version2['template_id']:
                logger.error(f"两个版本不属于同一模板")
                return None
            
            # 读取两个版本的模板内容
            template1_info = self.read_template(version1['file_path'])
            template2_info = self.read_template(version2['file_path'])
            
            if not template1_info or not template2_info:
                return None
            
            # 比较基本信息
            differences = {
                "version1": {
                    "version_id": version_id1,
                    "version_number": version1['version_number'],
                    "created_at": version1['created_at'],
                    "description": version1['description'],
                    "created_by": version1['created_by']
                },
                "version2": {
                    "version_id": version_id2,
                    "version_number": version2['version_number'],
                    "created_at": version2['created_at'],
                    "description": version2['description'],
                    "created_by": version2['created_by']
                },
                "content_differences": []
            }
            
            # 比较内容差异（简单比较行数和列数）
            data1 = template1_info.get("data", [])
            data2 = template2_info.get("data", [])
            
            max_rows = max(len(data1), len(data2))
            
            for i in range(max_rows):
                if i >= len(data1):
                    # 版本2比版本1多的行
                    differences["content_differences"].append({
                        "type": "row_added",
                        "row": i + 1,
                        "content": data2[i]
                    })
                    continue
                if i >= len(data2):
                    # 版本1比版本2多的行
                    differences["content_differences"].append({
                        "type": "row_removed",
                        "row": i + 1,
                        "content": data1[i]
                    })
                    continue
                
                row1 = data1[i]
                row2 = data2[i]
                max_cols = max(len(row1), len(row2))
                
                for j in range(max_cols):
                    if j >= len(row1):
                        # 版本2比版本1多的列
                        differences["content_differences"].append({
                            "type": "cell_added",
                            "cell": f"行{i+1},列{j+1}",
                            "value": row2[j]
                        })
                        continue
                    if j >= len(row2):
                        # 版本1比版本2多的列
                        differences["content_differences"].append({
                            "type": "cell_removed",
                            "cell": f"行{i+1},列{j+1}",
                            "value": row1[j]
                        })
                        continue
                    
                    if row1[j] != row2[j]:
                        # 单元格值不同
                        differences["content_differences"].append({
                            "type": "cell_changed",
                            "cell": f"行{i+1},列{j+1}",
                            "value1": row1[j],
                            "value2": row2[j]
                        })
            
            logger.info(f"比较两个版本的差异成功，版本ID1: {version_id1}, 版本ID2: {version_id2}")
            return differences
        except Exception as e:
            logger.error(f"比较两个版本的差异失败: {e}")
            return None
    
    def delete_template_version(self, version_id):
        """删除模板版本"""
        try:
            logger.info(f"删除模板版本，版本ID: {version_id}")
            
            # 获取版本信息
            version = TemplateModel.get_template_version(version_id)
            if not version:
                return False
            
            # 删除版本文件
            version_path = version.get("file_path")
            if version_path and os.path.exists(version_path):
                os.remove(version_path)
                logger.info(f"删除版本文件: {version_path}")
            
            # 删除数据库中的版本记录
            # 注意：这里需要在template_model.py中添加delete_template_version方法
            # 目前template_model.py中没有这个方法，我们需要先添加它
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM template_versions WHERE id = ?', (version_id,))
            conn.commit()
            conn.close()
            
            logger.info(f"删除模板版本成功，版本ID: {version_id}")
            return True
        except Exception as e:
            logger.error(f"删除模板版本失败: {e}")
            return False
    
    def set_default_version(self, template_id, version_id):
        """设置模板的默认版本"""
        try:
            logger.info(f"设置模板的默认版本，模板ID: {template_id}, 版本ID: {version_id}")
            
            # 获取版本信息
            version = TemplateModel.get_template_version(version_id)
            if not version:
                return False
            
            # 验证版本是否属于该模板
            if version['template_id'] != template_id:
                logger.error(f"版本 {version_id} 不属于模板 {template_id}")
                return False
            
            # 更新模板元数据，记录默认版本
            # 先检查是否已存在默认版本元数据
            metadata = TemplateModel.get_template_metadata(template_id)
            default_version_metadata = next((m for m in metadata if m['key'] == 'default_version'), None)
            
            if default_version_metadata:
                # 更新现有元数据
                TemplateModel.update_template_metadata(default_version_metadata['id'], str(version_id))
            else:
                # 添加新元数据
                TemplateModel.add_template_metadata(template_id, 'default_version', str(version_id))
            
            logger.info(f"设置模板的默认版本成功，模板ID: {template_id}, 版本ID: {version_id}")
            return True
        except Exception as e:
            logger.error(f"设置模板的默认版本失败: {e}")
            return False
    
    def get_default_version(self, template_id):
        """获取模板的默认版本"""
        try:
            logger.info(f"获取模板的默认版本，模板ID: {template_id}")
            
            # 获取模板元数据
            metadata = TemplateModel.get_template_metadata(template_id)
            default_version_metadata = next((m for m in metadata if m['key'] == 'default_version'), None)
            
            if not default_version_metadata:
                # 没有设置默认版本，返回最新版本
                versions = TemplateModel.get_template_versions(template_id)
                if versions:
                    return versions[0]  # 最新版本在第一个位置
                return None
            
            # 获取默认版本信息
            version_id = int(default_version_metadata['value'])
            default_version = TemplateModel.get_template_version(version_id)
            
            logger.info(f"获取模板的默认版本成功，模板ID: {template_id}, 版本ID: {version_id}")
            return default_version
        except Exception as e:
            logger.error(f"获取模板的默认版本失败: {e}")
            return None
    
    # 标签管理功能
    def create_tag(self, tag_name, description=''):
        """创建模板标签"""
        try:
            logger.info(f"创建模板标签: {tag_name}")
            
            # 检查标签是否已存在
            existing_tags = TemplateModel.get_template_tags()
            existing_tag = next((t for t in existing_tags if t['name'] == tag_name), None)
            if existing_tag:
                logger.warning(f"标签 {tag_name} 已存在")
                return existing_tag['id']
            
            # 创建新标签
            tag_id = TemplateModel.add_template_tag(tag_name, description)
            
            logger.info(f"创建模板标签成功: {tag_name}, 标签ID: {tag_id}")
            return tag_id
        except Exception as e:
            logger.error(f"创建模板标签失败: {e}")
            return None
    
    def get_tags(self):
        """获取所有模板标签"""
        try:
            logger.info("获取所有模板标签")
            
            tags = TemplateModel.get_template_tags()
            
            logger.info(f"获取到 {len(tags)} 个模板标签")
            return tags
        except Exception as e:
            logger.error(f"获取模板标签失败: {e}")
            return []
    
    def get_tag(self, tag_id):
        """获取特定模板标签"""
        try:
            logger.info(f"获取特定模板标签，标签ID: {tag_id}")
            
            tag = TemplateModel.get_template_tag(tag_id)
            if not tag:
                logger.warning(f"标签ID {tag_id} 不存在")
                return None
            
            logger.info(f"获取特定模板标签成功，标签ID: {tag_id}")
            return tag
        except Exception as e:
            logger.error(f"获取特定模板标签失败: {e}")
            return None
    
    def update_tag(self, tag_id, tag_name, description=''):
        """更新模板标签"""
        try:
            logger.info(f"更新模板标签，标签ID: {tag_id}, 新名称: {tag_name}")
            
            # 检查标签是否存在
            existing_tag = TemplateModel.get_template_tag(tag_id)
            if not existing_tag:
                logger.warning(f"标签ID {tag_id} 不存在")
                return False
            
            # 检查新名称是否已被其他标签使用
            all_tags = TemplateModel.get_template_tags()
            name_used = any(t for t in all_tags if t['name'] == tag_name and t['id'] != tag_id)
            if name_used:
                logger.warning(f"标签名称 {tag_name} 已被使用")
                return False
            
            # 更新标签
            result = TemplateModel.update_template_tag(tag_id, tag_name, description)
            
            logger.info(f"更新模板标签成功，标签ID: {tag_id}")
            return result
        except Exception as e:
            logger.error(f"更新模板标签失败: {e}")
            return False
    
    def delete_tag(self, tag_id):
        """删除模板标签"""
        try:
            logger.info(f"删除模板标签，标签ID: {tag_id}")
            
            # 检查标签是否存在
            existing_tag = TemplateModel.get_template_tag(tag_id)
            if not existing_tag:
                logger.warning(f"标签ID {tag_id} 不存在")
                return False
            
            # 删除标签
            result = TemplateModel.delete_template_tag(tag_id)
            
            logger.info(f"删除模板标签成功，标签ID: {tag_id}")
            return result
        except Exception as e:
            logger.error(f"删除模板标签失败: {e}")
            return False
    
    def batch_create_tags(self, tag_list):
        """批量创建模板标签"""
        try:
            logger.info(f"批量创建模板标签，数量: {len(tag_list)}")
            
            created_tags = []
            for tag_info in tag_list:
                tag_name = tag_info.get('name', '')
                description = tag_info.get('description', '')
                
                if tag_name:
                    tag_id = self.create_tag(tag_name, description)
                    if tag_id:
                        created_tags.append({
                            'name': tag_name,
                            'id': tag_id
                        })
            
            logger.info(f"批量创建模板标签成功，创建数量: {len(created_tags)}")
            return created_tags
        except Exception as e:
            logger.error(f"批量创建模板标签失败: {e}")
            return []
    
    def batch_add_template_tags(self, template_id, tag_ids):
        """为模板批量添加标签"""
        try:
            logger.info(f"为模板批量添加标签，模板ID: {template_id}, 标签数量: {len(tag_ids)}")
            
            added_tags = []
            for tag_id in tag_ids:
                # 检查标签是否存在
                tag = TemplateModel.get_template_tag(tag_id)
                if tag:
                    # 检查标签是否已关联到模板
                    existing_tags = TemplateModel.get_template_tags_by_template(template_id)
                    already_added = any(t for t in existing_tags if t['id'] == tag_id)
                    if not already_added:
                        TemplateModel.add_template_tag_relation(template_id, tag_id)
                        added_tags.append(tag_id)
            
            logger.info(f"为模板批量添加标签成功，模板ID: {template_id}, 添加数量: {len(added_tags)}")
            return added_tags
        except Exception as e:
            logger.error(f"为模板批量添加标签失败: {e}")
            return []
    
    def batch_remove_template_tags(self, template_id, tag_ids):
        """为模板批量移除标签"""
        try:
            logger.info(f"为模板批量移除标签，模板ID: {template_id}, 标签数量: {len(tag_ids)}")
            
            removed_tags = []
            for tag_id in tag_ids:
                # 检查标签是否存在
                tag = TemplateModel.get_template_tag(tag_id)
                if tag:
                    # 检查标签是否已关联到模板
                    existing_tags = TemplateModel.get_template_tags_by_template(template_id)
                    already_added = any(t for t in existing_tags if t['id'] == tag_id)
                    if already_added:
                        TemplateModel.delete_template_tag_relation(template_id, tag_id)
                        removed_tags.append(tag_id)
            
            logger.info(f"为模板批量移除标签成功，模板ID: {template_id}, 移除数量: {len(removed_tags)}")
            return removed_tags
        except Exception as e:
            logger.error(f"为模板批量移除标签失败: {e}")
            return []
    
    def get_templates_by_tags(self, tag_ids):
        """根据多个标签ID获取相关模板"""
        try:
            logger.info(f"根据多个标签ID获取相关模板，标签数量: {len(tag_ids)}")
            
            if not tag_ids:
                return []
            
            # 构建查询条件
            placeholders = ', '.join(['?'] * len(tag_ids))
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # 查询关联了所有指定标签的模板
            cursor.execute(f'''
            SELECT t.id, t.name, t.file_path, t.company_id, t.description, t.created_at, c.name as company_name
            FROM templates t
            LEFT JOIN companies c ON t.company_id = c.id
            JOIN template_tag_relations ttr ON t.id = ttr.template_id
            WHERE ttr.tag_id IN ({placeholders})
            GROUP BY t.id, t.name, t.file_path, t.company_id, t.description, t.created_at, c.name
            HAVING COUNT(DISTINCT ttr.tag_id) = {len(tag_ids)}
            ORDER BY t.name
            ''', tag_ids)
            
            templates = cursor.fetchall()
            conn.close()
            
            # 转换为字典列表
            template_list = []
            for template in templates:
                template_list.append({
                    "id": template[0],
                    "name": template[1],
                    "file_path": template[2],
                    "company_id": template[3],
                    "description": template[4],
                    "created_at": template[5],
                    "company_name": template[6]
                })
            
            logger.info(f"获取到 {len(template_list)} 个模板，标签数量: {len(tag_ids)}")
            return template_list
        except Exception as e:
            logger.error(f"根据多个标签ID获取相关模板失败: {e}")
            return []
    
    def search_tags(self, keyword):
        """根据关键词搜索标签"""
        try:
            logger.info(f"根据关键词搜索标签，关键词: {keyword}")
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
            SELECT id, name, description
            FROM template_tags
            WHERE name LIKE ? OR description LIKE ?
            ORDER BY name
            ''', (f'%{keyword}%', f'%{keyword}%'))
            
            tags = cursor.fetchall()
            conn.close()
            
            # 转换为字典列表
            tag_list = []
            for tag in tags:
                tag_list.append({
                    "id": tag[0],
                    "name": tag[1],
                    "description": tag[2]
                })
            
            logger.info(f"搜索到 {len(tag_list)} 个标签，关键词: {keyword}")
            return tag_list
        except Exception as e:
            logger.error(f"根据关键词搜索标签失败: {e}")
            return []
    
    def get_tag_statistics(self):
        """获取标签使用统计信息"""
        try:
            logger.info("获取标签使用统计信息")
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
            SELECT tt.id, tt.name, tt.description, COUNT(ttr.template_id) as template_count
            FROM template_tags tt
            LEFT JOIN template_tag_relations ttr ON tt.id = ttr.tag_id
            GROUP BY tt.id, tt.name, tt.description
            ORDER BY template_count DESC
            ''')
            
            tag_stats = cursor.fetchall()
            conn.close()
            
            # 转换为字典列表
            stats_list = []
            for stat in tag_stats:
                stats_list.append({
                    "id": stat[0],
                    "name": stat[1],
                    "description": stat[2],
                    "template_count": stat[3]
                })
            
            logger.info(f"获取到 {len(stats_list)} 个标签统计信息")
            return stats_list
        except Exception as e:
            logger.error(f"获取标签使用统计信息失败: {e}")
            return []
