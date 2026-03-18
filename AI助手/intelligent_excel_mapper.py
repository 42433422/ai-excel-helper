"""
智能Excel字段映射模块
专门用于解析三个指定Excel文件中的产品数据
"""

import pandas as pd
import openpyxl
import os
import re
import logging
from typing import Dict, List, Tuple, Optional

logger = logging.getLogger(__name__)

class IntelligentExcelMapper:
    """智能Excel文件字段映射器"""
    
    def __init__(self):
        # 预定义的Excel文件配置
        self.file_configs = {
            '七彩乐园.xlsx': {
                'sheet_name': '25出货',
                'columns': {
                    'product_model': 'C',     # 列C：产品型号
                    'product_name': 'F',      # 列F：产品名称
                    'specification': 'H',    # 列H：规格/KG
                    'price': 'J',            # 列J：单价/元
                    'date': 'A',             # 列A：日期
                    'order_number': 'B'      # 列B：单号
                },
                'header_row': 2,  # 第2行是标题行
                'data_start_row': 3  # 从第3行开始是数据
            },
            '宜榢.xlsx': {
                'sheet_name': '出货',
                'columns': {
                    'product_model': 'D',     # 列D：产品型号
                    'product_name': 'G',      # 列G：产品名称
                    'specification': 'I',    # 列I：规格/KG
                    'price': 'K',            # 列K：单价/元
                    'date': 'B',             # 列B：日期
                    'order_number': 'C'      # 列C：单号
                },
                'header_row': 2,  # 第2行是标题行
                'data_start_row': 3,  # 从第3行开始是数据
                'allow_empty_model': True  # 允许产品型号为空
            },
            '迎扬李总.xlsx': {
                'sheet_name': '出货',
                'columns': {
                    'product_model': 'C',     # 列C：产品型号
                    'product_name': 'F',      # 列F：产品名称
                    'specification': 'H',    # 列H：规格/KG
                    'price': 'J',            # 列J：单价/元
                    'date': 'A',             # 列A：日期
                    'order_number': 'B'      # 列B：单号
                },
                'header_row': 2,  # 第2行是标题行
                'data_start_row': 3  # 从第3行开始是数据
            },
            '侯雪梅.xlsx': {
                'sheet_name': '25年出货',
                'columns': {
                    'product_model': 'D',     # 列D：产品型号
                    'product_name': 'G',      # 列G：产品名称
                    'specification': 'I',    # 列I：规格/KG
                    'price': 'J',            # 列J：单价/元
                    'date': 'B',             # 列B：日期
                    'order_number': 'C'      # 列C：单号
                },
                'header_row': 2,  # 第2行是标题行
                'data_start_row': 3  # 从第3行开始是数据
            },
            '刘英.xlsx': {
                'sheet_name': '出货',
                'columns': {
                    'product_model': 'A',     # 列A：产品型号
                    'product_name': 'D',      # 列D：产品名称
                    'specification': 'F',    # 列F：规格/KG
                    'price': 'H',            # 列H：单价/元
                    'date': 'B',             # 列B：日期
                    'order_number': 'C'      # 列C：单号
                },
                'header_row': 3,  # 第3行是标题行
                'data_start_row': 4,  # 从第4行开始是数据
                'allow_empty_model': True  # 允许产品型号为空
            },
            '国圣化工.xlsx': {
                'sheet_name': '国圣',
                'columns': {
                    'product_model': 'A',     # 列A：产品型号
                    'product_name': 'D',      # 列D：产品名称
                    'specification': 'F',    # 列F：规格/KG
                    'price': 'H',            # 列H：单价/元
                    'date': 'B',             # 列B：日期
                    'order_number': 'C'      # 列C：单号
                },
                'header_row': 3,  # 第3行是标题行
                'data_start_row': 4,  # 从第4行开始是数据
                'allow_empty_model': True  # 允许产品型号为空
            },
            '宗南.xlsx': {
                'sheet_name': '出货',
                'columns': {
                    'product_model': 'C',     # 列C：产品型号
                    'product_name': 'F',      # 列F：产品名称
                    'specification': 'H',    # 列H：规格/KG
                    'price': 'J',            # 列J：单价/元
                    'date': 'A',             # 列A：日期
                    'order_number': 'B'      # 列B：单号
                },
                'header_row': 2,  # 第2行是标题行
                'data_start_row': 3  # 从第3行开始是数据
            },
            '现金.xlsx': {
                'sheet_name': '出货',
                'columns': {
                    'product_model': 'A',     # 列A：产品型号
                    'product_name': 'D',      # 列D：产品名称
                    'specification': 'F',    # 列F：规格/KG
                    'price': 'H',            # 列H：单价/元
                    'date': 'B',             # 列B：日期
                    'order_number': 'C'      # 列C：单号
                },
                'header_row': 3,  # 第3行是标题行
                'data_start_row': 4,  # 从第4行开始是数据
                'allow_empty_model': True  # 允许产品型号为空
            },
            '尹玉华1.xlsx': {
                'sheet_name': '出货',
                'columns': {
                    'product_model': 'D',     # 列D：产品型号
                    'product_name': 'F',      # 列F：产品名称
                    'specification': 'H',    # 列H：规格/KG
                    'price': 'I',            # 列I：单价/元
                    'date': 'B',             # 列B：日期
                    'order_number': 'C'      # 列C：单号
                },
                'header_row': 2,  # 第2行是标题行
                'data_start_row': 3  # 从第3行开始是数据
            },
            '志泓.xlsx': {
                'sheet_name': '出货',
                'columns': {
                    'product_model': 'C',     # 列C：产品型号
                    'product_name': 'F',      # 列F：产品名称
                    'specification': 'H',    # 列H：规格/KG
                    'price': 'I',            # 列I：单价/元
                    'date': 'A',             # 列A：日期
                    'order_number': 'B'      # 列B：单号
                },
                'header_row': 2,  # 第2行是标题行
                'data_start_row': 3  # 从第3行开始是数据
            },
            '新旺博旺.xlsx': {
                'sheet_name': '24出货',
                'columns': {
                    'product_model': 'C',     # 列C：产品型号
                    'product_name': 'F',      # 列F：产品名称
                    'specification': 'H',    # 列H：规格/KG
                    'price': 'I',            # 列I：单价/元
                    'date': 'A',             # 列A：日期
                    'order_number': 'B'      # 列B：单号
                },
                'header_row': 2,  # 第2行是标题行
                'data_start_row': 3  # 从第3行开始是数据
            },
            '温总.xlsx': {
                'sheet_name': '温总',
                'columns': {
                    'product_model': 'A',     # 列A：产品型号
                    'product_name': 'D',      # 列D：产品名称
                    'specification': 'F',    # 列F：规格/KG
                    'price': 'H',            # 列H：单价/元
                    'date': 'B',             # 列B：日期
                    'order_number': 'C'      # 列C：单号
                },
                'header_row': 3,  # 第3行是标题行
                'data_start_row': 4,  # 从第4行开始是数据
                'allow_empty_model': True  # 允许产品型号为空
            },
            '澜宇电视柜.xlsx': {
                'sheet_name': '出货',
                'columns': {
                    'product_model': 'D',     # 列D：产品型号
                    'product_name': 'F',      # 列F：产品名称
                    'specification': 'H',    # 列H：规格/KG
                    'price': 'I',            # 列I：单价/元
                    'date': 'B',             # 列B：日期
                    'order_number': 'C'      # 列C：单号
                },
                'header_row': 3,  # 第3行是标题行
                'data_start_row': 4  # 从第4行开始是数据
            },
            '小洋杨总、.xlsx': {
                'sheet_name': '出货',
                'columns': {
                    'product_model': 'D',     # 列D：产品型号
                    'product_name': 'F',      # 列F：产品名称
                    'specification': 'H',    # 列H：规格/KG
                    'price': 'I',            # 列I：单价/元
                    'date': 'B',             # 列B：日期
                    'order_number': 'C'      # 列C：单号
                },
                'header_row': 3,  # 第3行是标题行
                'data_start_row': 4,  # 从第4行开始是数据
                'allow_empty_model': True  # 允许产品型号为空
            },
            '迎扬李总(1).xlsx': {
                'sheet_name': '出货',
                'columns': {
                    'product_model': 'C',     # 列C：产品型号
                    'product_name': 'F',      # 列F：产品名称
                    'specification': 'H',    # 列H：规格/KG
                    'price': 'J',            # 列J：单价/元
                    'date': 'A',             # 列A：日期
                    'order_number': 'B'      # 列B：单号
                },
                'header_row': 2,  # 第2行是标题行
                'data_start_row': 3  # 从第3行开始是数据
            },
            '邻居杨总.xlsx': {
                'sheet_name': '出货',
                'columns': {
                    'product_model': 'C',     # 列C：产品型号
                    'product_name': 'F',      # 列F：产品名称
                    'specification': 'H',    # 列H：规格/KG
                    'price': 'J',            # 列J：单价/元
                    'date': 'A',             # 列A：日期
                    'order_number': 'B'      # 列B：单号
                },
                'header_row': 2,  # 第2行是标题行
                'data_start_row': 3  # 从第3行开始是数据
            },
            '邻居贾总.xlsx': {
                'sheet_name': '贾总',
                'columns': {
                    'product_model': 'C',     # 列C：产品型号
                    'product_name': 'F',      # 列F：产品名称
                    'specification': 'H',    # 列H：规格/KG
                    'price': 'J',            # 列J：单价/元
                    'date': 'A',             # 列A：日期
                    'order_number': 'B'      # 列B：单号
                },
                'header_row': 2,  # 第2行是标题行
                'data_start_row': 3  # 从第3行开始是数据
            },
            '邻居杨总.xlsx': {
                'sheet_name': '杨总提货明细',
                'columns': {
                    'product_model': 'B',     # 列B：产品名称（作为型号）
                    'product_name': 'B',      # 列B：产品名称
                    'specification': 'C',    # 列C：规格/KG
                    'price': 'G',            # 列G：没有价格，用备注作为价格
                    'date': 'A',             # 列A：日期
                    'order_number': 'F'      # 列F：经手人作为订单标识
                },
                'header_row': 2,  # 第2行是标题行
                'data_start_row': 4,  # 从第4行开始是数据
                'allow_empty_model': True,  # 允许产品型号为空
                'special_format': True  # 特殊格式标记
            }
        }
    
    def extract_column_index(self, column_letter: str) -> int:
        """将列字母转换为索引（从0开始）"""
        result = 0
        for char in column_letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1
    
    def process_excel_file(self, file_path: str) -> List[Dict]:
        """
        处理单个Excel文件，提取产品数据
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        file_name = os.path.basename(file_path)
        
        return self._process_file_with_config(file_path, file_name)
    
    def process_excel_file_with_name(self, file_path: str, original_file_name: str) -> List[Dict]:
        """
        处理Excel文件，使用原始文件名进行配置匹配
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        return self._process_file_with_config(file_path, original_file_name)
    
    def _process_file_with_config(self, file_path: str, file_name: str) -> List[Dict]:
        """
        内部方法：使用指定的文件名查找配置并处理文件
        """
        if file_name not in self.file_configs:
            raise ValueError(f"不支持的文件: {file_name}")
        
        config = self.file_configs[file_name]
        sheet_name = config['sheet_name']
        columns = config['columns']
        
        # 使用openpyxl读取文件以获取公式计算结果
        logger.info(f"使用openpyxl读取文件: {file_path}")
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在于文件 {file_name}")
        
        sheet = workbook[sheet_name]
        results = []
        
        # 遍历数据行
        for row_idx in range(config['data_start_row'], sheet.max_row + 1):
            # 检查是否为有效数据行
            product_model_cell = sheet[f"{columns['product_model']}{row_idx}"]
            product_name_cell = sheet[f"{columns['product_name']}{row_idx}"]
            
            product_model = str(product_model_cell.value or '').strip()
            product_name = str(product_name_cell.value or '').strip()
            
            # 如果产品型号为空，用空格代替，避免被映射到其他字段
            if not product_model:
                product_model = ' '  # 用空格代替空型号
            
            # 如果允许空型号（宜榢.xlsx的情况），只要有产品名称就认为有效
            allow_empty_model = config.get('allow_empty_model', False)
            if not allow_empty_model:
                # 原始逻辑：产品型号和产品名称都为空才跳过
                if not product_model.strip() and not product_name:
                    continue
            else:
                # 宜榢.xlsx逻辑：只要有产品名称就认为有效
                if not product_name:
                    continue
            
            # 提取数据
            record = {
                'file_name': file_name,
                'sheet_name': sheet_name,
                'row_number': row_idx,
                'product_model': product_model,
                'product_name': product_name,
                'specification': self._extract_specification(sheet[f"{columns['specification']}{row_idx}"].value),
                'price': self._extract_price(sheet[f"{columns['price']}{row_idx}"].value),
                'date': sheet[f"{columns['date']}{row_idx}"].value,
                'order_number': str(sheet[f"{columns['order_number']}{row_idx}"].value or '').strip()
            }
            
            # 数据验证和清洗
            if record['product_name']:  # 只要有产品名称就添加记录
                results.append(record)
        
        workbook.close()
        return results
    
    def _extract_specification(self, value) -> float:
        """提取规格数值"""
        if value is None:
            return 0.0  # 空值返回0.0，在数据库层面处理为NULL
        
        # 如果是数字类型
        if isinstance(value, (int, float)):
            return float(value)
        
        # 如果是字符串，尝试提取数字
        value_str = str(value).strip()
        if not value_str:  # 空字符串
            return 0.0
        
        if value_str.startswith('='):  # Excel公式
            # 简化处理，返回0，实际使用时会计算公式
            return 0.0
        
        # 使用正则表达式提取数字
        numbers = re.findall(r'\d+(?:\.\d+)?', value_str)
        if numbers:
            return float(numbers[0])
        
        return 0.0
    
    def _extract_price(self, value) -> float:
        """提取价格数值"""
        if value is None:
            return 0.0  # 空值返回0.0，在数据库层面处理为NULL
        
        # 如果是数字类型
        if isinstance(value, (int, float)):
            return float(value)
        
        # 如果是字符串，尝试提取数字
        value_str = str(value).strip()
        if not value_str:  # 空字符串
            return 0.0
        
        if value_str.startswith('='):  # Excel公式
            # 简化处理，返回0，实际使用时会计算公式
            return 0.0
        
        # 使用正则表达式提取数字
        numbers = re.findall(r'\d+(?:\.\d+)?', value_str)
        if numbers:
            return float(numbers[0])
        
        return 0.0
    
    def process_all_files(self, file_paths: List[str]) -> List[Dict]:
        """
        处理多个Excel文件
        """
        all_results = []
        
        for file_path in file_paths:
            try:
                print(f"正在处理文件: {os.path.basename(file_path)}")
                results = self.process_excel_file(file_path)
                all_results.extend(results)
                print(f"成功提取 {len(results)} 条产品记录")
            except Exception as e:
                print(f"处理文件 {file_path} 时出错: {str(e)}")
                continue
        
        return all_results
    
    def convert_to_product_format(self, records: List[Dict]) -> List[Dict]:
        """
        将提取的记录转换为前端产品格式
        """
        products = []
        
        for record in records:
            # 构建产品名称（只使用原始的产品名称，不要拼接型号）
            product_display_name = record['product_name']  # 不拼接product_model
            
            product = {
                'product_name': product_display_name,  # 前端期望的字段名
                'product_model': record['product_model'],  # 产品型号
                'specification': record['specification'],  # 前端期望的字段名
                'price': record['price'],  # 价格
                'description': f"来源: {record['file_name']} - {record['sheet_name']} - 行{record['row_number']}",  # 描述
                'quantity': 0,  # 数量，前端需要这个字段
                'unit_id': None  # 可以后续根据需要设置
            }
            
            # 只有有效的产品名称才添加
            if product['product_name'].strip():
                products.append(product)
        
        return products
    
    def export_to_csv(self, records: List[Dict], output_path: str):
        """导出记录到CSV文件"""
        if not records:
            print("没有记录可导出")
            return
        
        df = pd.DataFrame(records)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"已导出 {len(records)} 条记录到 {output_path}")

def main():
    """测试函数"""
    mapper = IntelligentExcelMapper()
    
    # 文件路径列表
    file_paths = [
        'f:\\guoshi\\七彩乐园.xlsx',
        'f:\\guoshi\\宜榢.xlsx', 
        'f:\\guoshi\\迎扬李总.xlsx'
    ]
    
    # 处理所有文件
    all_records = mapper.process_all_files(file_paths)
    
    print(f"\n总共提取了 {len(all_records)} 条产品记录")
    
    # 转换为产品格式
    products = mapper.convert_to_product_format(all_records)
    
    print(f"转换为产品格式: {len(products)} 个产品")
    
    # 显示前5个产品示例
    print("\n前5个产品示例:")
    for i, product in enumerate(products[:5], 1):
        print(f"{i}. 名称: {product['name']}")
        print(f"   型号: {product['model']}")
        print(f"   规格: {product['spec']}")
        print(f"   价格: {product['price']}")
        print(f"   描述: {product['description'][:50]}...")
        print()
    
    # 导出到CSV
    mapper.export_to_csv(all_records, 'f:\\guoshi\\extracted_products.csv')
    
    return products

if __name__ == "__main__":
    products = main()