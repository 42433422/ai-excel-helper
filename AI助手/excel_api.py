#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件上传和分析API
"""

import os
import uuid
import logging
from datetime import datetime
from flask import Blueprint, request, jsonify, current_app
from werkzeug.utils import secure_filename
import openpyxl

# 设置日志
logger = logging.getLogger(__name__)

# 创建Blueprint
excel_bp = Blueprint('excel_api', __name__)

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def analyze_excel_structure(file_path):
    """分析Excel文件结构"""
    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        
        # 获取基本信息
        workbook_name = os.path.basename(file_path)
        sheets = workbook.sheetnames
        
        analysis_result = {
            'workbook_name': workbook_name,
            'sheets': sheets,
            'analysis_time': datetime.now().isoformat()
        }
        
        # 分析主要工作表（通常是第一个）
        main_sheet_name = sheets[0] if sheets else None
        if main_sheet_name:
            worksheet = workbook[main_sheet_name]
            
            # 获取表头
            headers = []
            for col in range(1, min(21, worksheet.max_column + 1)):  # 最多分析20列
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value))
            
            analysis_result['headers'] = headers
            
            # 计算数据行数
            total_rows = 0
            for row in range(2, worksheet.max_row + 1):
                has_data = False
                for col in range(1, min(21, worksheet.max_column + 1)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip():
                        has_data = True
                        break
                if has_data:
                    total_rows += 1
            
            analysis_result['total_rows'] = total_rows
            
            # 智能列映射
            column_mapping = analyze_column_mapping(headers)
            analysis_result['column_mapping'] = column_mapping
            
            # 计算置信度
            confidence = calculate_mapping_confidence(column_mapping)
            analysis_result['confidence'] = confidence
            
        workbook.close()
        
        return analysis_result
        
    except Exception as e:
        logger.error(f"分析Excel文件失败: {e}")
        raise

def analyze_column_mapping(headers):
    """智能分析列映射"""
    mapping = {}
    
    # 常见的列名模式
    patterns = {
        'customer_column': ['客户', '买方', '购买单位', '采购单位'],
        'order_number_column': ['订单号', '单号', '编号', '订单'],
        'product_column': ['产品', '商品', '名称', '品名'],
        'quantity_column': ['数量', '数量(kg)', 'kg', '重量'],
        'unit_price_column': ['单价', '价格', '价格(元)', '元'],
        'amount_column': ['金额', '总计', '总价'],
        'date_column': ['日期', '时间'],
        'remarks_column': ['备注', '说明', '备注信息']
    }
    
    # 尝试匹配每个列
    for col_name, col_idx in enumerate(headers, 1):
        col_value = str(col_idx).strip()
        
        for target, keywords in patterns.items():
            for keyword in keywords:
                if keyword in col_value:
                    mapping[target] = col_idx
                    break
    
    return mapping

def calculate_mapping_confidence(mapping):
    """计算映射置信度"""
    if not mapping:
        return 0.0
    
    # 基础置信度
    confidence = 0.3
    
    # 根据映射的列数增加置信度
    confidence += len(mapping) * 0.1
    
    # 重要列的额外加分
    important_cols = ['customer_column', 'product_column', 'quantity_column']
    for col in important_cols:
        if col in mapping:
            confidence += 0.15
    
    # 确保置信度不超过1.0
    return min(confidence, 1.0)

@excel_bp.route('/upload', methods=['POST'])
def upload_excel():
    """上传并分析Excel文件"""
    try:
        # 检查是否有文件
        if 'excel_file' not in request.files:
            return jsonify({
                'success': False,
                'error': '未找到Excel文件'
            }), 400
        
        file = request.files['excel_file']
        
        # 检查文件名
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': '未选择文件'
            }), 400
        
        # 检查文件扩展名
        if not allowed_file(file.filename):
            return jsonify({
                'success': False,
                'error': '不支持的文件格式，请上传.xlsx或.xls文件'
            }), 400
        
        # 生成安全的文件名
        original_filename = secure_filename(file.filename)
        file_extension = original_filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{uuid.uuid4().hex}.{file_extension}"
        
        # 保存文件
        upload_dir = current_app.config.get('UPLOAD_FOLDER', 'uploads')
        os.makedirs(upload_dir, exist_ok=True)
        
        file_path = os.path.join(upload_dir, unique_filename)
        file.save(file_path)
        
        logger.info(f"Excel文件已保存: {file_path}")
        
        # 分析文件结构
        analysis_result = analyze_excel_structure(file_path)
        
        return jsonify({
            'success': True,
            'file_path': file_path,
            'file_name': original_filename,
            'analysis': analysis_result
        })
        
    except Exception as e:
        logger.error(f"Excel文件上传失败: {e}")
        return jsonify({
            'success': False,
            'error': f'文件上传失败: {str(e)}'
        }), 500

@excel_bp.route('/analyze', methods=['POST'])
def analyze_excel():
    """分析已上传的Excel文件"""
    try:
        data = request.get_json()
        file_path = data.get('file_path')
        
        if not file_path or not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'error': '文件路径无效或文件不存在'
            }), 400
        
        # 分析文件结构
        analysis_result = analyze_excel_structure(file_path)
        
        return jsonify({
            'success': True,
            'analysis': analysis_result
        })
        
    except Exception as e:
        logger.error(f"Excel文件分析失败: {e}")
        return jsonify({
            'success': False,
            'error': f'文件分析失败: {str(e)}'
        }), 500

@excel_bp.route('/validate', methods=['POST'])
def validate_excel_structure():
    """验证Excel文件是否适用于出货记录"""
    try:
        data = request.get_json()
        file_path = data.get('file_path')
        
        if not file_path or not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'error': '文件路径无效或文件不存在'
            }), 400
        
        # 加载工作簿进行验证
        workbook = openpyxl.load_workbook(file_path)
        main_sheet_name = workbook.sheetnames[0] if workbook.sheetnames else None
        
        validation_result = {
            'is_valid': False,
            'issues': [],
            'recommendations': []
        }
        
        if not main_sheet_name:
            validation_result['issues'].append('未找到工作表')
            return jsonify(validation_result)
        
        worksheet = workbook[main_sheet_name]
        
        # 检查表头
        headers = []
        for col in range(1, min(21, worksheet.max_column + 1)):
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value))
        
        if not headers:
            validation_result['issues'].append('未找到表头行')
            return jsonify(validation_result)
        
        # 检查是否包含基本列
        essential_columns = ['客户', '产品', '数量']
        found_essential = 0
        
        for essential in essential_columns:
            found = any(essential in header for header in headers)
            if found:
                found_essential += 1
            else:
                validation_result['issues'].append(f'缺少必需列: {essential}')
        
        # 检查数据行数
        data_rows = 0
        for row in range(2, worksheet.max_row + 1):
            has_data = False
            for col in range(1, min(21, worksheet.max_column + 1)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                    break
            if has_data:
                data_rows += 1
        
        if data_rows == 0:
            validation_result['issues'].append('未找到数据行')
        
        # 判断是否有效
        validation_result['is_valid'] = found_essential >= 2 and data_rows > 0
        
        if validation_result['is_valid']:
            validation_result['recommendations'].append('Excel文件结构良好，可以用于出货记录同步')
        else:
            validation_result['recommendations'].append('请检查Excel文件是否包含客户、产品、数量等基本信息')
        
        workbook.close()
        
        return jsonify(validation_result)
        
    except Exception as e:
        logger.error(f"Excel文件验证失败: {e}")
        return jsonify({
            'success': False,
            'error': f'文件验证失败: {str(e)}'
        }), 500

@excel_bp.route('/sync', methods=['POST'])
def excel_auto_sync():
    """执行Excel自动同步"""
    try:
        data = request.get_json()
        file_path = data.get('file_path')
        analysis = data.get('analysis')
        shipment_data = data.get('shipmentData')
        
        # 验证必需参数
        if not file_path or not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'error': '文件路径无效或文件不存在'
            }), 400
            
        if not shipment_data:
            return jsonify({
                'success': False,
                'error': '缺少发货单数据'
            }), 400
        
        # 导入出货记录同步模块
        from shipment_excel_sync import ShipmentRecordSyncManager, create_from_shipment_document
        
        # 初始化Excel同步管理器
        sync_manager = ShipmentRecordSyncManager(file_path)
        
        # 如果有分析结果，则应用列映射
        if analysis and analysis.get('column_mapping'):
            # 这里可以根据分析结果配置列映射
            logger.info("应用AI分析结果的列映射")
        
        # 将前端数据转换为ShipmentRecord格式
        try:
            # 创建发货记录对象
            record = create_shipment_record_from_data(shipment_data)
            
            # 执行同步
            success = sync_manager.insert_shipment_record(record)
            
            if success:
                # 计算同步结果
                products_count = len(getattr(record, 'products', [record]))
                
                return jsonify({
                    'success': True,
                    'excel_sync': {
                        'success': True,
                        'rows_written': products_count,
                        'start_row': sync_manager._last_inserted_row if hasattr(sync_manager, '_last_inserted_row') else None,
                        'products_count': products_count,
                        'details': {
                            'products': [{
                                'name': record.product_name,
                                'quantity': record.quantity,
                                'unit': '桶'
                            }] if not hasattr(record, 'products') else [
                                {
                                    'name': p.product_name if hasattr(p, 'product_name') else p.product_name,
                                    'quantity': p.quantity if hasattr(p, 'quantity') else record.quantity,
                                    'unit': '桶'
                                } for p in getattr(record, 'products', [])
                            ]
                        }
                    }
                })
            else:
                return jsonify({
                    'success': False,
                    'excel_sync': {
                        'success': False,
                        'error': 'Excel同步执行失败'
                    }
                })
                
        except Exception as e:
            logger.error(f"执行Excel同步失败: {e}")
            return jsonify({
                'success': False,
                'excel_sync': {
                    'success': False,
                    'error': f'Excel同步失败: {str(e)}'
                }
            })
            
    except Exception as e:
        logger.error(f"Excel自动同步API失败: {e}")
        return jsonify({
            'success': False,
            'error': f'自动同步失败: {str(e)}'
        }), 500

def create_shipment_record_from_data(shipment_data):
    """将前端数据转换为ShipmentRecord对象"""
    from shipment_excel_sync import ShipmentRecord
    
    record = ShipmentRecord()
    record.customer_name = shipment_data.get('customerName', '')
    record.order_number = shipment_data.get('orderNumber', '')
    record.product_name = shipment_data.get('products', [{}])[0].get('name', '') if shipment_data.get('products') else ''
    record.quantity = shipment_data.get('products', [{}])[0].get('quantity', 0) if shipment_data.get('products') else 0
    record.unit_price = shipment_data.get('products', [{}])[0].get('unitPrice', 0) if shipment_data.get('products') else 0
    record.amount = record.quantity * record.unit_price
    record.date = datetime.now().strftime("%Y-%m-%d")
    record.remarks = f"发货单: {record.order_number}"
    
    # 处理多产品情况
    products = shipment_data.get('products', [])
    if len(products) > 1:
        record.products = []
        for product in products:
            product_record = ShipmentRecord()
            product_record.customer_name = record.customer_name
            product_record.order_number = record.order_number
            product_record.product_name = product.get('name', '')
            product_record.quantity = product.get('quantity', 0)
            product_record.unit_price = product.get('unitPrice', 0)
            product_record.amount = product_record.quantity * product_record.unit_price
            product_record.date = record.date
            product_record.remarks = record.remarks
            record.products.append(product_record)
    
    return record
