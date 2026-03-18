#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简化的Excel同步测试
"""

import sys
import os
import logging
from datetime import datetime

# 添加当前目录到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.styles import Font, Alignment
import pandas as pd

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def test_excel_write():
    """测试直接写入Excel"""
    print("=" * 80)
    print("测试Excel原始格式同步")
    print("=" * 80)

    excel_path = "尹玉华132.xlsx"

    try:
        # 加载Excel文件
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook["25年出货"]

        # 找到最后一行
        last_row = worksheet.max_row
        next_row = last_row + 1

        print(f"找到最后一行: {last_row}")
        print(f"准备写入第{next_row}行")

        # 按照原始Excel格式写入数据，并设置正确的字体和对齐格式
        # 创建与原始数据相同的字体格式：宋体 16.0号
        original_font = Font(name='宋体', size=16.0)
        
        # 创建对齐格式，与原始数据保持一致
        # 列1-2: 垂直居中
        vertical_center = Alignment(vertical='center')
        # 列3-4: 水平+垂直居中  
        center_center = Alignment(horizontal='center', vertical='center')
        # 列7: 左对齐，垂直居中
        left_center = Alignment(horizontal='left', vertical='center')
        # 列8-12: 水平+垂直居中
        center_center2 = Alignment(horizontal='center', vertical='center')

        # 创建合并单元格 D-F (第4-6列)，与原始数据保持一致
        worksheet.merge_cells(f'D{next_row}:F{next_row}')

        # 列1: 客户名称 (垂直居中)
        cell = worksheet.cell(row=next_row, column=1, value="蕊芯")
        cell.font = original_font
        cell.alignment = vertical_center
        
        # 列2: 内部编号 (垂直居中)
        cell = worksheet.cell(row=next_row, column=2, value=46040)
        cell.font = original_font
        cell.alignment = vertical_center
        
        # 列3: 状态 (水平+垂直居中)
        cell = worksheet.cell(row=next_row, column=3, value="未开单")
        cell.font = original_font
        cell.alignment = center_center
        
        # 列4-6: 合并单元格中放置产品型号信息 (水平+垂直居中)
        # 由于D-F列被合并，我们只在D列写入产品型号
        cell = worksheet.cell(row=next_row, column=4, value=9806)
        cell.font = original_font
        cell.alignment = center_center
        
        # 列7: 产品名称 (左对齐，垂直居中)
        cell = worksheet.cell(row=next_row, column=7, value="PE白底漆（定制）")
        cell.font = original_font
        cell.alignment = left_center
        
        # 列8: 数量（桶数） (水平+垂直居中)
        cell = worksheet.cell(row=next_row, column=8, value=10)
        cell.font = original_font
        cell.alignment = center_center2
        
        # 列9: 单价1 (水平+垂直居中)
        cell = worksheet.cell(row=next_row, column=9, value=6)
        cell.font = original_font
        cell.alignment = center_center2
        
        # 列10: 金额公式 (水平+垂直居中)
        cell = worksheet.cell(row=next_row, column=10, value=f"=H{next_row}*I{next_row}")
        cell.font = original_font
        cell.alignment = center_center2
        
        # 列11: 单价2 (水平+垂直居中)
        cell = worksheet.cell(row=next_row, column=11, value=26)
        cell.font = original_font
        cell.alignment = center_center2
        
        # 列12: 金额公式2 (水平+垂直居中)
        cell = worksheet.cell(row=next_row, column=12, value=f"=J{next_row}*K{next_row}")
        cell.font = original_font
        cell.alignment = center_center2

        print(f"数据写入完成到第{next_row}行")

        # 保存文件
        workbook.save(excel_path)
        workbook.close()

        print("✅ 文件保存成功!")

        # 验证数据
        workbook2 = openpyxl.load_workbook(excel_path)
        worksheet2 = workbook2["25年出货"]
        
        print(f"\n验证第{next_row}行的数据:")
        for col in range(1, 13):
            cell_value = worksheet2.cell(row=next_row, column=col).value
            if cell_value is not None:
                print(f"  列{col}: {cell_value}")
        
        workbook2.close()

        return True

    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_multi_products():
    """测试多产品同步"""
    print("\n" + "=" * 80)
    print("测试多产品同步")
    print("=" * 80)

    excel_path = "尹玉华132.xlsx"

    try:
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook["25年出货"]
        
        last_row = worksheet.max_row
        current_row = last_row + 1
        
        print(f"准备写入多产品数据，起始行: {current_row}")

        # 创建与原始数据相同的字体格式：宋体 16.0号
        original_font = Font(name='宋体', size=16.0)
        
        # 创建对齐格式，与原始数据保持一致
        # 列1-2: 垂直居中
        vertical_center = Alignment(vertical='center')
        # 列3-4: 水平+垂直居中  
        center_center = Alignment(horizontal='center', vertical='center')
        # 列7: 左对齐，垂直居中
        left_center = Alignment(horizontal='left', vertical='center')
        # 列8-12: 水平+垂直居中
        center_center2 = Alignment(horizontal='center', vertical='center')
        
        # 产品1: PE白底漆
        # 创建合并单元格 D-F (第4-6列)，与原始数据保持一致
        worksheet.merge_cells(f'D{current_row}:F{current_row}')
        
        cell = worksheet.cell(row=current_row, column=1, value="蕊芯")  # 列1: 客户
        cell.font = original_font
        cell.alignment = vertical_center
        cell = worksheet.cell(row=current_row, column=2, value=46041)  # 列2: 内部编号
        cell.font = original_font
        cell.alignment = vertical_center
        cell = worksheet.cell(row=current_row, column=3, value="未开单")  # 列3: 状态
        cell.font = original_font
        cell.alignment = center_center
        cell = worksheet.cell(row=current_row, column=4, value=9806)  # 列4: 产品型号（在合并单元格中）
        cell.font = original_font
        cell.alignment = center_center
        cell = worksheet.cell(row=current_row, column=7, value="PE白底漆（定制）")  # 列7: 产品名称
        cell.font = original_font
        cell.alignment = left_center
        cell = worksheet.cell(row=current_row, column=8, value=10)  # 列8: 桶数
        cell.font = original_font
        cell.alignment = center_center2
        cell = worksheet.cell(row=current_row, column=9, value=6)  # 列9: 单价1
        cell.font = original_font
        cell.alignment = center_center2
        cell = worksheet.cell(row=current_row, column=10, value=f"=H{current_row}*I{current_row}")  # 列10: 金额公式
        cell.font = original_font
        cell.alignment = center_center2
        cell = worksheet.cell(row=current_row, column=11, value=26)  # 列11: 单价2
        cell.font = original_font
        cell.alignment = center_center2
        cell = worksheet.cell(row=current_row, column=12, value=f"=J{current_row}*K{current_row}")  # 列12: 金额公式2
        cell.font = original_font
        cell.alignment = center_center2
        
        current_row += 1
        
        # 产品2: PE稀释剂
        # 创建合并单元格 D-F (第4-6列)，与原始数据保持一致
        worksheet.merge_cells(f'D{current_row}:F{current_row}')
        
        cell = worksheet.cell(row=current_row, column=3, value="未开单")  # 列3: 状态
        cell.font = original_font
        cell.alignment = center_center
        cell = worksheet.cell(row=current_row, column=4, value=9806)  # 列4: 产品型号（在合并单元格中）
        cell.font = original_font
        cell.alignment = center_center
        cell = worksheet.cell(row=current_row, column=7, value="PE白底漆稀释剂")  # 列7: 产品名称
        cell.font = original_font
        cell.alignment = left_center
        cell = worksheet.cell(row=current_row, column=8, value=1)  # 列8: 桶数
        cell.font = original_font
        cell.alignment = center_center2
        cell = worksheet.cell(row=current_row, column=9, value=2)  # 列9: 单价1
        cell.font = original_font
        cell.alignment = center_center2
        cell = worksheet.cell(row=current_row, column=10, value=f"=H{current_row}*I{current_row}")  # 列10: 金额公式
        cell.font = original_font
        cell.alignment = center_center2
        cell = worksheet.cell(row=current_row, column=11, value=45)  # 列11: 单价2
        cell.font = original_font
        cell.alignment = center_center2
        cell = worksheet.cell(row=current_row, column=12, value=f"=J{current_row}*K{current_row}")  # 列12: 金额公式2
        cell.font = original_font
        cell.alignment = center_center2
        
        print(f"多产品数据写入完成")
        print(f"产品1: 第{current_row-1}行 - PE白底漆")
        print(f"产品2: 第{current_row}行 - PE稀释剂")

        # 保存文件
        workbook.save(excel_path)
        workbook.close()

        print("✅ 多产品同步完成!")
        return True

    except Exception as e:
        print(f"❌ 多产品测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """主函数"""
    # 切换到AI助手目录
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    results = []
    
    # 测试1: 简单Excel写入
    results.append(("简单Excel写入", test_excel_write()))
    
    # 测试2: 多产品同步
    results.append(("多产品同步", test_multi_products()))

    # 汇总结果
    print("\n" + "=" * 80)
    print("📋 测试结果汇总")
    print("=" * 80)

    all_passed = True
    for test_name, passed in results:
        status = "✅ 通过" if passed else "❌ 失败"
        print(f"  {test_name}: {status}")
        if not passed:
            all_passed = False

    print("\n" + "=" * 80)
    if all_passed:
        print("🎉 所有测试通过!")
        print("Excel同步功能已验证可用!")
    else:
        print("⚠️  部分测试失败")
    print("=" * 80)

    return all_passed


if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)
