#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
固定模板出货记录同步器
简化版：直接按固定列布局写入出货记录
"""

import os
import openpyxl
from datetime import datetime
from typing import Dict, List, Optional
import logging

# 配置日志编码
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', encoding='utf-8')
logger = logging.getLogger(__name__)


class FixedTemplateSyncManager:
    """固定模板出货记录同步管理器"""
    
    # 七彩乐园模板固定列布局
    TEMPLATE_COLUMNS = {
        'date': 1,           # A: 日期
        'order_no': 2,       # B: 单号
        'model': 3,          # C: 产品型号
        'product': 6,        # F: 产品名称
        'quantity_tins': 7,  # G: 数量/件
        'spec': 8,           # H: 规格/KG
        'quantity_kg': 9,    # I: 数量/KG (公式)
        'unit_price': 10,    # J: 单价/元
        'amount': 11,        # K: 金额/元 (公式)
    }
    
    def __init__(self, excel_path: str = None, worksheet_name: str = "25出货"):
        """初始化"""
        self.excel_path = excel_path
        self.worksheet_name = worksheet_name
        self.workbook = None
        self.worksheet = None
        
        if excel_path:
            self.load_excel(excel_path)
    
    def load_excel(self, excel_path: str):
        """加载Excel文件"""
        self.excel_path = excel_path
        if os.path.exists(excel_path):
            self.workbook = openpyxl.load_workbook(excel_path)
            # 使用指定的工作表名
            if self.worksheet_name in self.workbook.sheetnames:
                self.worksheet = self.workbook[self.worksheet_name]
                logger.info(f"已加载Excel: {excel_path}, 工作表: {self.worksheet_name}")
            else:
                logger.error(f"未找到工作表: {self.worksheet_name}")
                self.worksheet = self.workbook.active
        else:
            logger.error(f"文件不存在: {excel_path}")
    
    def load_excel_by_unit(self, unit_name: str) -> bool:
        """
        根据购买单位名称加载对应的出货记录文件
        
        查找路径: 发货单/{unit_name}.xlsx
        
        Args:
            unit_name: 购买单位名称
        
        Returns:
            是否成功加载
        """
        # 清理单位名称，移除非法字符
        import re
        safe_unit_name = re.sub(r'[<>:"/\\|?*]', '', unit_name)
        safe_unit_name = safe_unit_name.strip()
        
        # 构建文件路径 - 优先查找"出货记录"文件夹
        base_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 首先在"出货记录"文件夹中查找（上级目录）
        shipment_dir = os.path.join(base_dir, '..', '出货记录')
        found_path = None
        
        logger.info(f"正在查找出货记录文件...")
        logger.info(f"查找目录: {shipment_dir}")
        
        # 检查"出货记录"文件夹中的客户子文件夹
        if os.path.exists(shipment_dir):
            logger.info(f"检查出货记录文件夹: {shipment_dir}")
            
            # 收集所有候选匹配
            exact_matches = []
            partial_matches = []
            
            # 查找匹配的子文件夹
            for item in os.listdir(shipment_dir):
                item_path = os.path.join(shipment_dir, item)
                if os.path.isdir(item_path):
                    # 清理文件夹名称
                    clean_folder_name = re.sub(r'[^\w\s]', '', item)
                    clean_folder_name = clean_folder_name.strip()
                    
                    # 精确匹配（优先）
                    if safe_unit_name == clean_folder_name:
                        exact_matches.append((item, item_path))
                        logger.info(f"找到精确匹配: {item}")
                    
                    # 部分匹配
                    elif safe_unit_name in clean_folder_name or clean_folder_name in safe_unit_name:
                        partial_matches.append((item, item_path))
                        logger.info(f"找到部分匹配: {item}")
            
            # 选择最佳匹配：优先精确匹配，否则选择最相似的部分匹配
            best_match = None
            if exact_matches:
                best_match = exact_matches[0]
                logger.info(f"使用精确匹配: {best_match[0]}")
            elif partial_matches:
                # 优先选择包含完整单位名称的匹配
                for item, item_path in partial_matches:
                    clean_name = re.sub(r'[^\w\s]', '', item)
                    if safe_unit_name in clean_name:
                        best_match = (item, item_path)
                        logger.info(f"使用优先部分匹配: {item}")
                        break
                
                if not best_match:
                    best_match = partial_matches[0]
                    logger.info(f"使用默认部分匹配: {best_match[0]}")
            
            # 如果找到最佳匹配，查找Excel文件
            if best_match:
                item, item_path = best_match
                logger.info(f"找到匹配的文件夹: {item}")
                
                for file in os.listdir(item_path):
                    if file.endswith('.xlsx') and file != '出货记录模板.xlsx':
                        found_path = os.path.join(item_path, file)
                        logger.info(f"在出货记录文件夹中找到匹配: {found_path}")
                        break
        else:
            logger.info(f"出货记录文件夹不存在: {shipment_dir}")
        
        # 如果在"出货记录"中找不到，再查找"发货单"文件夹（向后兼容）
        if not found_path:
            logger.info(f"在出货记录中未找到，查找发货单文件夹...")
            legacy_shipment_dir = os.path.join(base_dir, '发货单')
            
            if os.path.exists(legacy_shipment_dir):
                # 尝试其他可能的文件名格式
                possible_names = [
                    f"{safe_unit_name}.xlsx",
                    f"{safe_unit_name}出货记录.xlsx",
                    f"{safe_unit_name}出货.xlsx"
                ]
                
                for name in possible_names:
                    test_path = os.path.join(legacy_shipment_dir, name)
                    if os.path.exists(test_path):
                        found_path = test_path
                        logger.info(f"在发货单文件夹中找到: {found_path}")
                        break
                
                # 如果都找不到，检查发货单文件夹下的所有文件，尝试模糊匹配
                if not found_path:
                    logger.info(f"正在查找与 {safe_unit_name} 相关的出货记录文件...")
                    for filename in os.listdir(legacy_shipment_dir):
                        if filename.endswith('.xlsx'):
                            # 移除扩展名进行匹配
                            name_without_ext = filename.replace('.xlsx', '')
                            name_without_ext = re.sub(r'[^\w\s]', '', name_without_ext)  # 移除非字母数字字符
                            if safe_unit_name in name_without_ext or name_without_ext in safe_unit_name:
                                found_path = os.path.join(legacy_shipment_dir, filename)
                                logger.info(f"找到匹配的出货记录文件: {filename}")
                                break
            else:
                logger.info(f"发货单文件夹不存在: {legacy_shipment_dir}")
        
        if not found_path:
            logger.info(f"未找到购买单位 {unit_name} 的出货记录文件")
            # 显示调试信息
            logger.info(f"查找路径: {shipment_dir}")
            if os.path.exists(shipment_dir):
                available_files = [f for f in os.listdir(shipment_dir) if f.endswith('.xlsx')]
                logger.info(f"出货记录文件夹中的文件: {available_files}")
            else:
                logger.info(f"出货记录文件夹不存在")
            
            # 尝试查找"发货单"文件夹
            legacy_shipment_dir = os.path.join(base_dir, '发货单')
            if os.path.exists(legacy_shipment_dir):
                available_files = [f for f in os.listdir(legacy_shipment_dir) if f.endswith('.xlsx')]
                logger.info(f"发货单文件夹中的文件: {available_files}")
            else:
                logger.info(f"发货单文件夹不存在")
            return False
        
        # 加载文件
        self.excel_path = found_path
        self.workbook = openpyxl.load_workbook(found_path)
        if self.worksheet_name in self.workbook.sheetnames:
            self.worksheet = self.workbook[self.worksheet_name]
            logger.info(f"已加载 {unit_name} 的出货记录: {found_path}")
            return True
        else:
            logger.error(f"未找到工作表: {self.worksheet_name}")
            # 尝试使用第一个工作表
            if self.workbook.sheetnames:
                self.worksheet = self.workbook.active
                logger.info(f"使用默认工作表: {self.worksheet.title}")
                return True
            return False
    
    def find_last_data_row(self) -> int:
        """找到最后有数据的行"""
        if not self.worksheet:
            raise ValueError("请先加载Excel文件")
        
        # 从一个较大的行号开始向前查找（避免max_row不更新的问题）
        # 或者使用max_row，但跳过完全为空的行
        for row in range(2000, 0, -1):
            cell_value = self.worksheet.cell(row=row, column=1).value
            if cell_value is not None and str(cell_value).strip():
                return row
        
        return 1
    
    def insert_records(self, order_no: str, order_date: str, products: List[Dict], unit_price_map: Dict[str, float] = None) -> bool:
        """
        插入出货记录
        
        Args:
            order_no: 订单编号
            order_date: 订单日期 (如 "2024-01-15")
            products: 产品列表 [{'name': 'PE白底漆', 'model': '9803', 'quantity_tins': 10, 'spec': 28}]
            unit_price_map: 产品单价映射 {'PE白底漆': 10, 'PU稀释剂': 11}
        
        Returns:
            是否成功
        """
        if not self.worksheet:
            raise ValueError("请先加载Excel文件")
        
        try:
            # 每次写入前重新加载文件，确保获取最新数据
            import os
            if os.path.exists(self.excel_path):
                self.workbook = openpyxl.load_workbook(self.excel_path)
                if self.worksheet_name in self.workbook.sheetnames:
                    self.worksheet = self.workbook[self.worksheet_name]
            
            # 转换日期格式为中文（无年），如 "1月30日"
            from datetime import datetime
            try:
                parsed_date = datetime.strptime(order_date, '%Y-%m-%d')
                chinese_date = f"{parsed_date.month}月{parsed_date.day}日"
            except:
                chinese_date = order_date
            
            # 提取发货单号后三位（去掉前导0，加"号"），如 "26-0100071A" -> "71号"
            import re
            order_short_no = ""
            if order_no and order_no.startswith('26-'):
                seq = order_no[4:]  # 去掉 '26-'
                num_part = re.sub(r'[A-Z]$', '', seq)  # 去掉末尾字母
                if len(num_part) >= 5:
                    num5 = num_part[-5:]  # 取后5位数字
                    order_short_no = str(int(num5)) + "号"
            
            # 找到最后一行
            last_row = self.find_last_data_row()
            start_row = last_row + 1  # 从下一行开始写入
            
            logger.info(f"最后数据行: {last_row}, 开始写入行: {start_row}")
            
            # 写入每个产品
            for i, product in enumerate(products):
                current_row = start_row + i
                product_name = product.get('name', '')
                model_number = product.get('model', '')
                # 如果型号为空，从产品名称中提取或使用默认值
                if not model_number:
                    model_number = product.get('model_number', '')
                quantity_tins = product.get('quantity_tins', 0)
                tin_spec = product.get('tin_spec', 10)
                quantity_kg = quantity_tins * tin_spec
                
                # 获取单价
                if unit_price_map and product_name in unit_price_map:
                    unit_price = unit_price_map[product_name]
                else:
                    unit_price = product.get('unit_price', 0)
                
                # 计算金额
                amount = round(quantity_kg * unit_price, 2)
                
                # 写入数据
                # A列: 中文日期（无年），如 "1月30日"
                self.worksheet.cell(row=current_row, column=self.TEMPLATE_COLUMNS['date'], value=chinese_date)
                # B列: 短单号（发货单后三位数字），如 "26-0100071A" -> "71"
                self.worksheet.cell(row=current_row, column=self.TEMPLATE_COLUMNS['order_no'], value=order_short_no)
                # C列: 型号
                self.worksheet.cell(row=current_row, column=self.TEMPLATE_COLUMNS['model'], value=model_number)
                self.worksheet.cell(row=current_row, column=self.TEMPLATE_COLUMNS['product'], value=product_name)
                self.worksheet.cell(row=current_row, column=self.TEMPLATE_COLUMNS['quantity_tins'], value=quantity_tins)
                self.worksheet.cell(row=current_row, column=self.TEMPLATE_COLUMNS['spec'], value=tin_spec)
                
                # 数量KG使用公式
                self.worksheet.cell(row=current_row, column=self.TEMPLATE_COLUMNS['quantity_kg'], 
                                   value=f"=G{current_row}*H{current_row}")
                
                # 单价
                self.worksheet.cell(row=current_row, column=self.TEMPLATE_COLUMNS['unit_price'], value=unit_price)
                
                # 金额使用公式
                self.worksheet.cell(row=current_row, column=self.TEMPLATE_COLUMNS['amount'], 
                                   value=f"=I{current_row}*J{current_row}")
                
                logger.info(f"写入第{current_row}行: {product_name} x {quantity_tins}桶")
            
            # 计算总金额
            total_amount = 0
            for i, product in enumerate(products):
                quantity_tins = product.get('quantity_tins', 0)
                tin_spec = product.get('tin_spec', 10)
                quantity_kg = quantity_tins * tin_spec
                product_name = product.get('name', '')
                
                if unit_price_map and product_name in unit_price_map:
                    unit_price = unit_price_map[product_name]
                else:
                    unit_price = product.get('unit_price', 0)
                
                total_amount += round(quantity_kg * unit_price, 2)
            
            logger.info(f"总金额: {total_amount}")
            
            # 在最后一个产品行添加金额合计（放在金额K列后面）
            if products:
                last_row = start_row + len(products) - 1  # 最后一行产品
                # L列（第12列）写入合计金额
                self.worksheet.cell(row=last_row, column=12, value=total_amount)
                logger.info(f"在第{last_row}行L列写入合计金额: {total_amount}")
            
            # 保存文件
            logger.info(f"保存文件: {self.excel_path}")
            self.workbook.save(self.excel_path)
            logger.info(f"文件保存完成")
            
            # 验证文件是否更新
            import os
            file_mtime = os.path.getmtime(self.excel_path)
            logger.info(f"文件修改时间: {file_mtime}")
            
            logger.info(f"成功写入 {len(products)} 条记录到Excel，合计金额: {total_amount}")
            return True
            
        except Exception as e:
            logger.error(f"写入出货记录失败: {e}")
            return False
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            try:
                self.workbook.close()
            except:
                pass


def sync_from_parsed_order(excel_path: str, order_no: str, order_date: str, products: List[Dict], 
                           unit_price_map: Dict[str, float] = None) -> bool:
    """
    从解析的订单同步到出货记录
    
    Args:
        excel_path: Excel文件路径
        order_no: 订单编号
        order_date: 订单日期
        products: 产品列表
        unit_price_map: 产品单价映射
    
    Returns:
        是否成功
    """
    manager = FixedTemplateSyncManager(excel_path)
    return manager.insert_records(order_no, order_date, products, unit_price_map)


if __name__ == "__main__":
    # 测试
    manager = FixedTemplateSyncManager(r"c:\Users\97088\Desktop\新建文件夹 (4)\AI助手\templates\七彩乐园.xlsx")
    
    test_products = [
        {'name': 'PE白底漆', 'model': '9803', 'quantity_tins': 10, 'spec': 28},
        {'name': 'PU净味哑光浅蓝色漆', 'model': 'ST_七彩_001', 'quantity_tins': 1, 'spec': 20},
        {'name': 'PE白底漆稀料', 'model': '608A', 'quantity_tins': 1, 'spec': 180},
    ]
    
    unit_prices = {
        'PE白底漆': 10,
        'PU净味哑光浅蓝色漆': 22,
        'PE白底漆稀料': 11,
    }
    
    success = manager.insert_records("20260130", "2026-01-30", test_products, unit_prices)
    print(f"同步结果: {'成功' if success else '失败'}")
