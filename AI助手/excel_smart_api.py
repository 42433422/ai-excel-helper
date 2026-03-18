"""
Excel智能解析API增强模块
扩展现有的upload_xlsx端点，支持智能字段映射
"""

import os
import openpyxl
import pandas as pd
from flask import jsonify, request
import logging
import json

# 导入我们的智能映射模块
from intelligent_excel_mapper import IntelligentExcelMapper

logger = logging.getLogger(__name__)

class ExcelSmartParser:
    """Excel智能解析器"""
    
    def __init__(self):
        self.mapper = IntelligentExcelMapper()
        # 支持的Excel文件名列表
        self.supported_files = [
            '七彩乐园.xlsx',
            '宜榢.xlsx', 
            '迎扬李总.xlsx'
        ]
    
    def is_supported_file(self, filename):
        """检查是否为支持的Excel文件"""
        return filename in self.supported_files
    
    def smart_parse_excel(self, file_path, file):
        """智能解析Excel文件"""
        filename = file.filename
        
        logger.info(f"智能解析尝试: filename={filename}")
        logger.info(f"supported_files: {self.supported_files}")
        logger.info(f"是否支持文件: {self.is_supported_file(filename)}")
        
        if self.is_supported_file(filename):
            # 使用智能映射器解析
            try:
                # 将上传的文件保存到临时位置，确保文件完整保存
                import tempfile
                import uuid
                # 使用UUID确保文件名唯一，避免冲突
                unique_id = uuid.uuid4().hex[:8]
                temp_file_path = os.path.join(tempfile.gettempdir(), f"smart_parser_{unique_id}_{filename}")
                logger.info(f"保存临时文件: {temp_file_path}")
                
                # 使用完整的文件内容保存
                file.stream.seek(0)  # 重置流位置
                file.save(temp_file_path)
                
                # 验证临时文件是否存在且有效
                logger.info(f"临时文件存在: {os.path.exists(temp_file_path)}")
                if os.path.exists(temp_file_path):
                    file_size = os.path.getsize(temp_file_path)
                    logger.info(f"临时文件大小: {file_size} bytes")
                    if file_size < 100:  # 如果文件太小，说明保存失败
                        logger.error(f"临时文件太小，可能是保存失败: {file_size} bytes")
                        raise ValueError(f"文件保存失败，文件大小异常: {file_size} bytes")
                
                logger.info(f"开始智能解析: temp_file_path={temp_file_path}, filename={filename}")
                
                # 使用智能映射器处理（传入原始文件名）
                records = self.mapper.process_excel_file_with_name(temp_file_path, filename)
                products = self.mapper.convert_to_product_format(records)
                
                logger.info(f"智能解析成功: {len(products)} 个产品")
                
                # 清理临时文件
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                
                return {
                    'success': True,
                    'data': products,
                    'headers': ['product_name', 'product_model', 'specification', 'price', 'quantity', 'description'],  # 使用英文字段名
                    'source': 'smart_mapping',
                    'total_records': len(products),
                    'message': f'智能解析成功，提取了 {len(products)} 个产品'
                }
                
            except Exception as e:
                logger.error(f"智能解析失败: {e}")
                # 智能解析失败，不回退到普通解析，直接返回错误
                return {
                    'success': False,
                    'message': f'智能解析失败，不支持该文件格式: {filename}。错误: {str(e)}'
                }
        else:
            # 不支持的Excel文件，返回错误
            return {
                'success': False,
                'message': f'不支持的文件格式: {filename}。当前仅支持: {", ".join(self.supported_files)}'
            }
    
    def fallback_parse_excel(self, file):
        """回退到普通的Excel解析方式"""
        try:
            workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            
            sheet_name = request.form.get('sheet_name')
            if not sheet_name or sheet_name not in sheet_names:
                sheet_name = workbook.active.title
            
            sheet = workbook[sheet_name]
            
            # 读取数据
            data = []
            for row in sheet.iter_rows(values_only=True):
                row_data = []
                for cell in row:
                    row_data.append(cell)
                data.append(row_data)
            
            return {
                'success': True,
                'data': data,
                'sheet_name': sheet_name,
                'sheet_names': sheet_names,
                'rows': len(data),
                'source': 'standard_parsing',
                'message': '标准解析完成'
            }
            
        except Exception as e:
            logger.error(f"回退解析失败: {e}")
            return {
                'success': False,
                'message': f'解析失败：{str(e)}'
            }

# 创建全局解析器实例
smart_parser = ExcelSmartParser()

def upload_xlsx_enhanced():
    """增强版的上传XLSX文件API"""
    try:
        if 'file' not in request.files:
            return jsonify({"success": False, "message": "没有上传文件"})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"success": False, "message": "没有选择文件"})
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({"success": False, "message": "只支持 Excel 文件 (.xlsx)"})
        
        # 获取解析模式参数
        parse_mode = request.form.get('parse_mode', 'auto')  # auto, smart, standard
        
        if parse_mode == 'smart':
            # 强制使用智能解析
            result = smart_parser.smart_parse_excel(None, file)
        elif parse_mode == 'standard':
            # 强制使用标准解析
            result = smart_parser.fallback_parse_excel(file)
        else:
            # 自动选择解析方式
            result = smart_parser.smart_parse_excel(None, file)
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"上传XLSX失败: {e}")
        return jsonify({"success": False, "message": f"上传失败：{str(e)}"})

def batch_add_products_enhanced():
    """增强版的批量添加产品API"""
    try:
        data = request.get_json()
        products = data.get('products', [])
        source = data.get('source', 'unknown')
        
        if not products:
            return jsonify({"success": False, "message": "没有产品数据"})
        
        # 导入ProductName模型
        from models import ProductName
        
        added_count = 0
        skipped_count = 0
        errors = []
        
        for product in products:
            try:
                # 数据验证
                name = str(product.get('name', '')).strip()
                if not name:
                    skipped_count += 1
                    continue
                
                # 检查是否已存在相同的产品
                existing = ProductName.query.filter_by(name=name).first()
                if existing:
                    skipped_count += 1
                    continue
                
                # 添加产品
                ProductName.add(
                    name=name,
                    model_number=str(product.get('model', '')),
                    description=str(product.get('description', '')),
                    price=float(product.get('price', 0)),
                    specification=float(product.get('spec', 0)),
                    purchase_unit_id=product.get('unit_id', None)
                )
                added_count += 1
                
            except Exception as e:
                error_msg = f"添加产品失败: {product.get('name', 'Unknown')}, 错误: {e}"
                logger.warn(error_msg)
                errors.append(error_msg)
        
        response = {
            "success": True,
            "message": f"成功添加 {added_count} 个产品",
            "count": added_count,
            "skipped": skipped_count,
            "source": source
        }
        
        if errors:
            response["errors"] = errors
        
        return jsonify(response)
        
    except Exception as e:
        logger.error(f"批量添加产品失败: {e}")
        return jsonify({"success": False, "message": f"批量添加失败：{str(e)}"})

def quick_import_products():
    """快速导入产品API - 专门用于这三个Excel文件"""
    try:
        if 'file' not in request.files:
            return jsonify({"success": False, "message": "没有上传文件"})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"success": False, "message": "没有选择文件"})
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({"success": False, "message": "只支持 Excel 文件 (.xlsx)"})
        
        # 检查是否为支持的Excel文件
        if not smart_parser.is_supported_file(file.filename):
            return jsonify({
                "success": False, 
                "message": f"此文件不被支持。支持的文件：{', '.join(smart_parser.supported_files)}"
            })
        
        # 保存文件到临时位置
        temp_file_path = f"temp_{file.filename}"
        file.save(temp_file_path)
        
        try:
            # 使用智能映射器处理
            records = smart_parser.mapper.process_excel_file(temp_file_path)
            products = smart_parser.mapper.convert_to_product_format(records)
            
            # 导入ProductName模型
            from models import ProductName
            
            added_count = 0
            skipped_count = 0
            
            for product in products:
                try:
                    name = str(product.get('name', '')).strip()
                    if not name:
                        skipped_count += 1
                        continue
                    
                    # 检查是否已存在相同的产品
                    existing = ProductName.query.filter_by(name=name).first()
                    if existing:
                        skipped_count += 1
                        continue
                    
                    # 添加产品
                    ProductName.add(
                        name=name,
                        model_number=str(product.get('model', '')),
                        description=str(product.get('description', '')),
                        price=float(product.get('price', 0)),
                        specification=float(product.get('spec', 0)),
                        purchase_unit_id=product.get('unit_id', None)
                    )
                    added_count += 1
                    
                except Exception as e:
                    logger.warn(f"添加产品失败: {product}, 错误: {e}")
                    continue
            
            return jsonify({
                "success": True,
                "message": f"快速导入完成！成功添加 {added_count} 个产品，跳过 {skipped_count} 个重复产品",
                "count": added_count,
                "skipped": skipped_count,
                "source": "quick_import",
                "extracted_products": products  # 返回提取的产品用于调试
            })
            
        finally:
            # 清理临时文件
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
        
    except Exception as e:
        logger.error(f"快速导入失败: {e}")
        return jsonify({"success": False, "message": f"快速导入失败：{str(e)}"})

if __name__ == "__main__":
    # 测试代码
    parser = ExcelSmartParser()
    
    # 测试智能解析
    test_files = [
        'f:\\guoshi\\七彩乐园.xlsx',
        'f:\\guoshi\\宜榢.xlsx',
        'f:\\guoshi\\迎扬李总.xlsx'
    ]
    
    for file_path in test_files:
        if os.path.exists(file_path):
            filename = os.path.basename(file_path)
            print(f"\\n=== 测试文件: {filename} ===")
            
            # 模拟文件上传
            class MockFile:
                def __init__(self, filename):
                    self.filename = filename
                
                def save(self, path):
                    # 复制文件到临时位置
                    import shutil
                    shutil.copy2(file_path, path)
            
            mock_file = MockFile(filename)
            result = parser.smart_parse_excel(file_path, mock_file)
            
            if result['success']:
                print(f"解析成功: {result['message']}")
                print(f"提取产品数量: {len(result['data'])}")
            else:
                print(f"解析失败: {result['message']}")