import openpyxl

wb = openpyxl.load_workbook(r'templates\七彩乐园.xlsx')
ws = wb['25出货']

last_row = ws.max_row
print('=== 七彩乐园.xlsx - 25出货工作表 ===')
print(f'总行数: {last_row}\n')

print('最后5行数据:')
for row in range(max(1, last_row - 4), last_row + 1):
    row_data = []
    for col in range(1, 12):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            if isinstance(val, str) and val.startswith('='):
                row_data.append('公式')
            else:
                row_data.append(str(val)[:20])
        else:
            row_data.append('')
    print(f'行{row}: A={row_data[0]} | B={row_data[1]} | C={row_data[2]} | F={row_data[5]} | G={row_data[6]} | H={row_data[7]} | I={row_data[8]} | J={row_data[9]} | K={row_data[10]}')

wb.close()
