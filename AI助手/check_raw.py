import openpyxl

wb = openpyxl.load_workbook(r'templates\七彩乐园.xlsx')
ws = wb['25出货']

print('=== 搜索有数据的行 ===')
for row in range(420, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val is not None:
        print(f'行{row} A列: {val}')
        # 检查B列
        b_val = ws.cell(row=row, column=2).value
        print(f'行{row} B列: {b_val}')
        # 检查C列
        c_val = ws.cell(row=row, column=3).value
        print(f'行{row} C列: {c_val}')
        print('---')

wb.close()
