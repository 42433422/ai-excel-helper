import openpyxl

# 检查原始文件的所有工作表
wb = openpyxl.load_workbook(r"c:\Users\97088\Desktop\新建文件夹 (4)\AI助手\templates\七彩乐园.xlsx")

print("=== 工作表列表 ===")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"  {sheet_name}: {ws.max_row} 行")

# 检查25出货工作表
print("\n=== 25出货工作表数据 ===")
ws = wb["25出货"]
print(f"行数: {ws.max_row}, 列数: {ws.max_column}")

# 显示最后10行A列
print("\n最后10行A列日期:")
for row in range(max(1, ws.max_row - 9), ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    print(f"  行 {row}: {val}")

wb.close()
