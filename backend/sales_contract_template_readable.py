"""销售合同 Excel 模板是否可被填充链路打开（无其它 backend 依赖，避免循环导入）。"""

from __future__ import annotations

import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def is_sales_contract_excel_template_readable(path: str | Path) -> bool:
    """能否以销售合同填充链路打开（非 Word、非损坏的表格）。"""
    p = Path(path)
    if not p.is_file():
        return False
    suf = p.suffix.lower()
    try:
        if suf in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook

            wb = load_workbook(str(p), read_only=True, data_only=True)
            try:
                ws = wb.active
                _ = ws.max_row
            finally:
                wb.close()
            return True
        if suf == ".xls":
            import xlrd

            book = xlrd.open_workbook(str(p))
            _ = book.sheet_by_index(0).nrows
            return True
    except Exception:
        logger.debug("is_sales_contract_excel_template_readable: failed for %s", p, exc_info=True)
    return False
