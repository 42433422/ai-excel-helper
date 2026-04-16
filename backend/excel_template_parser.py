"""
Excel template intelligent parser.
Automatically identifies sheet structure, tables, headers, and field mappings.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


def detect_table_regions(ws: Worksheet) -> list[dict[str, Any]]:
    """
    Detect table regions in worksheet.
    
    Returns list of table regions with:
    - start_row, end_row, start_col, end_col
    - header_row
    - data_rows
    """
    tables = []
    
    try:
        if hasattr(ws, "tables") and ws.tables:
            for table_ref, table in ws.tables.items():
                ref = table.ref
                if ":" in ref:
                    start_cell, end_cell = ref.split(":")
                    start_col, start_row = _cell_to_row_col(start_cell)
                    end_col, end_row = _cell_to_row_col(end_cell)
                    
                    tables.append({
                        "name": table_ref,
                        "start_row": start_row,
                        "end_row": end_row,
                        "start_col": start_col,
                        "end_col": end_col,
                        "header_row": start_row if table.headerRowCount else start_row + 1,
                        "is_official_table": True,
                    })
    except Exception:
        pass
    
    if not tables:
        tables.extend(_detect_tables_by_pattern(ws))
    
    return tables


def _cell_to_row_col(cell_ref: str) -> tuple[int, int]:
    """Convert cell reference (e.g., 'A1') to (column, row)."""
    match = re.match(r"([A-Z]+)(\d+)", cell_ref.upper())
    if not match:
        return 1, 1
    
    col_str, row_str = match.groups()
    col = sum((ord(c) - ord("A") + 1) * (26 ** i) for i, c in enumerate(reversed(col_str)))
    row = int(row_str)
    return col, row


def _detect_tables_by_pattern(ws: Worksheet) -> list[dict[str, Any]]:
    """Detect tables by scanning for header patterns."""
    tables = []
    max_row = ws.max_row
    max_col = ws.max_column
    
    if max_row < 2 or max_col < 1:
        return []
    
    for start_row in range(1, min(max_row, 50)):
        header_cells = []
        for col in range(1, min(max_col, 30)):
            cell = ws.cell(row=start_row, column=col)
            if cell.value and str(cell.value).strip():
                header_cells.append((col, cell.value))
        
        if len(header_cells) >= 2:
            has_empty_below = False
            for col, _ in header_cells:
                cell_below = ws.cell(row=start_row + 1, column=col)
                if not cell_below.value or not str(cell_below.value).strip():
                    has_empty_below = True
                    break
            
            if has_empty_below:
                continue
            
            end_row = start_row + 1
            while end_row <= max_row:
                row_has_data = False
                for col, _ in header_cells:
                    cell = ws.cell(row=end_row, column=col)
                    if cell.value and str(cell.value).strip():
                        row_has_data = True
                        break
                
                if not row_has_data:
                    break
                end_row += 1
            
            if end_row > start_row + 1:
                tables.append({
                    "name": f"Table_{start_row}",
                    "start_row": start_row,
                    "end_row": end_row - 1,
                    "start_col": min(col for col, _ in header_cells),
                    "end_col": max(col for col, _ in header_cells),
                    "header_row": start_row,
                    "is_official_table": False,
                })
    
    return tables


def extract_headers(ws: Worksheet, header_row: int, start_col: int = 1, end_col: int | None = None) -> list[dict[str, Any]]:
    """
    Extract headers from a row.
    
    Returns list of field definitions with:
    - name: field name
    - display_name: display name
    - column: column index
    - data_type: inferred data type
    """
    if end_col is None:
        end_col = ws.max_column
    
    headers = []
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=header_row, column=col)
        value = cell.value
        
        if value is None:
            continue
        
        header_text = str(value).strip()
        if not header_text:
            continue
        
        field_name = _normalize_field_name(header_text)
        data_type = _infer_data_type_from_header(header_text)
        
        headers.append({
            "name": field_name,
            "display_name": header_text,
            "column": col,
            "data_type": data_type,
            "required": _is_required_field(header_text),
        })
    
    return headers


def _normalize_field_name(text: str) -> str:
    """Normalize field name for programmatic use."""
    text = text.strip().lower()
    text = re.sub(r"[\s\-\(\)]+", "_", text)
    text = re.sub(r"[^\w\u4e00-\u9fff]+", "", text)
    text = re.sub(r"_+", "_", text)
    text = text.strip("_")
    return text or "field"


def _infer_data_type_from_header(header: str) -> str:
    """Infer data type from header text."""
    header_lower = header.lower()
    
    if any(kw in header_lower for kw in ["日期", "date", "时间", "time"]):
        return "date"
    if any(kw in header_lower for kw in ["金额", "价格", "单价", "总价", "price", "amount", "cost"]):
        return "number"
    if any(kw in header_lower for kw in ["数量", "qty", "count", "num"]):
        return "integer"
    if any(kw in header_lower for kw in ["电话", "手机", "contact", "phone"]):
        return "phone"
    if "@" in header or "email" in header_lower:
        return "email"
    if any(kw in header_lower for kw in ["状态", "status", "类型", "type"]):
        return "string"
    
    return "string"


def _is_required_field(header: str) -> bool:
    """Determine if field is required based on header."""
    required_keywords = ["必填", "必须", "required", "*"]
    return any(kw in header for kw in required_keywords)


def infer_field_types(ws: Worksheet, headers: list[dict[str, Any]], data_start_row: int, data_end_row: int) -> list[dict[str, Any]]:
    """
    Infer field types by analyzing actual data.
    """
    for header in headers:
        col = header["column"]
        values = []
        
        for row in range(data_start_row, min(data_end_row + 1, data_start_row + 100)):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value is not None and str(value).strip():
                values.append(value)
        
        if not values:
            continue
        
        inferred_type = _infer_type_from_values(values)
        if inferred_type != header["data_type"]:
            header["inferred_type"] = inferred_type
    
    return headers


def _infer_type_from_values(values: list[Any]) -> str:
    """Infer data type from sample values."""
    if not values:
        return "string"
    
    type_counts = {"integer": 0, "number": 0, "date": 0, "string": 0}
    
    for val in values[:50]:
        if isinstance(val, bool):
            type_counts["string"] += 1
        elif isinstance(val, int):
            type_counts["integer"] += 1
        elif isinstance(val, float):
            type_counts["number"] += 1
        elif isinstance(val, str):
            if _is_date_string(val):
                type_counts["date"] += 1
            elif _is_number_string(val):
                type_counts["number"] += 1
            else:
                type_counts["string"] += 1
    
    return max(type_counts, key=type_counts.get)


def _is_date_string(text: str) -> bool:
    """Check if string looks like a date."""
    patterns = [
        r"\d{4}-\d{2}-\d{2}",
        r"\d{4}/\d{2}/\d{2}",
        r"\d{2}-\d{2}-\d{4}",
        r"\d{2}/\d{2}/\d{4}",
    ]
    return any(re.match(p, text) for p in patterns)


def _is_number_string(text: str) -> bool:
    """Check if string represents a number."""
    try:
        float(text.replace(",", ""))
        return True
    except ValueError:
        return False


def extract_placeholder_patterns(ws: Worksheet) -> list[str]:
    """
    Extract placeholder patterns like {{field_name}} from cells.
    """
    placeholders = set()
    pattern = re.compile(r"\{\{([^}]+)\}\}")
    
    for row in range(1, min(ws.max_row + 1, 100)):
        for col in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                matches = pattern.findall(cell.value)
                placeholders.update(matches)
    
    return sorted(list(placeholders))


def parse_excel_template(file_path: Path) -> dict[str, Any]:
    """
    Main entry point for parsing Excel template.
    
    Returns comprehensive template metadata.
    """
    wb = load_workbook(file_path, data_only=True)
    
    result = {
        "file_name": file_path.name,
        "sheets": [],
        "placeholders": [],
        "metadata": {
            "total_sheets": len(wb.sheetnames),
            "workbook_properties": {},
        },
    }
    
    all_placeholders = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        tables = detect_table_regions(ws)
        
        sheet_info = {
            "name": sheet_name,
            "tables": [],
            "headers": [],
        }
        
        for table in tables:
            headers = extract_headers(
                ws,
                table["header_row"],
                table["start_col"],
                table["end_col"],
            )
            
            headers = infer_field_types(
                ws,
                headers,
                table["header_row"] + 1,
                table["end_row"],
            )
            
            table["headers"] = headers
            table["row_count"] = table["end_row"] - table["header_row"]
            sheet_info["tables"].append(table)
            sheet_info["headers"].extend(headers)
        
        placeholders = extract_placeholder_patterns(ws)
        all_placeholders.extend(placeholders)
        
        result["sheets"].append(sheet_info)
    
    result["placeholders"] = sorted(list(set(all_placeholders)))
    
    try:
        wb.close()
    except Exception:
        pass
    
    return result
