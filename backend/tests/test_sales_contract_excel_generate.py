"""销售合同 Excel 模板填充与预览（送货单.xls 版式）。"""

from pathlib import Path

import pytest

from backend.sales_contract_excel_generate import (
    fill_sales_contract_excel_template,
    read_excel_sales_contract_preview,
)


def _template_path() -> Path:
    return Path(__file__).resolve().parents[2] / "424" / "document_templates" / "送货单.xls"


def test_fill_sales_contract_excel_smoke(tmp_path: Path) -> None:
    tpl = _template_path()
    if not tpl.is_file():
        pytest.skip("缺少 424/document_templates/送货单.xls")
    out = tmp_path / "合同测试.xlsx"
    td = {
        "customer_name": "单元测试客户",
        "customer_phone": "13900000000",
        "contract_date": "2026年04月14日",
        "return_buckets_expected": 2,
        "return_buckets_actual": 1,
        "total_quantity": "3",
        "total_weight": "30KG",
        "total_amount": "300元",
        "total_amount_chinese": "叁佰元整",
        "products": [
            {
                "model_number": "T1",
                "name": "测试品",
                "specification": "5KG×1",
                "unit": "桶",
                "quantity": "3",
                "unit_price": "20",
                "amount": "300",
                "total_weight": 15.0,
            }
        ],
    }
    fill_sales_contract_excel_template(tpl, td, out)
    assert out.is_file() and out.stat().st_size > 1000
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=2, column=2).value == "单元测试客户"
    assert ws.cell(row=3, column=9).value == "2026年04月14日"
    assert ws.cell(row=5, column=1).value == "T1"


def test_read_excel_sales_contract_preview() -> None:
    tpl = _template_path()
    if not tpl.is_file():
        pytest.skip("缺少 424/document_templates/送货单.xls")
    data = read_excel_sales_contract_preview(tpl)
    assert data.get("success") is True
    assert isinstance(data.get("headers"), list)
    assert len(data["headers"]) >= 1
    assert "编号" in data["headers"]
    assert isinstance(data.get("sample_rows"), list)
    joined = " ".join(str(v) for v in data["sample_rows"][0].values())
    assert "KG" in joined.upper()


def test_resolve_sales_delivery_excel_never_returns_docx() -> None:
    """Word 模板行存在时，送货单解析器仍须落到 .xls/.xlsx（不可回落到默认 .docx）。"""
    from backend.document_template_service import resolve_sales_delivery_excel_template_with_meta

    p, slug = resolve_sales_delivery_excel_template_with_meta()
    if p is None:
        pytest.skip("无可用送货单模板（库或仓库）")
    assert p.suffix.lower() in (".xls", ".xlsx", ".xlsm"), p
    assert slug


def test_build_sales_contract_preview_json_excel_branch(monkeypatch: pytest.MonkeyPatch) -> None:
    from backend.price_list_docx_export import build_sales_contract_template_preview_json

    tpl = _template_path()
    if not tpl.is_file():
        pytest.skip("缺少 424/document_templates/送货单.xls")

    def fake_resolve(*, role: str, slug: str | None):
        return tpl, "fixture"

    monkeypatch.setattr(
        "backend.document_template_service.resolve_template_path_with_meta",
        fake_resolve,
    )
    body = build_sales_contract_template_preview_json("any")
    assert body.get("success") is True
    assert "headers" in body
    assert body.get("template_hint")


def test_fill_sales_contract_excel_numeric_cells_for_formula_inputs(tmp_path: Path) -> None:
    tpl = _template_path()
    if not tpl.is_file():
        pytest.skip("缺少 424/document_templates/送货单.xls")

    out = tmp_path / "合同数值测试.xlsx"
    td = {
        "customer_name": "数值单元格测试客户",
        "customer_phone": "13900000000",
        "contract_date": "2026年04月15日",
        "return_buckets_expected": 0,
        "return_buckets_actual": 0,
        "total_quantity": "3",
        "total_weight": "15KG",
        "total_amount": "300元",
        "total_amount_chinese": "叁佰元整",
        "products": [
            {
                "model_number": "T-NUM",
                "name": "测试品",
                "specification": "5KG×1",
                "spec_quantity": "5KG",
                "unit": "桶",
                "quantity": "3桶",
                "unit_price": "20元",
            }
        ],
    }
    fill_sales_contract_excel_template(tpl, td, out)

    from openpyxl import load_workbook

    wb = load_workbook(out, data_only=False)
    ws = wb.active
    # 规格与单件数量合并到第 4 列，第 3 列不再单独放数字
    assert ws.cell(row=5, column=3).value in ("", None)
    spec_cell = str(ws.cell(row=5, column=4).value or "")
    assert "KG" in spec_cell.upper()
    assert "5" in spec_cell
    assert isinstance(ws.cell(row=5, column=5).value, (int, float))
    assert isinstance(ws.cell(row=5, column=9).value, (int, float))
    # 空白产品行应清空词条内容（含公式列），避免出现 #VALUE!
    assert ws.cell(row=6, column=7).value in ("", None)
    assert ws.cell(row=6, column=11).value in ("", None)
    wb.close()
