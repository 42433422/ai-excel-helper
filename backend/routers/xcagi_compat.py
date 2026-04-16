"""
XCAGI 前端兼容 API（FastAPI APIRouter）。
与 Excel/对话核心分离；无持久化处返回 200 + 空或默认数据。
"""

from __future__ import annotations

import asyncio
import threading
import time
from typing import Any
import hashlib
import logging
import os
import json
import re
import sqlite3
import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, Body, File, Form, HTTPException, Query, Request, UploadFile
from openai import APIConnectionError, APIError, AuthenticationError, RateLimitError
from pydantic import AliasChoices, BaseModel, ConfigDict, Field
from fastapi.responses import Response, StreamingResponse

from backend.price_list_docx_export import (
    build_price_list_docx_bytes,
    build_price_list_template_preview_json,
    resolve_price_list_docx_template,
)

from backend.database import (
    get_db_status,
    get_sync_engine,
    resolve_mode,
    switch_to_production_mode,
    switch_to_test_mode,
)
from backend.llm_config import set_mode as set_llm_mode
from backend.ai_tier import (
    assert_p2_elevated_claim_or_raise,
    resolve_ai_tier,
    runtime_context_with_tier,
)
from backend.planner import chat as run_agent_chat, chat_stream_sse_events
from backend.runtime_context import (
    planner_workflow_interrupt_reply,
    runtime_context_after_workflow_interrupt,
)
from backend.tools import execute_workflow_tool
from backend.tools_directory_compat import get_tool_categories_payload, get_tools_payload
from backend.db_read_auth import verify_db_read_token_header
from backend.db_write_auth import verify_db_write_token_header
from backend.shell.mod_row_scope import (
    append_mod_scope_where,
    products_update_or_delete_mod_and,
    scoped_mod_id,
)

router = APIRouter(tags=["xcagi-compat"])
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# 轻量会话持久化（进程内）：供 XCAGI 前端「历史对话」列表与切回会话使用。
# 重启进程后清空；单用户最多保留 _XCAGI_CONVERSATION_MAX_SESSIONS 个会话。
# ---------------------------------------------------------------------------
_XCAGI_CONVERSATION_MAX_SESSIONS = 200
_conversation_lock = threading.Lock()
# user_id -> session_id -> {"messages": list[dict], "created_ts": float, "updated_ts": float}
_xcagi_user_sessions: dict[str, dict[str, dict[str, Any]]] = {}


def _xcagi_strip_html(text: str) -> str:
    s = str(text or "")
    return re.sub(r"<[^>]+>", "", s)


def _xcagi_iso_from_ts(ts: float) -> str:
    return (
        datetime.fromtimestamp(float(ts), tz=timezone.utc)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z")
    )


def _xcagi_normalize_chat_role(role: Any) -> str:
    r = str(role or "user").strip().lower()
    if r in ("assistant", "model"):
        return "ai"
    if r in ("user", "ai", "task"):
        return r
    return "ai"


def _xcagi_title_from_messages(msgs: list[dict]) -> str | None:
    for m in msgs:
        if m.get("role") != "user":
            continue
        t = _xcagi_strip_html(str(m.get("content") or "")).strip()
        if not t:
            continue
        return (t[:48] + "…") if len(t) > 48 else t
    return None


def _xcagi_summary_from_messages(msgs: list[dict]) -> str:
    if not msgs:
        return ""
    c = str(msgs[-1].get("content") or "")
    t = _xcagi_strip_html(c).strip().replace("\n", " ")
    if len(t) > 120:
        return t[:120] + "…"
    return t


def _xcagi_evict_oldest_session_if_needed(user_bucket: dict[str, dict[str, Any]], new_sid: str) -> None:
    if new_sid in user_bucket or len(user_bucket) < _XCAGI_CONVERSATION_MAX_SESSIONS:
        return
    oldest_sid = min(user_bucket.keys(), key=lambda k: float(user_bucket[k].get("updated_ts") or 0.0))
    user_bucket.pop(oldest_sid, None)


_DEFAULT_INDUSTRY = {"id": "general", "name": "通用", "code": "general"}

_DEFAULT_MEASURE_UNITS: list[dict] = [
    {"id": 1, "name": "件", "symbol": "pcs"},
    {"id": 2, "name": "千克", "symbol": "kg"},
]

# products.unit 常被填成计量词；合并进下拉时排除，与 XCAGI ProductRepository 一致。
TRIVIAL_MEASURE_UNITS: frozenset[str] = frozenset(
    {
        "件",
        "个",
        "只",
        "箱",
        "盒",
        "包",
        "袋",
        "瓶",
        "桶",
        "罐",
        "千克",
        "公斤",
        "克",
        "斤",
        "两",
        "吨",
        "米",
        "厘米",
        "毫米",
        "千米",
        "升",
        "毫升",
        "套",
        "组",
        "台",
        "条",
        "张",
        "根",
        "卷",
        "块",
        "片",
        "支",
        "双",
        "对",
        "副",
        "把",
        "捆",
        "扎",
    }
)


@router.get("/wechat_contacts/work_mode_feed")
async def wechat_work_mode_feed(per_contact: int = Query(default=1, ge=1, le=100)) -> dict:
    """
    获取微信会话工作模式数据，实时从已解密的 session.db 读取。
    返回每个联系人的最新消息状态。
    """
    try:
        import os
        import sys
        import json
        import sqlite3
        from contextlib import closing
        from datetime import datetime

        wechat_decrypt_path = os.environ.get("WECHAT_DECRYPT_PATH", r"e:\FHD\XCAGI\wechat-decrypt")
        if wechat_decrypt_path not in sys.path:
            sys.path.insert(0, wechat_decrypt_path)

        config_file = os.path.join(wechat_decrypt_path, "config.json")
        keys_file = os.path.join(wechat_decrypt_path, "all_keys.json")

        if not os.path.exists(config_file) or not os.path.exists(keys_file):
            return {"items": [], "per_contact": per_contact, "error": "wechat-decrypt not configured"}

        with open(config_file) as f:
            cfg = json.load(f)

        with open(keys_file) as f:
            all_keys = json.load(f)

        copy_db_dir = os.path.join(wechat_decrypt_path, "raw_db")
        db_dir = cfg.get("db_dir", "")

        PAGE_SZ = 4096
        RESERVE_SZ = 80
        SALT_SZ = 16
        SQLITE_HDR = b'SQLite format 3\x00'
        KEY_SZ = 32

        def strip_key_metadata(keys):
            if isinstance(keys, dict):
                return keys
            result = []
            for k in keys:
                if isinstance(k, dict):
                    if "enc_key" in k:
                        result.append(k)
                    elif "keys" in k:
                        result.extend(k["keys"])
            return result

        def derive_mac_key(enc_key, salt):
            import hashlib
            mac_salt = bytes(b ^ 0x3a for b in salt)
            return hashlib.pbkdf2_hmac("sha512", enc_key, mac_salt, 2, dklen=KEY_SZ)

        def decrypt_page(enc_key, page_data, pgno):
            from Crypto.Cipher import AES
            iv = page_data[PAGE_SZ - RESERVE_SZ : PAGE_SZ - RESERVE_SZ + 16]
            if pgno == 1:
                encrypted = page_data[SALT_SZ : PAGE_SZ - RESERVE_SZ]
                cipher = AES.new(enc_key, AES.MODE_CBC, iv)
                decrypted = cipher.decrypt(encrypted)
                page = bytearray(SQLITE_HDR + decrypted + b'\x00' * RESERVE_SZ)
                return bytes(page)
            else:
                encrypted = page_data[:PAGE_SZ - RESERVE_SZ]
                cipher = AES.new(enc_key, AES.MODE_CBC, iv)
                decrypted = cipher.decrypt(encrypted)
                return decrypted + b'\x00' * RESERVE_SZ

        def get_key_info(keys, rel_path):
            if isinstance(keys, dict):
                for path_key in keys:
                    if path_key == rel_path or path_key.replace("\\", "/") == rel_path.replace("\\", "/"):
                        info = keys[path_key].copy()
                        info["path"] = path_key
                        return info
                return None
            for k in keys:
                if k.get("path") == rel_path or k.get("path", "").replace("\\", "/") == rel_path.replace("\\", "/"):
                    return k
                if "keys" in k:
                    for sub in k["keys"]:
                        if sub.get("path") == rel_path:
                            return sub
            return None

        def full_decrypt(db_path, enc_key):
            file_size = os.path.getsize(db_path)
            total_pages = file_size // PAGE_SZ
            chunks = []
            with open(db_path, 'rb') as fin:
                for pgno in range(1, total_pages + 1):
                    page = fin.read(PAGE_SZ)
                    if len(page) < PAGE_SZ:
                        if len(page) > 0:
                            page = page + b'\x00' * (PAGE_SZ - len(page))
                        else:
                            break
                    chunks.append(decrypt_page(enc_key, page, pgno))
            return b''.join(chunks)

        stripped_keys = strip_key_metadata(all_keys)
        session_key_info = get_key_info(stripped_keys, os.path.join("session", "session.db"))
        if not session_key_info:
            return {"items": [], "per_contact": per_contact, "error": "session.db key not found"}

        enc_key = bytes.fromhex(session_key_info["enc_key"])
        session_db = os.path.join(copy_db_dir, "session", "session.db")
        if not os.path.exists(session_db):
            return {"items": [], "per_contact": per_contact, "error": "session.db not found in raw_db, run sync_raw_db.py first"}

        decrypted_data = full_decrypt(session_db, enc_key)

        import tempfile
        tmp_path = os.path.join(tempfile.gettempdir(), "wechat_work_mode_feed.db")
        with open(tmp_path, 'wb') as f:
            f.write(decrypted_data)

        decrypted_dir = cfg.get("decrypted_dir", os.path.join(wechat_decrypt_path, "decrypted"))
        if not os.path.isabs(decrypted_dir):
            decrypted_dir = os.path.join(wechat_decrypt_path, decrypted_dir)
        contact_cache = os.path.join(decrypted_dir, "contact", "contact.db")
        contact_names = {}
        if os.path.exists(contact_cache):
            try:
                cconn = sqlite3.connect(contact_cache)
                for r in cconn.execute("SELECT username, nick_name, remark FROM contact").fetchall():
                    uname, nick, remark = r
                    contact_names[uname] = remark if remark else nick if nick else uname
                cconn.close()
            except Exception:
                pass

        items = []
        zstd_dctx = None
        try:
            import zstandard as zstd  # type: ignore[import-untyped]

            zstd_dctx = zstd.ZstdDecompressor()
        except ImportError:
            logger.warning(
                "wechat_work_mode_feed: zstandard not installed; "
                "install with `pip install zstandard` for session summary text"
            )
        try:
            with closing(sqlite3.connect(tmp_path)) as conn:
                conn.row_factory = sqlite3.Row
                rows = conn.execute("""
                    SELECT username, unread_count, summary, last_timestamp,
                           last_msg_type, last_msg_sender, last_sender_display_name
                    FROM SessionTable
                    WHERE last_timestamp > 0
                    ORDER BY last_timestamp DESC
                    LIMIT ?
                """, (per_contact * 10,)).fetchall()

                for r in rows:
                    username = r["username"]
                    display = contact_names.get(username, username)
                    summary = r["summary"] or ""
                    if isinstance(summary, bytes):
                        if zstd_dctx is not None:
                            try:
                                summary = zstd_dctx.decompress(summary).decode(
                                    "utf-8", errors="replace"
                                )
                            except Exception:
                                summary = "(compressed)"
                        else:
                            try:
                                summary = summary.decode("utf-8", errors="replace")
                            except Exception:
                                summary = "(compressed; pip install zstandard)"
                    if isinstance(summary, str) and ':\n' in summary:
                        summary = summary.split(':\n', 1)[1]

                    ts = r["last_timestamp"]
                    time_str = datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S') if ts else ""

                    msg_types = {1: 'text', 3: 'image', 34: 'voice', 42: 'card', 43: 'video', 47: 'sticker', 10000: 'system'}
                    msg_type = msg_types.get(r["last_msg_type"], f'type={r["last_msg_type"]}')

                    is_group = '@chatroom' in username
                    sender = r["last_msg_sender"] or ""
                    sender_name = r["last_sender_display_name"] or ""
                    sender_display = contact_names.get(sender, sender_name or sender) if sender else ""

                    items.append({
                        "username": username,
                        "display_name": display,
                        "is_group": is_group,
                        "unread_count": r["unread_count"] or 0,
                        "summary": summary,
                        "timestamp": ts,
                        "time_str": time_str,
                        "msg_type": msg_type,
                        "sender": sender,
                        "sender_display": sender_display,
                    })
        finally:
            try:
                os.remove(tmp_path)
            except Exception:
                pass

        return {"items": items[:per_contact], "per_contact": per_contact, "total": len(items)}

    except Exception as e:
        logger.exception("wechat_work_mode_feed error")
        return {"items": [], "per_contact": per_contact, "error": str(e)}


class WechatStarredContact(BaseModel):
    model_config = ConfigDict(extra="ignore")

    type: str = Field(validation_alias=AliasChoices("type", "contactType"))
    nickname: str = Field(validation_alias=AliasChoices("nickname", "备注", "remark"))
    remark: str = Field(default="")
    wxid: str = Field(validation_alias=AliasChoices("wxid", "微信号"))
    starred: bool = Field(default=True)


_STARRED_CONTACTS_DB: dict[str, dict] = {}
_STARRED_NEXT_ID: int = 1


def _migrate_starred_contact_ids() -> None:
    """为内存星标表补齐整数 id，供前端 /api/wechat_contacts 删除、编辑使用。"""
    global _STARRED_NEXT_ID
    for _wxid, c in _STARRED_CONTACTS_DB.items():
        if "id" not in c:
            c["id"] = _STARRED_NEXT_ID
            _STARRED_NEXT_ID += 1


def _starred_row_for_frontend(c: dict) -> dict:
    ct = (c.get("type") or "contact").lower()
    return {
        "id": c.get("id"),
        "contact_name": c.get("nickname") or "",
        "remark": c.get("remark") or "",
        "wechat_id": c.get("wxid") or "",
        "contact_type": "group" if ct == "group" else "contact",
        "type": ct,
        "nickname": c.get("nickname"),
        "wxid": c.get("wxid"),
        "starred": bool(c.get("starred", True)),
    }


def _search_hit_for_frontend(c: dict) -> dict:
    row = _starred_row_for_frontend(c)
    row["already_starred"] = True
    dn = (row.get("contact_name") or "").strip() or (row.get("remark") or "").strip() or (row.get("wechat_id") or "").strip()
    row["display_name"] = dn or "-"
    row["username"] = (row.get("wechat_id") or "").strip()
    row["nick_name"] = (row.get("contact_name") or "").strip()
    return row


@router.get("/wechat_contacts/decrypt_status")
async def wechat_contacts_decrypt_status_compat() -> dict:
    """与 XCAGI 前端 wechatApi.getDecryptStatus 对齐的轻量探测。"""
    path = os.environ.get("WECHAT_CONTACT_DB_PATH", "").strip()
    exists = bool(path and os.path.isfile(path))
    if not exists:
        base = os.environ.get("WECHAT_DECRYPT_PATH", "").strip()
        if base:
            path = os.path.join(base, "decrypted", "contact", "contact.db")
            exists = os.path.isfile(path)
    return {
        "success": True,
        "plugin_available": True,
        "contact_db_path": path or None,
        "contact_db_exists": exists,
    }


@router.get("/wechat_contacts/search")
async def wechat_contacts_search_compat(
    q: str = Query("", description="搜索关键字"),
    keyword: str = Query("", description="与 q 等价"),
) -> dict:
    term = (q or keyword or "").strip().lower()
    if not term:
        return {"success": True, "results": []}
    _migrate_starred_contact_ids()
    hits: list[dict] = []
    for _wxid, c in _STARRED_CONTACTS_DB.items():
        blob = " ".join(
            str(x).lower()
            for x in (c.get("nickname"), c.get("remark"), c.get("wxid"), c.get("type"))
            if x
        )
        if term in blob:
            hits.append(_search_hit_for_frontend(c))
    return {"success": True, "results": hits}


@router.get("/wechat_contacts")
async def wechat_contacts_list_compat(
    type: str = Query("all", description="类型: all, contact, group"),
    keyword: str = Query("", description="昵称/备注/微信号筛选"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
) -> dict:
    """XCAGI 前端 wechatApi.getStarredContacts → GET /api/wechat_contacts?type=…"""
    _migrate_starred_contact_ids()
    all_items = list(_STARRED_CONTACTS_DB.values())
    t = (type or "all").strip().lower()
    if t and t != "all":
        filtered = [c for c in all_items if str(c.get("type", "")).lower() == t]
    else:
        filtered = all_items
    if keyword:
        kw = keyword.lower()
        filtered = [
            c
            for c in filtered
            if kw in str(c.get("nickname", "")).lower()
            or kw in str(c.get("remark", "")).lower()
            or kw in str(c.get("wxid", "")).lower()
        ]
    return {
        "success": True,
        "data": [_starred_row_for_frontend(c) for c in filtered],
        "page": page,
        "per_page": per_page,
    }


@router.post("/wechat_contacts/unstar_all")
async def wechat_contacts_unstar_all_compat() -> dict:
    return await wechat_starred_clear()


@router.post("/wechat_contacts/refresh_messages_cache")
async def wechat_contacts_refresh_messages_cache_compat() -> dict:
    return {"success": True, "message": "ok"}


@router.post("/wechat_contacts/refresh_contact_cache")
async def wechat_contacts_refresh_contact_cache_compat() -> dict:
    return {"success": True, "message": "ok", "data": {"sync": {"success": True, "message": "占位"}}}


@router.post("/wechat_contacts")
async def wechat_contacts_create_compat(body: dict = Body(default_factory=dict)) -> dict:
    """与 wechatApi.addStarredContact 对齐：wechat_id / contact_name / contact_type / remark。"""
    wxid = str(body.get("wechat_id") or body.get("wxid") or "").strip()
    if not wxid:
        raise HTTPException(status_code=400, detail="wechat_id 不能为空")
    nickname = str(body.get("contact_name") or body.get("nickname") or "").strip()
    remark = str(body.get("remark") or "").strip()
    contact_type = str(body.get("contact_type") or body.get("type") or "contact").strip() or "contact"
    global _STARRED_NEXT_ID
    _migrate_starred_contact_ids()
    cid = _STARRED_NEXT_ID
    _STARRED_NEXT_ID += 1
    contact = {
        "id": cid,
        "type": contact_type,
        "nickname": nickname,
        "remark": remark,
        "wxid": wxid,
        "starred": True,
    }
    _STARRED_CONTACTS_DB[wxid] = contact
    return {"success": True, "message": "ok", "data": {"id": cid}}


@router.get("/wechat_contacts/starred")
async def wechat_starred_list(
    type: str = Query(default="all", description="类型筛选: all, contact, group"),
    keyword: str = Query(default="", description="昵称/备注筛选关键字"),
) -> dict:
    """
    获取星标联系人列表。
    type: all-全部, contact-联系人, group-群聊
    keyword: 昵称/备注筛选
    """
    _migrate_starred_contact_ids()
    all_items = list(_STARRED_CONTACTS_DB.values())

    if type and type != "all":
        filtered = [c for c in all_items if c.get("type", "").lower() == type.lower()]
    else:
        filtered = all_items

    if keyword:
        kw = keyword.lower()
        filtered = [
            c for c in filtered
            if kw in str(c.get("nickname", "")).lower()
            or kw in str(c.get("remark", "")).lower()
            or kw in str(c.get("wxid", "")).lower()
        ]

    return {
        "success": True,
        "data": filtered,
        "total": len(filtered),
        "filter": {"type": type, "keyword": keyword},
    }


@router.delete("/wechat_contacts/starred/{wxid}")
async def wechat_starred_delete(wxid: str) -> dict:
    """
    删除指定星标联系人。
    """
    if wxid in _STARRED_CONTACTS_DB:
        del _STARRED_CONTACTS_DB[wxid]
        return {"success": True, "message": f"已删除星标联系人 {wxid}"}
    return {"success": False, "message": f"星标联系人 {wxid} 不存在"}


@router.delete("/wechat_contacts/starred")
async def wechat_starred_clear() -> dict:
    """
    一键解除全部星标。
    """
    count = len(_STARRED_CONTACTS_DB)
    _STARRED_CONTACTS_DB.clear()
    return {"success": True, "message": f"已清除 {count} 个星标"}


@router.post("/wechat_contacts/starred")
async def wechat_starred_add(body: dict = Body(...)) -> dict:
    """
    添加星标联系人。
    """
    global _STARRED_NEXT_ID
    wxid = str(body.get("wxid") or body.get("wechat_id") or "").strip()
    if not wxid:
        raise HTTPException(status_code=400, detail="wxid 不能为空")

    contact_type = body.get("type") or body.get("contact_type") or "contact"
    nickname = body.get("nickname") or body.get("contact_name") or ""
    remark = body.get("remark", "")

    _migrate_starred_contact_ids()
    cid = _STARRED_NEXT_ID
    _STARRED_NEXT_ID += 1
    contact = {
        "id": cid,
        "type": contact_type,
        "nickname": nickname,
        "remark": remark,
        "wxid": wxid,
        "starred": True,
    }
    _STARRED_CONTACTS_DB[wxid] = contact
    return {"success": True, "data": contact}


@router.delete("/wechat_contacts/{contact_id}")
async def wechat_contacts_delete_compat(contact_id: str) -> dict:
    """须注册在 ``/wechat_contacts/starred`` 等静态路径之后，避免将 ``starred`` 当成 id。"""
    _migrate_starred_contact_ids()
    for wxid, c in list(_STARRED_CONTACTS_DB.items()):
        if str(c.get("id")) == str(contact_id):
            del _STARRED_CONTACTS_DB[wxid]
            return {"success": True, "message": "已删除"}
    return {"success": False, "message": "联系人不存在"}


@router.put("/wechat_contacts/{contact_id}")
async def wechat_contacts_update_compat(contact_id: str, body: dict = Body(default_factory=dict)) -> dict:
    _migrate_starred_contact_ids()
    for _wxid, c in _STARRED_CONTACTS_DB.items():
        if str(c.get("id")) == str(contact_id):
            if "contact_name" in body:
                c["nickname"] = str(body.get("contact_name") or "")
            if "remark" in body:
                c["remark"] = str(body.get("remark") or "")
            if "wechat_id" in body:
                c["wxid"] = str(body.get("wechat_id") or "")
            if "contact_type" in body:
                c["type"] = str(body.get("contact_type") or "contact")
            return {"success": True, "data": _starred_row_for_frontend(c)}
    return {"success": False, "message": "联系人不存在"}


@router.get("/wechat_contacts/{contact_id}/context")
async def wechat_contacts_context_compat(contact_id: str) -> dict:
    _ = contact_id
    return {"success": True, "messages": []}


@router.post("/wechat_contacts/{contact_id}/refresh_messages")
async def wechat_contacts_refresh_messages_compat(contact_id: str) -> dict:
    _ = contact_id
    return {"success": True, "message": "FHD 精简后端未实现消息库同步"}


@router.get("/system/industries")
async def system_industries() -> dict:
    return {"success": True, "data": [_DEFAULT_INDUSTRY]}


@router.get("/system/industry")
async def system_industry_get() -> dict:
    return {"success": True, "data": _DEFAULT_INDUSTRY}


@router.post("/system/industry")
async def system_industry_post(body: dict = Body(default_factory=dict)) -> dict:
    ind = body.get("industry") if isinstance(body.get("industry"), dict) else None
    return {"success": True, "industry": ind or _DEFAULT_INDUSTRY}


@router.get("/system/openapi")
async def system_openapi(request: Request) -> dict:
    """在仅反代 /api 的部署下仍可提供 OpenAPI JSON（等价于根路径 /openapi.json）。"""
    return request.app.openapi()


def _sql_ident(name: str) -> str:
    """SQLite 标识符转义（支持中文列名等）。"""
    return '"' + name.replace('"', '""') + '"'


def _insp_table_exists(insp, table_name: str) -> bool:
    """Reflection：优先 has_table（PG 上比 get_table_names 更稳），失败则回退名称集合。"""
    ht = getattr(insp, "has_table", None)
    if callable(ht):
        try:
            return bool(ht(table_name))
        except Exception:
            pass
    try:
        return table_name in insp.get_table_names()
    except Exception:
        return False


def _exc_chain_has_undefined_table(e: BaseException) -> bool:
    """识别 psycopg UndefinedTable（常被 SQLAlchemy 包在 orig/__cause__ 里）。"""
    cur: BaseException | None = e
    for _ in range(20):
        if cur is None:
            break
        if cur.__class__.__name__ == "UndefinedTable":
            return True
        cur = cur.__cause__ or getattr(cur, "orig", None)
    return False


def _load_purchase_units_rows_pg() -> list[dict]:
    try:
        from backend.shell.mod_business_scope import business_data_exposed

        if not business_data_exposed():
            return []
    except Exception:
        pass
    try:
        from sqlalchemy import inspect, text

        eng = get_sync_engine()
        insp = inspect(eng)
    except Exception as e:
        logger.warning("purchase_units pg: no engine (%s)", e)
        return []
    if not _insp_table_exists(insp, "purchase_units"):
        return []
    pu_cols = {c["name"] for c in insp.get_columns("purchase_units")}
    where_parts: list[str] = []
    bind: dict[str, object] = {}
    append_mod_scope_where(where_parts, bind, pu_cols)
    where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
    try:
        with eng.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    SELECT id, unit_name, contact_person, contact_phone, address, is_active
                    FROM purchase_units
                    {where_sql}
                    ORDER BY unit_name
                    """
                ),
                bind,
            ).mappings().all()
    except Exception as e:
        if _exc_chain_has_undefined_table(e):
            logger.debug("purchase_units pg: relation missing at query time (%s)", e)
        else:
            logger.warning("purchase_units pg: query failed: %s", e)
        return []
    out: list[dict] = []
    for r in rows:
        ia = r.get("is_active")
        if ia in (0, False, "0", "false", "f", "F"):
            continue
        out.append(dict(r))
    return out


def _load_purchase_units_rows() -> list[dict]:
    """购买单位与产品共用主库（PostgreSQL）。"""
    return _load_purchase_units_rows_pg()


def _load_customers_pg_from_customers_table(eng, insp) -> list[dict]:
    from sqlalchemy import text

    cols = {c["name"] for c in insp.get_columns("customers")}
    id_col = "id" if "id" in cols else ("customer_id" if "customer_id" in cols else None)
    name_col = next(
        (c for c in ("customer_name", "name", "客户名称") if c in cols),
        None,
    )
    if not id_col or not name_col:
        return []
    sel = [
        f"{_sql_ident(id_col)} AS id",
        f"{_sql_ident(name_col)} AS customer_name",
    ]
    cp = next((c for c in ("contact_person", "联系人") if c in cols), None)
    if cp:
        sel.append(f"{_sql_ident(cp)} AS contact_person")
    ph = next((c for c in ("contact_phone", "phone", "电话") if c in cols), None)
    if ph:
        sel.append(f"{_sql_ident(ph)} AS contact_phone")
    ad = next((c for c in ("address", "地址") if c in cols), None)
    if ad:
        sel.append(f"{_sql_ident(ad)} AS address")
    if "is_active" in cols:
        sel.append(_sql_ident("is_active") + " AS is_active")
    where_parts: list[str] = []
    bind: dict[str, object] = {}
    if "is_active" in cols:
        where_parts.append(
            "("
            + _sql_ident("is_active")
            + " IS NULL OR "
            + _sql_ident("is_active")
            + " = true OR CAST("
            + _sql_ident("is_active")
            + " AS INTEGER) = 1)"
        )
    append_mod_scope_where(where_parts, bind, cols)
    where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
    order = f"{_sql_ident(name_col)}, {_sql_ident(id_col)}"
    sql = f"SELECT {', '.join(sel)} FROM {_sql_ident('customers')}{where_sql} ORDER BY {order}"
    try:
        with eng.connect() as conn:
            rows = conn.execute(text(sql), bind).mappings().all()
    except Exception as e:
        logger.warning("customers pg customers table: %s", e)
        return []
    out: list[dict] = []
    for r in rows:
        d = dict(r)
        if "is_active" not in d:
            d["is_active"] = 1
        out.append(d)
    return out


def _load_customers_pg_from_purchase_units(eng) -> list[dict]:
    from sqlalchemy import inspect, text

    try:
        insp = inspect(eng)
        pu_cols = {c["name"] for c in insp.get_columns("purchase_units")}
        where_parts: list[str] = []
        bind: dict[str, object] = {}
        append_mod_scope_where(where_parts, bind, pu_cols)
        where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
        with eng.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    SELECT id, unit_name, contact_person, contact_phone, address, is_active
                    FROM purchase_units
                    {where_sql}
                    ORDER BY unit_name
                    """
                ),
                bind,
            ).mappings().all()
    except Exception as e:
        logger.warning("customers pg purchase_units: %s", e)
        return []
    out: list[dict] = []
    for r in rows:
        d = dict(r)
        ia = d.get("is_active")
        if ia in (0, False, "0", "false", "f", "F"):
            continue
        d["customer_name"] = (d.pop("unit_name", None) or "") or ""
        out.append(d)
    return out


def _load_customers_rows_pg() -> list[dict]:
    try:
        from sqlalchemy import inspect

        eng = get_sync_engine()
        insp = inspect(eng)
    except Exception as e:
        logger.warning("customers pg: no engine (%s)", e)
        return []
    names = set(insp.get_table_names())
    if "customers" in names:
        rows = _load_customers_pg_from_customers_table(eng, insp)
        if rows:
            return rows
    if "purchase_units" in names:
        return _load_customers_pg_from_purchase_units(eng)
    return []


def _load_customers_rows() -> list[dict]:
    """
    供 /api/customers 与 /api/customers/list 使用。
    读 PostgreSQL 内 ``customers`` / ``purchase_units``；若为空则回退为合并单位名称（含仅写在 products.unit 上的客户）。
    """
    try:
        from backend.shell.mod_business_scope import business_data_exposed

        if not business_data_exposed():
            return []
    except Exception:
        pass
    rows = _load_customers_rows_pg()
    if rows:
        return rows
    return _customer_rows_from_merged_unit_entries()


def _business_mod_json_block() -> dict | None:
    """启用「业务依赖扩展 Mod」且当前无扩展 manifest 时，返回 {success, message} 供 JSON 接口短路。"""
    try:
        from backend.shell.mod_business_scope import business_data_exposed, business_data_hidden_reason

        if business_data_exposed():
            return None
        return {
            "success": False,
            "message": business_data_hidden_reason() or "扩展 Mod 未就绪，业务接口已关闭。",
        }
    except Exception:
        return None


def _load_products_list_impl_pg(
    page: int,
    per_page: int,
    keyword: str | None,
    unit: str | None,
) -> tuple[list[dict], int, str | None]:
    """
    第三项 ``schema_hint``：仅在「缺表 / 缺列 / 无法连接」时返回中文提示；正常空表为 ``None``。
    """
    try:
        from backend.shell.mod_business_scope import business_data_exposed, business_data_hidden_reason

        if not business_data_exposed():
            return [], 0, business_data_hidden_reason()
    except Exception:
        pass
    try:
        from sqlalchemy import inspect, text

        eng = get_sync_engine()
    except Exception as e:
        return [], 0, f"无法连接 PostgreSQL：{e}。请检查 DATABASE_URL 与数据库是否已启动。"

    insp = inspect(eng)
    if "products" not in insp.get_table_names():
        return (
            [],
            0,
            "当前库中不存在 public.products 表，产品列表为空。请在目标库执行仓库 scripts/pg_init_xcagi_core.sql 后重启后端。",
        )
    col_names = {c["name"] for c in insp.get_columns("products")}
    if not {"id", "model_number", "name"}.issubset(col_names):
        return (
            [],
            0,
            "products 表存在但缺少必要列（至少需要 id、model_number、name）。请对照 scripts/pg_init_xcagi_core.sql 修正表结构。",
        )

    where_parts: list[str] = []
    params: dict[str, object] = {}

    if "is_active" in col_names:
        where_parts.append("(is_active IS NULL OR CAST(is_active AS INTEGER) = 1)")

    kw = (keyword or "").strip()
    if kw:
        like = f"%{kw}%"
        or_parts: list[str] = []
        if "model_number" in col_names:
            or_parts.append("CAST(model_number AS TEXT) ILIKE :kw")
        if "name" in col_names:
            or_parts.append("CAST(name AS TEXT) ILIKE :kw")
        if "specification" in col_names:
            or_parts.append("CAST(specification AS TEXT) ILIKE :kw")
        if or_parts:
            where_parts.append("(" + " OR ".join(or_parts) + ")")
            params["kw"] = like

    un = (unit or "").strip()
    if un and "unit" in col_names:
        where_parts.append("unit = :uunit")
        params["uunit"] = un

    append_mod_scope_where(where_parts, params, col_names)

    where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""

    count_sql = f"SELECT COUNT(*) FROM products{where_sql}"
    with eng.connect() as conn:
        total = int(conn.execute(text(count_sql), params).scalar_one())

    sel: list[str] = ["id", "model_number", "name"]
    sel.append("specification" if "specification" in col_names else "NULL AS specification")
    sel.append("price" if "price" in col_names else "NULL AS price")
    sel.append("quantity" if "quantity" in col_names else "0 AS quantity")
    sel.append("description" if "description" in col_names else "NULL AS description")
    sel.append("category" if "category" in col_names else "NULL AS category")
    sel.append("brand" if "brand" in col_names else "NULL AS brand")
    sel.append("unit" if "unit" in col_names else "'' AS unit")
    sel.append("is_active" if "is_active" in col_names else "1 AS is_active")
    sel.append("created_at" if "created_at" in col_names else "NULL AS created_at")
    sel.append("updated_at" if "updated_at" in col_names else "NULL AS updated_at")

    offset = (page - 1) * per_page
    order = "created_at DESC NULLS LAST, id DESC" if "created_at" in col_names else "id DESC"

    data_sql = (
        f"SELECT {', '.join(sel)} FROM products{where_sql} ORDER BY {order} LIMIT :lim OFFSET :off"
    )
    qparams = {**params, "lim": per_page, "off": offset}
    with eng.connect() as conn:
        rows = conn.execute(text(data_sql), qparams).mappings().all()

    rows_out: list[dict] = []
    for r in rows:
        d = dict(r)
        if d.get("price") is None:
            d["price"] = 0
        if d.get("quantity") is None:
            d["quantity"] = 0
        if d.get("unit") is None:
            d["unit"] = ""
        if d.get("is_active") is None:
            d["is_active"] = 1
        rows_out.append(d)
    return rows_out, total, None


_EXPORT_MAX_ROWS = 50_000


def _load_products_all_for_export(
    keyword: str | None, unit: str | None
) -> list[dict]:
    rows, _, _hint = _load_products_list_impl_pg(1, _EXPORT_MAX_ROWS, keyword, unit)
    return rows


def _distinct_units_from_products_db() -> list[dict]:
    return _distinct_units_from_products_db_pg()


def _distinct_units_from_products_db_pg() -> list[dict]:
    try:
        from backend.shell.mod_business_scope import business_data_exposed

        if not business_data_exposed():
            return []
    except Exception:
        pass
    try:
        from sqlalchemy import inspect, text

        eng = get_sync_engine()
    except Exception:
        return []
    insp = inspect(eng)
    if "products" not in insp.get_table_names():
        return []
    col_names = {c["name"] for c in insp.get_columns("products")}
    if "unit" not in col_names:
        return []
    where_parts = ["unit IS NOT NULL AND TRIM(unit) != ''"]
    if "is_active" in col_names:
        where_parts.append("(is_active IS NULL OR CAST(is_active AS INTEGER) = 1)")
    bind: dict[str, object] = {}
    append_mod_scope_where(where_parts, bind, col_names)
    where_sql = "WHERE " + " AND ".join(where_parts)
    try:
        with eng.connect() as conn:
            rows = conn.execute(
                text(
                    f"SELECT DISTINCT unit FROM products {where_sql} ORDER BY unit"
                ),
                bind,
            ).fetchall()
    except Exception as e:
        logger.warning("distinct product units pg: %s", e)
        return []
    names = [str(row[0]).strip() for row in rows if row[0] is not None]
    return [{"id": i + 1, "name": u, "symbol": u} for i, u in enumerate(names)]


def _merged_purchase_unit_entries() -> list[dict]:
    """
    购买单位主数据 + 产品上已填写但尚未写入 purchase_units 的单位（去重、计量词过滤）。
    避免「主表里已有若干客户」时仅出现在 products.unit 的新单位在下拉中消失。
    """
    rows = _load_purchase_units_rows()
    seen = {str(r.get("unit_name") or "").strip().lower() for r in rows if str(r.get("unit_name") or "").strip()}
    out: list[dict] = [dict(r) for r in rows]
    distinct = _distinct_units_from_products_db()
    max_id = 0
    for r in out:
        rid = r.get("id")
        if isinstance(rid, int):
            max_id = max(max_id, rid)
    syn = 0
    for d in distinct:
        name = str(d.get("name") or "").strip()
        if not name:
            continue
        key = name.lower()
        if key in seen:
            continue
        if name in TRIVIAL_MEASURE_UNITS:
            continue
        seen.add(key)
        syn += 1
        out.append(
            {
                "id": max_id + syn,
                "unit_name": name,
                "contact_person": "",
                "contact_phone": "",
                "address": "",
                "is_active": 1,
            }
        )
    return out


def _customer_rows_from_merged_unit_entries() -> list[dict]:
    """
    与产品页「购买单位」下拉（_merged_purchase_unit_entries）同源：
    当库内无 customers / 无可用 purchase_units 行时，仍用 products.unit 等合并出的名称作为客户列表，
    使客户总数与单位筛选器一致。
    """
    out: list[dict] = []
    for r in _merged_purchase_unit_entries():
        name = str(r.get("unit_name") or "").strip()
        if not name:
            continue
        rid = r.get("id")
        oid = rid if isinstance(rid, int) else len(out) + 1
        out.append(
            {
                "id": oid,
                "customer_name": name,
                "contact_person": str(r.get("contact_person") or ""),
                "contact_phone": str(r.get("contact_phone") or ""),
                "address": str(r.get("address") or ""),
                "is_active": int(r.get("is_active") or 1),
            }
        )
    return out


def _customer_row_for_api(row: dict) -> dict:
    cn = (row.get("customer_name") or row.get("unit_name") or row.get("name") or "").strip()
    return {
        "id": row["id"],
        "name": cn,
        "customer_name": cn,
        "contact_person": row.get("contact_person") or "",
        "contact_phone": row.get("contact_phone") or "",
        "contact_address": row.get("address") or row.get("contact_address") or "",
        "address": row.get("address") or "",
        "is_active": int(row.get("is_active") or 1),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


def _customer_row_matches_keyword(row: dict, keyword: str) -> bool:
    """大小写不敏感子串匹配，供 /customers/list 可选 keyword 过滤。"""
    k = (keyword or "").strip().lower()
    if not k:
        return True
    for key in (
        "customer_name",
        "name",
        "unit_name",
        "contact_person",
        "contact_phone",
        "phone",
        "company",
        "address",
        "contact_address",
    ):
        v = row.get(key)
        if v is None:
            continue
        if k in str(v).lower():
            return True
    return False


def _customer_find_by_id(customer_id: int) -> dict | None:
    for row in _load_customers_rows():
        if int(row.get("id") or 0) == int(customer_id):
            return dict(row)
    return None


def _customer_body_name_contact(body: dict) -> tuple[str, str, str, str]:
    name = (
        body.get("customer_name")
        or body.get("name")
        or body.get("unit_name")
        or ""
    )
    name = str(name).strip()
    cp = str(body.get("contact_person") or "").strip()
    ph = str(body.get("contact_phone") or "").strip()
    addr = str(body.get("contact_address") or body.get("address") or "").strip()
    return name, cp, ph, addr


def _customers_write_raise() -> None:
    try:
        from sqlalchemy import inspect

        eng = get_sync_engine()
        insp = inspect(eng)
        if "purchase_units" not in insp.get_table_names():
            raise HTTPException(
                status_code=503,
                detail="PostgreSQL 库中缺少 purchase_units 表，无法写入客户。",
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=503,
            detail=f"无法校验 PostgreSQL 库结构: {e}",
        ) from e


def _products_write_raise(request: Request) -> None:
    """二级写入令牌 + 校验 PostgreSQL 中存在 ``products`` 表。"""
    verify_db_write_token_header(request)
    try:
        from sqlalchemy import inspect

        eng = get_sync_engine()
        insp = inspect(eng)
        if "products" not in insp.get_table_names():
            raise HTTPException(
                status_code=503,
                detail="PostgreSQL 库中缺少 products 表，无法写入产品。",
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=503,
            detail=f"无法校验 PostgreSQL 库结构: {e}",
        ) from e


def _product_parse_id(raw: object) -> int | None:
    if raw is None or raw is False:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, int):
        return raw if raw > 0 else None
    try:
        n = int(str(raw).strip())
        return n if n > 0 else None
    except (TypeError, ValueError):
        return None


def _product_parse_quantity(raw: object) -> int:
    if raw is None or raw == "":
        return 0
    try:
        return int(float(str(raw).strip()))
    except (TypeError, ValueError):
        return 0


def _product_parse_is_active(raw: object) -> bool | None:
    """``None``：调用方不显式更新 ``is_active``。"""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, (int, float)):
        return int(raw) != 0
    s = str(raw).strip().lower()
    if s in ("0", "false", "no", "off"):
        return False
    if s in ("1", "true", "yes", "on"):
        return True
    return None


def _pg_purchase_unit_active_sql(column: str = "is_active") -> str:
    """与读侧一致的「活跃」条件（兼容 boolean / smallint）。"""
    c = _sql_ident(column)
    return f"({c} IS NULL OR {c} = true OR CAST({c} AS INTEGER) = 1)"


def _customer_pg_engine_insp():
    from sqlalchemy import inspect

    eng = get_sync_engine()
    return eng, inspect(eng)


def _customer_pg_products_has_unit(insp) -> bool:
    if "products" not in insp.get_table_names():
        return False
    return "unit" in {c["name"] for c in insp.get_columns("products")}


def _pg_expr_norm_unit(column_sql: str) -> str:
    """PostgreSQL：与 strip 接近的 unit 规范化（全角空格 U+3000、NBSP、制表符）。"""
    return (
        f"lower(btrim(replace(replace(replace(coalesce({column_sql}, ''), CHR(12288), ' '), "
        f"CHR(160), ' '), CHR(9), ' ')))"
    )


def _products_delete_by_unit_pg(eng, unit_name: str) -> int:
    from sqlalchemy import inspect, text

    un = (unit_name or "").strip()
    if not un:
        return 0
    insp = inspect(eng)
    if not _customer_pg_products_has_unit(insp):
        return 0
    pcols = {c["name"] for c in insp.get_columns("products")}
    col = _pg_expr_norm_unit("CAST(unit AS TEXT)")
    prm = _pg_expr_norm_unit("CAST(:u AS TEXT)")
    where_parts = [f"{col} = {prm}"]
    params: dict[str, object] = {"u": un}
    append_mod_scope_where(where_parts, params, pcols)
    where_sql = " AND ".join(where_parts)
    with eng.begin() as conn:
        r = conn.execute(
            text(f"DELETE FROM products WHERE {where_sql}"),
            params,
        )
        return int(r.rowcount or 0)


def _purchase_units_delete_by_norm_unit_pg(eng, unit_name: str) -> int:
    """按规范化名称删除 purchase_units 中所有匹配行（与 products.unit 比对规则一致）。"""
    from sqlalchemy import inspect, text

    un = (unit_name or "").strip()
    if not un:
        return 0
    insp = inspect(eng)
    if "purchase_units" not in insp.get_table_names():
        return 0
    cols = {c["name"] for c in insp.get_columns("purchase_units")}
    if "unit_name" not in cols:
        return 0
    col = _pg_expr_norm_unit("CAST(unit_name AS TEXT)")
    prm = _pg_expr_norm_unit("CAST(:u AS TEXT)")
    where_parts = [f"{col} = {prm}"]
    params: dict[str, object] = {"u": un}
    append_mod_scope_where(where_parts, params, cols)
    where_sql = " AND ".join(where_parts)
    with eng.begin() as conn:
        r = conn.execute(
            text(f"DELETE FROM purchase_units WHERE {where_sql}"),
            params,
        )
        return int(r.rowcount or 0)


def _customers_delete_by_norm_name_pg(eng, insp, customer_name: str) -> int:
    """按规范化名称删除 customers 表中所有匹配行。"""
    from sqlalchemy import text

    cn = (customer_name or "").strip()
    if not cn or "customers" not in insp.get_table_names():
        return 0
    cols_names = {c["name"] for c in insp.get_columns("customers")}
    name_col = next(
        (c for c in ("customer_name", "name", "客户名称") if c in cols_names),
        None,
    )
    if not name_col:
        return 0
    col = _pg_expr_norm_unit(f"CAST({_sql_ident(name_col)} AS TEXT)")
    prm = _pg_expr_norm_unit("CAST(:u AS TEXT)")
    where_parts = [f"{col} = {prm}"]
    params: dict[str, object] = {"u": cn}
    append_mod_scope_where(where_parts, params, cols_names)
    where_sql = " AND ".join(where_parts)
    with eng.begin() as conn:
        r = conn.execute(
            text(f"DELETE FROM {_sql_ident('customers')} WHERE {where_sql}"),
            params,
        )
        return int(r.rowcount or 0)


def _purchase_units_delete_by_id_pg(eng, customer_id: int) -> int:
    """按主键删除 purchase_units 中对应记录（兜底清理）。"""
    from sqlalchemy import inspect, text

    insp = inspect(eng)
    if "purchase_units" not in insp.get_table_names():
        return 0
    cols = {c["name"] for c in insp.get_columns("purchase_units")}
    if "id" not in cols:
        return 0
    where_parts = ["id = :id"]
    params: dict[str, object] = {"id": int(customer_id)}
    append_mod_scope_where(where_parts, params, cols)
    where_sql = " AND ".join(where_parts)
    with eng.begin() as conn:
        r = conn.execute(
            text(f"DELETE FROM purchase_units WHERE {where_sql}"),
            params,
        )
        return int(r.rowcount or 0)


def _customers_delete_by_id_pg(eng, insp, customer_id: int) -> int:
    """按 customers 主键删除（兼容 id / customer_id）。"""
    from sqlalchemy import text

    if "customers" not in insp.get_table_names():
        return 0
    cols = {c["name"] for c in insp.get_columns("customers")}
    id_col = "id" if "id" in cols else ("customer_id" if "customer_id" in cols else None)
    if not id_col:
        return 0
    where_parts = [f"{_sql_ident(id_col)} = :id"]
    params: dict[str, object] = {"id": int(customer_id)}
    append_mod_scope_where(where_parts, params, cols)
    where_sql = " AND ".join(where_parts)
    with eng.begin() as conn:
        r = conn.execute(
            text(f"DELETE FROM {_sql_ident('customers')} WHERE {where_sql}"),
            params,
        )
        return int(r.rowcount or 0)


def _products_unit_replace_pg(eng, old_name: str, new_name: str) -> None:
    o = (old_name or "").strip()
    n = (new_name or "").strip()
    if not o or not n or o == n:
        return
    from sqlalchemy import inspect, text

    insp = inspect(eng)
    if not _customer_pg_products_has_unit(insp):
        return
    pcols = {c["name"] for c in insp.get_columns("products")}
    col = _pg_expr_norm_unit("CAST(unit AS TEXT)")
    prm = _pg_expr_norm_unit("CAST(:oo AS TEXT)")
    where_parts = [f"{col} = {prm}"]
    params: dict[str, object] = {"nn": n, "oo": o}
    append_mod_scope_where(where_parts, params, pcols)
    where_sql = " AND ".join(where_parts)
    with eng.connect() as conn:
        conn.execute(
            text(f"UPDATE products SET unit = :nn WHERE {where_sql}"),
            params,
        )
        conn.commit()


def _customer_pg_row_select_sql(insp) -> tuple[str, list[str]]:
    """返回 SELECT 片段与 purchase_units 存在的列名列表。"""
    have = {c["name"] for c in insp.get_columns("purchase_units")}
    want = [
        "id",
        "unit_name",
        "contact_person",
        "contact_phone",
        "address",
        "is_active",
        "created_at",
        "updated_at",
    ]
    sel = [c for c in want if c in have]
    if not sel or "id" not in sel:
        raise HTTPException(status_code=503, detail="purchase_units 表结构不完整（缺少 id）。")
    parts = []
    for c in sel:
        if c == "unit_name":
            parts.append(f"{_sql_ident(c)} AS unit_name")
        else:
            parts.append(_sql_ident(c))
    return ", ".join(parts), sel


def _customer_pg_fetch_by_id(eng, insp, customer_id: int) -> dict:
    from sqlalchemy import text

    sel_sql, _ = _customer_pg_row_select_sql(insp)
    pu_cols = {c["name"] for c in insp.get_columns("purchase_units")}
    where_parts = ["id = :id"]
    params: dict[str, object] = {"id": int(customer_id)}
    append_mod_scope_where(where_parts, params, pu_cols)
    where_sql = " AND ".join(where_parts)
    with eng.connect() as conn:
        r = conn.execute(
            text(f"SELECT {sel_sql} FROM purchase_units WHERE {where_sql}"),
            params,
        ).mappings().first()
    if not r:
        raise HTTPException(status_code=404, detail="客户不存在")
    d = dict(r)
    d["customer_name"] = d.get("unit_name") or ""
    return _customer_row_for_api(d)


def _customer_pg_insert(name: str, cp: str, ph: str, addr: str) -> dict:
    from sqlalchemy import inspect, text

    eng, insp = _customer_pg_engine_insp()
    pu_cols = {c["name"]: c for c in insp.get_columns("purchase_units")}
    if "unit_name" not in pu_cols:
        raise HTTPException(status_code=503, detail="purchase_units 缺少 unit_name 列。")
    now = datetime.utcnow()
    with eng.connect() as conn:
        dup_parts = ["unit_name = :n", _pg_purchase_unit_active_sql()]
        dup_bind: dict[str, object] = {"n": name}
        append_mod_scope_where(dup_parts, dup_bind, pu_cols)
        dup_sql = " AND ".join(dup_parts)
        dup = conn.execute(
            text(f"SELECT id FROM purchase_units WHERE {dup_sql}"),
            dup_bind,
        ).first()
        if dup:
            raise HTTPException(status_code=400, detail="客户名称已存在")
        col_pairs: list[tuple[str, str]] = [
            ("unit_name", "un"),
            ("contact_person", "cp"),
            ("contact_phone", "ph"),
            ("address", "addr"),
        ]
        bind: dict = {"un": name, "cp": cp, "ph": ph, "addr": addr}
        mid = scoped_mod_id()
        if "xcagi_mod_id" in pu_cols and mid:
            col_pairs.append(("xcagi_mod_id", "xmid"))
            bind["xmid"] = mid
        if "is_active" in pu_cols:
            col_pairs.append(("is_active", "ia"))
            t = str(pu_cols["is_active"].get("type") or "").lower()
            bind["ia"] = True if "bool" in t else 1
        if "created_at" in pu_cols:
            col_pairs.append(("created_at", "ca"))
            bind["ca"] = now
        if "updated_at" in pu_cols:
            col_pairs.append(("updated_at", "ua"))
            bind["ua"] = now
        cols_sql = ", ".join(_sql_ident(c) for c, _ in col_pairs)
        vals_sql = ", ".join(f":{bk}" for _, bk in col_pairs)
        r = conn.execute(
            text(
                f"INSERT INTO purchase_units ({cols_sql}) VALUES ({vals_sql}) RETURNING id"
            ),
            bind,
        )
        new_id = int(r.scalar_one())
        conn.commit()
    return _customer_pg_fetch_by_id(eng, insp, new_id)


def _customer_pg_update(customer_id: int, name: str, cp: str, ph: str, addr: str) -> dict:
    from sqlalchemy import inspect, text

    eng, insp = _customer_pg_engine_insp()
    pu_cols = {c["name"] for c in insp.get_columns("purchase_units")}
    with eng.connect() as conn:
        prev_parts = ["id = :id", _pg_purchase_unit_active_sql()]
        prev_bind: dict[str, object] = {"id": int(customer_id)}
        append_mod_scope_where(prev_parts, prev_bind, pu_cols)
        prev = conn.execute(
            text(f"SELECT id, unit_name FROM purchase_units WHERE {' AND '.join(prev_parts)}"),
            prev_bind,
        ).mappings().first()
        if not prev:
            raise HTTPException(status_code=404, detail="客户不存在或已删除")
        old_name = str(prev["unit_name"] or "").strip()
        clash_parts = [
            "unit_name = :n",
            "id != :id",
            _pg_purchase_unit_active_sql(),
        ]
        clash_bind: dict[str, object] = {"n": name, "id": int(customer_id)}
        append_mod_scope_where(clash_parts, clash_bind, pu_cols)
        clash = conn.execute(
            text(f"SELECT id FROM purchase_units WHERE {' AND '.join(clash_parts)}"),
            clash_bind,
        ).first()
        if clash:
            raise HTTPException(status_code=400, detail="客户名称与其他记录冲突")
        now = datetime.utcnow()
        upd_bind = {
            "un": name,
            "cp": cp,
            "ph": ph,
            "addr": addr,
            "id": int(customer_id),
        }
        mod_and = products_update_or_delete_mod_and(pu_cols, upd_bind)
        if "updated_at" in pu_cols:
            upd_bind["ua"] = now
            conn.execute(
                text(
                    "UPDATE purchase_units SET unit_name = :un, contact_person = :cp, "
                    "contact_phone = :ph, address = :addr, updated_at = :ua WHERE id = :id"
                    + mod_and
                ),
                upd_bind,
            )
        else:
            conn.execute(
                text(
                    "UPDATE purchase_units SET unit_name = :un, contact_person = :cp, "
                    "contact_phone = :ph, address = :addr WHERE id = :id" + mod_and
                ),
                upd_bind,
            )
        conn.commit()
    if old_name and old_name != name.strip():
        _products_unit_replace_pg(eng, old_name, name.strip())
    return _customer_pg_fetch_by_id(eng, insp, customer_id)


def _customer_pg_select_customers_name_by_id(
    eng, insp, customer_id: int
) -> tuple[str, str] | None:
    """若 customers 表存在且主键命中，返回 (客户名称, id列名)；否则 None。"""
    from sqlalchemy import text

    if "customers" not in insp.get_table_names():
        return None
    cols = {c["name"] for c in insp.get_columns("customers")}
    id_col = "id" if "id" in cols else ("customer_id" if "customer_id" in cols else None)
    name_col = next(
        (c for c in ("customer_name", "name", "客户名称") if c in cols),
        None,
    )
    if not id_col or not name_col:
        return None
    where_parts = [f"{_sql_ident(id_col)} = :id"]
    cbind: dict[str, object] = {"id": int(customer_id)}
    append_mod_scope_where(where_parts, cbind, cols)
    with eng.connect() as conn:
        r = conn.execute(
            text(
                f"SELECT {_sql_ident(name_col)} AS nm FROM {_sql_ident('customers')} "
                f"WHERE {' AND '.join(where_parts)}"
            ),
            cbind,
        ).mappings().first()
        if not r:
            return None
        nm = str(r["nm"] or "").strip()
        return (nm, id_col)


def _customer_pg_delete_anywhere(customer_id: int) -> None:
    """
    客户管理「删除」：按解析出的客户名称，同时清理
    products（该 unit）、purchase_units（同名）、customers（同名），
    避免只删 customers 行而 purchase_units 仍在、产品页/下拉仍出现该单位。
    """
    from sqlalchemy import text

    eng, insp = _customer_pg_engine_insp()
    cid = int(customer_id)
    resolved_name: str | None = None

    if "purchase_units" in insp.get_table_names():
        pu_cols = {c["name"] for c in insp.get_columns("purchase_units")}
        where_parts = ["id = :id"]
        pbind: dict[str, object] = {"id": cid}
        append_mod_scope_where(where_parts, pbind, pu_cols)
        with eng.connect() as conn:
            r = conn.execute(
                text(f"SELECT unit_name FROM purchase_units WHERE {' AND '.join(where_parts)}"),
                pbind,
            ).first()
            if r:
                resolved_name = str(r[0] or "").strip() or None

    if not resolved_name:
        csel = _customer_pg_select_customers_name_by_id(eng, insp, cid)
        if csel:
            nm, _id_col = csel
            resolved_name = (nm or "").strip() or None

    if not resolved_name:
        hint = _customer_find_by_id(cid)
        if hint and int(hint.get("id") or 0) == cid:
            resolved_name = str(hint.get("customer_name") or "").strip() or None

    n_prod = 0
    n_pu = 0
    n_cu = 0
    if resolved_name:
        n_prod += _products_delete_by_unit_pg(eng, resolved_name)
        n_pu += _purchase_units_delete_by_norm_unit_pg(eng, resolved_name)
        n_cu += _customers_delete_by_norm_name_pg(eng, insp, resolved_name)

    # 兜底：即使名称不一致/为空，也按 ID 清理主表，确保“删客户=数据库里该客户记录清空”。
    n_pu += _purchase_units_delete_by_id_pg(eng, cid)
    n_cu += _customers_delete_by_id_pg(eng, insp, cid)

    if n_prod == 0 and n_pu == 0 and n_cu == 0:
        raise HTTPException(status_code=404, detail="客户不存在")


def _customer_delete_unified(customer_id: int) -> None:
    _customer_pg_delete_anywhere(customer_id)


def _units_select_data_unified() -> list[dict]:
    """
    与客户列表 API（_load_customers_rows）同源的唯一名称，再并入 products.unit 中尚未出现的名称；
    供出货/产品单位下拉，避免与 /api/customers 两套口径漂移。
    """
    rows = _load_customers_rows()
    seen: set[str] = set()
    staged: list[tuple[str, int | None]] = []
    for row in rows:
        name = str(row.get("customer_name") or row.get("name") or "").strip()
        if not name:
            continue
        lk = name.lower()
        if lk in seen:
            continue
        seen.add(lk)
        rid = row.get("id")
        try:
            oid = int(rid)
        except (TypeError, ValueError):
            oid = None
        staged.append((name, oid))

    out: list[dict] = []
    max_id = 0
    for _, oid in staged:
        if oid is not None:
            max_id = max(max_id, oid)
    next_id = max_id
    for name, oid in staged:
        if oid is not None:
            out.append({"id": oid, "name": name, "symbol": name})
        else:
            next_id += 1
            out.append({"id": next_id, "name": name, "symbol": name})
            max_id = next_id

    distinct = _distinct_units_from_products_db()
    syn = 0
    for d in distinct:
        name = str(d.get("name") or "").strip()
        if not name:
            continue
        lk = name.lower()
        if lk in seen:
            continue
        if name in TRIVIAL_MEASURE_UNITS:
            continue
        seen.add(lk)
        syn += 1
        out.append({"id": max_id + syn, "name": name, "symbol": name})
    return out


def _products_units_for_select() -> dict:
    """
    出货/产品单位下拉：与客户列表同源（_units_select_data_unified）；
    若仍为空则用 products 去重单位；再否则返回空列表（不再注入计量词占位）。
    """
    data = _units_select_data_unified()
    if data:
        return {"success": True, "data": data}
    distinct = _distinct_units_from_products_db()
    if distinct:
        return {"success": True, "data": distinct}
    return {"success": True, "data": []}


@router.get("/products/units")
async def products_units(request: Request) -> dict:
    verify_db_read_token_header(request)
    return _products_units_for_select()


@router.get("/shipment/shipment-records/units")
@router.get("/shipment/shipment-records/units/")
async def shipment_records_units() -> dict:
    """出货记录页单位下拉：与产品单位同源（purchase_units / products.unit / 默认计量）。"""
    return _products_units_for_select()


@router.get("/purchase_units")
@router.get("/purchase_units/")
async def purchase_units_list() -> dict:
    """与产品页单位同源合并列表；含 success 供 CreateOrderView 等旧前端判断。"""
    return {"success": True, "data": _merged_purchase_unit_entries()}


@router.get("/customers")
@router.get("/customers/")
async def customers_all() -> dict:
    """兼容旧版 GET /api/customers（body 含 data 列表）。"""
    return {"success": True, "data": _load_customers_rows()}


@router.get("/customers/match")
@router.get("/customers/match/")
async def customers_match(
    customer_name: str | None = Query(None, description="已抽取的客户名或片段"),
    context: str | None = Query(None, description="可选：本轮用户原始话术，供 extract_customer_name"),
) -> dict:
    """
    将展示用客户名对齐到数据库 purchase_units / 客户合并列表（与 ``run_sales_contract_generation`` 同源逻辑）。
    须注册在 ``/customers/{customer_id}`` 之前，避免 ``match`` 被当成整数 id。
    """
    from backend.shared_utils import extract_customer_name, find_matching_customer

    field = (customer_name or "").strip()
    ctx = (context or "").strip()
    if not field and not ctx:
        return {"success": True, "matched": None, "input": "", "extracted": None, "display": None}

    if _business_mod_json_block():
        return {"success": True, "matched": None, "input": field or ctx, "extracted": None, "display": None}

    extracted = extract_customer_name(field) or extract_customer_name(ctx)
    if extracted:
        matched = find_matching_customer(extracted)
    else:
        matched = find_matching_customer(field or ctx)

    chosen = (matched or "").strip()
    base = field or ctx
    display = chosen or base
    return {
        "success": True,
        "input": base,
        "extracted": extracted,
        "matched": chosen or None,
        "display": display,
    }


@router.get("/customers/list")
async def customers_list(
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=20, ge=1, le=500),
    keyword: str | None = Query(None, description="按名称、电话、地址等子串过滤（大小写不敏感）"),
) -> dict:
    """CustomersView.vue 等使用的分页列表。"""
    rows = _load_customers_rows()
    kw = (keyword or "").strip()
    if kw:
        rows = [r for r in rows if _customer_row_matches_keyword(r, kw)]
    total = len(rows)
    offset = (page - 1) * per_page
    out = {
        "success": True,
        "data": rows[offset : offset + per_page],
        "total": total,
    }
    if total == 0:
        ch = _customers_schema_hint_if_empty()
        if ch:
            out["schema_hint"] = ch
    return out


@router.get("/customers/{customer_id}")
@router.get("/customers/{customer_id}/")
async def customers_get_one(customer_id: int) -> dict:
    row = _customer_find_by_id(customer_id)
    if not row:
        raise HTTPException(status_code=404, detail="客户不存在")
    return {"success": True, "data": _customer_row_for_api(row)}


@router.post("/customers")
@router.post("/customers/")
async def customers_create(body: dict = Body(default_factory=dict)) -> dict:
    _customers_write_raise()
    name, cp, ph, addr = _customer_body_name_contact(body)
    if not name:
        raise HTTPException(status_code=400, detail="客户名称不能为空")
    data = _customer_pg_insert(name, cp, ph, addr)
    return {"success": True, "data": data}


@router.put("/customers/{customer_id}")
@router.put("/customers/{customer_id}/")
async def customers_update(customer_id: int, body: dict = Body(default_factory=dict)) -> dict:
    _customers_write_raise()
    name, cp, ph, addr = _customer_body_name_contact(body)
    if not name:
        raise HTTPException(status_code=400, detail="客户名称不能为空")
    data = _customer_pg_update(customer_id, name, cp, ph, addr)
    return {"success": True, "data": data}


@router.delete("/customers/{customer_id}")
@router.delete("/customers/{customer_id}/")
async def customers_delete(customer_id: int) -> dict:
    _customers_write_raise()
    _customer_delete_unified(customer_id)
    return {"success": True, "message": "已删除"}


@router.post("/customers/batch-delete")
@router.post("/customers/batch-delete/")
async def customers_batch_delete(body: dict = Body(default_factory=dict)) -> dict:
    _customers_write_raise()
    ids = body.get("ids") or body.get("customer_ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="ids 须为非空数组")
    deleted = 0
    errors: list[str] = []
    for raw in ids:
        try:
            cid = int(raw)
        except (TypeError, ValueError):
            errors.append(str(raw))
            continue
        try:
            _customer_delete_unified(cid)
            deleted += 1
        except HTTPException as e:
            if e.status_code == 404:
                errors.append(str(cid))
            else:
                raise
    return {
        "success": True,
        "message": f"已删除 {deleted} 条",
        "deleted": deleted,
        "skipped": errors,
    }


@router.post("/customers/import")
@router.post("/customers/import/")
async def customers_import(file: UploadFile = File(...)) -> dict:
    """
    上传 ``.xlsx`` / ``.xls``：首工作表、首行表头；写入 ``customers``，并同步 ``purchase_units``（与写入校验一致）。
    """
    _customers_write_raise()
    gate = _business_mod_json_block()
    if gate:
        return gate
    try:
        content = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取上传文件失败：{e}") from e
    from backend.customers_excel_import import run_customers_excel_import_bytes

    out = run_customers_excel_import_bytes(content)
    if not out.get("success"):
        msg = str(out.get("message") or out.get("error") or "导入失败")
        raise HTTPException(status_code=400, detail=msg)
    return {"success": True, "data": out}


@router.get("/customers/export")
@router.get("/customers/export/")
async def customers_export_stub() -> dict:
    raise HTTPException(
        status_code=501,
        detail="客户 Excel 导出尚未在 FastAPI 兼容层实现；请接回 XCAGI 全量后端或使用本地模板导出。",
    )


@router.get("/products/list")
async def products_list(
    request: Request,
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=20, ge=1, le=500),
    keyword: str | None = Query(None),
    unit: str | None = Query(None),
) -> dict:
    verify_db_read_token_header(request)
    try:
        items, total, schema_hint = _load_products_list_impl_pg(page, per_page, keyword, unit)
        out: dict = {"success": True, "data": items, "total": total}
        if schema_hint:
            out["schema_hint"] = schema_hint
        return out
    except Exception as e:
        logger.exception("products list failed (postgresql)")
        return {"success": False, "message": str(e), "data": [], "total": 0}


@router.post("/products/resolve-name-hints")
@router.post("/products/resolve-name-hints/")
async def products_resolve_name_hints(request: Request, body: dict = Body(default_factory=dict)) -> dict:
    """
    按中文/口语品名提示在主数据 ``products`` 中解析型号（与销售合同 ``name_hint`` 规则一致）。

    body: ``{ "hints": ["慢干水", "PU稀释剂"], "unit"?: "购买单位（可选，与 list 的 unit 一致）" }``
    返回 ``{ success, data: [{ hint, status, product?, candidates? }] }``。
    """
    verify_db_read_token_header(request)
    raw = body.get("hints") or body.get("names") or []
    if not isinstance(raw, list):
        return {"success": False, "message": "hints 须为字符串数组", "data": []}
    hints = [str(x).strip() for x in raw if str(x).strip()]
    if not hints:
        return {"success": False, "message": "hints 不能为空", "data": []}
    unit = str(body.get("unit") or body.get("purchase_unit") or "").strip() or None

    gate = _business_mod_json_block()
    if gate:
        return {**gate, "data": []}

    from backend.product_name_resolve import resolve_product_name_hints

    data = resolve_product_name_hints(hints, purchase_unit=unit)
    return {"success": True, "data": data}


@router.post("/products/update")
@router.post("/products/update/")
async def products_update(request: Request, body: dict = Body(default_factory=dict)) -> dict:
    """
    单条更新（XCAGI 前端 ``productsApi.updateProduct`` → ``POST /api/products/update``）。
    body: ``id`` 必填；其余字段按表列存在情况写入。
    """
    _products_write_raise(request)
    gate = _business_mod_json_block()
    if gate:
        return gate
    from sqlalchemy import inspect, text

    from backend.products_bulk_import import _parse_price

    pid = _product_parse_id(body.get("id"))
    if pid is None:
        raise HTTPException(status_code=400, detail="id 无效或缺失")

    eng = get_sync_engine()
    insp = inspect(eng)
    col_names = {c["name"] for c in insp.get_columns("products")}
    if not {"id", "model_number", "name"}.issubset(col_names):
        raise HTTPException(
            status_code=503,
            detail="products 表缺少必要列（至少需要 id、model_number、name）。",
        )

    name = str(body.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="产品名称不能为空")

    sets: list[str] = []
    params: dict[str, object] = {"pid": pid}

    if "model_number" in col_names:
        mn = body.get("model_number")
        sets.append("model_number = :model_number")
        params["model_number"] = (str(mn).strip() if mn is not None else "")[:120]

    sets.append("name = :name")
    params["name"] = name[:500]

    if "specification" in col_names:
        sp = body.get("specification")
        sets.append("specification = :specification")
        params["specification"] = None if sp is None else str(sp)

    if "price" in col_names:
        sets.append("price = :price")
        params["price"] = _parse_price(body.get("price"))

    if "quantity" in col_names:
        sets.append("quantity = :quantity")
        params["quantity"] = _product_parse_quantity(body.get("quantity"))

    if "unit" in col_names:
        un = body.get("unit")
        sets.append("unit = :unit")
        params["unit"] = (str(un).strip() if un is not None else "")[:200]

    if "description" in col_names:
        dv = body.get("description")
        sets.append("description = :description")
        params["description"] = None if dv is None else str(dv)

    if "category" in col_names:
        cv = body.get("category")
        sets.append("category = :category")
        params["category"] = None if cv is None else str(cv)[:200]

    if "brand" in col_names:
        bv = body.get("brand")
        sets.append("brand = :brand")
        params["brand"] = None if bv is None else str(bv)[:200]

    if "is_active" in col_names:
        ia = _product_parse_is_active(body.get("is_active"))
        if ia is not None:
            sets.append("is_active = :is_active")
            params["is_active"] = ia

    if "updated_at" in col_names:
        sets.append("updated_at = NOW()")

    if not sets:
        raise HTTPException(status_code=400, detail="没有可更新的列")

    mod_and = products_update_or_delete_mod_and(col_names, params)
    sql = "UPDATE products SET " + ", ".join(sets) + " WHERE id = :pid" + mod_and
    try:
        with eng.begin() as conn:
            r = conn.execute(text(sql), params)
            if r.rowcount == 0:
                raise HTTPException(status_code=404, detail="产品不存在")
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("products update failed")
        raise HTTPException(status_code=500, detail=f"更新失败：{e}") from e

    return {"success": True, "data": {"id": pid}}


@router.post("/products/add")
@router.post("/products/add/")
async def products_add(request: Request, body: dict = Body(default_factory=dict)) -> dict:
    """添加产品（``POST /api/products/add``）。"""
    _products_write_raise(request)
    gate = _business_mod_json_block()
    if gate:
        return gate
    from sqlalchemy import inspect, text

    from backend.products_bulk_import import _norm_model, _parse_price

    name = str(body.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="产品名称不能为空")

    spec = str(body.get("specification") or "").strip()
    mn_raw = body.get("model_number")
    model_number = str(mn_raw).strip() if mn_raw is not None else ""
    if not model_number:
        model_number = _norm_model("", name, spec)

    eng = get_sync_engine()
    insp = inspect(eng)
    col_names = {c["name"] for c in insp.get_columns("products")}
    if not {"model_number", "name"}.issubset(col_names):
        raise HTTPException(
            status_code=503,
            detail="products 表缺少必要列（至少需要 model_number、name）。",
        )

    icols: list[str] = []
    params: dict[str, object] = {}

    def _add(col: str, val: object) -> None:
        if col in col_names:
            icols.append(col)
            params[col] = val

    _add("model_number", model_number[:120])
    _add("name", name[:500])
    _add("specification", spec or None)
    _add("price", _parse_price(body.get("price")))
    _add("quantity", _product_parse_quantity(body.get("quantity")))
    unit = str(body.get("unit") or "").strip()[:200]
    _add("unit", unit)
    _add("description", str(body.get("description") or "") if body.get("description") is not None else None)
    _add("category", str(body.get("category") or "")[:200] if body.get("category") is not None else None)
    _add("brand", str(body.get("brand") or "")[:200] if body.get("brand") is not None else None)
    ia = _product_parse_is_active(body.get("is_active"))
    if ia is not None and "is_active" in col_names:
        _add("is_active", ia)

    if not icols:
        raise HTTPException(status_code=500, detail="无法构造 INSERT 列")

    mid = scoped_mod_id()
    if "xcagi_mod_id" in col_names and mid:
        icols.append("xcagi_mod_id")
        params["xcagi_mod_id"] = mid

    quoted = ", ".join(_sql_ident(c) for c in icols)
    ph = ", ".join(f":{c}" for c in icols)
    sql = f"INSERT INTO products ({quoted}) VALUES ({ph}) RETURNING id"

    try:
        with eng.begin() as conn:
            new_id = conn.execute(text(sql), params).scalar_one()
    except Exception as e:
        logger.exception("products add failed")
        raise HTTPException(status_code=500, detail=f"添加失败：{e}") from e

    return {"success": True, "data": {"id": int(new_id)}}


@router.post("/products/delete")
@router.post("/products/delete/")
async def products_delete(request: Request, body: dict = Body(default_factory=dict)) -> dict:
    """删除单条（``POST /api/products/delete``，body: ``{ id }``）。"""
    _products_write_raise(request)
    gate = _business_mod_json_block()
    if gate:
        return gate
    from sqlalchemy import text

    pid = _product_parse_id(body.get("id"))
    if pid is None:
        raise HTTPException(status_code=400, detail="id 无效或缺失")

    eng = get_sync_engine()
    from sqlalchemy import inspect

    insp = inspect(eng)
    pcols = {c["name"] for c in insp.get_columns("products")}
    del_params: dict[str, object] = {"pid": pid}
    mod_and = products_update_or_delete_mod_and(pcols, del_params)
    try:
        with eng.begin() as conn:
            r = conn.execute(
                text("DELETE FROM products WHERE id = :pid" + mod_and),
                del_params,
            )
            if r.rowcount == 0:
                raise HTTPException(status_code=404, detail="产品不存在")
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("products delete failed")
        raise HTTPException(status_code=500, detail=f"删除失败：{e}") from e

    return {"success": True, "message": "已删除"}


@router.post("/products/batch-delete")
@router.post("/products/batch-delete/")
async def products_batch_delete(request: Request, body: dict = Body(default_factory=dict)) -> dict:
    """批量删除（body: ``{ ids: [...] }``）。"""
    _products_write_raise(request)
    gate = _business_mod_json_block()
    if gate:
        return gate
    from sqlalchemy import text

    ids = body.get("ids") or body.get("product_ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="ids 须为非空数组")

    eng = get_sync_engine()
    from sqlalchemy import inspect

    insp = inspect(eng)
    pcols = {c["name"] for c in insp.get_columns("products")}
    deleted = 0
    skipped: list[str] = []
    try:
        with eng.begin() as conn:
            for raw in ids:
                pid = _product_parse_id(raw)
                if pid is None:
                    skipped.append(str(raw))
                    continue
                del_params = {"pid": pid}
                mod_and = products_update_or_delete_mod_and(pcols, del_params)
                r = conn.execute(
                    text("DELETE FROM products WHERE id = :pid" + mod_and),
                    del_params,
                )
                if r.rowcount:
                    deleted += 1
                else:
                    skipped.append(str(raw))
    except Exception as e:
        logger.exception("products batch-delete failed")
        raise HTTPException(status_code=500, detail=f"批量删除失败：{e}") from e

    return {
        "success": True,
        "message": f"已删除 {deleted} 条",
        "deleted": deleted,
        "skipped": skipped,
    }


def _customers_schema_hint_if_empty() -> str | None:
    """客户列表无数据时，提示是否缺表（有表但无行则返回 None）。"""
    try:
        from sqlalchemy import inspect

        eng = get_sync_engine()
        names = set(inspect(eng).get_table_names())
    except Exception as e:
        return f"无法连接 PostgreSQL：{e}。请检查 DATABASE_URL。"

    has_c = "customers" in names
    has_pu = "purchase_units" in names
    has_p = "products" in names
    if not has_c and not has_pu:
        return (
            "当前库缺少 customers 与 purchase_units 表，无法展示客户。请执行 scripts/pg_init_xcagi_core.sql "
            "（会创建 purchase_units / products；可选 customers）。"
        )
    if not has_pu and has_p:
        return (
            "当前库缺少 purchase_units；客户页将仅从 customers 或 products.unit 推断。"
            "建议执行 scripts/pg_init_xcagi_core.sql 补齐 purchase_units。"
        )
    if not has_p and not has_c:
        return "当前库缺少 customers 与 products 表，客户与产品数据均为空。请执行 scripts/pg_init_xcagi_core.sql。"
    return None


def _products_price_list_word_response(
    unit: str | None,
    keyword: str | None,
    export_date: str | None,
    template_slug: str | None = None,
) -> Response:
    """
    使用 ``document_templates``（``price_list_docx``）或降级路径导出报价表：
    写入客户名称（购买单位）、报价日期，并在首表填入型号/名称/规格/单价。
    """
    from backend.shell.mod_business_scope import business_data_exposed, business_data_hidden_reason

    if not business_data_exposed():
        raise HTTPException(
            status_code=503,
            detail=business_data_hidden_reason() or "扩展 Mod 未就绪，无法导出价格表。",
        )
    tpl = resolve_price_list_docx_template(template_slug)
    if not tpl:
        raise HTTPException(
            status_code=404,
            detail="未找到 Word 模板：请放置 424/document_templates/price_list_default.docx 或通过模板库登记，或设置 FHD_PRICE_LIST_DOCX_TEMPLATE。",
        )
    rows = _load_products_all_for_export(keyword, unit)
    customer = (unit or "").strip()
    try:
        body = build_price_list_docx_bytes(
            tpl,
            customer_name=customer,
            quote_date=export_date,
            products=rows,
        )
    except Exception as e:
        logger.exception("products export docx failed")
        raise HTTPException(status_code=500, detail=f"生成 Word 失败：{e}") from e

    today = date.today().strftime("%Y-%m-%d")
    label = customer or "全部单位"
    utf8_name = f"产品价格表_{label}_{today}.docx"
    disp = (
        "attachment; filename=\"price-list.docx\"; filename*=UTF-8''"
        + quote(utf8_name, safe="")
    )
    return Response(
        content=body,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": disp},
    )


@router.get("/products/price-list-export")
@router.get("/products/price-list-export/")
async def products_price_list_export(
    request: Request,
    unit: str | None = Query(None),
    keyword: str | None = Query(None),
    export_date: str | None = Query(None, description="报价日期 YYYY-MM-DD，默认当天"),
    template_id: str | None = Query(
        None, description="模板 slug（GET /api/document-templates?role=price_list_docx）"
    ),
) -> Response:
    """与 ``export.docx`` 相同；路径不含点号，避免网关/静态规则误匹配导致 404。"""
    verify_db_read_token_header(request)
    return _products_price_list_word_response(unit, keyword, export_date, template_id)


@router.get("/products/export.docx")
@router.get("/products/export.docx/")
async def products_export_docx(
    request: Request,
    unit: str | None = Query(None),
    keyword: str | None = Query(None),
    export_date: str | None = Query(None, description="报价日期 YYYY-MM-DD，默认当天"),
    template_id: str | None = Query(
        None, description="模板 slug（GET /api/document-templates?role=price_list_docx）"
    ),
) -> Response:
    verify_db_read_token_header(request)
    return _products_price_list_word_response(unit, keyword, export_date, template_id)


@router.get("/products/price-list-template-preview")
@router.get("/products/price-list-template-preview/")
async def products_price_list_template_preview(
    request: Request,
    template_id: str | None = Query(
        None, description="模板 slug（与 price-list-export 一致）"
    ),
) -> dict:
    """返回与 ``price-list-export`` 相同的 Word 文件之首表表头与示例行，供模板预览页动态对齐。"""
    verify_db_read_token_header(request)
    from backend.shell.mod_business_scope import business_data_exposed, business_data_hidden_reason

    if not business_data_exposed():
        raise HTTPException(
            status_code=503,
            detail=business_data_hidden_reason() or "扩展 Mod 未就绪。",
        )
    return build_price_list_template_preview_json(template_id)


def _excel_cell_to_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _form_bool(v: str | None) -> bool:
    if v is None:
        return False
    return str(v).strip().lower() in ("1", "true", "yes", "on")


def _pick_header_row(values_rows: list[list[str]]) -> tuple[int, list[str]]:
    """
    从前几行中选择最像表头的一行：
    - 非空单元格数最多
    - 至少 2 列非空
    """
    best_idx = 0
    best_non_empty = -1
    for i, row in enumerate(values_rows[:10]):
        non_empty = sum(1 for c in row if str(c or "").strip())
        if non_empty > best_non_empty:
            best_non_empty = non_empty
            best_idx = i
    header = values_rows[best_idx] if values_rows else []
    if sum(1 for c in header if str(c or "").strip()) < 2 and values_rows:
        best_idx = 0
        header = values_rows[0]
    return best_idx, header


def _detect_effective_col_count(values_rows: list[list[str]], fallback_cols: int) -> int:
    """检测有效列数：取所有行最后一个非空单元格位置的最大值。"""
    max_used = 0
    for row in values_rows:
        last_non_empty = 0
        for idx, cell in enumerate(row, start=1):
            if str(cell or "").strip():
                last_non_empty = idx
        if last_non_empty > max_used:
            max_used = last_non_empty
    if max_used > 0:
        return max_used
    return max(1, int(fallback_cols or 1))


def _detect_effective_row_count(values_rows: list[list[str]], fallback_rows: int) -> int:
    """检测有效行数：最后一个包含非空单元格的行。"""
    max_used = 0
    for row_idx, row in enumerate(values_rows, start=1):
        if any(str(c or "").strip() for c in row):
            max_used = row_idx
    if max_used > 0:
        return max_used
    return max(1, int(fallback_rows or 1))


def _excel_col_width_to_px(width: float | int | None) -> int:
    # Excel 列宽(字符单位)约换算到像素：px ≈ width * 7 + 5
    w = float(width or 8.43)
    return max(40, int(w * 7 + 5))


def _excel_row_height_to_px(height: float | int | None) -> int:
    # Excel 行高是 point，屏幕像素近似 point * 96 / 72
    h = float(height or 15.0)
    return max(20, int(h * 96.0 / 72.0))


def _merge_anchor_and_skip(
    ws, row_count: int, col_count: int
) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    merge_anchor: dict[tuple[int, int], tuple[int, int]] = {}
    merge_skip: set[tuple[int, int]] = set()
    for rg in ws.merged_cells.ranges:
        min_r, min_c, max_r, max_c = rg.min_row, rg.min_col, rg.max_row, rg.max_col
        if min_r > row_count or min_c > col_count:
            continue
        rowspan = max(1, min(max_r, row_count) - min_r + 1)
        colspan = max(1, min(max_c, col_count) - min_c + 1)
        merge_anchor[(min_r, min_c)] = (rowspan, colspan)
        for rr in range(min_r, min(max_r, row_count) + 1):
            for cc in range(min_c, min(max_c, col_count) + 1):
                if rr == min_r and cc == min_c:
                    continue
                merge_skip.add((rr, cc))
    return merge_anchor, merge_skip


def _serialize_cell_style(cell) -> dict:
    """可 JSON 序列化的轻量样式，供前端网格/模板预览复用。"""
    d: dict = {}
    try:
        font = cell.font
        if font:
            fd: dict = {}
            nm = getattr(font, "name", None)
            if nm:
                fd["name"] = str(nm)
            if getattr(font, "sz", None) is not None:
                try:
                    fd["size"] = float(font.sz)
                except (TypeError, ValueError):
                    pass
            if getattr(font, "b", None) is not None:
                fd["bold"] = bool(font.b)
            if getattr(font, "i", None) is not None:
                fd["italic"] = bool(font.i)
            col = getattr(font, "color", None)
            rgb = getattr(col, "rgb", None) if col is not None else None
            if rgb:
                fd["color"] = str(rgb)
            if fd:
                d["font"] = fd
        fill = cell.fill
        if fill is not None:
            fg = getattr(fill, "fgColor", None)
            rgb2 = getattr(fg, "rgb", None) if fg is not None else None
            if rgb2:
                d["fill"] = {"fgColor": str(rgb2)}
        al = cell.alignment
        if al is not None:
            ad: dict = {}
            if getattr(al, "horizontal", None):
                ad["horizontal"] = str(al.horizontal)
            if getattr(al, "vertical", None):
                ad["vertical"] = str(al.vertical)
            if getattr(al, "wrapText", None) is not None:
                ad["wrapText"] = bool(al.wrapText)
            if getattr(al, "textRotation", None) not in (None, 0):
                try:
                    ad["textRotation"] = int(al.textRotation)
                except (TypeError, ValueError):
                    pass
            if ad:
                d["alignment"] = ad
        border = cell.border
        if border is not None:
            sides = ("left", "right", "top", "bottom")
            bd: dict = {}
            for side in sides:
                b = getattr(border, side, None)
                if not b or not getattr(b, "style", None):
                    continue
                st = str(b.style)
                rgb3 = None
                colo = getattr(b, "color", None)
                if colo is not None:
                    rgb3 = getattr(colo, "rgb", None)
                bd[side] = {"style": st, "color": str(rgb3) if rgb3 else None}
            if bd:
                d["border"] = bd
    except Exception:
        return {}
    return d


def _build_grid_style_cache(ws, row_count: int, col_count: int, merge_skip: set[tuple[int, int]]) -> dict:
    from openpyxl.utils import get_column_letter

    styles: dict[str, dict] = {}
    cell_style_refs: dict[str, str] = {}
    max_styles = 256
    for r in range(1, row_count + 1):
        for c in range(1, col_count + 1):
            if (r, c) in merge_skip:
                continue
            cell = ws.cell(row=r, column=c)
            sd = _serialize_cell_style(cell)
            if not sd:
                continue
            key = hashlib.sha256(
                json.dumps(sd, sort_keys=True, ensure_ascii=False).encode("utf-8")
            ).hexdigest()[:16]
            if key not in styles:
                if len(styles) >= max_styles:
                    continue
                styles[key] = sd
            coord = f"{get_column_letter(c)}{r}"
            cell_style_refs[coord] = key
    return {"styles": styles, "cell_style_refs": cell_style_refs}


def _matrix_to_real_grid_rows(ws, values_rows: list[list[str]], row_count: int, col_count: int) -> list[list[dict]]:
    """
    转为 ExcelPreview 期望结构：
    [
      [{"col":1,"text":"A1","rowspan":1,"colspan":1}, ...],
      ...
    ]
    """
    from openpyxl.utils import get_column_letter

    merge_anchor, merge_skip = _merge_anchor_and_skip(ws, row_count, col_count)

    col_width_px: dict[int, int] = {}
    for c in range(1, col_count + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        col_width_px[c] = _excel_col_width_to_px(getattr(dim, "width", None))

    row_height_px: dict[int, int] = {}
    for r in range(1, row_count + 1):
        dim = ws.row_dimensions.get(r)
        row_height_px[r] = _excel_row_height_to_px(getattr(dim, "height", None))

    out: list[list[dict]] = []
    for r_idx in range(1, row_count + 1):
        row = values_rows[r_idx - 1] if r_idx - 1 < len(values_rows) else []
        line: list[dict] = []
        for col_idx in range(1, col_count + 1):
            if (r_idx, col_idx) in merge_skip:
                continue
            text = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            rowspan, colspan = merge_anchor.get((r_idx, col_idx), (1, 1))
            width_px = sum(col_width_px.get(col_idx + k, col_width_px.get(col_idx, 92)) for k in range(colspan))
            height_px = sum(row_height_px.get(r_idx + k, row_height_px.get(r_idx, 20)) for k in range(rowspan))
            line.append(
                {
                    "col": col_idx,
                    "text": str(text or ""),
                    "rowspan": rowspan,
                    "colspan": colspan,
                    "width_px": width_px,
                    "height_px": height_px,
                }
            )
        out.append(line)
    return out


_EXTRACT_GRID_MAX_SHEETS = 30


def _extract_single_sheet_bundle(
    wb,
    sn: str,
    _sheet_names_full: list[str],
    _persisted_rel: str,
    _original_filename: str,
) -> dict:
    """单工作表：fields / sample_rows / grid_preview / grid_style_cache / tables。"""
    ws = wb[sn]
    max_row = int(ws.max_row or 0)
    max_col = int(ws.max_column or 0)
    read_rows = min(max_row, 60)
    read_cols = min(max_col, 30)

    matrix: list[list[str]] = []
    for r in range(1, read_rows + 1):
        row_vals: list[str] = []
        for c in range(1, read_cols + 1):
            row_vals.append(_excel_cell_to_text(ws.cell(row=r, column=c).value))
        matrix.append(row_vals)

    effective_cols = _detect_effective_col_count(matrix, read_cols)
    effective_rows = _detect_effective_row_count(matrix, read_rows)
    matrix = matrix[:effective_rows]
    matrix = [row[:effective_cols] for row in matrix]

    empty_preview = {
        "rows": [],
        "max_row": max_row,
        "max_col": 0,
        "header_row_index": 1,
    }
    empty_cache = {"styles": {}, "cell_style_refs": {}}

    if not matrix:
        return {
            "sheet_name": sn,
            "fields": [],
            "sample_rows": [],
            "grid_preview": empty_preview,
            "grid_style_cache": empty_cache,
            "tables": [],
        }

    header_idx, header_row = _pick_header_row(matrix)
    headers: list[str] = []
    for i, h in enumerate(header_row, start=1):
        txt = str(h or "").strip()
        headers.append(txt or f"列{i}")

    fields = [{"label": h, "name": h, "type": "dynamic"} for h in headers if str(h).strip()]

    sample_rows: list[dict] = []
    for row in matrix[header_idx + 1 : header_idx + 11]:
        if not any(str(x or "").strip() for x in row):
            continue
        item = {}
        for i, key in enumerate(headers):
            item[key] = row[i] if i < len(row) else ""
        sample_rows.append(item)

    preview_row_count = min(20, len(matrix))
    grid_rows = _matrix_to_real_grid_rows(ws, matrix[:preview_row_count], preview_row_count, effective_cols)
    _, merge_skip = _merge_anchor_and_skip(ws, preview_row_count, effective_cols)
    grid_style_cache = _build_grid_style_cache(ws, preview_row_count, effective_cols, merge_skip)

    tables = [
        {
            "table_index": 0,
            "header_row": header_idx + 1,
            "fields": list(fields),
            "sample_rows": list(sample_rows[:12]),
        }
    ]

    grid_preview = {
        "rows": grid_rows,
        "max_row": max_row,
        "max_col": effective_cols,
        "header_row_index": header_idx + 1,
    }

    return {
        "sheet_name": sn,
        "fields": fields,
        "sample_rows": sample_rows,
        "grid_preview": grid_preview,
        "grid_style_cache": grid_style_cache,
        "tables": tables,
    }


def _bundle_to_sheet_entry(bundle: dict, sheet_index: int) -> dict:
    st = bundle["grid_style_cache"]
    return {
        "sheet_index": sheet_index,
        "sheet_name": bundle["sheet_name"],
        "fields": bundle["fields"],
        "sample_rows": bundle["sample_rows"],
        "grid_preview": bundle["grid_preview"],
        "style_cache": st,
        "tables": bundle["tables"],
    }


@router.post("/templates/extract-grid")
@router.post("/templates/extract-grid/")
async def templates_extract_grid(
    file: UploadFile = File(...),
    sheet_name: str | None = Form(default=None),
    analyze_all_sheets: str | None = Form(default=None),
) -> dict:
    """
    兼容前端 BusinessDockingView / 聊天「上传分析」：
    - fields + preview_data（sheet_names、sample_rows、grid_preview）
    - preview_data.grid_style_cache：单元格样式字典 + cell -> styleId
    - tables：主表区块（表头行 + 字段 + 样例）
    - analyze_all_sheets=true 且无有效 sheet_name 时：一次返回多表（sheets + preview_data.all_sheets）
    """
    name = (file.filename or "").strip()
    suffix = Path(name).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xlsm 文件")

    try:
        from openpyxl import load_workbook

        raw = await file.read()
        workspace_root = Path(os.environ.get("WORKSPACE_ROOT", os.getcwd())).resolve()
        upload_dir = workspace_root / "uploads"
        upload_dir.mkdir(parents=True, exist_ok=True)
        persisted_path = upload_dir / f"{uuid.uuid4().hex}{suffix}"
        persisted_path.write_bytes(raw)
        try:
            persisted_rel = persisted_path.relative_to(workspace_root).as_posix()
        except ValueError:
            persisted_rel = str(persisted_path)
        from io import BytesIO

        wb = load_workbook(filename=BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 读取失败: {e}") from e

    sheet_names = list(wb.sheetnames or [])
    if not sheet_names:
        return {
            "success": True,
            "template_name": name,
            "file_path": persisted_rel,
            "fields": [],
            "sheets": [],
            "preview_data": {
                "sheet_names": [],
                "selected_sheet_name": "",
                "sheet_name": "",
                "file_path": persisted_rel,
                "sample_rows": [],
                "grid_preview": {"rows": [], "max_row": 0, "max_col": 0},
                "grid_style_cache": {"styles": {}, "cell_style_refs": {}},
                "tables": [],
                "all_sheets": [],
            },
        }

    analyze_all = _form_bool(analyze_all_sheets)
    sn_arg = (sheet_name or "").strip()
    multi_mode = analyze_all and (not sn_arg or sn_arg not in sheet_names)

    if multi_mode:
        targets = sheet_names[:_EXTRACT_GRID_MAX_SHEETS]
        bundles = [_extract_single_sheet_bundle(wb, sn, sheet_names, persisted_rel, name) for sn in targets]
        first = bundles[0]
        sheets_top = [_bundle_to_sheet_entry(b, i + 1) for i, b in enumerate(bundles)]
        all_sheets = [_bundle_to_sheet_entry(b, i + 1) for i, b in enumerate(bundles)]
        return {
            "success": True,
            "template_name": name,
            "file_path": persisted_rel,
            "fields": first["fields"],
            "sheets": sheets_top,
            "preview_data": {
                "sheet_names": sheet_names,
                "selected_sheet_name": first["sheet_name"],
                "sheet_name": first["sheet_name"],
                "file_path": persisted_rel,
                "sample_rows": first["sample_rows"],
                "grid_preview": first["grid_preview"],
                "grid_style_cache": first["grid_style_cache"],
                "tables": first["tables"],
                "all_sheets": all_sheets,
            },
        }

    chosen = sn_arg if sn_arg in sheet_names else sheet_names[0]
    bundle = _extract_single_sheet_bundle(wb, chosen, sheet_names, persisted_rel, name)
    one = _bundle_to_sheet_entry(bundle, 1)
    return {
        "success": True,
        "template_name": name,
        "file_path": persisted_rel,
        "fields": bundle["fields"],
        "sheets": [one],
        "preview_data": {
            "sheet_names": sheet_names,
            "selected_sheet_name": bundle["sheet_name"],
            "sheet_name": bundle["sheet_name"],
            "file_path": persisted_rel,
            "sample_rows": bundle["sample_rows"],
            "grid_preview": bundle["grid_preview"],
            "grid_style_cache": bundle["grid_style_cache"],
            "tables": bundle["tables"],
            "all_sheets": [one],
        },
    }


def _xcagi_chat_http_exc(exc: BaseException) -> HTTPException:
    if isinstance(exc, AuthenticationError):
        return HTTPException(status_code=401, detail=f"大模型鉴权失败: {exc}")
    if isinstance(exc, RateLimitError):
        return HTTPException(status_code=429, detail=f"大模型限流: {exc}")
    if isinstance(exc, APIConnectionError):
        return HTTPException(status_code=503, detail=f"无法连接大模型服务: {exc}")
    if isinstance(exc, APIError):
        return HTTPException(status_code=502, detail=f"大模型接口错误: {exc}")
    if isinstance(exc, RuntimeError):
        return HTTPException(status_code=503, detail=str(exc))
    logger.exception("xcagi ai chat compat unexpected error")
    return HTTPException(status_code=500, detail=f"对话处理失败: {exc}")


class XcagiCompatChatBody(BaseModel):
    model_config = ConfigDict(extra="ignore")

    message: str = Field(
        ...,
        min_length=1,
        validation_alias=AliasChoices("message", "user_message", "content", "text", "query"),
    )
    context: dict | None = Field(
        default=None,
        validation_alias=AliasChoices(
            "context",
            "runtime_context",
            "session_context",
            "ddd_context",
            "neuro_context",
            "neuro_ddd_context",
        ),
    )
    system_prompt: str | None = Field(
        default=None,
        validation_alias=AliasChoices("system_prompt", "system", "instructions"),
    )
    mode: str | None = Field(
        default=None,
        validation_alias=AliasChoices("mode", "llm_mode"),
    )
    db_write_token: str | None = Field(
        default=None,
        description="与 FHD_DB_WRITE_TOKEN 一致时允许 Planner 调用 products_bulk_import",
    )


class XcagiCompatChatBatchBody(BaseModel):
    model_config = ConfigDict(extra="ignore")

    messages: list[str] = Field(default_factory=list)
    context: dict | None = Field(
        default=None,
        validation_alias=AliasChoices(
            "context",
            "runtime_context",
            "session_context",
            "ddd_context",
            "neuro_context",
            "neuro_ddd_context",
        ),
    )
    system_prompt: str | None = Field(
        default=None,
        validation_alias=AliasChoices("system_prompt", "system", "instructions"),
    )
    mode: str | None = Field(
        default=None,
        validation_alias=AliasChoices("mode", "llm_mode"),
    )
    db_write_token: str | None = Field(default=None)
    user_id: str | None = None
    source: str | None = None


def _xcagi_compat_reply_payload(
    reply: str, *, runtime_context_update: dict[str, Any] | None = None
) -> dict:
    text = str(reply or "")

    # 获取最近的工具结果（含错误码），便于前端与气泡展示写库失败原因
    tool_data: dict = {}
    last_result: dict = {}
    try:
        from backend.planner import get_last_tool_result

        raw = get_last_tool_result()
        if isinstance(raw, dict) and raw:
            last_result = raw
            from backend.tools import flatten_tool_result_dict_for_client

            tool_data = flatten_tool_result_dict_for_client(raw)
            errs = raw.get("errors")
            if isinstance(errs, list) and errs:
                preview = errs[:5]
                joined = "; ".join(str(x) for x in preview if x is not None)
                tool_data["errors_preview"] = joined[:2000]
                if len(errs) > 5:
                    tool_data["errors_truncated"] = True
    except Exception:
        pass

    err_code = str(last_result.get("error") or "").strip()
    err_msg = str(last_result.get("message") or "").strip()
    tool_key = str(last_result.get("tool_key") or "").strip()
    if err_code or (last_result.get("success") is False):
        notice_lines = ["---", "**工具反馈**（最近一次）"]
        if tool_key:
            notice_lines.append(f"- 工具：`{tool_key}`")
        if err_code:
            notice_lines.append(f"- 错误码：`{err_code}`")
        if err_msg:
            notice_lines.append(f"- 说明：{err_msg}")
        ep = tool_data.get("errors_preview")
        if ep:
            notice_lines.append(f"- 明细摘要：{ep}")
        notice = "\n".join(notice_lines)
        if notice not in text:
            text = f"{text.rstrip()}\n\n{notice}".strip()

    data: dict[str, Any] = {
        "response": text,
        "text": text,
        **tool_data,
    }
    if runtime_context_update is not None:
        data["runtime_context"] = runtime_context_update

    result: dict[str, Any] = {
        "success": True,
        "response": text,
        "data": data,
    }
    return result


_EXCEL_PATH_PATTERN = re.compile(
    r"@?([^\s'\"<>]+?\.(?:xlsx|xlsm|xls))(?=$|[\s,，。.!！?？])",
    re.IGNORECASE,
)


def _extract_excel_paths_from_message(message: str) -> list[str]:
    paths: list[str] = []
    for m in _EXCEL_PATH_PATTERN.finditer(message or ""):
        p = m.group(1).strip().strip("`\"'[](){}<>")
        if not p:
            continue
        p = p.replace("\\", "/")
        if p not in paths:
            paths.append(p)
    return paths


def _extract_excel_paths_from_context(runtime_context: dict) -> list[str]:
    paths: list[str] = []

    def _push(raw: object) -> None:
        s = str(raw or "").strip().replace("\\", "/")
        if not s:
            return
        if not re.search(r"\.(xlsx|xlsm|xls)$", s, re.IGNORECASE):
            return
        if s not in paths:
            paths.append(s)

    existing_single = runtime_context.get("excel_file_path")
    if isinstance(existing_single, str):
        _push(existing_single)
    existing_multi = runtime_context.get("excel_file_paths")
    if isinstance(existing_multi, (list, tuple)):
        for p in existing_multi:
            _push(p)

    excel_analysis = runtime_context.get("excel_analysis")
    if isinstance(excel_analysis, dict):
        _push(excel_analysis.get("file_path"))
        preview = excel_analysis.get("preview_data")
        if isinstance(preview, dict):
            _push(preview.get("file_path"))

    return paths


def _merge_runtime_context_with_message_paths(
    runtime_context: dict | None,
    message: str,
) -> tuple[dict, list[str]]:
    merged_ctx = dict(runtime_context or {})
    found = _extract_excel_paths_from_message(message)
    ctx_paths = _extract_excel_paths_from_context(merged_ctx)
    if not found and not ctx_paths:
        return merged_ctx, []

    all_paths: list[str] = []
    message_basenames = {Path(p).name.lower(): p for p in found}

    # Prefer mapping @filename to an already-known real path from context.
    for cp in ctx_paths:
        base = Path(cp).name.lower()
        if base in message_basenames and cp not in all_paths:
            all_paths.append(cp)

    for p in found:
        if p not in all_paths:
            all_paths.append(p)
    for cp in ctx_paths:
        if cp not in all_paths:
            all_paths.append(cp)

    if all_paths:
        merged_ctx["excel_file_path"] = all_paths[0]
        merged_ctx["excel_file_paths"] = all_paths
    return merged_ctx, found


def _looks_like_vector_request(message: str) -> bool:
    text = (message or "").lower()
    keywords = ("向量", "索引", "语义检索", "embedding", "vector", "semantic search")
    return any(k in text for k in keywords)


def _ensure_vector_index_if_needed(message: str, runtime_context: dict) -> str | None:
    if not _looks_like_vector_request(message):
        return None

    file_path = str(runtime_context.get("excel_file_path") or "").strip()
    if not file_path:
        return "我识别到您在请求向量索引，但没有拿到 Excel 路径。请发送类似 `@424/26年出货单打印/鸿瑞达报价26年.xlsx` 的路径。"

    root = os.environ.get("WORKSPACE_ROOT", os.getcwd())
    try:
        raw = execute_workflow_tool(
            "excel_vector_index",
            {"file_path": file_path},
            workspace_root=root,
        )
        result = json.loads(raw)
    except Exception as e:
        logger.exception("xcagi vector pre-index failed")
        return f"我尝试为 `{file_path}` 建立向量索引时失败：{e}。请确认文件路径是否存在，或告诉我要索引的工作表名。"

    if isinstance(result, dict) and result.get("error"):
        msg = result.get("message") or result.get("error")
        return f"我尝试为 `{file_path}` 建立向量索引失败：{msg}。请确认路径正确，或把目标工作表名发我。"
    return None


def _xcagi_chat_timeout_seconds() -> float:
    raw = os.environ.get("XCAGI_CHAT_TIMEOUT_SEC", "120").strip()
    try:
        v = float(raw)
        return max(5.0, min(v, 600.0))
    except ValueError:
        return 120.0


def _xcagi_chat_timeout_error_payload(timeout: float) -> dict:
    msg = (
        f"对话处理超时（>{int(timeout)} 秒）。可缩短问题后重试，或由管理员调大环境变量 XCAGI_CHAT_TIMEOUT_SEC。"
    )
    return {
        "success": False,
        "message": msg,
        "response": msg,
        "data": {"text": msg, "response": msg},
    }


def _sse_event_line(payload: dict) -> bytes:
    return ("data: " + json.dumps(payload, ensure_ascii=False) + "\n\n").encode("utf-8")


def _xcagi_planner_stream_bytes(body: XcagiCompatChatBody, *, ai_tier: str):
    """Planner SSE：phase / tool_* / token 事件，最后一条 ``done`` 携带与 JSON 接口一致的 ``result``。"""
    m = (body.mode or "").strip().lower()
    if m in ("online", "offline"):
        set_llm_mode(m)
    runtime_context, _ = _merge_runtime_context_with_message_paths(body.context, body.message)
    runtime_context = runtime_context_with_tier(runtime_context, ai_tier)
    intr = planner_workflow_interrupt_reply(body.message)
    if intr is not None:
        cleared = runtime_context_after_workflow_interrupt(runtime_context)
        yield _sse_event_line({"type": "token", "text": intr})
        yield _sse_event_line(
            {"type": "done", "result": _xcagi_compat_reply_payload(intr, runtime_context_update=cleared)}
        )
        return
    vector_error = _ensure_vector_index_if_needed(body.message, runtime_context)
    if vector_error:
        yield _sse_event_line({"type": "error", "message": vector_error})
        return
    reply_parts: list[str] = []
    try:
        for ev in chat_stream_sse_events(
            body.message,
            runtime_context=runtime_context or None,
            system_prompt=body.system_prompt,
            db_write_token=body.db_write_token,
        ):
            if ev.get("type") == "token":
                reply_parts.append(str(ev.get("text") or ""))
            yield _sse_event_line(ev)
        merged = "".join(reply_parts)
        yield _sse_event_line({"type": "done", "result": _xcagi_compat_reply_payload(merged)})
    except Exception as e:
        exc = _xcagi_chat_http_exc(e)
        yield _sse_event_line(
            {
                "type": "error",
                "message": exc.detail if isinstance(exc.detail, str) else str(exc.detail),
                "status_code": exc.status_code,
            }
        )


@router.post("/ai/unified_chat/stream")
@router.post("/ai/chat/stream")
async def ai_unified_chat_stream(request: Request, body: XcagiCompatChatBody):
    assert_p2_elevated_claim_or_raise(request)
    tier = resolve_ai_tier(request)
    return StreamingResponse(
        _xcagi_planner_stream_bytes(body, ai_tier=tier),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


async def _xcagi_execute_chat(request: Request, body: XcagiCompatChatBody) -> dict:
    m = (body.mode or "").strip().lower()
    if m in ("online", "offline"):
        set_llm_mode(m)

    runtime_context, _ = _merge_runtime_context_with_message_paths(body.context, body.message)
    assert_p2_elevated_claim_or_raise(request)
    tier = resolve_ai_tier(request)
    runtime_context = runtime_context_with_tier(runtime_context, tier)
    intr = planner_workflow_interrupt_reply(body.message)
    if intr is not None:
        cleared = runtime_context_after_workflow_interrupt(runtime_context)
        return _xcagi_compat_reply_payload(intr, runtime_context_update=cleared)

    vector_error = _ensure_vector_index_if_needed(body.message, runtime_context)
    if vector_error:
        return _xcagi_compat_reply_payload(vector_error)

    timeout = _xcagi_chat_timeout_seconds()
    try:
        reply = await asyncio.wait_for(
            asyncio.to_thread(
                run_agent_chat,
                body.message,
                runtime_context=runtime_context or None,
                system_prompt=body.system_prompt,
                db_write_token=body.db_write_token,
            ),
            timeout=timeout,
        )
    except asyncio.TimeoutError:
        return _xcagi_chat_timeout_error_payload(timeout)
    except Exception as e:
        raise _xcagi_chat_http_exc(e) from e
    return _xcagi_compat_reply_payload(reply)


@router.post("/conversations/message")
async def conversations_save_message(body: dict = Body(default_factory=dict)) -> dict:
    """持久化单条消息到会话桶（进程内），供历史列表与 ``GET /conversations/{id}`` 回放。"""
    session_id = str(body.get("session_id") or "").strip()
    user_id = str(body.get("user_id") or "default").strip() or "default"
    role = _xcagi_normalize_chat_role(body.get("role"))
    content = str(body.get("content") or "")
    if not session_id or not content.strip():
        return {"success": True, "saved": False, "message": "empty session_id or content"}

    now = time.time()
    msg = {"role": role, "content": content, "timestamp": _xcagi_iso_from_ts(now)}

    with _conversation_lock:
        user_bucket = _xcagi_user_sessions.setdefault(user_id, {})
        is_new = session_id not in user_bucket
        if is_new:
            _xcagi_evict_oldest_session_if_needed(user_bucket, session_id)
        rec = user_bucket.get(session_id)
        if rec is None:
            rec = {"messages": [], "created_ts": now, "updated_ts": now}
            user_bucket[session_id] = rec
        rec["messages"].append(msg)
        rec["updated_ts"] = now
    return {"success": True, "saved": True}


@router.get("/conversations/sessions")
async def conversations_sessions_list(
    limit: int = Query(default=50, ge=1, le=500),
    user_id: str = Query(default="default"),
) -> dict:
    uid = str(user_id or "default").strip() or "default"
    with _conversation_lock:
        user_bucket = dict(_xcagi_user_sessions.get(uid) or {})
    rows: list[tuple[float, dict[str, Any]]] = []
    for sid, rec in user_bucket.items():
        msgs = list(rec.get("messages") or [])
        title = _xcagi_title_from_messages(msgs)
        uts = float(rec.get("updated_ts") or 0.0)
        rows.append(
            (
                uts,
                {
                    "session_id": sid,
                    "title": title,
                    "summary": _xcagi_summary_from_messages(msgs),
                    "message_count": len(msgs),
                    "created_at": _xcagi_iso_from_ts(float(rec.get("created_ts") or 0.0)),
                    "last_message_at": _xcagi_iso_from_ts(uts),
                },
            )
        )
    rows.sort(key=lambda x: x[0], reverse=True)
    return {"success": True, "sessions": [r[1] for r in rows[: int(limit)]]}


@router.post("/conversations/sessions/clear")
async def conversations_sessions_clear(body: dict = Body(default_factory=dict)) -> dict:
    uid = str((body or {}).get("user_id") or "default").strip() or "default"
    with _conversation_lock:
        removed = _xcagi_user_sessions.pop(uid, None)
        deleted = len(removed) if isinstance(removed, dict) else 0
    return {"success": True, "deleted": deleted, "message": "ok"}


@router.post("/ai/conversation/new")
async def ai_conversation_new(body: dict = Body(default_factory=dict)) -> dict:
    sid = str(body.get("session_id") or "").strip() or uuid.uuid4().hex
    return {"success": True, "data": {"session_id": sid}}


@router.post("/ai/chat")
@router.post("/ai/chat/v2")
@router.post("/ai/unified_chat")
async def ai_chat_unified_compat(request: Request, body: XcagiCompatChatBody) -> dict:
    return await _xcagi_execute_chat(request, body)


@router.post("/ai/chat/batch")
@router.post("/ai/chat/v2/batch")
@router.post("/ai/unified_chat/batch")
async def ai_chat_batch_compat(request: Request, body: XcagiCompatChatBatchBody) -> dict:
    msgs = [str(x).strip() for x in (body.messages or []) if str(x).strip()]
    if not msgs:
        raise HTTPException(status_code=400, detail="messages 须为非空字符串数组")
    assert_p2_elevated_claim_or_raise(request)
    batch_tier = resolve_ai_tier(request)
    m = (body.mode or "").strip().lower()
    if m in ("online", "offline"):
        set_llm_mode(m)
    results: list[dict] = []
    timeout = _xcagi_chat_timeout_seconds()
    rolling_ctx = body.context
    for txt in msgs:
        runtime_context, _ = _merge_runtime_context_with_message_paths(rolling_ctx, txt)
        runtime_context = runtime_context_with_tier(runtime_context, batch_tier)
        intr = planner_workflow_interrupt_reply(txt)
        if intr is not None:
            cleared = runtime_context_after_workflow_interrupt(runtime_context)
            rolling_ctx = cleared
            results.append(_xcagi_compat_reply_payload(intr, runtime_context_update=cleared))
            continue
        vector_error = _ensure_vector_index_if_needed(txt, runtime_context)
        if vector_error:
            results.append(_xcagi_compat_reply_payload(vector_error))
            continue
        try:
            reply = await asyncio.wait_for(
                asyncio.to_thread(
                    run_agent_chat,
                    txt,
                    runtime_context=runtime_context or None,
                    system_prompt=body.system_prompt,
                    db_write_token=body.db_write_token,
                ),
                timeout=timeout,
            )
            results.append(_xcagi_compat_reply_payload(reply))
        except asyncio.TimeoutError:
            results.append(_xcagi_chat_timeout_error_payload(timeout))
        except Exception as e:
            err = _xcagi_chat_http_exc(e)
            results.append(
                {
                    "success": False,
                    "message": err.detail if isinstance(err.detail, str) else str(err.detail),
                }
            )
    ok = all(r.get("success") for r in results)
    return {"success": ok, "batch": True, "results": results, "count": len(results)}


@router.get("/ai/context")
async def ai_context_get(user_id: str = Query(default="default")) -> dict:
    _ = user_id
    return {"success": True, "data": {}}


@router.post("/ai/context/clear")
async def ai_context_clear(body: dict = Body(default_factory=dict)) -> dict:
    _ = body
    return {"success": True}


@router.get("/ai/config")
async def ai_config_get() -> dict:
    return {"success": True, "data": {}}


@router.post("/tts/synthesize")
async def tts_synthesize_stub(body: dict = Body(default_factory=dict)) -> dict:
    """占位：无音频时前端跳过播放，仅消除 404。"""
    _ = body.get("text") or body.get("message")
    return {"success": False, "message": "TTS 未在 FHD 兼容层启用", "data": {}}


@router.get("/conversations/{conversation_id}")
async def conversations_get(
    conversation_id: str,
    user_id: str = Query(default="default"),
) -> dict:
    cid = (conversation_id or "").strip()
    uid = str(user_id or "default").strip() or "default"
    with _conversation_lock:
        rec = (_xcagi_user_sessions.get(uid) or {}).get(cid)
        msgs = [dict(m) for m in list((rec or {}).get("messages") or [])]
        title = _xcagi_title_from_messages(msgs) if rec else None
    return {
        "success": True,
        "id": cid,
        "title": title,
        "messages": msgs,
        "metadata": {},
    }


def _test_db_toggle_from_body(body: dict) -> bool | None:
    for key in (
        "enabled",
        "enable",
        "on",
        "test_mode",
        "test_db_enabled",
        "testDbEnabled",
        "value",
    ):
        if key not in body:
            continue
        v = body[key]
        if isinstance(v, bool):
            return v
        if isinstance(v, (int, float)):
            return bool(int(v))
        if isinstance(v, str):
            s = v.strip().lower()
            if s in ("true", "1", "yes", "on"):
                return True
            if s in ("false", "0", "no", "off"):
                return False
    return None


def _compat_current_db_display_label(info: dict) -> str:
    mode = info["mode"]
    if info.get("backend") == "postgresql":
        summ = info.get("postgresql_summary") or {}
        dbn = str(summ.get("database_name") or "").strip()
        hp = str(summ.get("host_port") or "").strip()
        if dbn and hp:
            core = f"{dbn} @ {hp}"
        else:
            core = dbn or hp or "PostgreSQL"
        return f"{core}（PostgreSQL · 与 XCAGI / Mod 共用 DATABASE_URL）"
    return f"{info['current_db_name']}（{'测试' if mode == 'test' else '真实'}）"


@router.get("/system/test-db/status")
@router.get("/system/test-db/status/")
async def system_test_db_status() -> dict:
    info = get_db_status()
    mode = info["mode"]
    label = _compat_current_db_display_label(info)
    return {
        "success": True,
        "data": {
            "enabled": mode == "test",
            "test_mode": mode == "test",
            "test_db_enabled": mode == "test",
            "current_db_display": label,
            **info,
        },
    }


@router.post("/system/test-db/enable")
@router.post("/system/test-db/enable/")
async def system_test_db_enable(body: dict | None = Body(default=None)) -> dict:
    body = body if isinstance(body, dict) else {}
    want = _test_db_toggle_from_body(body)
    if want is None:
        want = resolve_mode() == "production"
    if want:
        result = switch_to_test_mode()
    else:
        result = switch_to_production_mode()
    if result.get("error"):
        raise HTTPException(status_code=400, detail=result.get("message", str(result)))
    info = get_db_status()
    label = _compat_current_db_display_label(info)
    return {
        "success": True,
        "data": {
            "enabled": info["mode"] == "test",
            "test_mode": info["mode"] == "test",
            "test_db_enabled": info["mode"] == "test",
            "current_db_display": label,
            **info,
            "switch": result,
        },
    }


@router.post("/system/test-db/disable")
@router.post("/system/test-db/disable/")
async def system_test_db_disable(body: dict | None = Body(default=None)) -> dict:
    """前端 Settings 关闭测试模式时 POST 此路径；与 enable 共用逻辑，固定切回 production。"""
    merged: dict = dict(body) if isinstance(body, dict) else {}
    merged["enabled"] = False
    merged["test_db_enabled"] = False
    return await system_test_db_enable(merged)


@router.get("/preferences")
@router.get("/preferences/")
async def preferences_get(user_id: str = Query(default="default")) -> dict:
    return {
        "success": True,
        "data": {"user_id": user_id, "preferences": {}},
    }


@router.post("/preferences")
@router.post("/preferences/")
async def preferences_post(body: dict = Body(default_factory=dict)) -> dict:
    return {"success": True, "data": body or {}}


@router.get("/distillation/versions")
@router.get("/distillation/versions/")
async def distillation_versions() -> dict:
    return {"success": True, "data": []}


def _intent_packages_list_payload() -> dict:
    return {"success": True, "data": []}


@router.get("/intent-packages", operation_id="compat_intent_packages_hyphen")
async def compat_intent_packages_hyphen() -> dict:
    return _intent_packages_list_payload()


@router.get("/intent-packages/", operation_id="compat_intent_packages_hyphen_slash", include_in_schema=False)
async def compat_intent_packages_hyphen_slash() -> dict:
    return _intent_packages_list_payload()


@router.get("/intent_packages", operation_id="compat_intent_packages_underscore", include_in_schema=False)
async def compat_intent_packages_underscore() -> dict:
    return _intent_packages_list_payload()


@router.get("/intent_packages/", operation_id="compat_intent_packages_underscore_slash", include_in_schema=False)
async def compat_intent_packages_underscore_slash() -> dict:
    return _intent_packages_list_payload()


@router.get("/tools", summary="工具表目录（与 XCAGI ToolsView / pro-mode 对齐）")
@router.get("/tools/", summary="工具表目录（尾斜杠）")
async def compat_tools_list() -> dict:
    return get_tools_payload()


@router.get("/db-tools", summary="工具表目录别名（前端优先请求）")
@router.get("/db-tools/", summary="工具表目录别名（尾斜杠）")
async def compat_db_tools_list() -> dict:
    return get_tools_payload()


@router.get("/tool-categories", summary="工具分类列表")
@router.get("/tool-categories/", summary="工具分类列表（尾斜杠）")
async def compat_tool_categories_list() -> dict:
    return get_tool_categories_payload()
