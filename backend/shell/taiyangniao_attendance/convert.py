"""钉钉导出 xlsx → 月度统计表 xlsx（实现说明见原 Mod 文档）。"""

from __future__ import annotations

import logging
import os
import re
import shutil
import uuid
import warnings
from pathlib import Path
from typing import Any

import pandas as pd

from .mapping import DINGTALK_COLUMN_ALIASES, STATISTICS_OUTPUT_COLUMNS
from .paths import ensure_parent_dir
from .rules import process_attendance_dataframe, AttendanceRule

logger = logging.getLogger(__name__)


def _excel_cell_str(v: Any) -> str:
    """写入 Excel 单元格的短文本；pandas NaN / 缺失 转为空串。"""
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except TypeError:
        pass
    s = str(v).strip()
    if s.lower() in ("nan", "none", "<na>"):
        return ""
    return s


def _resolve_excel_path_on_disk(candidate: Path) -> Path | None:
    """``Path.is_file()`` 在部分中文路径下不可靠时，用目录枚举匹配真实文件名。"""
    if candidate.is_file():
        return candidate
    parent, name = candidate.parent, candidate.name
    norm = name.replace(" ", "")
    try:
        for fn in os.listdir(str(parent)):
            if fn.replace(" ", "") == norm:
                return parent / fn
    except OSError:
        pass
    return None


def _to_datetime_mixed(series: pd.Series) -> pd.Series:
    """钉钉列里常有字符串/时间戳混排；pandas 2.0+ 用 format='mixed'，旧版回退并抑制推断告警。"""
    s = series
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        try:
            return pd.to_datetime(s, errors="coerce", format="mixed")
        except TypeError:
            return pd.to_datetime(s, errors="coerce")


def _norm_header(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def _pick_column(
    df: pd.DataFrame,
    aliases: list[str],
    *,
    prefer_non_empty: bool = False,
) -> str | None:
    """按别名解析列名；``prefer_non_empty`` 时在多个同义列中选首个有非空样本值的列（如空「工号」+ 有值 UserId）。"""
    cols = list(df.columns)
    norm_map = {_norm_header(c): str(c) for c in cols}
    exact_hits: list[str] = []
    for a in aliases:
        key = _norm_header(a)
        if key in norm_map:
            exact_hits.append(norm_map[key])
    if exact_hits:
        if prefer_non_empty and not df.empty:
            sample_n = min(len(df), 300)
            for c in exact_hits:
                ser = df[c].head(sample_n)
                if ser.notna().any() and ser.dropna().astype(str).str.strip().ne("").any():
                    return c
        return exact_hits[0]
    for a in aliases:
        ak = _norm_header(a)
        for c in cols:
            cn = _norm_header(c)
            if ak and ak in cn:
                return str(c)
    return None


def _year_month_for_template_dates(month: str | None, dingtalk_detail: pd.DataFrame) -> tuple[int, int]:
    """模板按日填格子用的 (年, 月)；优先 ``YYYY-MM``，否则从明细首行日期推断，再否则当前月。"""
    raw = (month or "").strip()
    m = re.fullmatch(r"(\d{4})-(\d{2})", raw)
    if m:
        return int(m.group(1)), int(m.group(2))
    if "日期" in dingtalk_detail.columns and not dingtalk_detail.empty:
        v = dingtalk_detail["日期"].iloc[0]
        if isinstance(v, str):
            parts = v.split("-")
            if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
                y, mo = int(parts[0]), int(parts[1])
                if 1 <= mo <= 12:
                    return y, mo
    from datetime import datetime

    d = datetime.now()
    return d.year, d.month


def _extract_month_from_file(path: Path, sheet: str | int | None = 0) -> str | None:
    """从钉钉导出文件的标题行提取统计月份，如 '2026-03'。"""
    try:
        raw_df = pd.read_excel(path, sheet_name=sheet, header=None, engine="openpyxl")
        if raw_df.empty:
            return None
        title_val = raw_df.iloc[0, 0]
        if pd.isna(title_val):
            return None
        import re
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})\s+至\s+(\d{4})-(\d{2})-(\d{2})", str(title_val))
        if m:
            return f"{m.group(1)}-{m.group(2)}"
    except Exception:
        pass
    return None


def read_dingtalk_dataframe(
    path: Path,
    *,
    sheet: str | int | None = 0,
    header_row: int | None = None,
) -> pd.DataFrame:
    """``header_row`` 为 ``None`` 时自动检测。显式 ``0`` 若读不出「姓名」列（钉钉标题行被误当表头），会回退为自动检测。"""
    effective = _auto_detect_header_row(path, sheet) if header_row is None else int(header_row)
    df = pd.read_excel(path, sheet_name=sheet, header=effective, engine="openpyxl")
    if header_row is not None and int(header_row) == 0 and effective == 0:
        if _pick_column(df, DINGTALK_COLUMN_ALIASES["name"]) is None:
            alt = _auto_detect_header_row(path, sheet)
            if alt != 0:
                df = pd.read_excel(path, sheet_name=sheet, header=alt, engine="openpyxl")
    return df


def _auto_detect_header_row(path: Path, sheet: str | int | None = 0) -> int:
    """自动检测钉钉导出文件的表头行。

    钉钉导出的文件通常有2-4行表头（标题、生成时间、列名、日期行）。
    通过查找包含已知列名（如"姓名"、"工号"）的行来确定表头行。
    """
    try:
        raw_df = pd.read_excel(path, sheet_name=sheet, header=None, engine="openpyxl")
    except Exception:
        return 0

    if raw_df.empty:
        return 0

    known_columns = {"姓名", "工号", "部门", "打卡时间", "考勤日期"}
    max_scan_rows = min(10, len(raw_df))

    for row_idx in range(max_scan_rows):
        row_values = set(str(v).strip() for v in raw_df.iloc[row_idx].tolist() if pd.notna(v))
        matched = row_values & known_columns
        if len(matched) >= 2:
            return row_idx

    return 0


def _is_date_subheader_row(row: pd.Series) -> bool:
    """检测是否为钉钉日期子表头行（如 '日','2','3','4'...）。"""
    non_null = row.dropna()
    if len(non_null) == 0:
        return False
    str_vals = set(str(v).strip() for v in non_null)
    date_indicators = {"日", "一", "二", "三", "四", "五", "六", "天"}
    weekday_count = sum(1 for v in str_vals if v in date_indicators)
    if weekday_count >= 2:
        return True
    numeric_days = [v for v in str_vals if v.isdigit() and 1 <= int(v) <= 31]
    return len(numeric_days) >= 10


def _parse_date_columns(df: pd.DataFrame, header_row_idx: int) -> dict[int, str]:
    """解析日期子表头行，返回列索引到日期字符串的映射。"""
    date_map = {}
    if header_row_idx >= len(df):
        return date_map

    row = df.iloc[header_row_idx]
    for col_idx in range(len(row)):
        val = row.iloc[col_idx]
        if pd.isna(val):
            continue
        val_str = str(val).strip()
        if val_str.isdigit() and 1 <= int(val_str) <= 31:
            date_map[col_idx] = val_str
        elif val_str in {"日", "一", "二", "三", "四", "五", "六", "天"}:
            date_map[col_idx] = val_str
    return date_map


def _parse_check_times(val: Any) -> tuple[Any, Any]:
    """解析钉钉的打卡时间字符串，返回 (上班打卡, 下班打卡)。"""
    if pd.isna(val):
        return pd.NaT, pd.NaT
    val_str = str(val).strip()
    parts = val_str.split('\n')
    check_in = pd.NaT
    check_out = pd.NaT
    for i, p in enumerate(parts):
        p = p.strip()
        if not p:
            continue
        p_clean = p.replace(' ', '')
        try:
            if ':' in p_clean:
                h, m = p_clean.split(':')[:2]
                if h.isdigit() and m.isdigit():
                    hours = int(h)
                    minutes = int(m)
                    if 0 <= hours <= 23 and 0 <= minutes <= 59:
                        import datetime
                        dt = datetime.datetime(1900, 1, 1, hours, minutes)
                        if i == 0:
                            check_in = dt
                        else:
                            check_out = dt
        except Exception:
            pass
    return check_in, check_out


def _build_wide_to_long(df: pd.DataFrame, meta_cols: dict[str, str | None]) -> pd.DataFrame:
    """将钉钉宽表格式转换为长表格式（每行一条打卡记录）。"""
    rows = []
    for _, emp_row in df.iterrows():
        emp_name = str(emp_row[meta_cols["name"]]).strip() if meta_cols.get("name") else ""
        emp_no = str(emp_row[meta_cols["emp_no"]]).strip() if meta_cols.get("emp_no") else ""
        dept = str(emp_row[meta_cols["dept"]]).strip() if meta_cols.get("dept") else ""

        for col_idx, date_info in meta_cols["date_columns"].items():
            day_str, date_month = date_info
            date_val = f"{date_month}-{int(day_str):02d}" if date_month else day_str
            check_in_val = emp_row.iloc[col_idx] if col_idx < len(emp_row) else None
            check_in, check_out = _parse_check_times(check_in_val)

            rows.append({
                "姓名": emp_name,
                "工号": emp_no,
                "部门": dept,
                "日期": date_val,
                "上班打卡": check_in,
                "下班打卡": check_out,
            })
    return pd.DataFrame(rows)


def build_normalized_frame(df: pd.DataFrame, *, month: str | None = None) -> pd.DataFrame:
    # 检测日期标题行（可能在第 0 行或第 1 行）
    date_header_row_idx = None
    if len(df) > 0 and _is_date_subheader_row(df.iloc[0]):
        date_header_row_idx = 0
    elif len(df) > 1 and _is_date_subheader_row(df.iloc[1]):
        date_header_row_idx = 1

    if date_header_row_idx is not None:
        # 跳过日期标题行
        df_data = df.drop(df.index[date_header_row_idx]).reset_index(drop=True)
    else:
        df_data = df

    picks: dict[str, str | None] = {}
    for logical, aliases in DINGTALK_COLUMN_ALIASES.items():
        picks[logical] = _pick_column(
            df_data,
            aliases,
            prefer_non_empty=(logical == "emp_no"),
        )

    date_columns = {}
    if date_header_row_idx is not None:
        for col_idx in range(len(df.columns)):
            val = df.iloc[date_header_row_idx, col_idx]
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str.isdigit() and 1 <= int(val_str) <= 31:
                    date_columns[col_idx] = (val_str, month)

    meta_cols = {
        "name": picks.get("name"),
        "emp_no": picks.get("emp_no"),
        "dept": picks.get("dept"),
        "date_columns": date_columns,
    }

    return _build_wide_to_long(df_data, meta_cols)


def _join_unique_results(s: pd.Series) -> str:
    parts = [str(x).strip() for x in s.tolist() if str(x).strip() and str(x).lower() != "nan"]
    return "、".join(dict.fromkeys(parts))


def aggregate_monthly_stats(norm: pd.DataFrame, *, month: str | None = None) -> pd.DataFrame:
    if norm.empty:
        return pd.DataFrame(columns=STATISTICS_OUTPUT_COLUMNS)

    gcols = ["工号", "姓名", "部门", "日期"]

    agg_specs = [
        ("上班打卡", "上班打卡", "min"),
        ("下班打卡", "下班打卡", "max"),
    ]
    if "工作时长" in norm.columns:
        agg_specs.append(("工作时长", "工作时长", "first"))
    if "考勤结果" in norm.columns:
        agg_specs.append(("考勤结果", "考勤结果", _join_unique_results))

    agg_dict = {out_name: (in_name, agg_fn) for out_name, in_name, agg_fn in agg_specs}
    agg = norm.groupby(gcols, dropna=False, as_index=False).agg(**agg_dict)

    agg = process_attendance_dataframe(agg, base_date=month)

    for c in STATISTICS_OUTPUT_COLUMNS:
        if c not in agg.columns:
            agg[c] = pd.NA

    return agg[STATISTICS_OUTPUT_COLUMNS]


def _cell_value_for_openpyxl(v: Any) -> Any:
    """openpyxl 不接受 pd.NA / NaT；统一成 None 或原生类型。"""
    if v is pd.NA:
        return None
    try:
        if pd.isna(v):
            return None
    except TypeError:
        pass
    return v


def _workbook_append_dataframe_sheet(wb: Any, sheet_name: str, df: pd.DataFrame) -> None:
    """在已打开的工作簿中追加（或覆盖同名）由 DataFrame 构成的表。"""
    from openpyxl.utils.dataframe import dataframe_to_rows

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws_new = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws_new.append([_cell_value_for_openpyxl(v) for v in row])


def _write_workbook_to_path(
    stats: pd.DataFrame,
    target: Path,
    *,
    template_path: Path | None,
    dingtalk_detail: pd.DataFrame,
    month: str | None = None,
) -> None:
    """写入统计结果到模板。

    - 无模板时：仅「月度统计」「钉钉解析」两个工作表
    - 有模板时：复制考勤统计表 xlsx，按版式填写「明细」；并在同一文件中追加「月度统计」「钉钉解析」（与无模板分支数据一致）
    """
    use_template = False
    if template_path is not None:
        src = _resolve_excel_path_on_disk(template_path)
        if src is None:
            logger.warning("template file not found on disk: %s", template_path)
        else:
            try:
                ensure_parent_dir(target)
                shutil.copy2(str(src), str(target))
                use_template = True
            except OSError as e:
                logger.warning("template copy failed %s -> %s: %s", src, target, e)

    if use_template:
        from openpyxl import load_workbook

        wb = load_workbook(str(target))

        if "明细" in wb.sheetnames:
            ws = wb["明细"]
        else:
            ws = wb.active

        merged_cells: set[tuple[int, int]] = set()
        for mr in ws.merged_cells.ranges:
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    merged_cells.add((r, c))

        for row_idx in range(4, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                if (row_idx, col_idx) in merged_cells:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    continue
                cell.value = None

        if "姓名" in dingtalk_detail.columns:
            employees = dingtalk_detail["姓名"].dropna().unique()
        else:
            employees = []

        y_cal, m_cal = _year_month_for_template_dates(month, dingtalk_detail)
        current_row = 4
        for emp_name in employees:
            emp_data = dingtalk_detail[dingtalk_detail["姓名"] == emp_name]

            if emp_data.empty:
                continue

            emp_dept = (
                _excel_cell_str(emp_data["部门"].iloc[0])
                if "部门" in emp_data.columns
                else ""
            )

            for time_period_idx, time_period in enumerate(["上午", "下午", "晚上"]):
                label_row = current_row + time_period_idx * 2
                is_evening = time_period_idx == 2
                fill_rows = [label_row] if is_evening else [label_row, label_row + 1]

                if time_period_idx == 0:
                    ws.cell(row=label_row, column=1, value=emp_dept)
                    ws.cell(row=label_row, column=2, value="计时")
                    ws.cell(row=label_row, column=3, value=emp_name)

                ws.cell(row=label_row, column=4, value=time_period)

                for day in range(1, 32):
                    date_str = f"{y_cal}-{m_cal:02d}-{day:02d}"
                    day_data = emp_data[emp_data["日期"] == date_str]

                    col_base = 7 + (day - 1) * 3
                    if col_base + 2 > ws.max_column:
                        continue

                    left: str | None = None
                    mid: Any = None
                    right: str | None = None

                    if not day_data.empty:
                        check_in = (
                            day_data["上班打卡"].iloc[0] if "上班打卡" in day_data.columns else None
                        )
                        check_out = (
                            day_data["下班打卡"].iloc[0] if "下班打卡" in day_data.columns else None
                        )
                        work_hours = (
                            day_data["工作时长"].iloc[0] if "工作时长" in day_data.columns else 0
                        )
                        if pd.notna(check_in) or pd.notna(check_out):
                            left, right = "√", "√"
                            mid = work_hours if pd.notna(work_hours) else 0
                        else:
                            left, right = "☆", "☆"
                    else:
                        left, right = "☆", "☆"

                    for dr in fill_rows:
                        if left is not None:
                            ws.cell(row=dr, column=col_base, value=left)
                        if mid is not None:
                            ws.cell(row=dr, column=col_base + 1, value=mid)
                        if right is not None:
                            ws.cell(row=dr, column=col_base + 2, value=right)

            current_row += 6

        _workbook_append_dataframe_sheet(wb, "月度统计", stats)
        _workbook_append_dataframe_sheet(wb, "钉钉解析", dingtalk_detail)

        wb.save(str(target))
        wb.close()
    else:
        with pd.ExcelWriter(target, engine="openpyxl", mode="w") as writer:
            stats.to_excel(writer, sheet_name="月度统计", index=False)
            dingtalk_detail.to_excel(writer, sheet_name="钉钉解析", index=False)


def write_statistics_workbook(
    stats: pd.DataFrame,
    output_path: Path,
    *,
    template_path: Path | None = None,
    dingtalk_detail: pd.DataFrame | None = None,
    month: str | None = None,
) -> None:
    """先写到同目录临时文件再 ``os.replace``，避免目标已存在时部分环境直接打开失败。"""
    ensure_parent_dir(output_path)
    detail = dingtalk_detail if dingtalk_detail is not None else stats
    parent = output_path.parent
    parent.mkdir(parents=True, exist_ok=True)
    tmp = parent / f"{output_path.stem}_writing_{uuid.uuid4().hex[:12]}{output_path.suffix}"
    try:
        _write_workbook_to_path(
            stats,
            tmp,
            template_path=template_path,
            dingtalk_detail=detail,
            month=month,
        )
        os.replace(tmp, output_path)
    except PermissionError as e:
        try:
            tmp.unlink(missing_ok=True)
        except OSError:
            pass
        raise PermissionError(
            "无法写入输出文件：若该文件正被 Excel（或其它程序）打开，请先关闭后再试；"
            "或在前端修改「输出相对路径」为新的文件名。"
        ) from e
    except Exception:
        try:
            tmp.unlink(missing_ok=True)
        except OSError:
            pass
        raise


def convert_dingtalk_file(
    source: Path,
    output: Path,
    *,
    month: str | None = None,
    sheet: str | int | None = 0,
    header_row: int | None = None,
    template_path: Path | None = None,
) -> dict[str, Any]:
    raw = read_dingtalk_dataframe(source, sheet=sheet, header_row=header_row)
    if month is None:
        month = _extract_month_from_file(source, sheet)
    norm = build_normalized_frame(raw, month=month)
    stats = aggregate_monthly_stats(norm, month=month)

    write_statistics_workbook(
        stats,
        output,
        template_path=template_path,
        dingtalk_detail=norm,
        month=month,
    )

    return {
        "rows_in": len(raw),
        "rows_stats": len(stats),
        "output": str(output.resolve()),
        "month": month or "",
    }


def coerce_sheet_arg(raw: Any) -> str | int:
    if raw is None or raw == "":
        return 0
    if isinstance(raw, int) and not isinstance(raw, bool):
        return raw
    s = str(raw).strip()
    if s.isdigit():
        return int(s)
    return s
