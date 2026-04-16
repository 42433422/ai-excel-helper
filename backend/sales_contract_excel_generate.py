"""
销售合同 Excel（送货单版式）：可读性检测、模板填充、列预览。

源模板为 ``.xls/.xlsx/.xlsm``；填充产物为 ``.xlsx``（openpyxl）。
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def _import_xlrd_for_xls():
    """
    返回可用于读取 ``.xls`` (BIFF8) 的 xlrd 模块。

    xlrd 2.0 起不再支持 ``.xls``；本仓库在 ``requirements.txt`` 中固定 ``xlrd==1.2.0``。
    """
    try:
        import xlrd  # type: ignore[import-untyped]
    except ImportError as e:
        raise ImportError(
            "读取 .xls 销售合同模板需要安装 xlrd。请执行：pip install 'xlrd==1.2.0'。"
            "说明：xlrd>=2 已移除对 .xls 的支持。"
        ) from e

    ver = str(getattr(xlrd, "__version__", "") or "").strip()
    if ver:
        try:
            major = int(ver.split(".", 1)[0])
        except ValueError:
            major = 0
        if major >= 2:
            raise ImportError(
                f"当前 xlrd 版本为 {ver}，不支持 .xls 模板。请执行：pip install 'xlrd==1.2.0'"
            )
    return xlrd


def is_sales_contract_excel_template_readable(path: str | Path) -> bool:
    """能否以销售合同填充链路打开（非 Word、非损坏的表格）。"""
    p = Path(path)
    if not p.is_file():
        return False
    suf = p.suffix.lower()
    try:
        if suf in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook

            wb = load_workbook(str(p), read_only=True, data_only=True)
            try:
                ws = wb.active
                _ = ws.max_row
            finally:
                wb.close()
            return True
        if suf == ".xls":
            xlrd = _import_xlrd_for_xls()
            book = xlrd.open_workbook(str(p), formatting_info=False)
            _ = book.sheet_by_index(0).nrows
            return True
    except Exception:
        logger.debug("is_sales_contract_excel_template_readable: failed for %s", p, exc_info=True)
    return False


# 送货单版式中产品块默认占用的数据行数（与 ``424/document_templates/送货单.xls`` 一致）
_SALES_CONTRACT_TEMPLATE_PRODUCT_ROWS = 5
_SALES_CONTRACT_FIRST_PRODUCT_ROW = 5


def _parse_quantity_cell(value: str | int | float | None) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r"(\d+(?:\.\d+)?)", str(value))
    return float(m.group(1)) if m else 0.0


def _parse_spec_quantity(value: Any) -> float:
    """``spec_quantity`` 可能为 float 或 ``'5KG'`` 等字符串（与 ``sales_contract_generate_core`` 一致）。"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r"(\d+(?:\.\d+)?)", str(value))
    return float(m.group(1)) if m else 0.0


def _parse_money_cell(value: str | int | float | None) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r"[\d.]+", str(value))
    if not m:
        return 0.0
    try:
        return round(float(m.group(0)), 4)
    except ValueError:
        return 0.0


def _is_mass_unit_label(u: str) -> bool:
    """是否为质量类单位（第 7 列数值按重量理解时用）。"""
    s = (u or "").strip().lower()
    if not s:
        return False
    return s in ("kg", "k", "公斤", "千克", "克", "g", "吨", "t")


def _parse_weight_value_and_unit(text: str | None) -> tuple[float | None, str]:
    """
    从 ``'15KG'``、``'30 公斤'``、``'12'`` 等解析 (重量数值, 重量单位)。

    模板第 7 列为重量数、第 8 列为该列单位（与 ``送货单.xls`` 一致），缺省单位为 ``KG``。
    """
    t = str(text or "").strip()
    if not t:
        return None, "KG"
    m = re.match(
        r"^\s*([\d.]+)\s*(KG|公斤|kg|克|g|吨|t)\s*$",
        t,
        re.IGNORECASE,
    )
    if m:
        val = float(m.group(1))
        raw_u = m.group(2).lower()
        if raw_u in ("公斤",):
            return val, "公斤"
        if raw_u in ("克", "g"):
            return val, "克"
        if raw_u in ("吨", "t"):
            return val, "吨"
        return val, "KG"
    mnum = re.search(r"[\d.]+", t)
    if mnum:
        try:
            return float(mnum.group(0)), "KG"
        except ValueError:
            pass
    return None, "KG"


def _product_weight_unit_label(p: dict) -> str:
    """
    产品行第 8 列：与第 7 列数值配套的单位。

    - ``weight_unit`` 显式给出时优先。
    - ``total_weight`` 文本里带 KG/公斤/克/吨 等时，用解析出的质量单位。
    - 否则按 **销售单位**（桶、缶、箱等）填写，避免计件合同整列固定 ``KG``（与模板 F 列「桶」语义一致）。
    """
    wu = str(p.get("weight_unit") or "").strip()
    if wu:
        return wu
    tw_raw = str(p.get("total_weight") or "").strip()
    if tw_raw and re.search(r"[\d.]+\s*(KG|公斤|kg|千克|克|g|吨|t)\b", tw_raw, re.IGNORECASE):
        _, u = _parse_weight_value_and_unit(tw_raw)
        return u
    su = str(p.get("unit") or "").strip()
    if su and not _is_mass_unit_label(su):
        return su
    _, u = _parse_weight_value_and_unit(tw_raw)
    return u


def _normalize_spec_times_sign(text: str) -> str:
    """规格列统一用「×」连接，禁止「KG /」「KG/」等斜杠写法。"""
    t = str(text or "")
    t = re.sub(r"(?i)kg\s*[/／]+\s*", "KG×", t)
    t = re.sub(r"公斤\s*[/／]+\s*", "公斤×", t)
    t = re.sub(r"千克\s*[/／]+\s*", "千克×", t)
    t = re.sub(r"×{2,}", "×", t)
    return t.strip()


def _specification_cell_value(raw: str, spec_q: float) -> str:
    """
    第 4 列「规格」展示：若字符串以「与第 3 列相同的数字 + 质量单位」开头（如 ``20KG``、``5公斤×1``），
    则去掉该前缀，避免与第 3 列数值重复（前面已有数字）。
    """
    s = _normalize_spec_times_sign(str(raw or "").strip())
    if not s or spec_q <= 0:
        return _normalize_spec_times_sign(s) if s else s
    m = re.match(
        r"^(\d+(?:\.\d+)?)\s*(KG|公斤|kg|千克|克|g|吨|t)\s*(.*)$",
        s,
        re.IGNORECASE,
    )
    if not m:
        return _normalize_spec_times_sign(s)
    try:
        n = float(m.group(1))
    except ValueError:
        return _normalize_spec_times_sign(s)
    if abs(n - spec_q) > 1e-5:
        return _normalize_spec_times_sign(s)
    unit = m.group(2)
    tail = (m.group(3) or "").strip()
    u_disp = "KG" if str(unit).lower() in ("kg", "k") else unit
    if tail:
        sep = "" if tail[0] in "×xX*" else " "
        out = f"{u_disp}{sep}{tail}".replace("  ", " ").strip()
        return _normalize_spec_times_sign(out)
    return _normalize_spec_times_sign(str(u_disp))


def _merge_spec_unit_qty_single_cell(raw: str, spec_q: float) -> str:
    """
    将「单件重量/规格」与 ``spec_quantity`` 写在 **同一格**（第 4 列），避免第 3 列数字 + 第 4 列「KG×」拆成两格。

    例如：模板式 ``KG×`` + 数量 ``5`` → ``KG×5``；``5KG×1`` 去重后已为 ``KG×1`` 则不再追加。
    一律使用 ``×``，不出现 ``KG /``。
    """
    base = _specification_cell_value(str(raw or "").strip(), spec_q)
    if spec_q <= 0:
        return _normalize_spec_times_sign(base)
    qt = (
        str(int(round(spec_q)))
        if abs(spec_q - round(spec_q)) < 1e-6
        else str(spec_q).rstrip("0").rstrip(".")
    )
    b = base.strip()
    if not b:
        return f"KG×{qt}"
    b = _normalize_spec_times_sign(b)
    # 已有「×数字」或「/ 数字」等，不再重复缀数量
    if re.search(r"[×xX*]\s*\.?\d", b) or re.search(r"[/／]\s*\.?\d", b):
        return _normalize_spec_times_sign(b)
    if b[-1:] in ("×", "x", "X", "*"):
        return _normalize_spec_times_sign(b + qt)
    if b[-1:] in ("/", "／"):
        b = re.sub(r"[/／]\s*$", "×", b)
        return _normalize_spec_times_sign(b + qt)
    if re.fullmatch(r"(?i)kg", b):
        return f"KG×{qt}"
    return _normalize_spec_times_sign(f"{b}×{qt}")


def _price_per_unit_cell_label(p: dict) -> str:
    """第 10 列「元/…」表头：与第 8 列单位一致（元/桶、元/KG 等）。"""
    dim = _product_weight_unit_label(p)
    if _is_mass_unit_label(dim):
        return "元/KG" if dim.lower() in ("kg", "k", "") else f"元/{dim}"
    return f"元/{dim}" if dim else "元/件"


def _load_workbook_mutable(template_path: Path):
    """返回 (workbook, worksheet)，支持 ``.xls``（先转值为 openpyxl 工作簿）。"""
    from openpyxl import Workbook, load_workbook

    suf = template_path.suffix.lower()
    if suf in (".xlsx", ".xlsm"):
        return load_workbook(str(template_path)), None

    if suf == ".xls":
        xlrd = _import_xlrd_for_xls()

        book = xlrd.open_workbook(str(template_path), formatting_info=False)
        sh = book.sheet_by_index(0)
        wb = Workbook()
        ws = wb.active
        ws.title = sh.name[:31] or "Sheet1"
        xl_cell_date = getattr(xlrd, "XL_CELL_DATE", 3)
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                val = cell.value
                if cell.ctype == xl_cell_date and val:
                    try:
                        val = xlrd.xldate_as_datetime(val, book.datemode)
                    except Exception:
                        pass
                ws.cell(r + 1, c + 1, val)
        return wb, None

    raise ValueError(f"不支持的模板格式: {suf}")


def fill_sales_contract_excel_template(
    template_path: str | Path,
    template_data: dict[str, Any],
    output_path: str | Path,
) -> None:
    """按送货单版式写入 ``template_data``，保存为 ``.xlsx``。"""
    tpl = Path(template_path)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb, _ = _load_workbook_mutable(tpl)
    try:
        ws = wb.active

        ws.cell(2, 2, template_data.get("customer_name") or "")
        ws.cell(3, 9, template_data.get("contract_date") or "")

        products = list(template_data.get("products") or [])
        first_row = _SALES_CONTRACT_FIRST_PRODUCT_ROW

        for offset in range(_SALES_CONTRACT_TEMPLATE_PRODUCT_ROWS):
            r = first_row + offset
            if offset < len(products):
                p = products[offset]
                qty = _parse_quantity_cell(p.get("quantity"))
                spec_q = _parse_spec_quantity(p.get("spec_quantity"))
                price = _parse_money_cell(p.get("unit_price"))
                amount = _parse_money_cell(p.get("amount"))
                if amount == 0.0 and qty and spec_q and price:
                    amount = round(qty * spec_q * price, 2)

                ws.cell(r, 1, str(p.get("model_number") or "").strip())
                ws.cell(r, 2, str(p.get("name") or "").strip())
                # 单件规格与数量合并到第 4 列（与「KG×5」一体），第 3 列留空，避免两列拆开显示
                ws.cell(r, 3, "")
                ws.cell(
                    r,
                    4,
                    _merge_spec_unit_qty_single_cell(
                        str(p.get("specification") or "").strip(),
                        spec_q,
                    ),
                )
                ws.cell(r, 5, qty)
                ws.cell(r, 6, str(p.get("unit") or "").strip())
                tw_src = p.get("total_weight")
                if tw_src is not None and str(tw_src).strip() != "":
                    wn, _pu = _parse_weight_value_and_unit(str(tw_src))
                    row_w = float(wn) if wn is not None else float(qty * spec_q)
                else:
                    row_w = float(qty * spec_q)
                ws.cell(r, 7, round(row_w, 4) if row_w else 0.0)
                # 第 8 列为第 7 列重量的单位（模板为 KG；可按产品写 weight_unit 或 total_weight 如 15公斤）
                ws.cell(r, 8, _product_weight_unit_label(p))
                ws.cell(r, 9, price)
                # 模板自带「元/KG」；计件（桶/缶等）时必须改为「元/桶」等与第 8 列一致
                ws.cell(r, 10, _price_per_unit_cell_label(p))
                ws.cell(r, 11, amount if amount else 0.0)
                if ws.cell(r, 12).value in (None, ""):
                    ws.cell(r, 12, "元")
            else:
                # 用空字符串清空，避免 openpyxl 对 None 仍保留旧数值（如 0）导致公式列出现 #VALUE!
                # 勿在空行再写「KG」：模板第 18 行总重旁已有单位，空产品行保留 KG 会叠出多余单位列。
                for c in (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12):
                    ws.cell(r, c, "")

        tw = str(template_data.get("total_weight") or "").strip()
        w_val, w_unit = _parse_weight_value_and_unit(tw)
        if w_val is not None:
            ws.cell(18, 7, round(w_val, 4))
            ws.cell(18, 8, w_unit)

        # 送货单.xls：第 19 行 B 列为中文大写金额占位，L 列为金额数字占位（与模板一致）
        tac = str(template_data.get("total_amount_chinese") or "").strip()
        ws.cell(19, 2, tac)
        ta = template_data.get("total_amount")
        if ta not in (None, ""):
            num = _parse_money_cell(ta)
            if num > 0:
                ws.cell(19, 12, round(num, 2))

        wb.save(str(out))
    finally:
        wb.close()


def read_excel_sales_contract_preview(template_path: str | Path) -> dict[str, Any]:
    """抽取首表表头与示例行，供前端列预览。"""
    p = Path(template_path)
    if not p.is_file():
        return {"success": False, "message": f"文件不存在: {p}"}
    try:
        if p.suffix.lower() == ".xls":
            xlrd = _import_xlrd_for_xls()

            book = xlrd.open_workbook(str(p), formatting_info=False)
            sh = book.sheet_by_index(0)
            header_row_idx = None
            headers: list[str] = []
            for ri in range(min(20, sh.nrows)):
                row_vals = [str(sh.cell_value(ri, ci)).strip() for ci in range(sh.ncols)]
                if "编号" in row_vals:
                    header_row_idx = ri
                    headers = [str(sh.cell_value(ri, ci)) for ci in range(sh.ncols)]
                    break
            if header_row_idx is None:
                headers = [
                    "编号",
                    "型号",
                    "名称",
                    "规格",
                    "",
                    "单位",
                    "数量",
                    "单位",
                    "",
                    "计价",
                    "单价",
                    "金额",
                    "",
                ]
                data_start = 4
            else:
                data_start = header_row_idx + 1
            sample_rows: list[dict[str, Any]] = []
            for ri in range(data_start, min(data_start + 3, sh.nrows)):
                if ri >= sh.nrows:
                    break
                row_map: dict[str, Any] = {}
                for ci, h in enumerate(headers):
                    if not str(h).strip():
                        continue
                    row_map[str(h)] = sh.cell_value(ri, ci) if ci < sh.ncols else ""
                if any(v not in ("", None) for v in row_map.values()):
                    sample_rows.append(row_map)
            if not sample_rows and sh.nrows >= 5:
                ri = 4
                row_map = {}
                for ci, h in enumerate(headers):
                    if not str(h).strip():
                        continue
                    row_map[str(h)] = sh.cell_value(ri, ci) if ci < sh.ncols else ""
                sample_rows.append(row_map)
            return {"success": True, "headers": headers, "sample_rows": sample_rows}
        from openpyxl import load_workbook

        wb = load_workbook(str(p), read_only=True, data_only=True)
        try:
            ws = wb.active
            header_row_idx = None
            headers = []
            max_scan = min(ws.max_row or 1, 25)
            max_col = min(ws.max_column or 1, 30)
            for ri in range(1, max_scan + 1):
                row_vals = []
                for ci in range(1, max_col + 1):
                    v = ws.cell(ri, ci).value
                    row_vals.append("" if v is None else str(v).strip())
                if "编号" in row_vals:
                    header_row_idx = ri
                    headers = [
                        "" if ws.cell(ri, ci).value is None else str(ws.cell(ri, ci).value)
                        for ci in range(1, max_col + 1)
                    ]
                    break
            if not headers:
                headers = [
                    "编号",
                    "型号",
                    "名称",
                    "规格",
                    "",
                    "单位",
                    "数量",
                    "单位",
                    "",
                    "计价",
                    "单价",
                    "金额",
                    "",
                ]
                header_row_idx = 4
            sample_rows = []
            start = (header_row_idx or 4) + 1
            for ri in range(start, min(start + 3, (ws.max_row or start) + 1)):
                row_map = {}
                for ci, h in enumerate(headers, start=1):
                    hn = str(h).strip() if h is not None else ""
                    if not hn:
                        continue
                    row_map[hn] = ws.cell(ri, ci).value
                if any(v not in ("", None) for v in row_map.values()):
                    sample_rows.append(row_map)
            return {"success": True, "headers": headers, "sample_rows": sample_rows}
        finally:
            wb.close()
    except Exception as e:
        logger.exception("read_excel_sales_contract_preview failed")
        return {"success": False, "message": str(e)}


def update_preview_products(
    template_path,
    products=None,
    customer_name: str = "客户",
    contract_date: str = "",
    output_dir=None,
) -> dict:
    """
    前端删除/修改行后调用：用最新 products 重新生成 Excel，返回新 file_path 和预览数据。
    供右侧任务面板「销售合同 · Excel 列预览」实时同步。
    """
    from datetime import datetime
    from pathlib import Path

    if products is None:
        products = []

    if output_dir is None:
        from backend.sales_contract_generate_core import _get_output_dir

        output_dir = _get_output_dir()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(output_dir) / f"sales_contract_preview_{timestamp}.xlsx"

    template_data = {
        "customer_name": customer_name,
        "contract_date": contract_date or datetime.now().strftime("%Y-%m-%d"),
        "products": products,
    }

    try:
        fill_sales_contract_excel_template(template_path, template_data, output_path)
        preview = read_excel_sales_contract_preview(template_path)
        preview["products"] = products
        preview["file_path"] = str(output_path)
        preview["download_url"] = f"/api/sales-contract/download?filepath={output_path}"
        preview["success"] = True
        return preview
    except Exception as e:
        logger.exception("update_preview_products failed")
        return {"success": False, "message": str(e), "products": products}
