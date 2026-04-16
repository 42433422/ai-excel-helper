"""
Excel (.xlsx) template handling for price table export.
"""

from __future__ import annotations

import os
import re
import uuid
from datetime import date
from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter


_ALLOWED_SUFFIX = {".xlsx", ".xlsm"}


def _load_units_from_db() -> dict[str, dict]:
    """Load purchase units from database."""
    from backend.product_db_read import load_purchase_units_for_templates

    return load_purchase_units_for_templates()


def _load_products_from_db() -> dict[str, dict]:
    """Load products from database."""
    from backend.product_db_read import load_products_dict_for_templates

    return load_products_dict_for_templates()


def _normalize_text(text: str) -> str:
    """Text normalization: remove extra spaces and special characters."""
    if not text:
        return ""
    text = text.strip()
    text = re.sub(r'\s+', ' ', text)
    return text


def _fuzzy_match(text: str, candidates: dict[str, Any], threshold: float = 0.8) -> tuple[str, Any] | None:
    """Fuzzy match text to candidates."""
    if not text:
        return None
    normalized = _normalize_text(text).lower()
    normalized_key = normalized.replace(' ', '')

    for key, value in candidates.items():
        key_normalized = key.lower().replace(' ', '')
        if normalized == key or normalized_key == key_normalized:
            return (key, value)

        if normalized in key.lower() or key.lower() in normalized:
            return (key, value)

    text_chars = set(normalized.replace(' ', ''))
    best_match = None
    best_score = 0

    for key, value in candidates.items():
        key_chars = set(key.lower().replace(' ', ''))
        if not key_chars:
            continue
        intersection = text_chars & key_chars
        score = len(intersection) / max(len(text_chars), len(key_chars))
        if score >= threshold and score > best_score:
            best_score = score
            best_match = (key, value)

    return best_match


def cleanup_price_data(
    raw_data: list[dict[str, Any]],
    customer_name: str | None = None,
    quote_date: str | None = None,
) -> dict[str, Any]:
    """
    Cleanup price table data:
    1. Normalize customer name -> match database units
    2. Normalize product name -> match database products
    3. Get unit price (from database or calculate)
    4. Format date to today

    Returns:
        {
            "cleaned_data": [...],
            "unit_info": {...},
            "products_matched": {...},
            "unmatched_products": [...],
            "quote_date": "2026-04-11",
        }
    """
    units_db = _load_units_from_db()
    products_db = _load_products_from_db()

    today = date.today().strftime("%Y-%m-%d")
    final_date = quote_date or today

    matched_unit = None
    unmatched_unit = customer_name
    if customer_name:
        matched = _fuzzy_match(customer_name, units_db)
        if matched:
            matched_unit = matched[1]
            unmatched_unit = None

    cleaned_data = []
    unmatched_products = []
    products_matched = {}

    for row in raw_data:
        cleaned_row = dict(row)

        product_name_raw = str(row.get("产品名称") or row.get("产品") or row.get("name") or "")
        product_name = _normalize_text(product_name_raw)

        if not product_name:
            continue

        product_match = _fuzzy_match(product_name, products_db)

        if product_match:
            matched_key, matched_info = product_match
            products_matched[product_name] = matched_key
            cleaned_row["产品名称"] = matched_info["name"]
            cleaned_row["单位"] = matched_info.get("specification") or row.get("单位") or ""
            cleaned_row["单价"] = matched_info.get("price", 0)
            cleaned_row["产品ID"] = matched_info["id"]
            cleaned_row["matched"] = True
        else:
            unmatched_products.append(product_name)
            cleaned_row["matched"] = False

        cleaned_row["客户名称"] = matched_unit["unit_name"] if matched_unit else (customer_name or "")
        cleaned_row["报价日期"] = final_date

        cleaned_data.append(cleaned_row)

    result = {
        "cleaned_data": cleaned_data,
        "quote_date": final_date,
        "unit_info": matched_unit,
        "unmatched_unit": unmatched_unit,
        "products_matched": products_matched,
        "unmatched_products": list(set(unmatched_products)),
    }

    if matched_unit:
        result["unit_id"] = matched_unit.get("id")

    return result


def resolve_safe_excel_path(workspace_root: str, file_path: str) -> Path:
    """
    Resolve file_path to an absolute path that must stay under workspace_root.
    """
    root = Path(workspace_root).resolve()
    if not root.is_dir():
        raise FileNotFoundError(f"workspace root is not a directory: {root}")

    raw = Path(file_path)
    if raw.is_absolute():
        candidate = raw.resolve()
    else:
        candidate = (root / raw).resolve()

    try:
        candidate.relative_to(root)
    except ValueError as e:
        raise PermissionError("file_path must resolve inside workspace root") from e

    if ".." in Path(file_path).parts:
        raise PermissionError("file_path must not contain '..'")

    if candidate.suffix.lower() not in _ALLOWED_SUFFIX:
        raise ValueError("only Excel files (.xlsx, .xlsm) are supported")

    return candidate


def _find_placeholder_cells(wb) -> list[dict[str, Any]]:
    """Find all cells containing {{placeholder}} patterns."""
    placeholders = []
    placeholder_pattern = re.compile(r'\{\{(\w+)\}\}')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    matches = placeholder_pattern.findall(cell.value)
                    if matches:
                        placeholders.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "row": cell.row,
                            "col": cell.column,
                            "value": cell.value,
                            "fields": matches,
                        })
    return placeholders


def parse_xlsx_template(path: Path) -> dict[str, Any]:
    """
    Parse all sheets and tables from an Excel workbook.
    Returns structure with sheet info and placeholder locations.
    """
    wb = load_workbook(path, data_only=False)
    sheets_data = []
    placeholders = _find_placeholder_cells(wb)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        rows_data = []
        for row_idx in range(1, min(max_row + 1, 101)):
            row_cells = []
            for col_idx in range(1, min(max_col + 1, 20)):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_cells.append(str(cell.value) if cell.value is not None else "")
            rows_data.append(row_cells)

        sheets_data.append({
            "sheet_name": sheet_name,
            "max_row": max_row,
            "max_column": max_col,
            "rows": rows_data,
            "header": rows_data[0] if rows_data else [],
            "data_rows": rows_data[1:] if len(rows_data) > 1 else [],
        })

    return {
        "sheet_count": len(wb.sheetnames),
        "sheets": sheets_data,
        "placeholders": placeholders,
        "placeholder_fields": list({f for p in placeholders for f in p["fields"]}),
    }


def fill_excel_template(
    template_path: Path,
    data: Mapping[str, Any],
    output_path: Path | None = None,
) -> dict[str, Any]:
    """
    Fill an Excel template with data.
    - template_path: path to .xlsx template
    - data: dict with keys matching template placeholders {{placeholder}}
    - output_path: if provided, save filled workbook here

    Placeholder format: {{field_name}} in cell values.
    """
    wb = load_workbook(template_path)
    placeholder_pattern = re.compile(r'\{\{(\w+)\}\}')

    filled_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original = cell.value
                    new_value = cell.value

                    def replace_placeholder(m):
                        nonlocal filled_count
                        key = m.group(1)
                        if key in data:
                            filled_count += 1
                            return str(data[key])
                        return m.group(0)

                    new_value = placeholder_pattern.sub(replace_placeholder, new_value)
                    if new_value != original:
                        cell.value = new_value

    if output_path:
        wb.save(output_path)

    return {
        "filled_count": filled_count,
        "output_path": str(output_path) if output_path else None,
    }


def export_to_excel(
    df: pd.DataFrame,
    output_path: Path,
    sheet_name: str = "Sheet1",
    include_header: bool = True,
) -> dict[str, Any]:
    """
    Export a DataFrame to an Excel file.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=not include_header, header=include_header)

    return {
        "row_count": len(df),
        "column_count": len(df.columns),
        "output_path": str(output_path),
        "sheet_name": sheet_name,
    }


def _default_workspace_root() -> str:
    return os.environ.get("WORKSPACE_ROOT", os.getcwd())


def handle_excel_template(
    arguments: Mapping[str, Any],
    workspace_root: str | None = None,
) -> dict[str, Any]:
    """
    Main entry point for Excel template operations.
    Actions: parse, fill, export, cleanup
    """
    root = workspace_root or _default_workspace_root()
    action = (arguments.get("action") or "").lower()
    file_path = arguments.get("file_path")

    if action == "cleanup":
        raw_data = arguments.get("raw_data") or arguments.get("records") or []
        customer_name = arguments.get("customer_name")
        quote_date = arguments.get("quote_date")
        return cleanup_price_data(raw_data, customer_name=customer_name, quote_date=quote_date)

    if not file_path:
        return {"error": "missing_file_path"}

    try:
        path = resolve_safe_excel_path(root, str(file_path))
    except (OSError, ValueError, PermissionError) as e:
        return {"error": "path_resolution_failed", "message": str(e)}

    if not path.is_file():
        return {"error": "file_not_found", "path": str(path)}

    if action == "parse":
        return parse_xlsx_template(path)

    if action == "fill":
        data = arguments.get("data") or {}
        upload_dir = Path(root) / "uploads"
        upload_dir.mkdir(parents=True, exist_ok=True)
        output_filename = f"filled_{uuid.uuid4().hex}.xlsx"
        output_path = upload_dir / output_filename

        result = fill_excel_template(path, data, output_path)
        try:
            rel = output_path.relative_to(Path(root).resolve())
            result["output_path"] = rel.as_posix()
        except ValueError:
            result["output_path"] = str(output_path)

        return result

    if action == "export":
        records = arguments.get("records") or arguments.get("data_rows") or []
        columns = arguments.get("columns") or (list(records[0].keys()) if records else [])

        df = pd.DataFrame(records, columns=columns)

        upload_dir = Path(root) / "uploads"
        upload_dir.mkdir(parents=True, exist_ok=True)
        output_filename = f"export_{uuid.uuid4().hex}.xlsx"
        output_path = upload_dir / output_filename

        sheet_name = arguments.get("sheet_name", "Sheet1")
        result = export_to_excel(df, output_path, sheet_name=sheet_name)
        try:
            rel = output_path.relative_to(Path(root).resolve())
            result["output_path"] = rel.as_posix()
        except ValueError:
            result["output_path"] = str(output_path)

        return result

    return {
        "error": "invalid_action",
        "action": action,
        "allowed": ["parse", "fill", "export", "cleanup"],
    }


def get_excel_template_tool_registry() -> list[dict[str, Any]]:
    """
    Excel template tool for OpenAI Chat Completions tools list.
    """
    return [
        {
            "type": "function",
            "function": {
                "name": "excel_template",
                "description": (
                    "【Excel 模板工具】处理 .xlsx 价格表模板："
                    "parse=解析表格结构；fill=用数据填充模板占位符；export=DataFrame导出为Excel；"
                    "cleanup=清理价格表数据（标准化客户名称匹配数据库单位、产品名称匹配数据库产品、格式化日期）"
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "action": {
                            "type": "string",
                            "enum": ["parse", "fill", "export", "cleanup"],
                            "description": "parse: 解析文档表格结构; fill: 填充模板占位符; export: DataFrame导出为Excel; cleanup: 清理价格表数据",
                        },
                        "file_path": {"type": "string", "description": "Path to .xlsx file under workspace"},
                        "data": {
                            "type": "object",
                            "description": "For fill action: key-value pairs to replace {{placeholder}} in template",
                        },
                        "records": {
                            "type": "array",
                            "description": "For export/cleanup action: array of records",
                        },
                        "raw_data": {
                            "type": "array",
                            "description": "Alias for records, for cleanup action",
                        },
                        "customer_name": {
                            "type": "string",
                            "description": "For cleanup action: 客户名称 to match against database units",
                        },
                        "quote_date": {
                            "type": "string",
                            "description": "For cleanup action: 报价日期 (defaults to today)",
                        },
                        "columns": {
                            "type": "array",
                            "description": "For export action: column names for the Excel sheet",
                        },
                        "data_rows": {
                            "type": "array",
                            "description": "Alias for records",
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "For export action: sheet name (default 'Sheet1')",
                            "default": "Sheet1",
                        },
                    },
                    "required": ["action"],
                },
            },
        },
    ]
