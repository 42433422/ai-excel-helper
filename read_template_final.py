"""查看模板文件结构 - 直接匹配"""
import os
from openpyxl import load_workbook
import tempfile
import shutil

dir_path = r'e:\FHD\424'
target_bytes = b'\xe8\x80\x83\xe5\x8b\xa4-2026-3\xe6\x9c\x88\xe4\xbb\xbd\xe8\x80\x83\xe5\x8b\xa4\xe7\xbb\x9f\xe8\xae\xa1\xe8\xa1\xa8.xlsx'

# 使用 os.listdir 和完整路径
files = os.listdir(dir_path)

# 找到目标文件
for f in files:
    if f.encode('utf-8') == target_bytes:
        full_path = os.path.join(dir_path, f)
        print(f"找到目标文件：{full_path}")
        print(f"文件大小：{os.path.getsize(full_path)} 字节")
        
        # 复制到临时文件
        temp_file = tempfile.mktemp(suffix='.xlsx')
        shutil.copy2(full_path, temp_file)
        
        try:
            wb = load_workbook(temp_file, data_only=True)
            
            print(f"\n工作表：{wb.sheetnames}\n")
            
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                print(f"=== 工作表：{ws_name} ===")
                print(f"最大行：{ws.max_row}, 最大列：{ws.max_column}")
                merged = list(ws.merged_cells.ranges) if ws.merged_cells.ranges else []
                print(f"合并单元格：{len(merged)} 个")
                
                # 显示前 10 行
                print(f"\n前 10 行:")
                for row_idx in range(1, min(11, ws.max_row + 1)):
                    row_data = []
                    for col_idx in range(1, min(15, ws.max_column + 1)):
                        val = ws.cell(row=row_idx, column=col_idx).value
                        row_data.append(str(val)[:20] if val else '')
                    print(f"  行{row_idx}: {' | '.join(row_data)}")
                print()
            
            wb.close()
            
        finally:
            if os.path.exists(temp_file):
                os.unlink(temp_file)
        
        break
