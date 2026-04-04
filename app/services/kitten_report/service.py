# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .plugins import (
    AnalysisPlugin,
    ForecastHeuristicPlugin,
    IndustryStrategyPlugin,
    PluginResult,
    RuleStatsPlugin,
    TimeSeriesModelPlugin,
)
from .financial_plugins import FinancialReportPlugin, InventoryValuationPlugin


class KittenReportExportService:
    """
    小猫分析报告导出（混合模式）：
    - 前端提供会话与数据摘要
    - 后端执行插件算法（规则/预测/行业策略/财务报表/库存评估）
    - 后端生成并返回可下载 XLSX
    """

    def __init__(self) -> None:
        self.plugins: List[AnalysisPlugin] = [
            RuleStatsPlugin(),
            TimeSeriesModelPlugin(),
            ForecastHeuristicPlugin(),
            IndustryStrategyPlugin(),
            FinancialReportPlugin(),
            InventoryValuationPlugin(),
        ]

    def build_report(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        dataset = payload.get("dataset") or {}
        messages = payload.get("messages") or []
        result = payload.get("result") or {}
        phase = str(payload.get("phase") or "unknown")
        industry = str(payload.get("industry") or "通用")

        plugin_payload = {
            "dataset": dataset,
            "messages": messages,
            "result": result,
            "phase": phase,
            "industry": industry,
        }
        plugin_results = [self._plugin_to_dict(plugin.run(plugin_payload)) for plugin in self.plugins]
        xlsx_bytes = self._build_workbook(
            dataset=dataset,
            messages=messages,
            result=result,
            phase=phase,
            industry=industry,
            plugin_results=plugin_results,
        )
        file_name = f"小猫分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return {"file_name": file_name, "content": xlsx_bytes, "plugins": plugin_results}

    @staticmethod
    def _plugin_to_dict(item: PluginResult) -> Dict[str, Any]:
        return {
            "key": item.key,
            "title": item.title,
            "level": item.level,
            "summary": item.summary,
            "details": item.details,
        }

    def _build_workbook(
        self,
        *,
        dataset: Dict[str, Any],
        messages: List[Dict[str, Any]],
        result: Dict[str, Any],
        phase: str,
        industry: str,
        plugin_results: List[Dict[str, Any]],
    ) -> bytes:
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "报告摘要"
        ws_summary.append(["字段", "内容"])
        ws_summary.append(["报告标题", str(result.get("title") or "AI 分析")])
        ws_summary.append(["报告时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_summary.append(["分析阶段", phase])
        ws_summary.append(["行业", industry])
        ws_summary.append(["摘要", str(result.get("summary") or "")])

        if dataset:
            ws_summary.append(["数据文件", str(dataset.get("name") or "")])
            ws_summary.append(["数据规模", f"{int(dataset.get('rows') or 0)} 行 / {int(dataset.get('columns') or 0)} 列"])

        ws_alg = wb.create_sheet("算法洞察")
        ws_alg.append(["插件", "级别", "结论", "详情"])
        for item in plugin_results:
            ws_alg.append([
                item.get("title", ""),
                item.get("level", ""),
                item.get("summary", ""),
                str(item.get("details", {})),
            ])

        ws_chat = wb.create_sheet("对话记录")
        ws_chat.append(["序号", "角色", "时间", "内容"])
        if messages:
            for idx, msg in enumerate(messages, start=1):
                ws_chat.append([
                    idx,
                    "AI" if str(msg.get("role", "")) == "ai" else "用户",
                    str(msg.get("time") or ""),
                    self._html_to_text(str(msg.get("content") or "")),
                ])
        else:
            ws_chat.append([1, "系统", "", "暂无对话记录"])

        if dataset:
            ws_dataset = wb.create_sheet("数据摘要")
            ws_dataset.append(["字段", "内容"])
            ws_dataset.append(["文件名", str(dataset.get("name") or "")])
            ws_dataset.append(["总行数", int(dataset.get("rows") or 0)])
            ws_dataset.append(["总列数", int(dataset.get("columns") or 0)])
            ws_dataset.append(["字段列表", "、".join(dataset.get("fieldNames") or [])])
            ws_dataset.append(["预览文本", str(dataset.get("previewText") or "")])

        self._add_financial_sheet(wb, plugin_results)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _add_financial_sheet(self, wb: Workbook, plugin_results: List[Dict[str, Any]]) -> None:
        financial_plugin = next((p for p in plugin_results if p.get("key") == "financial_report"), None)
        inventory_plugin = next((p for p in plugin_results if p.get("key") == "inventory_valuation"), None)

        if not financial_plugin and not inventory_plugin:
            return

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws_fin = wb.create_sheet("财务报表")
        current_row = 1
        ws_fin.cell(row=current_row, column=1, value="📊 小猫财务分析报告").font = Font(bold=True, size=16)
        ws_fin.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 2

        if financial_plugin:
            details = financial_plugin.get("details", {})
            metrics = details.get("metrics", {})

            ws_fin.cell(row=current_row, column=1, value="核心财务指标").font = Font(bold=True, size=12)
            current_row += 1

            headers = ['指标', '数值', '单位']
            for col, header in enumerate(headers, 1):
                cell = ws_fin.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            current_row += 1

            financial_data = [
                ('总营收', metrics.get('total_revenue', 0), '¥'),
                ('成本估算', metrics.get('total_cost', 0), '¥'),
                ('毛利润', metrics.get('gross_profit', 0), '¥'),
                ('毛利率', metrics.get('profit_margin', 0), '%'),
                ('订单数量', metrics.get('order_count', 0), '笔'),
                ('平均订单额', metrics.get('avg_order_value', 0), '¥'),
            ]

            for label, value, unit in financial_data:
                ws_fin.cell(row=current_row, column=1, value=label).border = thin_border
                cell_val = ws_fin.cell(row=current_row, column=2, value=round(value, 2) if isinstance(value, (int, float)) else value)
                cell_val.border = thin_border
                cell_val.alignment = Alignment(horizontal="right")
                ws_fin.cell(row=current_row, column=3, value=unit).border = thin_border
                current_row += 1

            monthly_data = details.get("monthly_breakdown", [])
            if monthly_data:
                current_row += 2
                ws_fin.cell(row=current_row, column=1, value="月度营收趋势").font = Font(bold=True, size=12)
                current_row += 1

                month_headers = ['月份', '营收 (¥)', '订单数']
                for col, header in enumerate(month_headers, 1):
                    cell = ws_fin.cell(row=current_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                current_row += 1

                for month in monthly_data:
                    ws_fin.cell(row=current_row, column=1, value=month.get("month")).border = thin_border
                    ws_fin.cell(row=current_row, column=2, value=round(month.get("revenue", 0), 2)).border = thin_border
                    ws_fin.cell(row=current_row, column=3, value=month.get("order_count", 0)).border = thin_border
                    current_row += 1

            product_analysis = details.get("product_analysis", [])
            if product_analysis:
                current_row += 2
                ws_fin.cell(row=current_row, column=1, value="产品销售排行 (Top 10)").font = Font(bold=True, size=12)
                current_row += 1

                prod_headers = ['产品名称', '总营收 (¥)', '销量', '订单数']
                for col, header in enumerate(prod_headers, 1):
                    cell = ws_fin.cell(row=current_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                current_row += 1

                for product in product_analysis[:10]:
                    ws_fin.cell(row=current_row, column=1, value=product.get("product_name")).border = thin_border
                    ws_fin.cell(row=current_row, column=2, value=round(product.get("total_revenue", 0), 2)).border = thin_border
                    ws_fin.cell(row=current_row, column=3, value=round(product.get("total_qty", 0), 2)).border = thin_border
                    ws_fin.cell(row=current_row, column=4, value=product.get("order_count", 0)).border = thin_border
                    current_row += 1

            customer_analysis = details.get("customer_analysis", [])
            if customer_analysis:
                current_row += 2
                ws_fin.cell(row=current_row, column=1, value="客户销售排行 (Top 10)").font = Font(bold=True, size=12)
                current_row += 1

                cust_headers = ['客户名称', '总金额 (¥)', '订单数', '平均订单额']
                for col, header in enumerate(cust_headers, 1):
                    cell = ws_fin.cell(row=current_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                current_row += 1

                for customer in customer_analysis[:10]:
                    ws_fin.cell(row=current_row, column=1, value=customer.get("customer")).border = thin_border
                    ws_fin.cell(row=current_row, column=2, value=round(customer.get("total_amount", 0), 2)).border = thin_border
                    ws_fin.cell(row=current_row, column=3, value=customer.get("order_count", 0)).border = thin_border
                    ws_fin.cell(row=current_row, column=4, value=round(customer.get("avg_order_value", 0), 2)).border = thin_border
                    current_row += 1

        if inventory_plugin:
            current_row += 2
            details = inventory_plugin.get("details", {})

            ws_fin.cell(row=current_row, column=1, value="📦 库存价值评估").font = Font(bold=True, size=12)
            current_row += 1

            materials = details.get("materials", {})
            products = details.get("products", {})

            inv_headers = ['类别', '项目数', '估值 (¥)']
            for col, header in enumerate(inv_headers, 1):
                cell = ws_fin.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            current_row += 1

            ws_fin.cell(row=current_row, column=1, value="原材料").border = thin_border
            ws_fin.cell(row=current_row, column=2, value=materials.get("total_items", 0)).border = thin_border
            ws_fin.cell(row=current_row, column=3, value=round(materials.get("total_value", 0), 2)).border = thin_border
            current_row += 1

            ws_fin.cell(row=current_row, column=1, value="成品/产品").border = thin_border
            ws_fin.cell(row=current_row, column=2, value=products.get("total_items", 0)).border = thin_border
            ws_fin.cell(row=current_row, column=3, value=round(products.get("total_value", 0), 2)).border = thin_border
            current_row += 1

            total_inv = materials.get("total_value", 0) + products.get("total_value", 0)
            ws_fin.cell(row=current_row, column=1, value="合计").font = Font(bold=True)
            ws_fin.cell(row=current_row, column=1).border = thin_border
            ws_fin.cell(row=current_row, column=2).border = thin_border
            ws_fin.cell(row=current_row, column=3, value=round(total_inv, 2)).font = Font(bold=True)
            ws_fin.cell(row=current_row, column=3).border = thin_border

            low_stock = details.get("low_stock_alerts", [])
            if low_stock:
                current_row += 2
                ws_fin.cell(row=current_row, column=1, value="⚠️ 低库存预警").font = Font(bold=True, color="FF0000")
                current_row += 1

                alert_headers = ['物料名称', '当前库存', '安全库存', '单价 (¥)']
                for col, header in enumerate(alert_headers, 1):
                    cell = ws_fin.cell(row=current_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    cell.alignment = header_alignment
                    cell.border = thin_border
                current_row += 1

                for item in low_stock[:10]:
                    ws_fin.cell(row=current_row, column=1, value=item.get("name")).border = thin_border
                    ws_fin.cell(row=current_row, column=2, value=item.get("current")).border = thin_border
                    ws_fin.cell(row=current_row, column=3, value=item.get("min_required")).border = thin_border
                    ws_fin.cell(row=current_row, column=4, value=item.get("unit_price")).border = thin_border
                    current_row += 1

        for col in range(1, 5):
            ws_fin.column_dimensions[get_column_letter(col)].width = 22

    @staticmethod
    def _html_to_text(content: str) -> str:
        text = content.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
        text = text.replace("<strong>", "").replace("</strong>", "")
        text = text.replace("&nbsp;", " ").replace("&amp;", "&")
        return text
