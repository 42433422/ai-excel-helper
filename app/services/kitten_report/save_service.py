# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

logger = __import__("logging").getLogger(__name__)


class AnalysisSaveService:
    def __init__(self, save_dir: Optional[str] = None):
        self.save_dir = save_dir or os.path.join(os.getcwd(), "saved_analyses")
        os.makedirs(self.save_dir, exist_ok=True)

    def save_analysis(
        self,
        analysis_type: str,
        data: Dict[str, Any],
        metadata: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{analysis_type}_{timestamp}.json"
            filepath = os.path.join(self.save_dir, filename)

            save_data = {
                "id": f"analysis_{timestamp}",
                "type": analysis_type,
                "created_at": datetime.now().isoformat(),
                "metadata": metadata or {},
                "data": data,
            }

            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2, default=str)

            logger.info("Analysis saved: %s", filepath)

            return {
                "success": True,
                "id": save_data["id"],
                "filename": filename,
                "filepath": filepath,
                "created_at": save_data["created_at"],
            }
        except Exception as e:
            logger.exception("Failed to save analysis: %s", e)
            return {"success": False, "error": str(e)}

    def list_saved_analyses(self, analysis_type: Optional[str] = None) -> List[Dict[str, Any]]:
        try:
            analyses = []

            for filename in os.listdir(self.save_dir):
                if not filename.endswith(".json"):
                    continue

                if analysis_type and not filename.startswith(f"{analysis_type}_"):
                    continue

                filepath = os.path.join(self.save_dir, filename)

                try:
                    with open(filepath, "r", encoding="utf-8") as f:
                        data = json.load(f)

                    analyses.append({
                        "id": data.get("id"),
                        "type": data.get("type"),
                        "created_at": data.get("created_at"),
                        "filename": filename,
                        "metadata": data.get("metadata", {}),
                        "filepath": filepath,
                    })
                except Exception:
                    continue

            analyses.sort(key=lambda x: x.get("created_at", ""), reverse=True)

            return analyses
        except Exception as e:
            logger.exception("Failed to list analyses: %s", e)
            return []

    def get_analysis(self, analysis_id: str) -> Optional[Dict[str, Any]]:
        try:
            for filename in os.listdir(self.save_dir):
                if not filename.endswith(".json"):
                    continue

                filepath = os.path.join(self.save_dir, filename)

                with open(filepath, "r", encoding="utf-8") as f:
                    data = json.load(f)

                if data.get("id") == analysis_id:
                    return data

            return None
        except Exception as e:
            logger.exception("Failed to get analysis: %s", e)
            return None

    def delete_analysis(self, analysis_id: str) -> Dict[str, Any]:
        try:
            analysis = self.get_analysis(analysis_id)

            if not analysis:
                return {"success": False, "error": "Analysis not found"}

            filepath = analysis.get("filepath")

            if filepath and os.path.exists(filepath):
                os.remove(filepath)
                return {"success": True, "message": "Analysis deleted"}

            return {"success": False, "error": "File not found"}
        except Exception as e:
            logger.exception("Failed to delete analysis: %s", e)
            return {"success": False, "error": str(e)}

    def export_analysis_to_xlsx(self, analysis_id: str) -> Dict[str, Any]:
        try:
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            analysis = self.get_analysis(analysis_id)

            if not analysis:
                return {"success": False, "error": "Analysis not found"}

            wb = Workbook()
            ws = wb.active
            ws.title = "财务报表"

            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            ws['A1'] = "小猫财务分析报告"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:D1')

            ws['A3'] = "生成时间:"
            ws['B3'] = analysis.get('created_at', '')
            ws['A4'] = "分析类型:"
            ws['B4'] = analysis.get('type', '')

            data = analysis.get('data', {})
            metrics = data.get('metrics', {})

            row = 6
            headers = ['指标', '数值']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            financial_metrics = [
                ('总营收 (¥)', metrics.get('total_revenue', 0)),
                ('总成本 (¥)', metrics.get('total_cost', 0)),
                ('毛利润 (¥)', metrics.get('gross_profit', 0)),
                ('毛利率 (%)', metrics.get('profit_margin', 0)),
                ('订单数量', metrics.get('order_count', 0)),
                ('平均订单金额 (¥)', metrics.get('avg_order_value', 0)),
            ]

            for i, (label, value) in enumerate(financial_metrics, row + 1):
                ws.cell(column=1, row=i, value=label).border = thin_border
                ws.cell(column=2, row=i, value=value).border = thin_border

            monthly_data = data.get('monthly_breakdown', [])

            if monthly_data:
                start_row = len(financial_metrics) + row + 3
                ws.cell(row=start_row, column=1, value="月度营收趋势").font = Font(bold=True, size=12)

                start_row += 1
                month_headers = ['月份', '营收 (¥)', '订单数']
                for col, header in enumerate(month_headers, 1):
                    cell = ws.cell(row=start_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                for i, month in enumerate(monthly_data, start_row + 1):
                    ws.cell(column=1, row=i, value=month.get('month', '')).border = thin_border
                    ws.cell(column=2, row=i, value=month.get('revenue', 0)).border = thin_border
                    ws.cell(column=3, row=i, value=month.get('order_count', 0)).border = thin_border

            product_data = data.get('product_analysis', [])
            if product_data:
                start_row = ws.max_row + 3
                ws.cell(row=start_row, column=1, value="产品销售排行").font = Font(bold=True, size=12)

                start_row += 1
                prod_headers = ['产品名称', '总营收 (¥)', '销量', '订单数', '平均单价']
                for col, header in enumerate(prod_headers, 1):
                    cell = ws.cell(row=start_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                for i, product in enumerate(product_data[:10], start_row + 1):
                    ws.cell(column=1, row=i, value=product.get('product_name', '')).border = thin_border
                    ws.cell(column=2, row=i, value=product.get('total_revenue', 0)).border = thin_border
                    ws.cell(column=3, row=i, value=product.get('total_qty', 0)).border = thin_border
                    ws.cell(column=4, row=i, value=product.get('order_count', 0)).border = thin_border
                    ws.cell(column=5, row=i, value=product.get('avg_price', 0)).border = thin_border

            customer_data = data.get('customer_analysis', [])
            if customer_data:
                start_row = ws.max_row + 3
                ws.cell(row=start_row, column=1, value="客户销售排行").font = Font(bold=True, size=12)

                start_row += 1
                cust_headers = ['客户名称', '总金额 (¥)', '订单数', '平均订单额']
                for col, header in enumerate(cust_headers, 1):
                    cell = ws.cell(row=start_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                for i, customer in enumerate(customer_data[:10], start_row + 1):
                    ws.cell(column=1, row=i, value=customer.get('customer', '')).border = thin_border
                    ws.cell(column=2, row=i, value=customer.get('total_amount', 0)).border = thin_border
                    ws.cell(column=3, row=i, value=customer.get('order_count', 0)).border = thin_border
                    ws.cell(column=4, row=i, value=customer.get('avg_order_value', 0)).border = thin_border

            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 20

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            file_name = f"小猫财务报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            return {
                "success": True,
                "file_name": file_name,
                "content": output.read(),
            }
        except Exception as e:
            logger.exception("Export to XLSX failed: %s", e)
            return {"success": False, "error": str(e)}

    def get_statistics_summary(self) -> Dict[str, Any]:
        try:
            analyses = self.list_saved_analyses()

            type_counts = {}
            for a in analyses:
                atype = a.get('type', 'unknown')
                type_counts[atype] = type_counts.get(atype, 0) + 1

            return {
                "total_analyses": len(analyses),
                "by_type": type_counts,
                "latest": analyses[:5] if analyses else [],
            }
        except Exception as e:
            logger.exception("Statistics failed: %s", e)
            return {"total_analyses": 0, "by_type": {}, "latest": []}


analysis_save_service = AnalysisSaveService()
