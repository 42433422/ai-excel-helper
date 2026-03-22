# -*- coding: utf-8 -*-
"""
Excel 模板分析工具模块

提供 Excel 模板结构分析和词条提取功能。
"""

import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Fill, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter


@dataclass
class CellStyle:
    font_name: Optional[str] = None
    font_size: Optional[float] = None
    font_bold: Optional[bool] = None
    font_color: Optional[str] = None
    fill_pattern: Optional[str] = None
    fill_fg_color: Optional[str] = None
    fill_bg_color: Optional[str] = None
    alignment_horizontal: Optional[str] = None
    alignment_vertical: Optional[str] = None
    border_style: Optional[str] = None
    border_color: Optional[str] = None
    number_format: Optional[str] = None


@dataclass
class CellInfo:
    address: str
    row: int
    col: int
    value: Any
    type: str
    formula: Optional[str] = None
    style: Optional[CellStyle] = None
    is_merged: bool = False
    merged_range: Optional[str] = None


@dataclass
class MergedCellInfo:
    range: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    purpose: str = ""


@dataclass
class ContentZone:
    name: str
    rows: List[int]
    type: str
    description: str = ""


@dataclass
class EditableRange:
    range: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    description: str


class ExcelTemplateAnalyzer:
    TEMPLATE_KEYWORDS = [
        '送货单', '收据', '发票', '订单', '清单', '表', '单', '合计', '汇总',
        '签名', '日期', '编号', '单位', '联系人', '电话', '地址', '备注',
        '产 品', '型 号', '名 称', '数 量', '规 格', '单 价', '金 额'
    ]

    HEADER_INDICATORS = [
        '产品型号', '产品名称', '数量', '规格', '单价', '金额', '备注',
        '型号', '名称', '件数', '千克', '公斤'
    ]

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.file_name = Path(file_path).name
        self.workbook: Optional[openpyxl.Workbook] = None
        self.worksheet: Optional[openpyxl.Workbook] = None
        self.workbook_name: str = ""
        self.max_row: int = 0
        self.max_col: int = 0
        self.merged_cells_list: List[MergedCellInfo] = []
        self.merged_ranges_set: set = set()
        self.cells: Dict[str, CellInfo] = {}
        self.zones: List[ContentZone] = []
        self.editable_ranges: List[EditableRange] = []

    def analyze(self, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        self._load_workbook(sheet_name)
        self._extract_structure()
        self._extract_merged_cells()
        self._extract_cells()
        self._identify_zones()
        self._identify_editable_ranges()

        return self._build_output()

    def _load_workbook(self, sheet_name: Optional[str] = None):
        self.workbook = openpyxl.load_workbook(self.file_path)

        if sheet_name:
            self.worksheet = self.workbook[sheet_name]
        else:
            self.worksheet = self.workbook.active

        self.workbook_name = self.worksheet.title

    def _extract_structure(self):
        self.max_row = self.worksheet.max_row
        self.max_col = self.worksheet.max_column

        for col in range(1, self.max_col + 1):
            col_letter = get_column_letter(col)
            width = self.worksheet.column_dimensions[col_letter].width

        for row in range(1, self.max_row + 1):
            height = self.worksheet.row_dimensions[row].height

    def _extract_merged_cells(self):
        for merged_range in self.worksheet.merged_cells.ranges:
            range_str = str(merged_range)
            self.merged_ranges_set.add(range_str)

            min_row = merged_range.min_row
            max_row = merged_range.max_row
            min_col = merged_range.min_col
            max_col = merged_range.max_col

            purpose = self._guess_merged_purpose(min_row, max_row, min_col, max_col)

            self.merged_cells_list.append(MergedCellInfo(
                range=range_str,
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                purpose=purpose
            ))

    def _guess_merged_purpose(self, min_row: int, max_row: int, min_col: int, max_col: int) -> str:
        cell_value = self.worksheet.cell(min_row, min_col).value

        if cell_value:
            if '送货单' in str(cell_value) or '收据' in str(cell_value):
                return '标题'
            elif '购货单位' in str(cell_value) or '乙方' in str(cell_value):
                return '购货单位信息'
            elif any(kw in str(cell_value) for kw in ['型号', '名称', '数量', '规格', '单价', '金额']):
                return '表头'
            elif '合计' in str(cell_value) or 'SUM' in str(cell_value):
                return '汇总'
            elif '签名' in str(cell_value):
                return '签名区'
            else:
                return '内容区'

        if max_row - min_row > 3:
            return '大标题区'

        return '合并内容'

    def _extract_cells(self):
        for row in range(1, self.max_row + 1):
            for col in range(1, self.max_col + 1):
                cell = self.worksheet.cell(row, col)
                address = cell.coordinate

                merged_range = self._get_merged_range(address)
                is_merged = merged_range is not None

                cell_type = self._classify_cell_type(cell, row, col, is_merged)

                formula = None
                if cell.data_type == 'f':
                    formula = cell.value
                    value = None
                else:
                    value = cell.value

                style = self._extract_style(cell)

                self.cells[address] = CellInfo(
                    address=address,
                    row=row,
                    col=col,
                    value=value,
                    type=cell_type,
                    formula=formula,
                    style=style,
                    is_merged=is_merged,
                    merged_range=merged_range
                )

    def _get_merged_range(self, address: str) -> Optional[str]:
        for merged_range in self.worksheet.merged_cells.ranges:
            if address in merged_range:
                return str(merged_range)
        return None

    def _classify_cell_type(self, cell, row: int, col: int, is_merged: bool) -> str:
        if cell.data_type == 'f':
            return 'formula'

        value = cell.value
        if value is None:
            return 'empty'

        value_str = str(value).strip()

        if row <= 3:
            return 'template'

        if col == 1 or col == 2:
            if value_str and not value_str.startswith('='):
                return 'editable'

        if any(keyword in value_str for keyword in ['合计', '小计', '总计', '签名', '日期：', '单位：']):
            return 'template'

        if re.match(r'^[\d\s\-/.]+$', value_str) and col > 1:
            return 'editable'

        if re.match(r'^[\u4e00-\u9fa5]+$', value_str) and len(value_str) > 2:
            return 'template'

        if col >= 2 and value_str and not value_str.startswith('='):
            return 'editable'

        return 'template'

    def _extract_style(self, cell) -> Optional[CellStyle]:
        style = CellStyle()

        if cell.font:
            style.font_name = cell.font.name
            style.font_size = cell.font.size
            style.font_bold = cell.font.bold
            if cell.font.color and cell.font.color.rgb:
                style.font_color = cell.font.color.rgb

        if cell.fill and cell.fill.patternType:
            style.fill_pattern = cell.fill.patternType
            if cell.fill.fgColor and cell.fill.fgColor.rgb:
                style.fill_fg_color = cell.fill.fgColor.rgb
            if cell.fill.bgColor and cell.fill.bgColor.rgb:
                style.fill_bg_color = cell.fill.bgColor.rgb

        if cell.alignment:
            style.alignment_horizontal = cell.alignment.horizontal
            style.alignment_vertical = cell.alignment.vertical

        if cell.border:
            border_styles = []
            for side in ['left', 'right', 'top', 'bottom']:
                side_obj = getattr(cell.border, side)
                if side_obj and side_obj.style:
                    border_styles.append(f"{side}:{side_obj.style}")
            if border_styles:
                style.border_style = ','.join(border_styles)

        style.number_format = cell.number_format

        return style

    def _identify_zones(self):
        header_rows = []
        data_rows = []
        summary_rows = []

        for row in range(1, self.max_row + 1):
            row_type = self._classify_row_type(row)

            if row_type == 'template':
                header_rows.append(row)
            elif row_type == 'editable':
                data_rows.append(row)
            else:
                summary_rows.append(row)

        if header_rows:
            self.zones.append(ContentZone(
                name='header',
                rows=header_rows,
                type='template',
                description='表头和标题区域'
            ))

        if data_rows:
            self.zones.append(ContentZone(
                name='data',
                rows=data_rows,
                type='editable',
                description='数据输入区域'
            ))

        if summary_rows:
            self.zones.append(ContentZone(
                name='summary',
                rows=summary_rows,
                type='template',
                description='汇总和签名区域'
            ))

    def _classify_row_type(self, row: int) -> str:
        has_formula = False
        has_editable = False
        has_template = False

        for col in range(1, self.max_col + 1):
            cell = self.cells.get(get_column_letter(col) + str(row))
            if cell:
                if cell.type == 'formula':
                    has_formula = True
                elif cell.type == 'editable':
                    has_editable = True
                elif cell.type == 'template':
                    has_template = True

        if row <= 3:
            return 'template'

        if has_editable:
            return 'editable'

        if has_template and not has_editable:
            return 'template'

        return 'template'

    def _identify_editable_ranges(self):
        editable_cells = [
            (cell.row, cell.col)
            for cell in self.cells.values()
            if cell.type == 'editable'
        ]

        if not editable_cells:
            return

        rows = sorted(set([r for r, c in editable_cells]))
        cols = sorted(set([c for r, c in editable_cells]))

        if len(rows) > 5:
            row_groups = self._group_consecutive(rows)
        else:
            row_groups = [rows]

        for row_group in row_groups:
            min_row = min(row_group)
            max_row = max(row_group)

            col_group = [c for c in cols if any(c == cell_col for r, cell_col in editable_cells if r in row_group)]

            if col_group:
                min_col = min(col_group)
                max_col = max(col_group)

                desc = self._describe_range(min_row, max_row, min_col, max_col)

                self.editable_ranges.append(EditableRange(
                    range=f"{get_column_letter(min_row)}{min_row}:{get_column_letter(max_col)}{max_row}",
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    description=desc
                ))

    def _group_consecutive(self, nums: List[int]) -> List[List[int]]:
        if not nums:
            return []

        groups = []
        current_group = [nums[0]]

        for i in range(1, len(nums)):
            if nums[i] - nums[i-1] <= 1:
                current_group.append(nums[i])
            else:
                groups.append(current_group)
                current_group = [nums[i]]

        groups.append(current_group)
        return groups

    def _describe_range(self, min_row: int, max_row: int, min_col: int, max_col: int) -> str:
        min_col_letter = get_column_letter(min_col)
        max_col_letter = get_column_letter(max_col)

        first_cell = self.worksheet.cell(min_row, min_col).value
        last_cell = self.worksheet.cell(min_row, max_col).value

        if first_cell:
            return f"{min_col_letter}{min_row}-{max_col_letter}{max_row} ({first_cell}...)"

        return f"{min_col_letter}{min_row}-{max_col_letter}{max_row}"

    def _build_output(self) -> Dict[str, Any]:
        cells_output = {}
        for address, cell in self.cells.items():
            cell_dict = {
                'address': cell.address,
                'row': cell.row,
                'col': cell.col,
                'value': cell.value,
                'type': cell.type,
                'is_merged': cell.is_merged,
                'merged_range': cell.merged_range
            }

            if cell.formula:
                cell_dict['formula'] = cell.formula

            if cell.style:
                style_dict = {}
                for key, value in asdict(cell.style).items():
                    if value is not None:
                        style_dict[key] = value
                cell_dict['style'] = style_dict

            cells_output[address] = cell_dict

        output = {
            'file': self.file_name,
            'sheet': self.workbook_name,
            'structure': {
                'max_row': self.max_row,
                'max_col': self.max_col,
                'max_col_letter': get_column_letter(self.max_col)
            },
            'zones': [asdict(zone) for zone in self.zones],
            'merged_cells': [asdict(mc) for mc in self.merged_cells_list],
            'editable_ranges': [asdict(er) for er in self.editable_ranges],
            'cells': cells_output
        }

        return output

    def save_json(self, output_path: str, pretty: bool = True):
        result = self.analyze()

        with open(output_path, 'w', encoding='utf-8') as f:
            if pretty:
                json.dump(result, f, ensure_ascii=False, indent=2)
            else:
                json.dump(result, f, ensure_ascii=False)

        return result


def analyze_template(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """分析 Excel 模板并返回结构信息"""
    analyzer = ExcelTemplateAnalyzer(file_path)
    return analyzer.analyze(sheet_name)


def analyze_to_json(file_path: str, output_path: str, sheet_name: Optional[str] = None):
    """分析 Excel 模板并保存为 JSON 文件"""
    analyzer = ExcelTemplateAnalyzer(file_path)
    return analyzer.save_json(output_path)


def extract_entries(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """提取 Excel 模板中的可编辑词条"""
    result = analyze_template(file_path, sheet_name)

    editable_entries = []
    for cell in result.get('cells', {}).values():
        if cell.get('type') == 'editable':
            editable_entries.append({
                'address': cell.get('address'),
                'row': cell.get('row'),
                'col': cell.get('col'),
                'value': cell.get('value'),
            })

    return {
        'file': result.get('file'),
        'sheet': result.get('sheet'),
        'editable_entries': editable_entries,
        'zones': result.get('zones', []),
    }


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Excel 模板分析工具')
    parser.add_argument('-f', '--file', required=True, help='Excel 文件路径')
    parser.add_argument('-o', '--output', help='输出 JSON 文件路径')
    parser.add_argument('-s', '--sheet', help='工作表名称（默认第一个）')
    parser.add_argument('-v', '--verbose', action='store_true', help='详细输出模式')

    args = parser.parse_args()

    result = analyze_template(args.file, args.sheet)

    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"分析结果已保存到: {args.output}")
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))
