from __future__ import annotations

from datetime import date, datetime
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import load_workbook


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().replace(" ", "").lower()


def _to_header_lookup(field_alias_map: Dict[str, Any]) -> List[Tuple[str, str, str]]:
    """
    Build a normalized lookup list:
    (normalized_header_text, display_header_text, record_key)
    """
    lookup: List[Tuple[str, str, str]] = []
    for key, value in (field_alias_map or {}).items():
        record_key = str(key).strip()
        if not record_key:
            continue

        aliases: Iterable[str]
        if isinstance(value, (list, tuple, set)):
            aliases = [str(v).strip() for v in value if str(v).strip()]
        elif isinstance(value, str):
            aliases = [value.strip()] if value.strip() else []
        else:
            aliases = []

        if not aliases:
            aliases = [record_key]

        for alias in aliases:
            normalized = _normalize_header(alias)
            if normalized:
                lookup.append((normalized, alias, record_key))
    return lookup


def _format_cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return value


def fill_workbook_from_template(
    template_path: str,
    records: List[Dict[str, Any]],
    field_alias_map: Dict[str, Any],
    sheet_name: str | None = None,
    max_row_scan: int = 30,
    max_col_scan: int = 40,
    append_missing_field_columns: bool = False,
    clear_existing_data_rows_all_columns: bool = False,
    truncate_rows_after_data_area: bool = False,
    clear_rows_above_header: bool = False,
):
    """
    Fill template workbook by scanning header row and mapping aliases to record keys.

    If ``append_missing_field_columns`` is True, any ``field_alias_map`` key that has
    no matching column in the template gets a new column appended (so e.g. 规格 is not
    dropped when the template only has 型号/名称/价格).

    If ``clear_rows_above_header`` is True and a header row was found by scanning (not
    the synthetic fallback), rows above that header are removed so preamble/sample
    text does not appear in exported files.
    """
    wb = load_workbook(template_path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb.active

    alias_lookup = _to_header_lookup(field_alias_map)
    alias_to_key = {normalized: key for normalized, _, key in alias_lookup}

    header_row = None
    header_cols: List[Tuple[int, str]] = []
    scanned_max_row = min(ws.max_row or 1, max_row_scan)
    scanned_max_col = min(ws.max_column or 1, max_col_scan)

    for row_idx in range(1, scanned_max_row + 1):
        cols: List[Tuple[int, str]] = []
        for col_idx in range(1, scanned_max_col + 1):
            raw_text = ws.cell(row_idx, col_idx).value
            normalized = _normalize_header(raw_text)
            if normalized in alias_to_key:
                cols.append((col_idx, normalized))
        if len(cols) >= 3:
            header_row = row_idx
            header_cols = cols
            break

    found_natural_header = header_row is not None

    if not header_row:
        header_row = 1
        seen_keys = set()
        ordered_headers: List[Tuple[str, str]] = []
        for _, display_header, record_key in alias_lookup:
            if record_key in seen_keys:
                continue
            seen_keys.add(record_key)
            ordered_headers.append((display_header, record_key))
        header_cols = [(index + 1, _normalize_header(header)) for index, (header, _) in enumerate(ordered_headers)]
        for index, (header, _) in enumerate(ordered_headers, start=1):
            ws.cell(header_row, index, header)

    # Remove title / instruction rows above the detected header so exports do not
    # carry template preamble (sample text, merged titles, etc.).
    if clear_rows_above_header and found_natural_header and header_row > 1:
        try:
            ws.delete_rows(1, header_row - 1)
            header_row = 1
        except Exception:
            for clear_r in range(1, header_row):
                for col_idx in range(1, (ws.max_column or 1) + 1):
                    ws.cell(clear_r, col_idx).value = None

    # Clear old sample/business data before writing new rows.
    # Optional mode clears every column under header row to avoid residual values
    # leaking from template history.
    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        if clear_existing_data_rows_all_columns:
            for col_idx in range(1, (ws.max_column or 1) + 1):
                ws.cell(row_idx, col_idx, None)
            continue
        for col_idx, _ in header_cols:
            ws.cell(row_idx, col_idx, None)

    for row_idx, record in enumerate(records or [], start=header_row + 1):
        for col_idx, normalized_header in header_cols:
            record_key = alias_to_key.get(normalized_header)
            value = record.get(record_key) if record_key else ""
            ws.cell(row_idx, col_idx, _format_cell_value(value) if value is not None else "")

    if append_missing_field_columns and field_alias_map:
        mapped_keys = {alias_to_key.get(norm) for _, norm in header_cols}
        mapped_keys.discard(None)
        max_col = max((c for c, _ in header_cols), default=0)
        for record_key, aliases in field_alias_map.items():
            if record_key in mapped_keys:
                continue
            if isinstance(aliases, (list, tuple, set)):
                display = next(
                    (str(a).strip() for a in aliases if str(a).strip()),
                    str(record_key),
                )
            else:
                display = str(aliases or "").strip() or str(record_key)
            max_col += 1
            ws.cell(header_row, max_col, display)
            for row_idx, record in enumerate(records or [], start=header_row + 1):
                value = record.get(record_key) if record else ""
                ws.cell(
                    row_idx,
                    max_col,
                    _format_cell_value(value) if value is not None else "",
                )

    # Strict mode: keep only fixed area above header + header row + data rows.
    # This removes any residual template tail area.
    if truncate_rows_after_data_area:
        data_end_row = header_row + max(len(records or []), 0)
        if data_end_row < (ws.max_row or data_end_row):
            ws.delete_rows(data_end_row + 1, (ws.max_row or data_end_row) - data_end_row)

    return wb
