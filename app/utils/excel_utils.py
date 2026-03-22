"""
Excel 处理工具函数模块

提供 Excel 文件读写、合并单元格处理等工具函数。
"""

from typing import Any, Iterable, Tuple


def get_header_indices(header: Iterable[str]) -> Tuple[int | None, int | None, int | None, int | None]:
    """
    获取列索引
    
    Args:
        header: 表头列表
        
    Returns:
        (col_unit, col_contact, col_phone, col_addr) 元组
    """
    col_unit = col_contact = col_phone = col_addr = None
    
    for i, h in enumerate(header):
        h = h.strip()
        if h == "单位名称":
            col_unit = i + 1
        elif h == "购买单位" and not col_unit:
            col_unit = i + 1
        elif h == "联系人":
            col_contact = i + 1
        elif h == "联系电话":
            col_phone = i + 1
        elif h == "地址":
            col_addr = i + 1
    
    return col_unit, col_contact, col_phone, col_addr


def merged_cell_value(ws, row: int, col: int) -> Any:
    """
    处理合并单元格，返回合并区域左上角的值
    
    Args:
        ws: openpyxl worksheet 对象
        row: 行号（从 1 开始）
        col: 列号（从 1 开始）
        
    Returns:
        单元格的值
    """
    try:
        merged = getattr(ws, "merged_cells", None) or getattr(ws, "merged_cell_ranges", None)
        if merged is None:
            return ws.cell(row=row, column=col).value
        
        ranges = getattr(merged, "ranges", None)
        if ranges is None:
            ranges = merged if isinstance(merged, (list, tuple)) else []
        
        for rng in (ranges or []):
            try:
                if hasattr(rng, "min_row"):
                    if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                        return ws.cell(row=rng.min_row, column=rng.min_col).value
                elif isinstance(rng, str):
                    try:
                        from openpyxl.utils.cell import range_boundaries
                    except ImportError:
                        from openpyxl.utils import range_boundaries
                    
                    min_col, min_row, max_col, max_row = range_boundaries(rng)
                    if min_row <= row <= max_row and min_col <= col <= max_col:
                        return ws.cell(row=min_row, column=min_col).value
            except Exception:
                continue
    except Exception:
        pass
    
    return ws.cell(row=row, column=col).value


def cell_str(ws, row: int, col: int | None) -> str:
    """
    获取单元格的字符串值
    
    Args:
        ws: openpyxl worksheet 对象
        row: 行号（从 1 开始）
        col: 列号（从 1 开始）
        
    Returns:
        单元格的字符串值
    """
    if col is None:
        return ""
    
    v = merged_cell_value(ws, row, col)
    if v is None:
        return ""
    
    s = str(v).strip()
    if isinstance(v, (int, float)) and not s:
        return str(int(v)) if isinstance(v, float) and v == int(v) else str(v)
    
    return s


def normalize_unit_name(raw) -> str:
    """
    标准化单位名称（去除空白字符，包括全角空格）
    
    Args:
        raw: 原始字符串
        
    Returns:
        标准化后的字符串
    """
    if raw is None or not isinstance(raw, str):
        raw = str(raw) if raw is not None else ""
    
    # 替换全角空格为普通空格，然后去除所有空白
    s = raw.replace("\u3000", " ").strip()
    return " ".join(s.split()) if s else ""
