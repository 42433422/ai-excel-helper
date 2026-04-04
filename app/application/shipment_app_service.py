import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from app.application.ports import (
    PurchaseUnitQueryPort,
    ShipmentDocumentGeneratorPort,
    ShipmentRecordCommandPort,
    ShipmentRecordQueryPort,
    ShipmentRecordStorePort,
    ShipmentRepository,
)
from app.domain.shipment.aggregates import Shipment, ShipmentItem
from app.domain.value_objects import ContactInfo, Money, OrderNumber, Quantity

logger = logging.getLogger(__name__)


class ShipmentApplicationService:
    """发货单应用服务 - 用例编排"""
    
    def __init__(
        self,
        repository: ShipmentRepository,
        document_generator: ShipmentDocumentGeneratorPort | None = None,
        record_store: ShipmentRecordStorePort | None = None,
        record_query: ShipmentRecordQueryPort | None = None,
        record_command: ShipmentRecordCommandPort | None = None,
        purchase_unit_query: PurchaseUnitQueryPort | None = None,
    ):
        self._repository = repository
        self._document_generator = document_generator
        self._record_store = record_store
        self._record_query = record_query
        self._record_command = record_command
        self._purchase_unit_query = purchase_unit_query
    
    def create_shipment(
        self,
        unit_name: str,
        items_data: List[Dict[str, Any]],
        contact_person: str = "",
        contact_phone: str = "",
    ) -> Dict[str, Any]:
        """
        创建发货单用例
        """
        try:
            contact_info = ContactInfo(person=contact_person, phone=contact_phone)
            shipment = Shipment.create(unit_name=unit_name, contact_info=contact_info)
            
            for item_data in items_data:
                try:
                    item = ShipmentItem.from_dict(item_data)
                    shipment.add_item(item)
                except ValueError as e:
                    logger.warning(f"跳过无效产品: {e}")
                    continue
            
            if not shipment.is_valid():
                return {"success": False, "message": "发货单无效：缺少购买单位或产品"}
            
            saved_shipment = self._repository.save(shipment)

            try:
                from app.infrastructure.mods.hooks import trigger
                trigger("shipment.created", shipment=saved_shipment)
            except Exception as hook_err:
                logger.warning(f"Hook trigger failed: {hook_err}")

            return {
                "success": True,
                "message": "发货单创建成功",
                "shipment": saved_shipment.to_dict(),
            }
            
        except Exception as e:
            logger.exception(f"创建发货单失败: {e}")
            return {"success": False, "message": f"创建失败: {str(e)}"}
    
    def get_shipment(self, shipment_id: int) -> Optional[Shipment]:
        """获取发货单"""
        return self._repository.find_by_id(shipment_id)
    
    def list_shipments(
        self,
        unit_name: Optional[str] = None,
        page: int = 1,
        per_page: int = 20,
    ) -> Dict[str, Any]:
        """查询发货单列表"""
        try:
            shipments = self._repository.find_all(page=page, per_page=per_page)
            
            if unit_name:
                shipments = self._repository.find_by_unit(unit_name)
            
            total = self._repository.count()
            
            return {
                "success": True,
                "data": [s.to_dict() for s in shipments],
                "total": total,
                "page": page,
                "per_page": per_page,
            }
            
        except Exception as e:
            logger.exception(f"查询发货单失败: {e}")
            return {"success": False, "message": str(e), "data": []}

    def query_shipment_orders(
        self,
        unit_name: Optional[str] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        page: int = 1,
        per_page: int = 20,
    ) -> Dict[str, Any]:
        """
        出货记录列表查询（read side），保持与旧接口返回结构一致。
        """
        if not self._record_query:
            return {
                "success": False,
                "message": "record_query 未配置",
                "data": [],
                "total": 0,
                "page": page,
                "per_page": per_page,
            }

        return self._record_query.query_shipments(
            unit_name=unit_name,
            start_date=start_date,
            end_date=end_date,
            page=page,
            per_page=per_page,
        )

    def search_orders(self, query: str) -> List[Dict[str, Any]]:
        """搜索出货记录（read side）。"""
        if not self._record_query:
            return []
        return self._record_query.search_shipments(query)

    def get_order(self, order_number: str) -> Optional[Dict[str, Any]]:
        """根据 id 查询出货记录（用于 GET /orders/<order_number>）。"""
        if not self._record_query:
            return None
        return self._record_query.get_shipment_by_id(order_number)

    def get_orders(self, limit: int = 10) -> List[Dict[str, Any]]:
        """获取最近创建的出货记录（用于 GET /orders/latest）。"""
        if not self._record_query:
            return []
        return self._record_query.get_latest_shipments(limit)

    def get_purchase_units(self) -> List[str]:
        """获取所有购买单位列表（用于 /orders/purchase-units）。"""
        if not self._purchase_unit_query:
            return []
        return self._purchase_unit_query.list_purchase_units()

    def clear_shipment_by_unit(self, purchase_unit: str) -> Dict[str, Any]:
        """清理指定购买单位的出货记录。"""
        if not self._record_command:
            return {"success": False, "message": "record_command 未配置"}
        return self._record_command.clear_by_unit(purchase_unit)

    def clear_all_orders(self) -> Dict[str, Any]:
        """清空所有出货记录。"""
        if not self._record_command:
            return {"success": False, "message": "record_command 未配置"}
        return self._record_command.clear_all()

    def get_shipment_records(self, unit_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """后台：获取出货记录列表（/shipment-records/records）。"""
        if not self._record_query:
            return []
        return self._record_query.get_shipment_records(unit_name, limit=100)

    def update_shipment_record(
        self,
        record_id: int,
        *,
        unit_name: Optional[str] = None,
        products: Optional[List[Dict[str, Any]]] = None,
        date: Optional[str] = None,
        **kwargs,
    ) -> Dict[str, Any]:
        """
        后台：更新出货记录。
        兼容旧接口：products 参数保留但当前不会用于修改 parsed/products 字段（沿用旧实现的行为）。
        """
        if not self._record_command:
            return {"success": False, "message": "record_command 未配置"}

        # 将旧实现里的 kwargs 全量传给 record 字段（排除 products/date/unit_name）
        fields = dict(kwargs)
        return self._record_command.update_record(
            record_id,
            unit_name=unit_name,
            date=date,
            fields=fields,
        )

    def delete_shipment_record(self, record_id: int) -> Dict[str, Any]:
        """后台：删除出货记录（/shipment-records/record DELETE）。"""
        if not self._record_command:
            return {"success": False, "message": "record_command 未配置"}
        return self._record_command.delete_record(record_id)

    def export_shipment_records(
        self,
        unit_name: Optional[str] = None,
        template_id: Optional[str] = None,
        status_filter: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        导出出货记录为 Excel 文件。
        说明：该功能偏 I/O/报表层，仍放在 application 层做编排。
        """
        try:
            from openpyxl import Workbook

            from app.utils.path_utils import get_data_dir
            from app.utils.template_export_utils import fill_workbook_from_template

            records = self.get_shipment_records(unit_name)
            normalized_status = str(status_filter or "").strip().lower()
            if normalized_status:
                if normalized_status in ("printed", "已打印"):
                    records = [r for r in records if str(r.get("status") or "").strip().lower() == "printed"]
                elif normalized_status in ("pending", "未打印"):
                    records = [r for r in records if str(r.get("status") or "").strip().lower() in ("pending", "")]

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            unit_prefix = unit_name if unit_name else "all"
            filename = f"shipment_records_{unit_prefix}_{timestamp}.xlsx"

            export_dir = os.path.join(get_data_dir(), "exports")
            os.makedirs(export_dir, exist_ok=True)
            file_path = os.path.join(export_dir, filename)
            template_path = None
            if template_id:
                try:
                    from app.application import get_template_app_service
                    templates = (get_template_app_service().get_templates() or {}).get("templates") or []
                    target = next((t for t in templates if str(t.get("id")) == str(template_id)), None)
                    if not target:
                        return {
                            "success": False,
                            "message": "导出失败：所选模板不存在，请重新选择模板",
                            "file_path": None,
                            "filename": None,
                            "count": 0,
                        }
                    target_scope = str(target.get("business_scope") or "").strip()
                    target_type = str(target.get("template_type") or "").strip()
                    if target_scope and target_scope != "shipmentRecords" and target_type != "出货记录":
                        return {
                            "success": False,
                            "message": "导出失败：所选模板不属于出货记录范围，请重新选择",
                            "file_path": None,
                            "filename": None,
                            "count": 0,
                        }
                    candidate_path = str(target.get("path") or target.get("file_path") or "").strip()
                    if not candidate_path:
                        return {
                            "success": False,
                            "message": "导出失败：所选模板未绑定 Excel 文件，请先在业务对接中上传并替换",
                            "file_path": None,
                            "filename": None,
                            "count": 0,
                        }
                    if not os.path.exists(candidate_path):
                        return {
                            "success": False,
                            "message": "导出失败：所选模板文件不存在，请重新上传模板后重试",
                            "file_path": None,
                            "filename": None,
                            "count": 0,
                        }
                    template_path = candidate_path
                    preview_data = target.get("preview_data") if isinstance(target.get("preview_data"), dict) else {}
                    business_rules = target.get("business_rules") if isinstance(target.get("business_rules"), dict) else {}
                    template_sheet_name = str(
                        preview_data.get("selected_sheet_name")
                        or business_rules.get("selected_sheet_name")
                        or preview_data.get("sheet_name")
                        or ""
                    ).strip()
                except Exception as e:
                    return {
                        "success": False,
                        "message": f"导出失败：读取模板信息异常（{str(e)}）",
                        "file_path": None,
                        "filename": None,
                        "count": 0,
                    }

            if template_path:
                header_alias = {
                    "purchase_unit": ["购买单位", "单位"],
                    "product_name": ["产品名称", "品名"],
                    "model_number": ["型号", "产品型号"],
                    "quantity_kg": ["数量", "数量/KG", "数量(kg)"],
                    "quantity_tins": ["数量/件", "数量/桶"],
                    "tin_spec": ["规格"],
                    "unit_price": ["单价", "单价/元"],
                    "amount": ["金额", "金额/元"],
                    "status": ["状态"],
                    "created_at": ["创建时间"],
                    "printed_at": ["打印时间"],
                    "printer_name": ["打印机"],
                }
                wb = fill_workbook_from_template(
                    template_path=template_path,
                    records=records,
                    field_alias_map=header_alias,
                    sheet_name=template_sheet_name or "出货记录",
                    clear_existing_data_rows_all_columns=True,
                    truncate_rows_after_data_area=True,
                    clear_rows_above_header=True,
                )
                # Export is delivered as shipment records only. Remove other worksheets
                # to avoid carrying historical/template auxiliary data into result file.
                used_sheet = (template_sheet_name or "出货记录").strip()
                if used_sheet not in wb.sheetnames and wb.sheetnames:
                    used_sheet = wb.sheetnames[0]
                for sheet in list(wb.sheetnames):
                    if sheet != used_sheet and len(wb.sheetnames) > 1:
                        wb.remove(wb[sheet])
                wb.save(file_path)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "出货记录"

                headers = [
                    "ID",
                    "购买单位",
                    "产品名称",
                    "型号",
                    "数量 (KG)",
                    "数量 (桶)",
                    "规格",
                    "单价",
                    "金额",
                    "状态",
                    "创建时间",
                    "打印时间",
                    "打印机",
                ]
                ws.append(headers)

                for r in records:
                    ws.append(
                        [
                            r.get("id"),
                            r.get("purchase_unit") or "",
                            r.get("product_name") or "",
                            r.get("model_number") or "",
                            r.get("quantity_kg") or 0,
                            r.get("quantity_tins") or 0,
                            r.get("tin_spec") or "",
                            r.get("unit_price") or 0,
                            r.get("amount") or 0,
                            r.get("status") or "",
                            r.get("created_at").strftime("%Y-%m-%d %H:%M:%S")
                            if r.get("created_at")
                            else "",
                            r.get("printed_at").strftime("%Y-%m-%d %H:%M:%S")
                            if r.get("printed_at")
                            else "",
                            r.get("printer_name") or "",
                        ]
                    )
                wb.save(file_path)

            return {
                "success": True,
                "file_path": str(file_path),
                "filename": filename,
                "count": len(records),
                "template_used": template_path or "",
            }
        except Exception as e:
            return {"success": False, "message": f"导出失败：{str(e)}", "file_path": None, "filename": None, "count": 0}

    def set_order_sequence(self, sequence: int) -> Dict[str, Any]:
        """设置订单序号（兼容旧接口，无状态实现）。"""
        return {"success": True, "message": "序号已设置", "sequence": int(sequence)}

    def reset_order_sequence(self) -> Dict[str, Any]:
        """重置订单序号（兼容旧接口，无状态实现）。"""
        return {"success": True, "message": "序号已重置", "sequence": 1}

    def download_shipment_order(self, filename: str) -> Dict[str, Any]:
        """检查发货单文件是否存在（兼容旧接口）。"""
        from app.utils.path_utils import get_app_data_dir
        output_dir = os.path.join(get_app_data_dir(), "shipment_outputs")
        file_path = os.path.join(output_dir, filename)
        if not os.path.exists(file_path):
            return {"success": False, "message": f"文件不存在：{filename}", "file_path": None}
        return {"success": True, "file_path": file_path, "message": "文件存在"}

    def mark_as_printed(self, shipment_id: int, printer_name: str = "") -> Dict[str, Any]:
        """标记发货单为已打印"""
        try:
            shipment = self._repository.find_by_id(shipment_id)
            if not shipment:
                return {"success": False, "message": "发货单不存在"}
            
            shipment.mark_as_printed(printer_name)
            self._repository.save(shipment)
            
            return {
                "success": True,
                "message": "已标记为已打印",
                "printed_at": datetime.now().isoformat(),
            }
            
        except Exception as e:
            logger.exception(f"标记打印失败: {e}")
            return {"success": False, "message": str(e)}
    
    def cancel_shipment(self, shipment_id: int) -> Dict[str, Any]:
        """取消发货单"""
        try:
            shipment = self._repository.find_by_id(shipment_id)
            if not shipment:
                return {"success": False, "message": "发货单不存在"}
            
            shipment.cancel()
            self._repository.save(shipment)
            
            return {"success": True, "message": "发货单已取消"}
            
        except Exception as e:
            logger.exception(f"取消发货单失败: {e}")
            return {"success": False, "message": str(e)}
    
    def delete_shipment(self, shipment_id: int) -> Dict[str, Any]:
        """删除发货单"""
        try:
            success = self._repository.delete(shipment_id)
            if success:
                return {"success": True, "message": "发货单已删除"}
            return {"success": False, "message": "发货单不存在"}
            
        except Exception as e:
            logger.exception(f"删除发货单失败: {e}")
            return {"success": False, "message": str(e)}
    
    def calculate_totals(self, items_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """计算发货单汇总"""
        total_amount = 0.0
        total_tins = 0
        total_kg = 0.0
        
        for item_data in items_data:
            tins = item_data.get("quantity_tins", 0)
            spec = item_data.get("tin_spec", 10.0)
            kg = tins * spec
            price = item_data.get("unit_price", 0)
            
            total_tins += tins
            total_kg += kg
            total_amount += price * kg
        
        return {
            "total_tins": total_tins,
            "total_kg": total_kg,
            "total_amount": total_amount,
        }

    def generate_shipment_document(
        self,
        *,
        unit_name: str,
        products: List[Dict[str, Any]],
        date: Optional[str] = None,
        template_name: Optional[str] = None,
        order_number: Optional[str] = None,
    ) -> Dict[str, Any]:
        """生成发货单文档（用例编排）。"""
        if not self._document_generator:
            return {"success": False, "message": "document_generator 未配置", "doc_name": None, "file_path": None}
        result = self._document_generator.generate(
            unit_name=unit_name,
            products=products,
            date=date,
            template_name=template_name,
            order_number=order_number,
        )
        if result.get("success") and self._record_store:
            try:
                record_products = result.get("parsed_products") or products
                record_result = self._record_store.record_document_generation(
                    unit_name=result.get("purchase_unit") or unit_name,
                    unit_id=result.get("unit_id"),
                    products=record_products,
                    document_result=result,
                    raw_text="",
                )
                record_id = (record_result or {}).get("record_id")
                if record_id:
                    # 向前兼容：历史前端把 order_id 当作 shipment_records 主键使用。
                    result["record_id"] = record_id
                    result["order_id"] = record_id
            except Exception:
                # 记录写入失败不影响文档生成返回
                pass
        return result


_shipment_app_service_instance = None


def get_shipment_application_service() -> ShipmentApplicationService:
    """获取发货服务单例"""
    global _shipment_app_service_instance
    if _shipment_app_service_instance is None:
        _shipment_app_service_instance = ShipmentApplicationService()
    return _shipment_app_service_instance
