import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.domain.customer.entities import PurchaseUnit

logger = logging.getLogger(__name__)

_customers_engine = None
_CustomersSessionLocal = None


def _get_customers_engine():
    global _customers_engine, _CustomersSessionLocal
    if _customers_engine is None:
        # 统一使用 products.db（已合并 purchase_units 表）
        from app.db.init_db import get_db_path

        database_url = (
            os.environ.get("CUSTOMERS_DATABASE_URL")
            or os.environ.get("CUSTOMERS_DB_URL")
            or f"sqlite:///{get_db_path('products.db')}"
        )
        _customers_engine = create_engine(
            database_url,
            pool_pre_ping=True,
            echo=False,
        )
        _CustomersSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=_customers_engine)
    return _customers_engine


def get_customers_session():
    """获取 customers.db 的 session"""
    _get_customers_engine()
    return _CustomersSessionLocal()


class CustomerApplicationService:
    """客户应用服务 - 用例编排

    统一管理客户和购买单位数据，使用 products.db 数据库（已合并 purchase_units 表）。
    """

    def __init__(self):
        self._engine = _get_customers_engine()
        self._SessionLocal = _CustomersSessionLocal

    def _get_session(self):
        return self._SessionLocal()

    def get_all(
        self,
        keyword: Optional[str] = None,
        page: int = 1,
        per_page: int = 20
    ) -> Dict[str, Any]:
        """获取所有购买单位（分页）"""
        try:
            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                query = session.query(PurchaseUnitModel).filter(PurchaseUnitModel.is_active == True)

                if keyword:
                    pattern = f"%{keyword}%"
                    query = query.filter(PurchaseUnitModel.unit_name.like(pattern))

                total = query.count()
                units = query.order_by(PurchaseUnitModel.unit_name).offset((page - 1) * per_page).limit(per_page).all()

                return {
                    "success": True,
                    "data": [
                        {
                            "id": unit.id,
                            "customer_name": unit.unit_name,
                            "contact_person": unit.contact_person or "",
                            "contact_phone": unit.contact_phone or "",
                            "contact_address": unit.address or "",
                            "created_at": unit.created_at.isoformat() if unit.created_at else None,
                            "updated_at": unit.updated_at.isoformat() if unit.updated_at else None
                        }
                        for unit in units
                    ],
                    "total": total,
                    "page": page,
                    "per_page": per_page
                }
            finally:
                session.close()

        except Exception as e:
            logger.exception(f"获取客户列表失败: {e}")
            return {"success": False, "message": str(e), "data": [], "total": 0}

    def get_by_id(self, customer_id: int) -> Dict[str, Any]:
        """根据 ID 获取单个购买单位"""
        try:
            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                unit = session.query(PurchaseUnitModel).filter(PurchaseUnitModel.id == customer_id).first()

                if not unit:
                    return {"success": False, "message": "客户不存在", "data": None}

                return {
                    "success": True,
                    "data": {
                        "id": unit.id,
                        "customer_name": unit.unit_name,
                        "contact_person": unit.contact_person or "",
                        "contact_phone": unit.contact_phone or "",
                        "contact_address": unit.address or "",
                        "created_at": unit.created_at.isoformat() if unit.created_at else None,
                        "updated_at": unit.updated_at.isoformat() if unit.updated_at else None
                    }
                }
            finally:
                session.close()

        except Exception as e:
            logger.exception(f"获取客户失败: {e}")
            return {"success": False, "message": str(e), "data": None}

    def create(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """创建购买单位"""
        try:
            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                customer_name = data.get("customer_name")
                if not customer_name:
                    return {"success": False, "message": "客户名称不能为空"}

                existing = session.query(PurchaseUnitModel).filter(
                    PurchaseUnitModel.unit_name == customer_name
                ).first()

                if existing:
                    return {"success": False, "message": "客户名称已存在"}

                unit = PurchaseUnitModel(
                    unit_name=customer_name,
                    contact_person=data.get("contact_person", ""),
                    contact_phone=data.get("contact_phone", ""),
                    address=data.get("contact_address", "")
                )

                session.add(unit)
                session.commit()
                session.refresh(unit)

                return {
                    "success": True,
                    "message": "客户创建成功",
                    "data": {
                        "id": unit.id,
                        "customer_name": unit.unit_name,
                        "contact_person": unit.contact_person or "",
                        "contact_phone": unit.contact_phone or "",
                        "contact_address": unit.address or "",
                        "created_at": unit.created_at.isoformat() if unit.created_at else None,
                        "updated_at": unit.updated_at.isoformat() if unit.updated_at else None
                    }
                }
            finally:
                session.close()

        except Exception as e:
            logger.exception(f"创建客户失败: {e}")
            return {"success": False, "message": str(e)}

    def update(self, customer_id: int, data: Dict[str, Any]) -> Dict[str, Any]:
        """更新购买单位"""
        try:
            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                unit = session.query(PurchaseUnitModel).filter(PurchaseUnitModel.id == customer_id).first()

                if not unit:
                    return {"success": False, "message": "客户不存在"}

                if "customer_name" in data:
                    existing = session.query(PurchaseUnitModel).filter(
                        PurchaseUnitModel.unit_name == data["customer_name"],
                        PurchaseUnitModel.id != customer_id
                    ).first()
                    if existing:
                        return {"success": False, "message": "客户名称已存在"}
                    unit.unit_name = data["customer_name"]

                if "contact_person" in data:
                    unit.contact_person = data["contact_person"]
                if "contact_phone" in data:
                    unit.contact_phone = data["contact_phone"]
                if "contact_address" in data:
                    unit.address = data["contact_address"]

                session.commit()
                session.refresh(unit)

                return {
                    "success": True,
                    "message": "客户更新成功",
                    "data": {
                        "id": unit.id,
                        "customer_name": unit.unit_name,
                        "contact_person": unit.contact_person or "",
                        "contact_phone": unit.contact_phone or "",
                        "contact_address": unit.address or "",
                        "created_at": unit.created_at.isoformat() if unit.created_at else None,
                        "updated_at": unit.updated_at.isoformat() if unit.updated_at else None
                    }
                }
            finally:
                session.close()

        except Exception as e:
            logger.exception(f"更新客户失败: {e}")
            return {"success": False, "message": str(e)}

    def _check_shipment_associations(self, unit_name: str) -> Dict[str, Any]:
        """检查购买单位是否有关联的发货记录

        Returns:
            {
                "has_associations": bool,
                "shipment_count": int,
                "sample_records": list - 最近3条发货记录示例
            }
        """
        try:
            from app.db.session import get_db
            from app.db.models.shipment import ShipmentRecord

            with get_db() as db:
                records = (
                    db.query(ShipmentRecord)
                    .filter(ShipmentRecord.purchase_unit == unit_name)
                    .order_by(ShipmentRecord.created_at.desc())
                    .limit(3)
                    .all()
                )

                total_count = (
                    db.query(ShipmentRecord)
                    .filter(ShipmentRecord.purchase_unit == unit_name)
                    .count()
                )

                sample_records = []
                for r in records:
                    sample_records.append({
                        "id": r.id,
                        "product_name": r.product_name,
                        "quantity_kg": r.quantity_kg,
                        "created_at": r.created_at.isoformat() if r.created_at else None
                    })

                return {
                    "has_associations": total_count > 0,
                    "shipment_count": total_count,
                    "sample_records": sample_records
                }
        except Exception as e:
            logger.warning(f"检查发货记录关联失败: {e}")
            return {
                "has_associations": False,
                "shipment_count": 0,
                "sample_records": [],
                "error": str(e)
            }

    def delete(self, customer_id: int, force: bool = False) -> Dict[str, Any]:
        """删除购买单位

        Args:
            customer_id: 客户ID
            force: 是否强制删除（忽略关联检查）

        Returns:
            删除结果，包含关联检查信息
        """
        try:
            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                unit = session.query(PurchaseUnitModel).filter(PurchaseUnitModel.id == customer_id).first()

                if not unit:
                    return {"success": False, "message": "客户不存在", "deleted_count": 0}

                unit_name = unit.unit_name

                association_check = self._check_shipment_associations(unit_name)

                if association_check.get("has_associations") and not force:
                    return {
                        "success": False,
                        "message": f"无法删除客户「{unit_name}」，存在 {association_check['shipment_count']} 条关联发货记录",
                        "deleted_count": 0,
                        "has_associations": True,
                        "association_details": {
                            "shipment_count": association_check["shipment_count"],
                            "sample_records": association_check["sample_records"]
                        },
                        "suggestion": "请先删除关联的发货记录，或使用 force=True 强制删除"
                    }

                session.delete(unit)
                session.commit()

                return {
                    "success": True,
                    "message": "客户删除成功",
                    "deleted_count": 1,
                    "has_associations": False
                }
            finally:
                session.close()

        except Exception as e:
            logger.exception(f"删除客户失败: {e}")
            return {"success": False, "message": str(e), "deleted_count": 0}

    def batch_delete(self, ids: List[int], force: bool = False) -> Dict[str, Any]:
        """批量删除购买单位

        Args:
            ids: 客户ID列表
            force: 是否强制删除（忽略关联检查）

        Returns:
            删除结果，包含关联检查信息
        """
        try:
            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                units = session.query(PurchaseUnitModel).filter(PurchaseUnitModel.id.in_(ids)).all()

                if not units:
                    return {"success": False, "message": "未找到要删除的客户", "deleted_count": 0}

                if not force:
                    affected_units = []
                    for unit in units:
                        check = self._check_shipment_associations(unit.unit_name)
                        if check.get("has_associations"):
                            affected_units.append({
                                "id": unit.id,
                                "unit_name": unit.unit_name,
                                "shipment_count": check["shipment_count"],
                                "sample_records": check["sample_records"]
                            })

                    if affected_units:
                        return {
                            "success": False,
                            "message": f"存在 {len(affected_units)} 个客户关联发货记录，无法批量删除",
                            "deleted_count": 0,
                            "has_associations": True,
                            "affected_units": affected_units,
                            "suggestion": "请先删除关联的发货记录，或使用 force=True 强制删除"
                        }

                for unit in units:
                    session.delete(unit)

                session.commit()

                return {"success": True, "message": f"成功删除 {len(units)} 条记录", "deleted_count": len(units)}
            finally:
                session.close()

        except Exception as e:
            logger.exception(f"批量删除失败: {e}")
            return {"success": False, "message": str(e), "deleted_count": 0}

    def import_data(
        self,
        data: List[Dict[str, Any]],
        skip_duplicates: bool = True,
        validate_before_import: bool = True,
        clean_data: bool = True
    ) -> Dict[str, Any]:
        """导入客户数据（从解析后的数据列表）

        Args:
            data: 客户数据列表，每个元素包含 customer_name 等字段
            skip_duplicates: 是否跳过重复数据
            validate_before_import: 是否导入前验证
            clean_data: 是否清理数据

        Returns:
            导入结果统计
        """
        try:
            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                imported = 0
                skipped = 0
                failed = 0
                failed_items = []
                skipped_items = []

                for item in data:
                    try:
                        customer_name = item.get("customer_name") or item.get("unit_name") or item.get("name", "").strip()

                        if not customer_name:
                            skipped += 1
                            skipped_items.append({"reason": "客户名称为空", "item": item})
                            continue

                        if clean_data:
                            customer_name = customer_name.strip()

                        existing = session.query(PurchaseUnitModel).filter(
                            PurchaseUnitModel.unit_name == customer_name
                        ).first()

                        if existing:
                            if skip_duplicates:
                                skipped += 1
                                skipped_items.append({"reason": "客户已存在", "customer_name": customer_name})
                                continue
                            else:
                                existing.contact_person = item.get("contact_person") or existing.contact_person
                                existing.contact_phone = item.get("contact_phone") or existing.contact_phone
                                existing.address = item.get("address") or item.get("contact_address") or existing.address
                        else:
                            unit = PurchaseUnitModel(
                                unit_name=customer_name,
                                contact_person=item.get("contact_person") or "",
                                contact_phone=item.get("contact_phone") or "",
                                address=item.get("address") or item.get("contact_address") or ""
                            )
                            session.add(unit)

                        imported += 1

                    except Exception as item_error:
                        failed += 1
                        failed_items.append({"reason": str(item_error), "item": item})

                session.commit()

                return {
                    "success": True,
                    "imported": imported,
                    "skipped": skipped,
                    "failed": failed,
                    "details": {
                        "failed_items": failed_items,
                        "skipped_items": skipped_items
                    }
                }
            finally:
                session.close()

        except Exception as e:
            logger.exception(f"导入客户数据失败: {e}")
            return {
                "success": False,
                "imported": 0,
                "skipped": 0,
                "failed": 0,
                "details": {"failed_items": [], "skipped_items": []}
            }

    def import_from_excel(self, file) -> Dict[str, Any]:
        """从 Excel 导入购买单位"""
        try:
            import os

            from openpyxl import load_workbook

            from app.utils.path_utils import get_data_dir

            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                wb = load_workbook(file)
                ws = wb.active

                updated = 0
                inserted = 0
                skipped = 0

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue

                    unit_name = str(row[0]).strip()
                    if not unit_name:
                        skipped += 1
                        continue

                    existing = session.query(PurchaseUnitModel).filter(
                        PurchaseUnitModel.unit_name == unit_name
                    ).first()

                    if existing:
                        if row[1]:
                            existing.contact_person = str(row[1])
                        if row[2]:
                            existing.contact_phone = str(row[2])
                        if row[3]:
                            existing.address = str(row[3])
                        updated += 1
                    else:
                        unit = PurchaseUnitModel(
                            unit_name=unit_name,
                            contact_person=str(row[1]) if row[1] else "",
                            contact_phone=str(row[2]) if row[2] else "",
                            address=str(row[3]) if row[3] else ""
                        )
                        session.add(unit)
                        inserted += 1

                session.commit()

                return {
                    "success": True,
                    "message": f"导入完成",
                    "updated": updated,
                    "inserted": inserted,
                    "skipped": skipped
                }
            finally:
                session.close()

        except Exception as e:
            logger.exception(f"导入失败: {e}")
            return {"success": False, "message": str(e), "updated": 0, "inserted": 0, "skipped": 0}

    def export_to_excel(self, keyword: Optional[str] = None, template_id: Optional[str] = None) -> Dict[str, Any]:
        """导出购买单位到 Excel"""
        try:
            import os

            from openpyxl import Workbook

            from app.utils.path_utils import get_data_dir
            from app.utils.template_export_utils import fill_workbook_from_template

            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                query = session.query(PurchaseUnitModel).filter(PurchaseUnitModel.is_active == True)

                if keyword:
                    pattern = f"%{keyword}%"
                    query = query.filter(PurchaseUnitModel.unit_name.like(pattern))

                units = query.order_by(PurchaseUnitModel.unit_name).all()

                records = [
                    {
                        "id": unit.id,
                        "customer_name": unit.unit_name or "",
                        "contact_person": unit.contact_person or "",
                        "contact_phone": unit.contact_phone or "",
                        "address": unit.address or "",
                    }
                    for unit in units
                ]

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"customers_{timestamp}.xlsx"

                export_dir = os.path.join(get_data_dir(), "exports")
                os.makedirs(export_dir, exist_ok=True)
                file_path = os.path.join(export_dir, filename)

                template_path = None
                if template_id:
                    try:
                        from app.application import get_template_app_service

                        templates = (get_template_app_service().get_templates() or {}).get("templates") or []
                        target = next((t for t in templates if str(t.get("id")) == str(template_id)), None)
                        if target:
                            candidate_path = str(target.get("path") or target.get("file_path") or "").strip()
                            if candidate_path and os.path.exists(candidate_path):
                                template_path = candidate_path
                    except Exception:
                        template_path = None

                if template_path:
                    header_alias = {
                        "id": ["ID", "编号"],
                        "customer_name": ["客户名称", "购买单位", "单位名称"],
                        "contact_person": ["联系人", "联系人姓名"],
                        "contact_phone": ["电话", "联系电话", "手机号"],
                        "address": ["地址", "联系地址"],
                    }
                    wb = fill_workbook_from_template(
                        template_path=template_path,
                        records=records,
                        field_alias_map=header_alias,
                        sheet_name="客户列表",
                    )
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "客户列表"
                    ws.append(["ID", "客户名称", "联系人", "电话", "地址"])
                    for row in records:
                        ws.append([
                            row["id"],
                            row["customer_name"],
                            row["contact_person"],
                            row["contact_phone"],
                            row["address"],
                        ])

                wb.save(file_path)

                return {
                    "success": True,
                    "message": f"成功导出 {len(units)} 条记录",
                    "file_path": str(file_path),
                    "filename": filename
                }
            finally:
                session.close()

        except Exception as e:
            logger.exception(f"导出失败: {e}")
            return {"success": False, "message": str(e)}

    def get_purchase_unit_by_name(self, name: str) -> Optional[PurchaseUnit]:
        """根据名称获取购买单位（用于内部业务）"""
        try:
            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                unit = session.query(PurchaseUnitModel).filter(
                    PurchaseUnitModel.unit_name == name,
                    PurchaseUnitModel.is_active == True
                ).first()

                if unit:
                    return PurchaseUnit(
                        id=unit.id,
                        unit_name=unit.unit_name,
                        contact_person=unit.contact_person or "",
                        contact_phone=unit.contact_phone or "",
                        address=unit.address or "",
                        discount_rate=unit.discount_rate or 1.0,
                        is_active=bool(unit.is_active),
                        created_at=unit.created_at,
                        updated_at=unit.updated_at,
                    )
                return None
            finally:
                session.close()

        except Exception as e:
            logger.exception(f"查询购买单位失败: {e}")
            return None

    def match_purchase_unit(self, input_name: str) -> Optional[PurchaseUnit]:
        """智能匹配购买单位（模糊匹配）"""
        try:
            name = str(input_name or "").strip()
            if not name:
                # 空串在 Python 中属于任意字符串的子串，若参与子串匹配会误命中第一条记录
                return None

            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                exact = session.query(PurchaseUnitModel).filter(
                    PurchaseUnitModel.unit_name == name,
                    PurchaseUnitModel.is_active == True
                ).first()

                if exact:
                    return PurchaseUnit(
                        id=exact.id,
                        unit_name=exact.unit_name,
                        contact_person=exact.contact_person or "",
                        contact_phone=exact.contact_phone or "",
                        address=exact.address or "",
                    )

                all_units = session.query(PurchaseUnitModel).filter(
                    PurchaseUnitModel.is_active == True
                ).all()

                # 子串匹配仅用于较长名称，避免单字/短串误命中多个客户
                if len(name) >= 2:
                    for unit in all_units:
                        un = unit.unit_name or ""
                        if name in un or un in name:
                            return PurchaseUnit(
                                id=unit.id,
                                unit_name=unit.unit_name,
                                contact_person=unit.contact_person or "",
                                contact_phone=unit.contact_phone or "",
                                address=unit.address or "",
                            )

                return None
            finally:
                session.close()

        except Exception as e:
            logger.exception(f"匹配购买单位失败：{e}")
            return None


_customer_app_service_instance = None


def get_customer_app_service() -> CustomerApplicationService:
    """获取客户服务单例"""
    global _customer_app_service_instance
    if _customer_app_service_instance is None:
        _customer_app_service_instance = CustomerApplicationService()
    return _customer_app_service_instance
