# -*- coding: utf-8 -*-
"""
Excel Toolkit Skill - Excel文件查看工具

用于分析Excel文件结构、内容、合并单元格和位置信息。
"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def view_excel_content(file_path: str, sheet_name: Optional[str] = None, max_rows: int = 100) -> Dict[str, Any]:
    """
    查看Excel文件内容

    Args:
        file_path: Excel文件路径
        sheet_name: Sheet名称（默认为活动Sheet）
        max_rows: 最大读取行数

    Returns:
        包含文件内容的字典
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)

        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                return {
                    "success": False,
                    "error": f"Sheet '{sheet_name}' 不存在"
                }
        else:
            ws = wb.active

        content = []
        row_count = 0

        for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row)):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    row_data.append({
                        "coordinate": cell.coordinate,
                        "value": cell.value,
                        "type": str(type(cell.value).__name__)
                    })
            if row_data:
                content.append({
                    "row": row[0].row,
                    "cells": row_data
                })
                row_count += 1

        return {
            "success": True,
            "file": Path(file_path).name,
            "sheet": ws.title,
            "structure": {
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "dimensions": ws.dimensions
            },
            "content": content,
            "row_count": row_count
        }

    except ImportError:
        return {
            "success": False,
            "error": "需要安装 openpyxl 库: pip install openpyxl"
        }
    except FileNotFoundError:
        return {
            "success": False,
            "error": f"文件不存在: {file_path}"
        }
    except Exception as e:
        logger.error(f"查看Excel内容失败: {e}")
        return {
            "success": False,
            "error": f"读取失败: {str(e)}"
        }


def get_merged_cells(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    获取Excel文件中的合并单元格信息

    Args:
        file_path: Excel文件路径
        sheet_name: Sheet名称

    Returns:
        包含合并单元格信息的字典
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path)

        if sheet_name:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        else:
            ws = wb.active

        merged_list = []
        for merged_range in ws.merged_cells.ranges:
            master_cell = ws.cell(merged_range.min_row, merged_range.min_col)
            merged_list.append({
                "range": str(merged_range),
                "min_row": merged_range.min_row,
                "max_row": merged_range.max_row,
                "min_col": merged_range.min_col,
                "max_col": merged_range.max_col,
                "value": master_cell.value,
                "master": merged_range.coord
            })

        return {
            "success": True,
            "file": Path(file_path).name,
            "sheet": ws.title,
            "merged_cells": merged_list,
            "count": len(merged_list)
        }

    except ImportError:
        return {
            "success": False,
            "error": "需要安装 openpyxl 库"
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


def get_cell_styles(file_path: str, sheet_name: Optional[str] = None, max_rows: int = 10) -> Dict[str, Any]:
    """
    获取单元格样式信息

    Args:
        file_path: Excel文件路径
        sheet_name: Sheet名称
        max_rows: 最大读取行数

    Returns:
        包含样式信息的字典
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        styles = []
        for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row)):
            for cell in row:
                if cell.value is not None:
                    style_info = {
                        "coordinate": cell.coordinate,
                        "value": cell.value,
                        "font": {
                            "name": cell.font.name if cell.font else None,
                            "size": cell.font.size if cell.font else None,
                            "bold": cell.font.bold if cell.font else None,
                            "color": str(cell.font.color) if cell.font and cell.font.color else None
                        },
                        "alignment": {
                            "horizontal": cell.alignment.horizontal if cell.alignment else None,
                            "vertical": cell.alignment.vertical if cell.alignment else None
                        },
                        "fill": {
                            "pattern": cell.fill.patternType if cell.fill else None,
                            "fg_color": str(cell.fill.fgColor) if cell.fill and cell.fill.fgColor else None
                        }
                    }
                    styles.append(style_info)

        return {
            "success": True,
            "file": Path(file_path).name,
            "sheet": ws.title,
            "styles": styles
        }

    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


def analyze_structure(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    分析Excel文件结构

    Args:
        file_path: Excel文件路径
        sheet_name: Sheet名称

    Returns:
        包含结构分析的字典
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        wb = load_workbook(file_path, data_only=True)

        sheet_names = wb.sheetnames
        target_sheet = sheet_name if sheet_name and sheet_name in sheet_names else wb.active
        ws = wb[target_sheet]

        column_info = []
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            sample_values = []
            for row in range(1, min(4, ws.max_row + 1)):
                val = ws.cell(row, col).value
                if val is not None:
                    sample_values.append(str(val)[:30])

            column_info.append({
                "column": col_letter,
                "index": col,
                "sample_values": sample_values
            })

        return {
            "success": True,
            "file": Path(file_path).name,
            "sheet_names": sheet_names,
            "current_sheet": ws.title,
            "structure": {
                "total_rows": ws.max_row,
                "total_columns": ws.max_column,
                "dimensions": ws.dimensions
            },
            "columns": column_info
        }

    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


class ExcelToolkitSkill:
    """Excel工具技能类"""

    def __init__(self):
        self.name = "excel_toolkit"
        self.description = "Excel文件查看工具，用于分析结构、内容、合并单元格和位置信息"

    def execute(self, file_path: str, action: str = "view", sheet_name: Optional[str] = None, **kwargs) -> Dict[str, Any]:
        """
        执行Excel工具操作

        Args:
            file_path: Excel文件路径
            action: 操作类型 (view/merged/styles/structure)
            sheet_name: Sheet名称
            **kwargs: 其他参数

        Returns:
            操作结果
        """
        if action == "view":
            return view_excel_content(file_path, sheet_name, kwargs.get("max_rows", 100))
        elif action == "merged":
            return get_merged_cells(file_path, sheet_name)
        elif action == "styles":
            return get_cell_styles(file_path, sheet_name, kwargs.get("max_rows", 10))
        elif action == "structure":
            return analyze_structure(file_path, sheet_name)
        else:
            return {
                "success": False,
                "error": f"未知操作: {action}"
            }

    def get_skill_info(self) -> Dict[str, Any]:
        """获取技能信息"""
        return {
            "name": self.name,
            "description": self.description,
            "actions": {
                "view": "查看Excel内容",
                "merged": "获取合并单元格信息",
                "styles": "获取单元格样式",
                "structure": "分析Excel结构"
            },
            "parameters": {
                "file_path": {"type": "string", "required": True, "description": "Excel文件路径"},
                "action": {"type": "string", "required": True, "description": "操作类型"},
                "sheet_name": {"type": "string", "required": False, "description": "Sheet名称"}
            }
        }


_skill_instance: Optional[ExcelToolkitSkill] = None


def get_excel_toolkit_skill() -> ExcelToolkitSkill:
    """获取Excel工具技能单例"""
    global _skill_instance
    if _skill_instance is None:
        _skill_instance = ExcelToolkitSkill()
    return _skill_instance