# -*- coding: utf-8 -*-
"""
Excel Analyzer Skill - Excel模板结构分析工具

基于预训练的Excel模板分析器，用于：
- 识别Excel模板结构、样式、可编辑内容
- 提取表头、可编辑区域、合并单元格信息
- 分析单元格类型（模板内容/可编辑内容/公式）
"""

import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def analyze_template(file_path: str, sheet_name: Optional[str] = None, verbose: bool = False) -> Dict[str, Any]:
    """
    分析Excel模板结构

    Args:
        file_path: Excel文件路径
        sheet_name: 要分析的Sheet名称（默认为第一个Sheet）
        verbose: 是否输出详细信息

    Returns:
        包含模板结构分析结果的字典
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        wb = load_workbook(file_path, data_only=False)

        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                return {
                    "success": False,
                    "error": f"Sheet '{sheet_name}' 不存在，可用Sheet: {wb.sheetnames}"
                }
        else:
            ws = wb.active

        result = {
            "success": True,
            "file": Path(file_path).name,
            "sheet": ws.title,
            "structure": {
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "max_col_letter": get_column_letter(ws.max_column)
            },
            "zones": [],
            "merged_cells": [],
            "editable_ranges": [],
            "cells": {}
        }

        merged_ranges = list(ws.merged_cells.ranges)
        for merge_range in merged_ranges:
            merge_info = {
                "range": str(merge_range),
                "min_row": merge_range.min_row,
                "max_row": merge_range.max_row,
                "min_col": merge_range.min_col,
                "max_col": merge_range.max_col
            }

            master_cell = ws.cell(merge_range.min_row, merge_range.min_col)
            if master_cell.value:
                merge_info["value"] = str(master_cell.value)
                if "标题" in str(master_cell.value) or "表头" in str(master_cell.value):
                    merge_info["purpose"] = "标题/表头"
                elif "合计" in str(master_cell.value) or "汇总" in str(master_cell.value):
                    merge_info["purpose"] = "汇总区域"
                elif "签名" in str(master_cell.value):
                    merge_info["purpose"] = "签名区域"
                else:
                    merge_info["purpose"] = "数据区域"

            result["merged_cells"].append(merge_info)

        zones = _identify_zones(ws, merged_ranges)
        result["zones"] = zones

        editable_ranges = _identify_editable_ranges(ws, zones)
        result["editable_ranges"] = editable_ranges

        if verbose:
            cells_info = {}
            for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row)):
                for cell in row:
                    if cell.value is not None or cell.data_type != 'n':
                        cells_info[cell.coordinate] = {
                            "address": cell.coordinate,
                            "row": cell.row,
                            "col": cell.column,
                            "value": cell.value,
                            "type": _classify_cell(ws, cell, zones),
                            "is_merged": cell.coordinate in [str(m) for m in merged_ranges]
                        }
            result["cells"] = cells_info

        return result

    except ImportError:
        return {
            "success": False,
            "error": "需要安装 openpyxl 库: pip install openpyxl"
        }
    except FileNotFoundError:
        return {
            "success": False,
            "error": f"文件不存在: {file_path}"
        }
    except Exception as e:
        logger.error(f"分析Excel模板失败: {e}")
        return {
            "success": False,
            "error": f"分析失败: {str(e)}"
        }


def analyze_to_json(file_path: str, output_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    分析Excel模板并保存结果到JSON文件

    Args:
        file_path: Excel文件路径
        output_path: 输出JSON文件路径
        sheet_name: 要分析的Sheet名称

    Returns:
        包含分析结果的字典
    """
    result = analyze_template(file_path, sheet_name, verbose=True)

    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        result["output_file"] = output_path
        return result
    except Exception as e:
        return {
            "success": False,
            "error": f"保存JSON失败: {str(e)}"
        }


def _identify_zones(ws, merged_ranges) -> List[Dict[str, Any]]:
    """识别模板中的区域（表头、数据区、汇总区等）"""
    zones = []
    max_row = ws.max_row

    header_rows = []
    for row_idx in range(1, min(6, max_row + 1)):
        row_values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if any(row_values):
            header_rows.append(row_idx)

    if header_rows:
        zones.append({
            "name": "header",
            "rows": header_rows,
            "type": "template",
            "description": "表头和标题区域"
        })

    data_start = max(header_rows) + 1 if header_rows else 4
    data_rows = list(range(data_start, max_row - 3)) if max_row > data_start + 3 else []

    if data_rows:
        zones.append({
            "name": "data",
            "rows": [min(data_rows), max(data_rows)],
            "type": "editable",
            "description": "数据输入区域"
        })

    if max_row > 5:
        summary_row = max_row - 2
        zones.append({
            "name": "summary",
            "rows": [summary_row, max_row],
            "type": "template",
            "description": "汇总和签名区域"
        })

    return zones


def _identify_editable_ranges(ws, zones) -> List[Dict[str, Any]]:
    """识别可编辑区域"""
    editable_ranges = []
    data_zone = next((z for z in zones if z["name"] == "data"), None)

    if data_zone:
        row_range = data_zone["rows"]
        editable_ranges.append({
            "range": f"A{row_range[0]}:{ws.max_column}{row_range[1]}",
            "min_row": row_range[0],
            "max_row": row_range[1],
            "min_col": 1,
            "max_col": ws.max_column,
            "description": "产品/数据输入区"
        })

    return editable_ranges


def _classify_cell(ws, cell, zones) -> str:
    """分类单元格类型"""
    if cell.data_type == 'f':
        return "formula"

    for zone in zones:
        if zone["type"] == "editable":
            if zone["rows"][0] <= cell.row <= zone["rows"][1]:
                return "editable"

    return "template"


class ExcelAnalyzerSkill:
    """Excel分析技能类"""

    def __init__(self):
        self.name = "excel_analyzer"
        self.description = "分析Excel模板结构，提取表头、可编辑区域、样式信息"

    def execute(self, file_path: str, sheet_name: Optional[str] = None, output_json: Optional[str] = None) -> Dict[str, Any]:
        """
        执行Excel模板分析

        Args:
            file_path: Excel文件路径
            sheet_name: Sheet名称（可选）
            output_json: 输出JSON路径（可选）

        Returns:
            分析结果
        """
        if output_json:
            return analyze_to_json(file_path, output_json, sheet_name)
        else:
            return analyze_template(file_path, sheet_name)

    def get_skill_info(self) -> Dict[str, Any]:
        """获取技能信息"""
        return {
            "name": self.name,
            "description": self.description,
            "parameters": {
                "file_path": {"type": "string", "required": True, "description": "Excel文件路径"},
                "sheet_name": {"type": "string", "required": False, "description": "Sheet名称"},
                "output_json": {"type": "string", "required": False, "description": "输出JSON路径"}
            }
        }


_skill_instance: Optional[ExcelAnalyzerSkill] = None


def get_excel_analyzer_skill() -> ExcelAnalyzerSkill:
    """获取Excel分析技能单例"""
    global _skill_instance
    if _skill_instance is None:
        _skill_instance = ExcelAnalyzerSkill()
    return _skill_instance