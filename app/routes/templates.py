# -*- coding: utf-8 -*-
"""
通用模板 API 路由
提供 /api/templates/* 路径的 API 端点
"""

from flask import Blueprint, jsonify, request
from flask_cors import CORS

from app.utils.json_safe import json_safe
import os
import json
import threading
import logging
import uuid

templates_bp = Blueprint('templates', __name__, url_prefix='/api/templates')
CORS(templates_bp)
logger = logging.getLogger(__name__)

# 全局进度跟踪字典
analysis_progress = {}
progress_lock = threading.Lock()

_TEMPLATE_SCOPE_REQUIRED_TERMS_CACHE = None
_DEFAULT_TEMPLATE_SCOPE_RULES = {
    "orders": {
        "templateType": "发货单",
        "requiredTerms": ["产品型号", "产品名称", "数量", "单价", "金额"],
    },
    "shipmentRecords": {
        "templateType": "出货记录",
        "requiredTerms": ["购买单位", "产品名称", "型号", "数量", "单价", "金额"],
    },
    "products": {
        "templateType": "产品清单",
        "requiredTerms": ["产品型号", "产品名称", "规格", "价格"],
    },
    "materials": {
        "templateType": "原材料清单",
        "requiredTerms": ["原材料编码", "名称", "分类", "规格", "单位", "库存数量", "单价", "供应商"],
    },
    "customers": {
        "templateType": "客户清单",
        "requiredTerms": ["客户名称", "联系人", "电话", "地址"],
    },
}


def _load_template_scope_required_terms():
    """从共享 JSON 配置加载业务词条规则。"""
    default_rules = {
        scope_key: list((meta or {}).get("requiredTerms") or [])
        for scope_key, meta in _DEFAULT_TEMPLATE_SCOPE_RULES.items()
    }
    config_path = str(os.environ.get("XCAGI_TEMPLATE_SCOPE_RULES_PATH") or "").strip()
    if not config_path:
        return default_rules
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config_data = json.load(f) or {}
        loaded_rules = {}
        for scope_key, rule in config_data.items():
            loaded_rules[scope_key] = list((rule or {}).get("requiredTerms") or [])
        return loaded_rules or default_rules
    except Exception:
        return default_rules


def _get_template_scope_required_terms():
    global _TEMPLATE_SCOPE_REQUIRED_TERMS_CACHE
    if _TEMPLATE_SCOPE_REQUIRED_TERMS_CACHE is None:
        _TEMPLATE_SCOPE_REQUIRED_TERMS_CACHE = _load_template_scope_required_terms()
    return _TEMPLATE_SCOPE_REQUIRED_TERMS_CACHE


def _normalize_term(value):
    return str(value or "").strip().replace(" ", "").lower()


_TERM_EQUIVALENTS = {
    "产品型号": ["产品型号", "型号", "产品编码"],
    "型号": ["型号", "产品型号", "产品编码"],
    "规格": ["规格", "规格型号", "规格/kg"],
    "规格型号": ["规格型号", "规格", "规格/kg"],
    "价格": ["价格", "单价", "单价/元"],
    "单价": ["单价", "价格", "单价/元"],
    "金额": ["金额", "金额/元", "金额合计", "总金额"],
    "数量": ["数量", "数量(kg)", "数量/kg", "数量/件", "数量/桶", "库存数量"],
    "电话": ["电话", "联系电话", "手机号"],
    "购买单位": ["购买单位", "单位", "单位名称", "客户名称", "厂名"],
    "客户名称": ["客户名称", "购买单位", "单位名称", "厂名"],
}


def _get_equivalent_normalized_terms(term: str):
    key = str(term or "").strip()
    aliases = _TERM_EQUIVALENTS.get(key) or [key]
    normalized = [_normalize_term(v) for v in aliases if _normalize_term(v)]
    normalized_key = _normalize_term(key)
    if normalized_key and normalized_key not in normalized:
        normalized.append(normalized_key)
    return normalized


def _has_equivalent_term(extracted_terms: set, required_term: str) -> bool:
    if not isinstance(extracted_terms, set):
        return False
    candidates = _get_equivalent_normalized_terms(required_term)
    return any(candidate in extracted_terms for candidate in candidates)


def _validate_required_terms(cells: dict, fields: list, template_scope: str):
    required_terms = _get_template_scope_required_terms().get(template_scope) or []
    if not required_terms:
        return True, []

    extracted = set()
    for f in fields or []:
        extracted.add(_normalize_term(f.get("label")))
        extracted.add(_normalize_term(f.get("name")))
        extracted.add(_normalize_term(f.get("value")))
    for cell_info in (cells or {}).values():
        extracted.add(_normalize_term((cell_info or {}).get("value")))

    missing_terms = [term for term in required_terms if not _has_equivalent_term(extracted, term)]
    return len(missing_terms) == 0, missing_terms


def _is_unreadable_workbook_error(error_message: str) -> bool:
    """判断是否为 Excel 文件损坏/不可读导致的已知错误。"""
    text = str(error_message or "").lower()
    markers = [
        "unable to read workbook",
        "could not read worksheets",
        "invalid xml",
        "badzipfile",
    ]
    return any(m in text for m in markers)


def _extract_structured_excel_preview(file_path: str, sheet_name: str = None, sample_limit: int = 6):
    """
    从 Excel 中提取更接近真实业务的预览数据：
    - 自动识别表头行
    - 基于表头提取样例数据行
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
            else:
                return {"fields": [], "sample_rows": [], "sheet_name": ""}

            header_row_idx = None
            header_entries = []
            max_scan_rows = min(ws.max_row or 0, 30)
            max_scan_cols = min(ws.max_column or 0, 25)
            best_candidate = None

            def _norm(text):
                return str(text or "").strip().replace(" ", "").lower()

            def _is_number_like(text: str) -> bool:
                t = str(text or "").strip().replace(",", "")
                if not t:
                    return False
                try:
                    float(t)
                    return True
                except Exception:
                    return False

            def _is_date_like(text: str) -> bool:
                t = str(text or "").strip()
                if not t:
                    return False
                if any(ch in t for ch in ("年", "月", "日", "-")) and any(ch.isdigit() for ch in t):
                    return True
                if len(t) >= 8 and t.isdigit():
                    return True
                return False

            for r in range(1, max_scan_rows + 1):
                row_headers = []
                norm_texts = []
                str_like = 0
                num_like = 0
                date_like = 0
                for c in range(1, max_scan_cols + 1):
                    value = ws.cell(r, c).value
                    if value is None:
                        continue
                    text = str(value).strip()
                    if not text:
                        continue
                    row_headers.append({
                        "name": text,
                        "column_index": c
                    })
                    nt = _norm(text)
                    norm_texts.append(nt)
                    if _is_number_like(text):
                        num_like += 1
                    elif _is_date_like(text):
                        date_like += 1
                    else:
                        str_like += 1

                if len(row_headers) < 3:
                    continue

                # 自适应评分（无业务关键词硬编码）：
                # 1) 表头通常“文本比例高、纯数字比例低”
                # 2) 表头字段重复率低（唯一性高）
                # 3) 下一行对这些列有较高承接（像数据行）
                # 4) 过于靠前的标题行略降权
                unique_ratio = (len(set(norm_texts)) / len(norm_texts)) if norm_texts else 0.0
                text_ratio = (str_like / len(row_headers)) if row_headers else 0.0
                number_ratio = (num_like / len(row_headers)) if row_headers else 0.0
                date_ratio = (date_like / len(row_headers)) if row_headers else 0.0

                score = 0.0
                score += text_ratio * 12.0
                score += unique_ratio * 8.0
                score -= number_ratio * 10.0
                score -= date_ratio * 4.0
                score -= (1.2 if r <= 2 else 0.0)

                # 次级校验：下一行在这些列里有值，说明更像数据表头
                next_row_non_empty = 0
                if r + 1 <= (ws.max_row or 0):
                    for h in row_headers[: min(len(row_headers), 16)]:
                        nv = ws.cell(r + 1, h["column_index"]).value
                        if nv is not None and str(nv).strip() != "":
                            next_row_non_empty += 1
                score += min(next_row_non_empty, 8)

                if best_candidate is None or score > best_candidate["score"]:
                    best_candidate = {
                        "row": r,
                        "headers": row_headers,
                        "score": score,
                    }

            if best_candidate is not None:
                header_row_idx = best_candidate["row"]
                header_entries = best_candidate["headers"]

            if header_row_idx is None:
                return {"fields": [], "sample_rows": [], "sheet_name": ws.title}

            fields = [
                {"label": h["name"], "value": "", "type": "dynamic"}
                for h in header_entries
            ]

            sample_rows = []
            for r in range(header_row_idx + 1, min((ws.max_row or 0), header_row_idx + sample_limit + 15) + 1):
                row_data = {}
                has_non_empty = False
                for h in header_entries:
                    value = ws.cell(r, h["column_index"]).value
                    if value is not None and str(value).strip() != "":
                        has_non_empty = True
                    row_data[h["name"]] = value
                if has_non_empty:
                    sample_rows.append(row_data)
                if len(sample_rows) >= sample_limit:
                    break

            return {
                "fields": fields,
                "sample_rows": sample_rows,
                "sheet_name": ws.title
            }
        finally:
            wb.close()
    except Exception:
        return {"fields": [], "sample_rows": [], "sheet_name": sheet_name or ""}


def _extract_excel_grid_preview(file_path: str, sheet_name: str = None, max_rows: int = 18, max_cols: int = 12):
    """
    提取真实 Excel 网格（包含合并单元格）用于前端预览。
    返回 rows: [[{text,rowspan,colspan,row,col}, ...], ...]
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries

        wb = load_workbook(file_path, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
            else:
                return {"sheet_name": "", "rows": []}

            row_limit = min(ws.max_row or 1, max_rows)
            col_limit = min(ws.max_column or 1, max_cols)

            merged_top_left = {}
            merged_covered = set()
            for merged in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(merged))
                if min_row > row_limit or min_col > col_limit:
                    continue
                span_row = max(1, min(max_row, row_limit) - min_row + 1)
                span_col = max(1, min(max_col, col_limit) - min_col + 1)
                merged_top_left[(min_row, min_col)] = (span_row, span_col)
                for r in range(min_row, min(max_row, row_limit) + 1):
                    for c in range(min_col, min(max_col, col_limit) + 1):
                        if (r, c) != (min_row, min_col):
                            merged_covered.add((r, c))

            rows = []
            for r in range(1, row_limit + 1):
                row_cells = []
                for c in range(1, col_limit + 1):
                    if (r, c) in merged_covered:
                        continue
                    rowspan, colspan = merged_top_left.get((r, c), (1, 1))
                    value = ws.cell(r, c).value
                    row_cells.append({
                        "row": r,
                        "col": c,
                        "text": "" if value is None else str(value),
                        "rowspan": rowspan,
                        "colspan": colspan
                    })
                rows.append(row_cells)

            return {"sheet_name": ws.title, "rows": rows}
        finally:
            wb.close()
    except Exception:
        return {"sheet_name": sheet_name or "", "rows": []}


def _extract_excel_grid_style_cache(file_path: str, sheet_name: str = None, max_rows: int = 24, max_cols: int = 14):
    """
    轻量提取网格样式缓存（前端仅缓存，不直接展示）。
    返回:
    {
      "sheet_name": "...",
      "styles": {"style_key": {...}},
      "cell_style_refs": {"r,c": "style_key"}
    }
    """
    try:
        from copy import copy
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
            else:
                return {"sheet_name": "", "styles": {}, "cell_style_refs": {}}

            row_limit = min(ws.max_row or 1, max_rows)
            col_limit = min(ws.max_column or 1, max_cols)

            styles = {}
            cell_style_refs = {}

            def _style_snapshot(cell):
                f = copy(cell.font)
                fill = copy(cell.fill)
                border = copy(cell.border)
                align = copy(cell.alignment)
                return {
                    "font": {
                        "name": f.name,
                        "size": f.size,
                        "bold": bool(f.bold),
                        "italic": bool(f.italic),
                        "color": str(getattr(f.color, "rgb", "") or ""),
                    },
                    "fill": {
                        "fill_type": fill.fill_type,
                        "fg_color": str(getattr(fill.fgColor, "rgb", "") or ""),
                    },
                    "border": {
                        "left": str(getattr(border.left, "style", "") or ""),
                        "right": str(getattr(border.right, "style", "") or ""),
                        "top": str(getattr(border.top, "style", "") or ""),
                        "bottom": str(getattr(border.bottom, "style", "") or ""),
                    },
                    "alignment": {
                        "horizontal": align.horizontal,
                        "vertical": align.vertical,
                        "wrap_text": bool(align.wrap_text),
                    },
                    "number_format": cell.number_format or "",
                }

            def _style_key(snapshot):
                import json
                return str(abs(hash(json.dumps(snapshot, ensure_ascii=False, sort_keys=True))))

            for r in range(1, row_limit + 1):
                for c in range(1, col_limit + 1):
                    cell = ws.cell(r, c)
                    snapshot = _style_snapshot(cell)
                    key = _style_key(snapshot)
                    if key not in styles:
                        styles[key] = snapshot
                    cell_style_refs[f"{r},{c}"] = key

            return {
                "sheet_name": ws.title,
                "styles": styles,
                "cell_style_refs": cell_style_refs,
            }
        finally:
            wb.close()
    except Exception:
        return {"sheet_name": sheet_name or "", "styles": {}, "cell_style_refs": {}}


def _extract_excel_all_sheets_preview(
    file_path: str,
    sample_limit: int = 8,
    max_rows: int = 24,
    max_cols: int = 14,
):
    """提取多 sheet 的结构 + 样例 + 网格 + 样式缓存。"""
    sheets = []
    for idx, name in enumerate(_list_excel_sheet_names(file_path), start=1):
        structured = _extract_structured_excel_preview(file_path, sheet_name=name, sample_limit=sample_limit)
        grid_preview = _extract_excel_grid_preview(file_path, sheet_name=name, max_rows=max_rows, max_cols=max_cols)
        style_cache = _extract_excel_grid_style_cache(file_path, sheet_name=name, max_rows=max_rows, max_cols=max_cols)
        logical_tables = _extract_logical_tables_from_sheet(file_path, sheet_name=name)
        sheets.append(
            {
                "sheet_index": idx,
                "sheet_name": name,
                "fields": structured.get("fields") or [],
                "sample_rows": structured.get("sample_rows") or [],
                "grid_preview": grid_preview,
                "style_cache": style_cache,
                "tables": logical_tables,
            }
        )
    return sheets


def _extract_logical_tables_from_sheet(file_path: str, sheet_name: str, max_scan_rows: int = 400, max_scan_cols: int = 25):
    """
    在单个 sheet 中识别多个“逻辑表块”（以空行分段 + 表头行判定）。
    返回:
    [
      {"table_index": 1, "header_row": 5, "fields": [...], "sample_rows": [...]},
      ...
    ]
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return []
            ws = wb[sheet_name]
            row_limit = min(ws.max_row or 0, max_scan_rows)
            col_limit = min(ws.max_column or 0, max_scan_cols)
            tables = []
            idx = 0
            r = 1
            while r <= row_limit:
                non_empty_cells = []
                for c in range(1, col_limit + 1):
                    v = ws.cell(r, c).value
                    if v is None:
                        continue
                    t = str(v).strip()
                    if t:
                        non_empty_cells.append({"name": t, "column_index": c})
                # 认为 >=3 列有值的行是候选表头
                if len(non_empty_cells) >= 3:
                    header_entries = non_empty_cells
                    sample_rows = []
                    rr = r + 1
                    while rr <= row_limit:
                        row_data = {}
                        has_non_empty = False
                        for h in header_entries:
                            v = ws.cell(rr, h["column_index"]).value
                            if v is not None and str(v).strip() != "":
                                has_non_empty = True
                            row_data[h["name"]] = v
                        if not has_non_empty:
                            break
                        sample_rows.append(row_data)
                        if len(sample_rows) >= 8:
                            break
                        rr += 1
                    fields = [{"label": h["name"], "value": "", "type": "dynamic"} for h in header_entries]
                    idx += 1
                    tables.append(
                        {
                            "table_index": idx,
                            "header_row": r,
                            "fields": fields,
                            "sample_rows": sample_rows,
                        }
                    )
                    r = max(rr + 1, r + 1)
                    continue
                r += 1
            return tables
        finally:
            wb.close()
    except Exception:
        return []


def _list_excel_sheet_names(file_path: str):
    """读取 Excel 的全部工作表名称。"""
    # 兼容不同 Excel 特性：先尝试 openpyxl 普通模式，再尝试 read_only，最后 pandas 兜底。
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True, read_only=False)
        try:
            names = [str(n).strip() for n in (wb.sheetnames or []) if str(n).strip()]
            if names:
                return names
        finally:
            wb.close()
    except Exception:
        pass

    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            names = [str(n).strip() for n in (wb.sheetnames or []) if str(n).strip()]
            if names:
                return names
        finally:
            wb.close()
    except Exception:
        pass

    try:
        import pandas as pd

        excel_file = pd.ExcelFile(file_path)
        names = [str(n).strip() for n in (excel_file.sheet_names or []) if str(n).strip()]
        return names
    except Exception:
        return []


def _parse_json_dict(raw_value):
    if isinstance(raw_value, dict):
        return raw_value
    if not raw_value:
        return {}
    try:
        data = json.loads(raw_value)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _parse_json_list(raw_value):
    if isinstance(raw_value, list):
        return raw_value
    if not raw_value:
        return []
    try:
        data = json.loads(raw_value)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _build_scope_template_type_map():
    map_result = {}
    for scope_key, meta in _DEFAULT_TEMPLATE_SCOPE_RULES.items():
        template_type = str((meta or {}).get("templateType") or "").strip()
        if template_type:
            map_result[template_type] = scope_key
    # 保底映射，避免规则文件缺失时无法推断
    if "发货单" not in map_result:
        map_result["发货单"] = "orders"
    if "出货记录" not in map_result:
        map_result["出货记录"] = "shipmentRecords"
    if "产品清单" not in map_result:
        map_result["产品清单"] = "products"
    if "原材料清单" not in map_result:
        map_result["原材料清单"] = "materials"
    if "客户清单" not in map_result:
        map_result["客户清单"] = "customers"
    return map_result


def _infer_business_scope(template_type: str):
    text = str(template_type or "").strip()
    if not text:
        return ""
    return _build_scope_template_type_map().get(text, "")


def _normalize_db_template_id(raw_id):
    value = str(raw_id or "").strip()
    if not value:
        return None
    if value.startswith("db:"):
        value = value.split(":", 1)[1].strip()
    if value.isdigit():
        return int(value)
    return None


def _ensure_template_tables_ready():
    """
    懒加载保障：在模板相关写操作前确保表存在。
    兼容未经过完整应用初始化流程的运行方式。
    """
    try:
        from app.db.init_db import init_template_tables
        init_template_tables()
    except Exception as e:
        logger.warning(f"确保模板表结构失败: {e}")


def _build_template_payload_from_row(row):
    analyzed_data = _parse_json_dict(getattr(row, "analyzed_data", None))
    business_rules = _parse_json_dict(getattr(row, "business_rules", None))
    business_scope = str(
        business_rules.get("business_scope")
        or analyzed_data.get("business_scope")
        or _infer_business_scope(getattr(row, "template_type", ""))
        or ""
    ).strip()
    fields = analyzed_data.get("fields")
    if not isinstance(fields, list):
        fields = _parse_json_list(getattr(row, "editable_config", None))
    preview_data = analyzed_data.get("preview_data")
    if not isinstance(preview_data, dict):
        preview_data = {}
    return {
        "id": f"db:{row.id}",
        "db_id": row.id,
        "name": getattr(row, "template_name", "") or "",
        "template_type": getattr(row, "template_type", "") or "",
        "business_scope": business_scope,
        "category": "excel",
        "source": business_rules.get("source") or analyzed_data.get("source") or "db",
        "file_path": getattr(row, "original_file_path", None),
        "fields": fields if isinstance(fields, list) else [],
        "preview_data": preview_data,
    }


@templates_bp.route('/extract-grid', methods=['POST'])
def extract_excel_grid():
    """
    轻量 Excel 网格提取接口：
    - 上传 Excel
    - 返回真实网格（含合并单元格）、字段和样例行
    """
    try:
        import uuid

        file = request.files.get('file')
        if not file or not file.filename:
            return jsonify({
                "success": False,
                "message": "请先上传 Excel 文件"
            }), 400

        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in ['.xlsx', '.xls']:
            return jsonify({
                "success": False,
                "message": "仅支持 .xlsx / .xls"
            }), 400

        # 保护性限制，避免超大文件导致接口长时间占用（前端会提示用户）
        file.stream.seek(0, os.SEEK_END)
        file_size = file.stream.tell()
        file.stream.seek(0)
        max_size = 10 * 1024 * 1024  # 10MB
        if file_size > max_size:
            return jsonify({
                "success": False,
                "message": "文件过大（>10MB），请先精简后上传分析"
            }), 400

        upload_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads', 'templates')
        os.makedirs(upload_dir, exist_ok=True)
        tmp_path = os.path.join(upload_dir, f"{uuid.uuid4().hex}{file_ext}")
        file.save(tmp_path)

        requested_sheet = request.form.get('sheet_name') or None
        analyze_all_sheets = str(request.form.get('analyze_all_sheets', 'true')).strip().lower() in {'1', 'true', 'yes', 'on'}
        sheet_names = _list_excel_sheet_names(tmp_path)
        structured = _extract_structured_excel_preview(tmp_path, sheet_name=requested_sheet, sample_limit=8)
        grid_preview = _extract_excel_grid_preview(tmp_path, sheet_name=requested_sheet, max_rows=24, max_cols=14)
        style_cache = _extract_excel_grid_style_cache(tmp_path, sheet_name=requested_sheet, max_rows=24, max_cols=14)
        selected_sheet_name = (
            structured.get("sheet_name")
            or grid_preview.get("sheet_name")
            or requested_sheet
            or (sheet_names[0] if sheet_names else "")
        )
        logical_tables = _extract_logical_tables_from_sheet(
            tmp_path,
            sheet_name=selected_sheet_name or (requested_sheet or (sheet_names[0] if sheet_names else "")),
        )
        all_sheets = _extract_excel_all_sheets_preview(tmp_path, sample_limit=8, max_rows=24, max_cols=14) if analyze_all_sheets else []

        response = jsonify({
            "success": True,
            "template_name": os.path.splitext(file.filename)[0],
            "template_type": "excel",
            "file_path": tmp_path,
            "fields": structured.get("fields") or [],
            "sheets": all_sheets,
            "preview_data": {
                "sample_rows": structured.get("sample_rows") or [],
                "sheet_name": structured.get("sheet_name") or requested_sheet or "",
                "selected_sheet_name": selected_sheet_name,
                "sheet_names": sheet_names,
                "grid_preview": grid_preview,
                "grid_style_cache": style_cache,
                "tables": logical_tables,
                "all_sheets": all_sheets,
                "file_path": tmp_path,
            }
        })
        return response
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"提取失败：{str(e)}"
        }), 500


@templates_bp.route('/progress/<task_id>', methods=['GET'])
def get_analysis_progress(task_id):
    """获取分析任务进度"""
    with progress_lock:
        progress = analysis_progress.get(task_id, {})
    
    return jsonify({
        "success": True,
        "task_id": task_id,
        "progress": progress.get('percent', 0),
        "step": progress.get('step', 1),
        "message": progress.get('message', '准备中...'),
        "completed": progress.get('completed', False)
    })


@templates_bp.route('/list', methods=['GET'])
def list_templates():
    """获取模板列表（兼容前端调用）"""
    try:
        from app.routes.excel_templates import _get_template_list

        templates = _get_template_list()
        include_scan = str(request.args.get('include_scan') or '').strip().lower() in {'1', 'true', 'yes'}
        if not include_scan:
            allowed_sources = {
                'db',
                'generated',
                'business-docking',
                'template-preview-replace',
                'system-default'
            }
            templates = [
                tpl for tpl in templates
                if str((tpl or {}).get('source') or '').strip() in allowed_sources
                or str((tpl or {}).get('id') or '').strip().startswith('db:')
            ]

        return jsonify(json_safe({"success": True, "templates": templates}))
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"获取模板列表失败：{e}")

        return jsonify({
            "success": False,
            "message": str(e),
            "templates": []
        }), 500


@templates_bp.route('/create', methods=['POST'])
def create_template():
    """
    创建模板（兼容前端 /api/templates/create）。
    """
    try:
        import uuid
        from datetime import datetime
        from sqlalchemy import text
        from app.db.session import get_db
        _ensure_template_tables_ready()

        payload = request.get_json(silent=True) or {}
        template_name = str(payload.get("name") or payload.get("template_name") or "").strip()
        if not template_name:
            return jsonify({
                "success": False,
                "message": "模板名称不能为空"
            }), 400

        template_type = str(payload.get("template_type") or "Excel").strip()
        business_scope = str(payload.get("business_scope") or _infer_business_scope(template_type) or "").strip()
        fields = payload.get("fields") if isinstance(payload.get("fields"), list) else []
        preview_data = payload.get("preview_data") if isinstance(payload.get("preview_data"), dict) else {}
        source = str(payload.get("source") or "generated").strip() or "generated"
        file_path = str(payload.get("file_path") or payload.get("original_file_path") or "").strip() or None
        if business_scope:
            valid, missing_terms = _validate_required_terms({}, fields, business_scope)
            if not valid:
                return jsonify({
                    "success": False,
                    "message": "必填字段未匹配，不能保存模板",
                    "business_scope": business_scope,
                    "missing_terms": missing_terms
                }), 400

        analyzed_data = {
            "category": "excel",
            "source": source,
            "business_scope": business_scope,
            "fields": fields,
            "preview_data": preview_data,
        }
        editable_config = fields
        business_rules = {
            "business_scope": business_scope,
            "source": source,
            "selected_sheet_name": preview_data.get("selected_sheet_name") or preview_data.get("sheet_name") or "",
        }

        template_key = f"TPL_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8].upper()}"
        with get_db() as db:
            result = db.execute(
                text("""
                    INSERT INTO templates (
                        template_key, template_name, template_type,
                        original_file_path, analyzed_data, editable_config,
                        zone_config, merged_cells_config, style_config,
                        business_rules, is_active
                    ) VALUES (
                        :template_key, :template_name, :template_type,
                        :original_file_path, :analyzed_data, :editable_config,
                        :zone_config, :merged_cells_config, :style_config,
                        :business_rules, :is_active
                    )
                """),
                {
                    "template_key": template_key,
                    "template_name": template_name,
                    "template_type": template_type,
                    "original_file_path": file_path,
                    "analyzed_data": json.dumps(analyzed_data, ensure_ascii=False),
                    "editable_config": json.dumps(editable_config, ensure_ascii=False),
                    "zone_config": json.dumps({}, ensure_ascii=False),
                    "merged_cells_config": json.dumps({}, ensure_ascii=False),
                    "style_config": json.dumps({}, ensure_ascii=False),
                    "business_rules": json.dumps(business_rules, ensure_ascii=False),
                    "is_active": 1,
                }
            )
            template_id = result.lastrowid
            db.commit()
            try:
                db.execute(
                    text("""
                        INSERT INTO template_usage_log (template_id, action, result)
                        VALUES (:template_id, 'create', :result)
                    """),
                    {"template_id": template_id, "result": f"创建模板：{template_name}"}
                )
                db.commit()
            except Exception as e:
                logger.warning(f"记录模板创建日志失败: {e}")

        return jsonify({
            "success": True,
            "message": "模板创建成功",
            "template": {
                "id": f"db:{template_id}",
                "db_id": template_id,
                "name": template_name,
                "template_type": template_type,
                "business_scope": business_scope,
                "category": "excel",
                "source": source,
                "file_path": file_path,
                "fields": fields,
                "preview_data": preview_data,
            }
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"创建失败：{str(e)}"
        }), 500


@templates_bp.route('/update', methods=['POST'])
def update_template():
    """
    更新模板（兼容前端 /api/templates/update）。
    - 支持基础编辑（名称等）
    - 支持网格替换并校验业务范围一致
    """
    try:
        from datetime import datetime
        from sqlalchemy import text
        from app.db.session import get_db
        _ensure_template_tables_ready()

        payload = request.get_json(silent=True) or {}
        db_id = _normalize_db_template_id(payload.get("id"))
        if db_id is None:
            return jsonify({
                "success": False,
                "message": "模板 id 无效"
            }), 400

        with get_db() as db:
            row = db.execute(
                text("""
                    SELECT id, template_name, template_type, original_file_path,
                           analyzed_data, editable_config, business_rules
                    FROM templates
                    WHERE id = :id
                """),
                {"id": db_id}
            ).fetchone()
            if not row:
                return jsonify({
                    "success": False,
                    "message": "模板不存在"
                }), 404

            existing_analyzed_data = _parse_json_dict(getattr(row, "analyzed_data", None))
            existing_business_rules = _parse_json_dict(getattr(row, "business_rules", None))
            existing_scope = str(
                existing_business_rules.get("business_scope")
                or existing_analyzed_data.get("business_scope")
                or _infer_business_scope(getattr(row, "template_type", ""))
                or ""
            ).strip()

            incoming_template_type = str(payload.get("template_type") or getattr(row, "template_type", "") or "").strip()
            incoming_scope = str(
                payload.get("business_scope")
                or existing_scope
                or _infer_business_scope(incoming_template_type)
                or ""
            ).strip()
            enforce_scope_match = bool(payload.get("enforce_scope_match") or payload.get("replace_mode"))
            if enforce_scope_match and existing_scope and incoming_scope and existing_scope != incoming_scope:
                return jsonify({
                    "success": False,
                    "message": f"仅允许替换同业务范围模板：当前为 {existing_scope}，目标为 {incoming_scope}"
                }), 400

            updates = []
            params = {"id": db_id, "updated_at": datetime.now()}

            new_name = str(payload.get("name") or payload.get("template_name") or "").strip()
            if new_name:
                updates.append("template_name = :template_name")
                params["template_name"] = new_name

            if incoming_template_type:
                updates.append("template_type = :template_type")
                params["template_type"] = incoming_template_type

            file_path = str(payload.get("file_path") or payload.get("original_file_path") or "").strip()
            if file_path:
                updates.append("original_file_path = :original_file_path")
                params["original_file_path"] = file_path

            incoming_fields = payload.get("fields")
            incoming_preview_data = payload.get("preview_data")
            source = str(payload.get("source") or existing_business_rules.get("source") or existing_analyzed_data.get("source") or "db").strip()
            fields_for_validation = incoming_fields if isinstance(incoming_fields, list) else (
                existing_analyzed_data.get("fields") if isinstance(existing_analyzed_data.get("fields"), list) else []
            )
            if incoming_scope:
                valid, missing_terms = _validate_required_terms({}, fields_for_validation, incoming_scope)
                if not valid:
                    return jsonify({
                        "success": False,
                        "message": "必填字段未匹配，不能替换模板",
                        "business_scope": incoming_scope,
                        "missing_terms": missing_terms
                    }), 400

            new_analyzed_data = {
                **existing_analyzed_data,
                "source": source,
                "business_scope": incoming_scope,
            }
            if isinstance(incoming_fields, list):
                new_analyzed_data["fields"] = incoming_fields
                updates.append("editable_config = :editable_config")
                params["editable_config"] = json.dumps(incoming_fields, ensure_ascii=False)
            if isinstance(incoming_preview_data, dict):
                merged_preview = {
                    **(existing_analyzed_data.get("preview_data") if isinstance(existing_analyzed_data.get("preview_data"), dict) else {}),
                    **incoming_preview_data
                }
                new_analyzed_data["preview_data"] = merged_preview

            updates.append("analyzed_data = :analyzed_data")
            params["analyzed_data"] = json.dumps(new_analyzed_data, ensure_ascii=False)

            new_business_rules = {
                **existing_business_rules,
                "business_scope": incoming_scope,
                "source": source,
            }
            if isinstance(incoming_preview_data, dict):
                selected_sheet_name = incoming_preview_data.get("selected_sheet_name") or incoming_preview_data.get("sheet_name")
                if selected_sheet_name:
                    new_business_rules["selected_sheet_name"] = selected_sheet_name
            updates.append("business_rules = :business_rules")
            params["business_rules"] = json.dumps(new_business_rules, ensure_ascii=False)

            updates.append("updated_at = :updated_at")

            # 验证字段名白名单，防止SQL注入
            allowed_fields = {
                "template_name", "template_type", "original_file_path",
                "editable_config", "analyzed_data", "business_rules", "updated_at"
            }
            for update_clause in updates:
                field_name = update_clause.split("=")[0].strip()
                if field_name not in allowed_fields:
                    return jsonify({
                        "success": False,
                        "message": f"无效的更新字段: {field_name}"
                    }), 400

            db.execute(
                text(f"UPDATE templates SET {', '.join(updates)} WHERE id = :id"),
                params
            )
            db.commit()

            try:
                db.execute(
                    text("""
                        INSERT INTO template_usage_log (template_id, action, result)
                        VALUES (:template_id, 'update', :result)
                    """),
                    {"template_id": db_id, "result": "更新模板配置"}
                )
                db.commit()
            except Exception as e:
                logger.warning(f"记录模板更新日志失败: {e}")

            refreshed = db.execute(
                text("""
                    SELECT id, template_name, template_type, original_file_path,
                           analyzed_data, business_rules
                    FROM templates
                    WHERE id = :id
                """),
                {"id": db_id}
            ).fetchone()

        return jsonify({
            "success": True,
            "message": "模板更新成功",
            "template": _build_template_payload_from_row(refreshed)
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"更新失败：{str(e)}"
        }), 500


@templates_bp.route('/delete', methods=['POST', 'DELETE'])
def delete_template():
    """
    删除模板（兼容前端旧调用）:
    - fs:<filename> -> 删除文件模板
    - db:<id> / 数字 id -> templates 表软删除
    """
    try:
        payload = request.get_json(silent=True) or {}
        template_id = str(
            payload.get('id')
            or request.args.get('id')
            or ''
        ).strip()

        if not template_id:
            return jsonify({
                "success": False,
                "message": "缺少模板 id"
            }), 400

        # 1) 文件扫描模板：fs:<filename>
        if template_id.startswith('fs:'):
            filename = template_id.split(':', 1)[1].strip()
            if not filename:
                return jsonify({
                    "success": False,
                    "message": "模板文件名无效"
                }), 400

            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            candidates = [
                os.path.join(base_dir, filename),
                os.path.join(base_dir, 'templates', filename),
                os.path.join(base_dir, 'resources', 'templates', filename),
            ]

            target_path = None
            for p in candidates:
                if os.path.isfile(p):
                    target_path = p
                    break

            if not target_path:
                return jsonify({
                    "success": False,
                    "message": f"模板文件不存在: {filename}"
                }), 404

            os.remove(target_path)
            return jsonify({
                "success": True,
                "message": "模板删除成功",
                "deleted": {
                    "id": template_id,
                    "path": target_path
                }
            })

        # 2) 数据库模板：db:<id> 或纯数字 id
        db_id = None
        if template_id.startswith('db:'):
            raw_db_id = template_id.split(':', 1)[1].strip()
            if raw_db_id.isdigit():
                db_id = int(raw_db_id)
        elif template_id.isdigit():
            db_id = int(template_id)

        if db_id is not None:
            from sqlalchemy import text
            from app.db.session import get_db
            from datetime import datetime
            _ensure_template_tables_ready()
            with get_db() as db:
                row = db.execute(
                    text("SELECT id FROM templates WHERE id = :id"),
                    {"id": db_id}
                ).fetchone()
                if not row:
                    return jsonify({
                        "success": False,
                        "message": "模板不存在"
                    }), 404

                db.execute(
                    text("UPDATE templates SET is_active = 0, updated_at = :updated_at WHERE id = :id"),
                    {"id": db_id, "updated_at": datetime.now()}
                )
                db.commit()

            return jsonify({
                "success": True,
                "message": "模板删除成功",
                "deleted": {
                    "id": template_id,
                    "db_id": db_id
                }
            })

        return jsonify({
            "success": False,
            "message": f"暂不支持删除该模板类型: {template_id}"
        }), 400
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"删除失败：{str(e)}"
        }), 500


@templates_bp.route('/detail/<template_id>', methods=['GET'])
def get_template_detail(template_id):
    """获取模板详情（包括预览数据）"""
    try:
        from app.infrastructure.skills.excel_analyzer.excel_template_analyzer import get_excel_analyzer_skill
        from app.routes.excel_templates import _get_template_list
        from openpyxl import load_workbook
        import logging
        logger = logging.getLogger(__name__)

        templates = _get_template_list()
        target = next((t for t in templates if str(t.get("id")) == str(template_id)), None)
        if not target:
            return jsonify({
                "success": False,
                "message": "模板不存在"
            }), 404

        template_path = target.get("path") or target.get("file_path")
        if not template_path or not os.path.exists(template_path):
            return jsonify({
                "success": False,
                "message": "模板文件不存在"
            }), 404

        sheet_name = "出货"
        try:
            wb = load_workbook(template_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames and wb.sheetnames:
                sheet_name = wb.sheetnames[0]
            wb.close()
        except Exception as e:
            logger.debug(f"读取工作表名称失败: {e}")

        skill = get_excel_analyzer_skill()
        analyze_result = skill.execute(file_path=template_path, sheet_name=sheet_name)

        if analyze_result.get('success'):
            cells = analyze_result.get('cells', {})
            editable_entries = analyze_result.get('editable_ranges', [])
            structured = _extract_structured_excel_preview(template_path, sheet_name=sheet_name, sample_limit=8)
            grid_preview = _extract_excel_grid_preview(template_path, sheet_name=sheet_name, max_rows=18, max_cols=12)
            fields = structured.get("fields") or []
            sample_rows = structured.get("sample_rows") or []

            if not fields:
                # 回退：从 analyzer cells 取词条
                for _, cell_info in list(cells.items())[:25]:
                    if cell_info.get('value') and cell_info.get('purpose') != '系统保留':
                        fields.append({
                            "label": str(cell_info.get('value', '')),
                            "value": '',
                            "type": "dynamic"
                        })

            return jsonify({
                "success": True,
                "template": {
                    "id": template_id,
                    "name": target.get("name") or target.get("template_name") or os.path.basename(template_path),
                    "category": target.get("category") or "excel",
                    "template_type": target.get("template_type") or "Excel",
                    "business_scope": target.get("business_scope"),
                    "source": target.get("source") or "fs_scan",
                    "fields": fields,
                    "preview_data": {
                        "cells": cells,
                        "editable_ranges": editable_entries,
                        "sample_rows": sample_rows,
                        "sheet_name": structured.get("sheet_name") or sheet_name,
                        "grid_preview": grid_preview
                    }
                }
            })
        else:
            error_message = analyze_result.get('error', '获取详情失败')
            if _is_unreadable_workbook_error(error_message):
                return jsonify({
                    "success": False,
                    "message": "模板文件损坏或格式异常，无法读取。请重新导出或另存为 .xlsx 后重试。",
                    "error_code": "UNREADABLE_WORKBOOK"
                }), 500
            return jsonify({
                "success": False,
                "message": error_message
            }), 500

    except Exception as e:
        logger.error(f"获取模板详情失败：{e}")
        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@templates_bp.route('/analyze', methods=['POST'])
def analyze_template():
    """分析上传的文件，自动识别为 Excel 模板或标签模板"""
    try:
        logger.info(f"收到 analyze 请求，files: {request.files.keys()}")
        logger.info(f"收到 analyze 请求，form: {request.form}")
        
        if 'file' not in request.files:
            logger.error("没有上传文件")
            return jsonify({
                "success": False,
                "message": "没有上传文件"
            }), 400

        file = request.files['file']
        template_name = request.form.get('template_name', '')
        template_scope = request.form.get('template_scope', '')

        logger.info(f"文件名：{file.filename}, 模板名：{template_name}")

        if file.filename == '':
            logger.error("文件名为空")
            return jsonify({
                "success": False,
                "message": "文件名为空"
            }), 400

        file_ext = os.path.splitext(file.filename)[1].lower()

        upload_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads', 'templates')
        os.makedirs(upload_dir, exist_ok=True)

        unique_filename = f"{uuid.uuid4().hex}{file_ext}"
        file_path = os.path.join(upload_dir, unique_filename)
        file.save(file_path)

        # 生成任务 ID 并初始化进度
        task_id = uuid.uuid4().hex
        with progress_lock:
            analysis_progress[task_id] = {
                'percent': 0,
                'step': 1,
                'message': '准备上传文件...',
                'completed': False
            }

        try:
            if file_ext in ['.xlsx', '.xls']:
                return _analyze_excel_template(file_path, template_name, file.filename, task_id, template_scope)
            elif file_ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp']:
                return _analyze_label_template(file_path, template_name, file.filename, task_id)
            else:
                with progress_lock:
                    if task_id in analysis_progress:
                        del analysis_progress[task_id]
                return jsonify({
                    "success": False,
                    "message": f"不支持的文件类型：{file_ext}"
                }), 400
        finally:
            # 清理临时文件
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as e:
                logger.warning(f"清理临时文件失败: {file_path}, 错误: {e}")

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"分析模板失败：{e}")
        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "message": f"分析失败：{str(e)}"
        }), 500


def _analyze_excel_template(
    file_path: str,
    template_name: str,
    original_filename: str,
    task_id: str,
    template_scope: str = ""
):
    """分析 Excel 模板 - 结合 excel_analyzer 和 excel_toolkit 技能"""
    try:
        from app.infrastructure.skills.excel_analyzer.excel_template_analyzer import get_excel_analyzer_skill

        # 步骤 1: 上传完成 (10%)
        _update_progress(task_id, 10, 1, '文件上传成功')
        
        skill = get_excel_analyzer_skill()
        
        # 步骤 2: 开始分析 Excel (50%)
        _update_progress(task_id, 50, 2, '分析 Excel 结构...')

        analyze_result = skill.execute(file_path=file_path, sheet_name='出货')
        
        # 步骤 3: 完成 (100%)
        _update_progress(task_id, 100, 3, '分析完成！')
        with progress_lock:
            if task_id in analysis_progress:
                analysis_progress[task_id]['completed'] = True

        if analyze_result.get('success'):
            cells = analyze_result.get('cells', {})
            editable_entries = analyze_result.get('editable_ranges', [])
            merged_cells = analyze_result.get('merged_cells', [])
            structure = analyze_result.get('structure', {})

            structured = _extract_structured_excel_preview(file_path, sheet_name='出货', sample_limit=8)
            grid_preview = _extract_excel_grid_preview(file_path, sheet_name='出货', max_rows=18, max_cols=12)
            fields = structured.get("fields") or []

            if not fields:
                for _, cell_info in list(cells.items())[:25]:
                    if cell_info.get('value') and cell_info.get('purpose') != '系统保留':
                        fields.append({
                            "label": str(cell_info.get('value', '')),
                            "value": '',
                            "type": "dynamic"
                        })

            valid, missing_terms = _validate_required_terms(cells, fields, template_scope)
            if not valid:
                return jsonify({
                    "success": False,
                    "message": "模板缺少必备词条，请补全后重试",
                    "required_terms": _get_template_scope_required_terms().get(template_scope, []),
                    "missing_terms": missing_terms
                }), 400

            sample_rows = structured.get("sample_rows") or []

            name = template_name if template_name else original_filename.replace('.xlsx', '').replace('.xls', '')

            return jsonify({
                "success": True,
                "task_id": task_id,
                "template_name": name,
                "template_type": "excel",
                "fields": fields,
                "preview_data": {
                    "cells": cells,
                    "editable_ranges": editable_entries,
                    "merged_cells": merged_cells,
                    "sample_rows": sample_rows,
                    "structure": structure,
                    "sheet_name": structured.get("sheet_name") or "出货",
                    "grid_preview": grid_preview,
                    "file_path": file_path
                }
            })
        else:
            return jsonify({
                "success": False,
                "message": analyze_result.get('error', 'Excel 分析失败')
            }), 500

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"分析 Excel 模板失败：{e}")
        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "message": f"分析 Excel 失败：{str(e)}"
        }), 500


def _update_progress(task_id: str, percent: int, step: int, message: str):
    """更新任务进度"""
    with progress_lock:
        if task_id in analysis_progress:
            analysis_progress[task_id].update({
                'percent': percent,
                'step': step,
                'message': message
            })

def _analyze_label_template(file_path: str, template_name: str, original_filename: str, task_id: str):
    """
    分析标签模板（图片）：PIL 分析尺寸/颜色 + OCR 抽字段 + 生成技能代码。

    OCR 与 /api/ocr 共用 OCRService（默认 PaddleOCR，可回退 EasyOCR/Tesseract）；
    标签网格仍用 OpenCV。说明见 RECOGNITION_TEMPLATE_OCR.md。
    """
    try:
        from app.services.skills.label_template_generator.label_template_generator import (
            LabelTemplateGeneratorSkill
        )
        import logging
        logger = logging.getLogger(__name__)

        # 步骤 1: 上传完成 (10%)
        _update_progress(task_id, 10, 1, '文件上传成功')
        
        skill = LabelTemplateGeneratorSkill()
        
        # 步骤 2: 开始检测网格 (25%)
        _update_progress(task_id, 25, 2, '检测表格网格...')

        ocr_result = skill.execute(
            image_path=file_path,
            class_name=template_name.replace(' ', '') + 'Generator' if template_name else 'LabelGenerator',
            enable_ocr=True,
            verbose=True
        )
        
        # 步骤 3: OCR 识别完成 (75%)
        _update_progress(task_id, 75, 3, 'OCR 识别文字...')
        
        # 步骤 4: 分析字段 (90%)
        _update_progress(task_id, 90, 4, '分析字段...')

        logger.info(f"OCR 识别结果：{ocr_result}")

        if ocr_result.get('success'):
            fields = []
            
            # 提前获取 ocr_data 以便后续使用 grid 信息
            image_analysis = ocr_result.get('analysis', {})
            ocr_data = ocr_result.get('ocr_result', {})
            
            if ocr_data and ocr_data.get('fields'):
                ocr_fields = ocr_data['fields']
                logger.info(f"OCR 提取到 {len(ocr_fields)} 个字段")
                
                import uuid
                for field in ocr_fields:
                    logger.info(f"字段：{field}")
                    fields.append({
                        "id": uuid.uuid4().hex,
                        "label": field.get('label', ''),
                        "value": field.get('value', ''),
                        "type": 'fixed' if field.get('type') == 'fixed_label' else 'dynamic',
                        "position": field.get('position', {}),
                        "confidence": field.get('confidence', 0)
                    })
            else:
                logger.warning("OCR 未提取到字段，使用默认字段")
                fields = [
                    {"label": "品名", "value": "示例产品", "type": "fixed"},
                    {"label": "货号", "value": "00000", "type": "dynamic"},
                    {"label": "颜色", "value": "黑色", "type": "dynamic"},
                    {"label": "码段", "value": "00000", "type": "dynamic"},
                    {"label": "等级", "value": "合格品", "type": "fixed"},
                    {"label": "执行标准", "value": "QB/Txxxx-xxxx", "type": "fixed"},
                    {"label": "统一零售价", "value": "¥0", "type": "dynamic"}
                ]

            name = template_name if template_name else original_filename.replace('.' + original_filename.split('.')[-1], '')
            
            # 标记任务完成 (100%)
            _update_progress(task_id, 100, 4, '分析完成！')
            with progress_lock:
                if task_id in analysis_progress:
                    analysis_progress[task_id]['completed'] = True

            # 构建 preview_data
            preview_data = {
                "image_path": file_path,
                "image_size": image_analysis.get('size', {}),
                "colors": image_analysis.get('colors', {}),
                "ocr_fields": fields
            }
            
            # 添加网格信息
            if ocr_data and ocr_data.get('grid'):
                preview_data['grid'] = ocr_data['grid']
                logger.info(f"网格信息：{ocr_data['grid']}")
            else:
                logger.warning("未找到网格信息")

            return jsonify({
                "success": True,
                "task_id": task_id,
                "template_name": name,
                "template_type": "label",
                "fields": fields,
                "generated_code": ocr_result.get('code', ''),
                "preview_data": preview_data
            })
        else:
            return jsonify({
                "success": False,
                "message": ocr_result.get('error', '标签生成失败')
            }), 500

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"分析标签模板失败：{e}")
        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "message": f"分析标签失败：{str(e)}"
        }), 500
