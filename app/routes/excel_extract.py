# -*- coding: utf-8 -*-
"""
Excel 提取与生成路由模块

提供 Excel 数据提取和生成相关接口，包括：
- /api/excel/extract: 从 Excel 中提取数据
- /api/excel/generate: 生成 Excel 文件
"""

import logging
import os
from datetime import datetime

from flasgger import swag_from
from flask import Blueprint, jsonify, request, send_file

logger = logging.getLogger(__name__)

excel_extract_bp = Blueprint("excel_extract", __name__, url_prefix="/api/excel/data")

TEMP_EXCEL_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "temp_excel")
os.makedirs(TEMP_EXCEL_DIR, exist_ok=True)


def get_base_dir():
    """获取项目根目录"""
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def _get_ai_product_parser():
    """延迟导入 AI 产品解析服务，避免循环依赖。"""
    from app.services import get_ai_product_parser

    return get_ai_product_parser()


def _extract_from_excel(file_path, sheet_name=None, header_row=1):
    """从 Excel 文件中提取数据"""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        if not os.path.exists(file_path):
            return {"success": False, "message": f"文件不存在: {file_path}"}, 404

        wb = load_workbook(file_path, data_only=True)

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v is not None:
                headers.append({
                    "column": get_column_letter(c),
                    "column_index": c,
                    "value": str(v).strip() if v else ""
                })

        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            row_data = {}
            for h in headers:
                v = ws.cell(r, h["column_index"]).value
                row_data[h["value"]] = v
            if any(v is not None and v != "" for v in row_data.values()):
                rows.append(row_data)

        return {
            "success": True,
            "file": os.path.basename(file_path),
            "sheet": ws.title,
            "header_row": header_row,
            "headers": headers,
            "rows": rows,
            "total_rows": len(rows),
        }, 200

    except Exception as e:
        logger.error(f"提取 Excel 数据失败: {e}")
        return {"success": False, "message": str(e)}, 500


def _generate_excel(data, filename=None, sheet_name="Sheet1"):
    """生成 Excel 文件"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{timestamp}.xlsx"

        file_path = os.path.join(TEMP_EXCEL_DIR, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data or not isinstance(data, list):
            return {"success": False, "message": "数据格式错误，需要数组类型"}, 400

        headers = list(data[0].keys()) if data else []

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row_idx, col_idx, row_data.get(header))

        wb.save(file_path)

        return {
            "success": True,
            "file_path": file_path,
            "filename": filename,
            "sheet": sheet_name,
            "rows": len(data),
        }, 200

    except Exception as e:
        logger.error(f"生成 Excel 文件失败: {e}")
        return {"success": False, "message": str(e)}, 500


@excel_extract_bp.route("/extract", methods=["POST"])
@swag_from({
    'summary': '提取Excel数据',
    'description': '从 Excel 文件中提取数据',
    'parameters': [
        {
            'name': 'body',
            'in': 'body',
            'required': True,
            'schema': {
                'type': 'object',
                'properties': {
                    'file_path': {'type': 'string', 'description': '文件路径'}
                },
                'required': ['file_path']
            }
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'data': {'type': 'object'}
                }
            }
        }
    }
})
def extract_from_excel():
    """
    从 Excel 文件中提取数据

    Request:
    {
        "file_path": "/path/to/file.xlsx",
        "sheet_name": "Sheet1",
        "header_row": 1
    }

    Response:
    {
        "success": True,
        "file": "filename.xlsx",
        "sheet": "Sheet1",
        "headers": [...],
        "rows": [...],
        "total_rows": 10
    }
    """
    try:
        data = request.json or {}
        file_path = data.get("file_path")
        sheet_name = data.get("sheet_name")
        header_row = data.get("header_row", 1)

        if not file_path:
            return jsonify({
                "success": False,
                "message": "请提供 file_path 参数"
            }), 400

        result, status = _extract_from_excel(file_path, sheet_name, header_row)
        return jsonify(result), status

    except Exception as e:
        logger.error(f"提取 Excel 数据失败: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@excel_extract_bp.route("/extract/upload", methods=["POST"])
@swag_from({
    'summary': '上传并提取',
    'description': '上传文件并提取数据',
    'parameters': [
        {
            'name': 'excel_file',
            'in': 'formData',
            'type': 'file',
            'required': True,
            'description': 'Excel 文件'
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'data': {'type': 'object'}
                }
            }
        }
    }
})
def extract_upload():
    """
    上传文件并提取数据

    Request:
        - excel_file: Excel 文件
        - sheet_name: 工作表名称（可选）
        - header_row: 表头行号（可选，默认 1）

    Response:
    {
        "success": True,
        "file": "filename.xlsx",
        "headers": [...],
        "rows": [...],
        "total_rows": 10
    }
    """
    try:
        if "excel_file" not in request.files:
            return jsonify({
                "success": False,
                "message": "请上传 Excel 文件"
            }), 400

        file = request.files["excel_file"]
        sheet_name = request.form.get("sheet_name")
        header_row = int(request.form.get("header_row", 1))

        if not file.filename:
            return jsonify({
                "success": False,
                "message": "请选择文件"
            }), 400

        if not file.filename.lower().endswith((".xlsx", ".xls")):
            return jsonify({
                "success": False,
                "message": "只支持 .xlsx 和 .xls 格式"
            }), 400

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"extract_{timestamp}_{file.filename}"
        file_path = os.path.join(TEMP_EXCEL_DIR, filename)
        file.save(file_path)

        result, status = _extract_from_excel(file_path, sheet_name, header_row)

        if os.path.exists(file_path):
            os.remove(file_path)

        return jsonify(result), status

    except Exception as e:
        logger.error(f"上传并提取 Excel 数据失败: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@excel_extract_bp.route("/generate", methods=["POST"])
@swag_from({
    'summary': '生成Excel',
    'description': '生成 Excel 文件',
    'parameters': [
        {
            'name': 'body',
            'in': 'body',
            'required': True,
            'schema': {
                'type': 'object',
                'properties': {
                    'data': {'type': 'object', 'description': '生成数据'},
                    'template_id': {'type': 'string', 'description': '模板 ID'}
                },
                'required': ['data']
            }
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'file_path': {'type': 'string'}
                }
            }
        }
    }
})
def generate_excel():
    """
    生成 Excel 文件

    Request:
    {
        "data": [
            {"列1": "值1", "列2": "值2"},
            {"列1": "值3", "列2": "值4"}
        ],
        "filename": "export.xlsx",
        "sheet_name": "Sheet1"
    }

    Response:
    {
        "success": True,
        "file_path": "/path/to/file.xlsx",
        "filename": "export.xlsx",
        "rows": 2
    }
    """
    try:
        data = request.json or {}
        excel_data = data.get("data", [])
        filename = data.get("filename")
        sheet_name = data.get("sheet_name", "Sheet1")

        if not excel_data:
            return jsonify({
                "success": False,
                "message": "请提供数据 data 参数"
            }), 400

        result, status = _generate_excel(excel_data, filename, sheet_name)
        return jsonify(result), status

    except Exception as e:
        logger.error(f"生成 Excel 文件失败: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@excel_extract_bp.route("/generate/download", methods=["POST"])
@swag_from({
    'summary': '生成并下载Excel',
    'description': '生成并下载 Excel 文件',
    'parameters': [
        {
            'name': 'body',
            'in': 'body',
            'required': True,
            'schema': {
                'type': 'object',
                'properties': {
                    'data': {'type': 'object', 'description': '生成数据'},
                    'template_id': {'type': 'string', 'description': '模板 ID'}
                },
                'required': ['data']
            }
        }
    ],
    'responses': {
        '200': {
            'description': '文件下载'
        }
    }
})
def download_generated_excel():
    """
    生成并下载 Excel 文件

    请求参数同上 /generate，响应为文件下载
    """
    try:
        data = request.json or {}
        excel_data = data.get("data", [])
        filename = data.get("filename")
        sheet_name = data.get("sheet_name", "Sheet1")

        if not excel_data:
            return jsonify({
                "success": False,
                "message": "请提供数据 data 参数"
            }), 400

        result, status = _generate_excel(excel_data, filename, sheet_name)

        if status != 200:
            return jsonify(result), status

        file_path = result.get("file_path")
        download_filename = result.get("filename")

        if not file_path or not os.path.exists(file_path):
            return jsonify({
                "success": False,
                "message": "生成的文件不存在"
            }), 500

        return send_file(
            file_path,
            as_attachment=True,
            download_name=download_filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        logger.error(f"生成并下载 Excel 文件失败: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@excel_extract_bp.route("/extract/test", methods=["GET"])
@swag_from({
    'summary': '测试Excel提取服务',
    'description': '测试 Excel 提取服务是否正常运行',
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'message': {'type': 'string'}
                }
            }
        }
    }
})
def test():
    """测试接口"""
    return jsonify({
        "success": True,
        "message": "Excel 提取服务运行正常",
        "timestamp": datetime.now().isoformat(),
    })


@excel_extract_bp.route("/import/products", methods=["POST"])
@swag_from({
    'summary': '导入产品',
    'description': '导入产品数据',
    'parameters': [
        {
            'name': 'body',
            'in': 'body',
            'required': True,
            'schema': {
                'type': 'object',
                'properties': {
                    'data': {'type': 'array', 'description': '产品数据'}
                },
                'required': ['data']
            }
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'message': {'type': 'string'}
                }
            }
        }
    }
})
def import_products():
    """
    导入产品数据
    
    Request:
    {
        "data": [...],
        "options": {
            "skip_duplicates": true,
            "validate_before_import": true,
            "clean_data": true
        }
    }
    
    Response:
    {
        "success": True,
        "imported": 10,
        "skipped": 2,
        "failed": 0,
        "details": {
            "skipped_items": [...],
            "failed_items": [...]
        }
    }
    """
    try:
        from app.application import (
            ExtractLogApplicationService,
            ProductImportApplicationService,
            get_extract_log_app_service,
            get_product_import_application_service,
        )

        data = request.json or {}
        extracted_data = data.get('data', [])
        options = data.get('options', {})
        
        if not extracted_data:
            return jsonify({'success': False, 'message': '缺少数据'}), 400
        
        # 创建日志记录
        log_service = get_extract_log_app_service()
        log_id = log_service.create_log(
            file_name=data.get('file_name', 'products_import'),
            data_type='products',
            total_rows=len(extracted_data),
            field_mapping=data.get('field_mapping')
        )
        
        # 如启用 AI 解析，则先用 AIProductParser 将原始文本标准化为产品字段
        use_ai_parse = bool(options.get("use_ai_parse", False))
        ai_source_field = options.get("ai_source_field") or ""
        if use_ai_parse and ai_source_field:
            parser = _get_ai_product_parser()
            normalized_rows = []
            for row in extracted_data:
                raw_text = str(row.get(ai_source_field, "") or "")
                if not raw_text.strip():
                    normalized_rows.append(row)
                    continue
                parsed = parser.parse_single(raw_text, use_ai=True, fallback_to_rule=True)
                if parsed.get("success"):
                    row = dict(row)
                    if parsed.get("product_code"):
                        row["product_code"] = parsed["product_code"]
                    if parsed.get("product_name"):
                        row["product_name"] = parsed["product_name"]
                    if parsed.get("specification"):
                        row["specification"] = parsed["specification"]
                    if parsed.get("unit") and "unit" not in row:
                        row["unit"] = parsed["unit"]
                    if parsed.get("unit_price") is not None and "unit_price" not in row:
                        row["unit_price"] = parsed.get("unit_price")
                normalized_rows.append(row)
            extracted_data = normalized_rows

        # 导入数据
        service = get_product_import_service()
        result = service.import_data(
            data=extracted_data,
            skip_duplicates=options.get('skip_duplicates', True),
            validate_before_import=options.get('validate_before_import', True),
            clean_data=options.get('clean_data', True),
        )
        
        # 更新日志
        log_service.update_log(
            log_id=log_id,
            status='completed' if result['failed'] == 0 else 'partial',
            imported_rows=result['imported'],
            skipped_rows=result['skipped'],
            failed_rows=result['failed']
        )
        
        return jsonify({
            'success': True,
            'log_id': log_id,
            'imported': result['imported'],
            'skipped': result['skipped'],
            'failed': result['failed'],
            'details': result['details']
        })
        
    except Exception as e:
        logger.error(f"导入产品数据失败：{e}")
        return jsonify({
            'success': False,
            'message': f'导入失败：{str(e)}'
        }), 500


@excel_extract_bp.route("/import/customers", methods=["POST"])
@swag_from({
    'summary': '导入客户',
    'description': '导入客户数据',
    'parameters': [
        {
            'name': 'body',
            'in': 'body',
            'required': True,
            'schema': {
                'type': 'object',
                'properties': {
                    'data': {'type': 'array', 'description': '客户数据'}
                },
                'required': ['data']
            }
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'message': {'type': 'string'}
                }
            }
        }
    }
})
def import_customers():
    """
    导入客户数据
    
    Request:
    {
        "data": [...],
        "options": {
            "skip_duplicates": true,
            "validate_before_import": true,
            "clean_data": true
        }
    }
    
    Response:
    {
        "success": True,
        "imported": 10,
        "skipped": 2,
        "failed": 0,
        "details": {
            "skipped_items": [...],
            "failed_items": [...]
        }
    }
    """
    try:
        from app.application import (
            CustomerApplicationService,
            ExtractLogApplicationService,
            get_customer_app_service,
            get_extract_log_app_service,
        )

        data = request.json or {}
        extracted_data = data.get('data', [])
        options = data.get('options', {})
        
        if not extracted_data:
            return jsonify({'success': False, 'message': '缺少数据'}), 400
        
        # 创建日志记录
        log_service = get_extract_log_app_service()
        log_id = log_service.create_log(
            file_name=data.get('file_name', 'customers_import'),
            data_type='customers',
            total_rows=len(extracted_data),
            field_mapping=data.get('field_mapping')
        )
        
        # 导入数据
        service = get_customer_app_service()
        result = service.import_data(
            data=extracted_data,
            skip_duplicates=options.get('skip_duplicates', True),
            validate_before_import=options.get('validate_before_import', True),
            clean_data=options.get('clean_data', True)
        )
        
        # 更新日志
        log_service.update_log(
            log_id=log_id,
            status='completed' if result['failed'] == 0 else 'partial',
            imported_rows=result['imported'],
            skipped_rows=result['skipped'],
            failed_rows=result['failed']
        )
        
        return jsonify({
            'success': True,
            'log_id': log_id,
            'imported': result['imported'],
            'skipped': result['skipped'],
            'failed': result['failed'],
            'details': result['details']
        })
        
    except Exception as e:
        logger.error(f"导入客户数据失败：{e}")
        return jsonify({
            'success': False,
            'message': f'导入失败：{str(e)}'
        }), 500


@excel_extract_bp.route("/logs", methods=["GET"])
@swag_from({
    'summary': '获取提取日志',
    'description': '获取提取日志列表',
    'parameters': [
        {
            'name': 'data_type',
            'in': 'query',
            'type': 'string',
            'required': False,
            'description': '数据类型过滤 (products/customers/orders)'
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'logs': {'type': 'array'}
                }
            }
        }
    }
})
def get_extract_logs():
    """
    获取提取日志列表
    
    Query Parameters:
    - data_type: 数据类型过滤 (products/customers/orders)
    - status: 状态过滤 (pending/completed/failed/partial)
    - limit: 限制数量 (默认 50)
    - offset: 偏移量 (默认 0)
    
    Response:
    {
        "success": True,
        "logs": [...],
        "total": 100
    }
    """
    try:
        from app.application import ExtractLogApplicationService, get_extract_log_app_service

        data_type = request.args.get('data_type')
        status = request.args.get('status')
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))
        
        log_service = get_extract_log_app_service()
        logs = log_service.get_logs(
            data_type=data_type,
            status=status,
            limit=limit,
            offset=offset
        )
        
        return jsonify({
            'success': True,
            'logs': logs,
            'total': len(logs)
        })
        
    except Exception as e:
        logger.error(f"获取提取日志失败：{e}")
        return jsonify({
            'success': False,
            'message': f'获取失败：{str(e)}'
        }), 500


@excel_extract_bp.route("/logs/<int:log_id>", methods=["GET"])
@swag_from({
    'summary': '获取日志详情',
    'description': '获取单个提取日志详情',
    'parameters': [
        {
            'name': 'log_id',
            'in': 'path',
            'type': 'integer',
            'required': True,
            'description': '日志 ID'
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'log': {'type': 'object'}
                }
            }
        }
    }
})
def get_extract_log(log_id: int):
    """
    获取单个提取日志详情
    
    Response:
    {
        "success": True,
        "log": {...}
    }
    """
    try:
        from app.application import ExtractLogApplicationService, get_extract_log_app_service

        log_service = get_extract_log_app_service()
        log = log_service.get_log(log_id)
        
        if not log:
            return jsonify({
                'success': False,
                'message': '日志不存在'
            }), 404
        
        return jsonify({
            'success': True,
            'log': log
        })
        
    except Exception as e:
        logger.error(f"获取提取日志详情失败：{e}")
        return jsonify({
            'success': False,
            'message': f'获取失败：{str(e)}'
        }), 500


@excel_extract_bp.route("/preview/<int:log_id>", methods=["GET"])
@swag_from({
    'summary': '获取预览',
    'description': '获取提取数据预览',
    'parameters': [
        {
            'name': 'log_id',
            'in': 'path',
            'type': 'integer',
            'required': True,
            'description': '日志 ID'
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'preview': {'type': 'object'}
                }
            }
        }
    }
})
def get_preview(log_id: int):
    """
    获取提取数据预览
    
    Response:
    {
        "success": True,
        "log": {...},
        "data": [...]  # 提取的原始数据（如果需要，可以从 field_mapping 或其他地方获取）
    }
    """
    try:
        from app.application import ExtractLogApplicationService, get_extract_log_app_service

        log_service = get_extract_log_app_service()
        log = log_service.get_log(log_id)

        if not log:
            return jsonify({
                'success': False,
                'message': '日志不存在'
            }), 404
        
        # 返回日志信息和预览数据
        # 注意：实际数据需要从其他地方获取（比如缓存或临时存储）
        # 这里简化处理，只返回日志信息
        
        return jsonify({
            'success': True,
            'log': log,
            'message': '预览数据需要从提取源获取'
        })
        
    except Exception as e:
        logger.error(f"获取预览失败：{e}")
        return jsonify({
            'success': False,
            'message': f'获取失败：{str(e)}'
        }), 500
