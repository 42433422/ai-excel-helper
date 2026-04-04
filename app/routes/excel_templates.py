# -*- coding: utf-8 -*-
"""
Excel 模板管理路由模块

提供 Excel 模板相关接口，包括：
- /api/excel/templates: 模板列表
- /api/excel/template/save: 保存模板
- /api/excel/template/decompose: 分解模板提取词条
"""

import logging
import os
from datetime import datetime

from flasgger import swag_from
from flask import Blueprint, jsonify, request, send_file

logger = logging.getLogger(__name__)

excel_templates_bp = Blueprint("excel_templates", __name__, url_prefix="/api/excel")

TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "templates")
TEMP_EXCEL_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "temp_excel")

os.makedirs(TEMPLATE_DIR, exist_ok=True)
os.makedirs(TEMP_EXCEL_DIR, exist_ok=True)


def get_base_dir():
    """获取项目根目录"""
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def _map_template_category(template_type: str) -> str:
    t = (template_type or "").strip().lower()
    if any(k in t for k in ["标签", "label", "print", "打印"]):
        return "label_print"
    return "excel"


def _normalize_template_dto(template: dict) -> dict:
    tpl = dict(template or {})
    template_type = tpl.get("template_type", "")
    category = tpl.get("category") or _map_template_category(template_type)
    file_path = tpl.get("file_path") or tpl.get("path")
    normalized = {
        **tpl,
        "category": category,
        "file_path": file_path,
        "is_active": bool(tpl.get("is_active", True)),
        "preview_capable": bool(file_path and tpl.get("exists", False)),
    }
    return normalized


def _resolve_template_path(filename):
    """解析模板文件路径"""
    base_dir = get_base_dir()
    path = os.path.join(base_dir, filename)
    if os.path.exists(path):
        return path

    alt_path = os.path.join(TEMPLATE_DIR, filename)
    if os.path.exists(alt_path):
        return alt_path

    return path if os.path.exists(path) else None


def _get_template_list():
    """获取模板列表"""
    from app.application import TemplateApplicationService, get_template_app_service

    return get_template_app_service().get_templates().get('templates', [])


def _is_unreadable_workbook_error(error_message: str) -> bool:
    """判断是否为 Excel 文件损坏/不可读导致的已知错误。"""
    text = str(error_message or "").lower()
    markers = [
        "unable to read workbook",
        "could not read worksheets",
        "invalid xml",
        "badzipfile",
    ]
    return any(m in text for m in markers)


@excel_templates_bp.route('/templates', methods=['GET'])
@swag_from({
    'summary': '获取模板列表',
    'description': '获取所有模板列表（支持 Excel 和标签打印）',
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'templates': {'type': 'array', 'items': {'type': 'object'}}
                }
            }
        }
    }
})
def get_templates():
    """获取模板列表"""
    try:
        templates = _get_template_list()
        return jsonify({
            "success": True,
            "templates": templates
        })
    except Exception as e:
        logger.error(f"获取模板列表失败：{e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


# 兼容旧 API 路径 /api/templates/list
@excel_templates_bp.route('/list', methods=['GET'])
def get_templates_list():
    """获取模板列表（兼容旧路径）"""
    try:
        templates = _get_template_list()
        return jsonify({
            "success": True,
            "templates": templates
        })
    except Exception as e:
        logger.error(f"获取模板列表失败：{e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


def _decompose_template(file_path, sheet_name=None, sample_rows=5):
    """分解 Excel 模板，提取词条信息"""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        if not os.path.exists(file_path):
            return {"success": False, "message": f"模板文件不存在: {file_path}"}, 404

        wb = load_workbook(file_path, data_only=True)

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif '出货' in wb.sheetnames:
            ws = wb['出货']
        else:
            ws = wb[wb.sheetnames[0]]

        header_row_idx = None
        header_cells = []

        for r in range(1, min(ws.max_row, 30) + 1):
            row_cells = []
            for c in range(1, min(ws.max_column, 25) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip():
                    row_cells.append({
                        "name": v.strip(),
                        "column": get_column_letter(c),
                        "column_index": c
                    })
            if len(row_cells) >= 4:
                header_row_idx = r
                header_cells = row_cells
                break

        if header_row_idx is None:
            header_row_idx = 1

        samples = []
        data_start = header_row_idx + 1
        data_end = min(ws.max_row, data_start + max(int(sample_rows), 1) - 1)

        for r in range(data_start, data_end + 1):
            row_data = {}
            non_empty = False
            for h in header_cells:
                v = ws.cell(r, h["column_index"]).value
                if v is not None and v != "":
                    non_empty = True
                row_data[h["name"]] = v
            if non_empty and row_data:
                samples.append(row_data)

        amount_related = [
            h for h in header_cells
            if any(k in h["name"] for k in ['金额', '单价', '价格', '数量'])
        ]

        result = {
            "success": True,
            "template": {
                "name": os.path.basename(file_path),
                "path": file_path,
                "sheet": ws.title,
                "dimension": ws.calculate_dimension(),
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            },
            "decomposition": {
                "header_row": header_row_idx,
                "editable_entries": header_cells,
                "amount_related_entries": amount_related,
                "sample_rows": samples,
                "merged_cells_count": len(ws.merged_cells.ranges),
            }
        }

        return result, 200

    except Exception as e:
        logger.error(f"分解 Excel 模板失败: {e}")
        if _is_unreadable_workbook_error(str(e)):
            return {
                "success": False,
                "message": "模板文件损坏或格式异常，无法读取。请重新导出或另存为 .xlsx 后重试。",
                "error_code": "UNREADABLE_WORKBOOK",
            }, 200
        return {"success": False, "message": str(e)}, 500


@excel_templates_bp.route("/templates", methods=["GET"])
@swag_from({
    'summary': '获取模板列表',
    'description': '获取 Excel 模板列表',
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'templates': {'type': 'array'}
                }
            }
        }
    }
})
def list_templates():
    """
    获取 Excel 模板列表

    Response:
    {
        "success": True,
        "templates": [
            {"id": "shipment", "name": "发货单模板", "filename": "发货单模板.xlsx", "exists": true}
        ]
    }
    """
    try:
        templates = [_normalize_template_dto(t) for t in _get_template_list()]
        return jsonify({
            "success": True,
            "templates": templates
        }), 200
    except Exception as e:
        logger.error(f"获取模板列表失败: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@excel_templates_bp.route("/template/<template_id>/file", methods=["GET"])
@swag_from({
    'summary': '获取模板文件',
    'description': '下载 Excel 模板文件',
    'parameters': [
        {
            'name': 'template_id',
            'in': 'path',
            'type': 'string',
            'required': True,
            'description': '模板 ID'
        }
    ],
    'responses': {
        '200': {
            'description': '文件下载'
        }
    }
})
def get_template_file(template_id):
    """获取模板文件"""
    try:
        templates = _get_template_list()
        template = next((t for t in templates if t["id"] == template_id), None)

        if not template:
            return jsonify({"success": False, "message": "模板不存在"}), 404

        if not template.get("exists") or not template.get("path"):
            return jsonify({"success": False, "message": "模板文件不存在"}), 404

        return send_file(
            template["path"],
            as_attachment=False,
            download_name=template["filename"],
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        logger.error(f"获取模板文件失败: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@excel_templates_bp.route("/template/save", methods=["POST"])
@swag_from({
    'summary': '保存模板',
    'description': '保存 Excel 模板',
    'parameters': [
        {
            'name': 'body',
            'in': 'body',
            'required': True,
            'schema': {
                'type': 'object',
                'properties': {
                    'template_id': {'type': 'string', 'description': '模板 ID'},
                    'name': {'type': 'string', 'description': '模板名称'},
                    'data': {'type': 'object', 'description': '模板数据'}
                },
                'required': ['template_id', 'name', 'data']
            }
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'message': {'type': 'string'}
                }
            }
        }
    }
})
def save_template():
    """
    保存 Excel 模板

    Request:
    {
        "source_name": "源文件名",
        "target_name": "目标文件名",
        "overwrite": false
    }

    Response:
    {
        "success": True,
        "message": "模板保存成功",
        "saved": true,
        "template_name": "发货单模板.xlsx"
    }
    """
    try:
        data = request.json or {}
        source_name = data.get("source_name", "尹玉华132.xlsx")
        target_name = data.get("target_name", "发货单模板.xlsx")
        overwrite = bool(data.get("overwrite", False))
        from app.application import TemplateApplicationService, get_template_app_service

        result = get_template_app_service().save_template_file(source_name, target_name, overwrite)
        status = 200 if result.get("success") else 404
        return jsonify(result), status

    except Exception as e:
        logger.error(f"保存模板失败: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@excel_templates_bp.route("/template/decompose", methods=["POST"])
@swag_from({
    'summary': '分解模板',
    'description': '分解 Excel 模板，提取词条',
    'parameters': [
        {
            'name': 'body',
            'in': 'body',
            'required': True,
            'schema': {
                'type': 'object',
                'properties': {
                    'template_id': {'type': 'string', 'description': '模板 ID'}
                },
                'required': ['template_id']
            }
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'entries': {'type': 'array'}
                }
            }
        }
    }
})
def decompose_template():
    """
    分解 Excel 模板，提取词条

    Request:
    {
        "filename": "发货单模板.xlsx",
        "file_path": "/path/to/file.xlsx",
        "sheet_name": "出货",
        "sample_rows": 5
    }

    Response:
    {
        "success": True,
        "template": {...},
        "decomposition": {...}
    }
    """
    try:
        data = request.json or {}
        filename = data.get("filename")
        file_path = data.get("file_path")
        sheet_name = data.get("sheet_name")
        sample_rows = data.get("sample_rows", 5)

        if file_path:
            target_path = file_path
        elif filename:
            target_path = _resolve_template_path(filename)
        else:
            return jsonify({
                "success": False,
                "message": "请提供 filename 或 file_path"
            }), 400

        if not target_path or not os.path.exists(target_path):
            return jsonify({
                "success": False,
                "message": f"模板文件不存在"
            }), 404

        result, status = _decompose_template(target_path, sheet_name, sample_rows)
        return jsonify(result), status

    except Exception as e:
        logger.error(f"分解模板失败: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@excel_templates_bp.route("/upload", methods=["POST"])
@swag_from({
    'summary': '上传Excel',
    'description': '上传 Excel 文件',
    'parameters': [
        {
            'name': 'excel_file',
            'in': 'formData',
            'type': 'file',
            'required': True,
            'description': 'Excel 文件'
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'message': {'type': 'string'}
                }
            }
        }
    }
})
def upload_excel():
    """
    上传 Excel 文件

    Request:
        - excel_file: Excel 文件

    Response:
    {
        "success": True,
        "file_path": "/path/to/file.xlsx",
        "filename": "原始文件名"
    }
    """
    try:
        if "excel_file" not in request.files:
            return jsonify({
                "success": False,
                "message": "请上传 Excel 文件"
            }), 400

        file = request.files["excel_file"]

        if not file.filename:
            return jsonify({
                "success": False,
                "message": "请选择文件"
            }), 400

        if not file.filename.lower().endswith((".xlsx", ".xls")):
            return jsonify({
                "success": False,
                "message": "只支持 .xlsx 和 .xls 格式"
            }), 400

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"excel_{timestamp}_{file.filename}"
        file_path = os.path.join(TEMP_EXCEL_DIR, filename)
        file.save(file_path)

        logger.info(f"Excel 文件已上传: {file_path}")

        return jsonify({
            "success": True,
            "file_path": file_path,
            "filename": file.filename,
            "message": "文件上传成功"
        })

    except Exception as e:
        logger.error(f"上传 Excel 文件失败: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@excel_templates_bp.route("/test", methods=["GET"])
@swag_from({
    'summary': '测试Excel模板服务',
    'description': '测试 Excel 模板服务是否正常运行',
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'message': {'type': 'string'}
                }
            }
        }
    }
})
def test():
    """测试接口"""
    return jsonify({
        "success": True,
        "message": "Excel 模板服务运行正常",
        "timestamp": datetime.now().isoformat(),
    })


@excel_templates_bp.route("/templates/<int:template_id>", methods=["GET"])
@swag_from({
    'summary': '获取模板详情',
    'description': '获取单个模板详情',
    'parameters': [
        {
            'name': 'template_id',
            'in': 'path',
            'type': 'integer',
            'required': True,
            'description': '模板 ID'
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'template': {'type': 'object'}
                }
            }
        }
    }
})
def get_template(template_id: int):
    """
    获取单个模板详情
    
    Response:
    {
        "success": True,
        "template": {
            "id": 1,
            "template_key": "TPL_xxx",
            "template_name": "模板名称",
            "template_type": "发货单",
            "analyzed_data": {...},
            "editable_config": {...},
            ...
        }
    }
    """
    try:
        from sqlalchemy import text

        from app.db.session import get_db
        
        with get_db() as db:
            result = db.execute(
                text("SELECT * FROM templates WHERE id = :id AND is_active = 1"),
                {'id': template_id}
            )
            row = result.fetchone()
            
            if not row:
                return jsonify({
                    "success": False,
                    "message": "模板不存在"
                }), 404
            
            import json
            template = {
                'id': row.id,
                'template_key': row.template_key,
                'template_name': row.template_name,
                'template_type': row.template_type,
                'original_file_path': row.original_file_path,
                'analyzed_data': json.loads(row.analyzed_data) if row.analyzed_data else None,
                'editable_config': json.loads(row.editable_config) if row.editable_config else None,
                'zone_config': json.loads(row.zone_config) if row.zone_config else None,
                'merged_cells_config': json.loads(row.merged_cells_config) if row.merged_cells_config else None,
                'style_config': json.loads(row.style_config) if row.style_config else None,
                'business_rules': json.loads(row.business_rules) if row.business_rules else None,
                'created_at': row.created_at,
                'updated_at': row.updated_at
            }
            
            return jsonify({
                "success": True,
                "template": template
            })
            
    except Exception as e:
        logger.error(f"获取模板详情失败：{e}")
        return jsonify({
            "success": False,
            "message": f"获取失败：{str(e)}"
        }), 500


@excel_templates_bp.route("/templates/<int:template_id>", methods=["PUT"])
@swag_from({
    'summary': '更新模板',
    'description': '更新模板配置',
    'parameters': [
        {
            'name': 'template_id',
            'in': 'path',
            'type': 'integer',
            'required': True,
            'description': '模板 ID'
        },
        {
            'name': 'body',
            'in': 'body',
            'required': True,
            'schema': {
                'type': 'object',
                'properties': {
                    'name': {'type': 'string', 'description': '模板名称'},
                    'description': {'type': 'string', 'description': '模板描述'}
                }
            }
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'message': {'type': 'string'}
                }
            }
        }
    }
})
def update_template(template_id: int):
    """
    更新模板配置
    
    Request:
    {
        "template_name": "新名称",
        "editable_config": {...},
        "business_rules": {...}
    }
    
    Response:
    {
        "success": True,
        "message": "模板更新成功"
    }
    """
    try:
        import json

        from sqlalchemy import text

        from app.db.session import get_db
        
        data = request.json or {}
        
        with get_db() as db:
            # 检查模板是否存在
            result = db.execute(
                text("SELECT id FROM templates WHERE id = :id"),
                {'id': template_id}
            )
            if not result.fetchone():
                return jsonify({
                    "success": False,
                    "message": "模板不存在"
                }), 404
            
            # 构建更新字段
            updates = []
            params = {'id': template_id}
            
            if 'template_name' in data:
                updates.append('template_name = :template_name')
                params['template_name'] = data['template_name']
            
            if 'template_type' in data:
                updates.append('template_type = :template_type')
                params['template_type'] = data['template_type']
            
            if 'editable_config' in data:
                updates.append('editable_config = :editable_config')
                params['editable_config'] = json.dumps(data['editable_config'], ensure_ascii=False)
            
            if 'zone_config' in data:
                updates.append('zone_config = :zone_config')
                params['zone_config'] = json.dumps(data['zone_config'], ensure_ascii=False)
            
            if 'business_rules' in data:
                updates.append('business_rules = :business_rules')
                params['business_rules'] = json.dumps(data['business_rules'], ensure_ascii=False)
            
            # 更新时间
            updates.append('updated_at = :updated_at')
            params['updated_at'] = datetime.now()
            
            sql = f"UPDATE templates SET {', '.join(updates)} WHERE id = :id"
            db.execute(text(sql), params)
            db.commit()
            
            # 记录使用日志
            db.execute(
                text("""
                    INSERT INTO template_usage_log (template_id, action, result)
                    VALUES (:template_id, 'update', :result)
                """),
                {'template_id': template_id, 'result': '更新模板配置'}
            )
            db.commit()
        
        return jsonify({
            "success": True,
            "message": "模板更新成功"
        })
        
    except Exception as e:
        logger.error(f"更新模板失败：{e}")
        return jsonify({
            "success": False,
            "message": f"更新失败：{str(e)}"
        }), 500


@excel_templates_bp.route("/templates/<int:template_id>", methods=["DELETE"])
@swag_from({
    'summary': '删除模板',
    'description': '删除模板（软删除）',
    'parameters': [
        {
            'name': 'template_id',
            'in': 'path',
            'type': 'integer',
            'required': True,
            'description': '模板 ID'
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'message': {'type': 'string'}
                }
            }
        }
    }
})
def delete_template(template_id: int):
    """
    删除模板（软删除）
    
    Response:
    {
        "success": True,
        "message": "模板删除成功"
    }
    """
    try:
        from sqlalchemy import text

        from app.db.session import get_db
        
        with get_db() as db:
            # 检查模板是否存在
            result = db.execute(
                text("SELECT id FROM templates WHERE id = :id"),
                {'id': template_id}
            )
            if not result.fetchone():
                return jsonify({
                    "success": False,
                    "message": "模板不存在"
                }), 404
            
            # 软删除
            db.execute(
                text("UPDATE templates SET is_active = 0, updated_at = :updated_at WHERE id = :id"),
                {'id': template_id, 'updated_at': datetime.now()}
            )
            
            # 记录使用日志
            db.execute(
                text("""
                    INSERT INTO template_usage_log (template_id, action, result)
                    VALUES (:template_id, 'delete', :result)
                """),
                {'template_id': template_id, 'result': '删除模板'}
            )
            db.commit()
        
        return jsonify({
            "success": True,
            "message": "模板删除成功"
        })
        
    except Exception as e:
        logger.error(f"删除模板失败：{e}")
        return jsonify({
            "success": False,
            "message": f"删除失败：{str(e)}"
        }), 500


@excel_templates_bp.route("/templates/by_type", methods=["GET"])
@swag_from({
    'summary': '按类型获取模板列表',
    'description': '按模板类型（例如 发货单）获取模板列表，可选仅返回启用模板',
    'parameters': [
        {
            'name': 'type',
            'in': 'query',
            'type': 'string',
            'description': '模板类型，例如 发货单，默认 发货单'
        },
        {
            'name': 'active_only',
            'in': 'query',
            'type': 'boolean',
            'description': '仅返回启用模板，默认 true'
        },
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'templates': {'type': 'array'},
                    'count': {'type': 'integer'},
                }
            }
        }
    }
})
def list_templates_by_type():
    """
    GET /api/excel/templates/by_type?type=发货单&active_only=true
    """
    try:
        from app.application import TemplateApplicationService, get_template_app_service

        template_type = request.args.get("type", "发货单")
        active_only = request.args.get("active_only", "true").lower() == "true"

        svc = get_template_app_service()
        templates = [_normalize_template_dto(t) for t in svc.list_by_type(template_type, active_only=active_only)]
        return jsonify({
            "success": True,
            "templates": templates,
            "count": len(templates),
        }), 200
    except Exception as e:
        logger.error(f"按类型获取模板列表失败：{e}")
        return jsonify({
            "success": False,
            "message": f"获取失败：{str(e)}"
        }), 500


@excel_templates_bp.route("/templates/default", methods=["GET"])
@swag_from({
    'summary': '获取某类型默认模板',
    'description': '获取指定类型当前默认模板的元数据（优先使用数据库模板，失败时回退到固定文件模板）',
    'parameters': [
        {
            'name': 'type',
            'in': 'query',
            'type': 'string',
            'description': '模板类型，例如 发货单，默认 发货单'
        }
    ],
    'responses': {
        '200': {
            'description': '成功响应',
            'schema': {
                'type': 'object',
                'properties': {
                    'success': {'type': 'boolean'},
                    'template': {'type': 'object'}
                }
            }
        },
        '404': {
            'description': '无可用模板'
        }
    }
})
def get_default_template():
    """
    GET /api/excel/templates/default?type=发货单
    返回当前类型默认模板的元数据。
    """
    try:
        from app.application import TemplateApplicationService, get_template_app_service

        template_type = request.args.get("type", "发货单")
        svc = get_template_app_service()
        tpl = svc.get_default_for_type(template_type)
        if not tpl:
            return jsonify({
                "success": False,
                "message": "暂无可用模板"
            }), 404
        return jsonify({
            "success": True,
            "template": _normalize_template_dto(tpl)
        }), 200
    except Exception as e:
        logger.error(f"获取默认模板失败：{e}")
        return jsonify({
            "success": False,
            "message": f"获取失败：{str(e)}"
        }), 500
