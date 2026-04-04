# -*- coding: utf-8 -*-

import base64
import json
import logging
import os
import queue
import shutil
import threading
from datetime import datetime, timezone

from flask import Blueprint, jsonify, request, Response, send_file
from werkzeug.utils import secure_filename

logger = logging.getLogger(__name__)

traditional_mode_bp = Blueprint('traditional_mode', __name__, url_prefix='/api/traditional-mode')

ROOT_DIR = r"e:\FHD\bang"

IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp', '.ico', '.tiff', '.svg'}
EXCEL_EXTENSIONS = {'.xlsx', '.xls'}

_watch_clients = []
_watch_clients_lock = threading.Lock()
_last_snapshot: dict[str, str] = {}
_snapshot_lock = threading.Lock()
_watchdog_running = False
_watchdog_thread = None


def resolve_safe_path(relative_path=""):
    safe = os.path.normpath(os.path.join(ROOT_DIR, relative_path or ""))
    if not os.path.abspath(safe).startswith(os.path.abspath(ROOT_DIR)):
        return None
    return safe


def _get_file_type(filename):
    name_lower = filename.lower()
    if os.path.splitext(name_lower)[1] in EXCEL_EXTENSIONS:
        return os.path.splitext(name_lower)[1][1:]
    if os.path.splitext(name_lower)[1] in IMAGE_EXTENSIONS:
        return os.path.splitext(name_lower)[1][1:]
    return os.path.splitext(name_lower)[1][1:] if '.' in name_lower else "文件"


def _format_time(ts):
    return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()


@traditional_mode_bp.route('/list', methods=['GET'])
def list_files():
    try:
        rel_path = request.args.get('path', '')
        full_path = resolve_safe_path(rel_path)
        if full_path is None:
            return jsonify({"success": False, "error": "路径越权访问被拒绝"}), 403
        if not os.path.exists(full_path):
            return jsonify({"success": True, "data": {"path": rel_path, "files": []}})
        if not os.path.isdir(full_path):
            return jsonify({"success": False, "error": "指定路径不是目录"}), 400

        entries = []
        dirs = []
        files = []
        for name in sorted(os.listdir(full_path)):
            entry_path = os.path.join(full_path, name)
            try:
                stat_info = os.stat(entry_path)
                is_dir = os.path.isdir(entry_path)
                entry = {
                    "name": name,
                    "is_dir": is_dir,
                    "size": 0 if is_dir else stat_info.st_size,
                    "modified_time": _format_time(stat_info.st_mtime),
                    "type": "文件夹" if is_dir else _get_file_type(name),
                }
                if is_dir:
                    dirs.append(entry)
                else:
                    files.append(entry)
            except (OSError, IOError) as e:
                logger.warning(f"无法获取文件信息: {entry_path}, 错误: {e}")
                continue

        entries = dirs + files
        return jsonify({
            "success": True,
            "data": {
                "path": rel_path,
                "files": entries,
            }
        })
    except Exception as e:
        logger.error(f"列出目录失败: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@traditional_mode_bp.route('/read', methods=['GET'])
def read_file():
    try:
        rel_file = request.args.get('file', '')
        full_path = resolve_safe_path(rel_file)
        if full_path is None:
            return jsonify({"success": False, "error": "路径越权访问被拒绝"}), 403
        if not os.path.exists(full_path):
            return jsonify({"success": False, "error": "文件不存在"}), 404
        if os.path.isdir(full_path):
            return jsonify({"success": False, "error": "指定路径是目录而非文件"}), 400

        ext = os.path.splitext(full_path)[1].lower()

        if ext in EXCEL_EXTENSIONS:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(full_path, data_only=True)
                sheets_data = {}
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    rows_data = []
                    for row in sheet.iter_rows(values_only=True):
                        row_list = []
                        for cell in row:
                            if cell is None:
                                row_list.append(None)
                            elif isinstance(cell, datetime):
                                row_list.append(cell.isoformat())
                            else:
                                row_list.append(cell)
                        rows_data.append(row_list)
                    sheets_data[sheet_name] = {
                        "rows": rows_data,
                        "max_row": sheet.max_row,
                        "max_column": sheet.max_column,
                    }
                wb.close()
                return jsonify({
                    "success": True,
                    "data": {
                        "type": "excel",
                        "sheets": list(wb.sheetnames),
                        "content": sheets_data,
                    }
                })
            except ImportError:
                return jsonify({
                    "success": False,
                    "error": "openpyxl 未安装，无法读取 Excel 文件"
                }), 500
            except Exception as e:
                logger.error(f"读取 Excel 文件失败: {e}", exc_info=True)
                return jsonify({"success": False, "error": f"读取 Excel 失败: {str(e)}"}), 500

        elif ext in IMAGE_EXTENSIONS:
            with open(full_path, 'rb') as f:
                img_base64 = base64.b64encode(f.read()).decode('utf-8')
            mime_map = {
                '.png': 'image/png',
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.gif': 'image/gif',
                '.webp': 'image/webp',
                '.bmp': 'image/bmp',
                '.ico': 'image/x-icon',
                '.tiff': 'image/tiff',
                '.svg': 'image/svg+xml',
            }
            return jsonify({
                "success": True,
                "data": {
                    "type": "image",
                    "mime": mime_map.get(ext, 'application/octet-stream'),
                    "content": img_base64,
                }
            })

        else:
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                return jsonify({
                    "success": True,
                    "data": {
                        "type": "text",
                        "content": content,
                    }
                })
            except UnicodeDecodeError:
                with open(full_path, 'rb') as f:
                    content = base64.b64encode(f.read()).decode('utf-8')
                return jsonify({
                    "success": True,
                    "data": {
                        "type": "binary",
                        "content": content,
                    }
                })

    except Exception as e:
        logger.error(f"读取文件失败: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@traditional_mode_bp.route('/write', methods=['POST'])
def write_excel():
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({"success": False, "error": "请求体为空或格式错误"}), 400

        rel_file = data.get('file', '')
        file_data = data.get('data', {})
        file_type = data.get('type', '')

        full_path = resolve_safe_path(rel_file)
        if full_path is None:
            return jsonify({"success": False, "error": "路径越权访问被拒绝"}), 403

        if file_type != 'excel':
            return jsonify({"success": False, "error": f"不支持的写入类型: {file_type}"})

        try:
            import openpyxl
        except ImportError:
            return jsonify({
                "success": False,
                "error": "openpyxl 未安装，无法写入 Excel 文件"
            }), 500

        parent_dir = os.path.dirname(full_path)
        if parent_dir and not os.path.exists(parent_dir):
            os.makedirs(parent_dir, exist_ok=True)

        wb = openpyxl.Workbook()
        default_sheet = wb.active
        default_sheet.title = file_data.get('active_sheet', 'Sheet')

        sheets_content = file_data.get('content', {})
        if isinstance(sheets_content, dict):
            for sheet_name, sheet_data_item in sheets_content.items():
                if sheet_name == default_sheet.title:
                    ws = default_sheet
                else:
                    ws = wb.create_sheet(title=sheet_name)

                rows = sheet_data_item.get('rows', []) if isinstance(sheet_data_item, dict) else []
                for r_idx, row in enumerate(rows, start=1):
                    for c_idx, cell_value in enumerate(row, start=1):
                        if cell_value is not None:
                            ws.cell(row=r_idx, column=c_idx, value=cell_value)

        if len(wb.sheetnames) > 1 and default_sheet.title in wb.sheetnames:
            if not sheets_content or default_sheet.title not in sheets_content:
                wb.remove(default_sheet)

        wb.save(full_path)
        wb.close()

        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"写入 Excel 文件失败: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@traditional_mode_bp.route('/mkdir', methods=['POST'])
def make_directory():
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({"success": False, "error": "请求体为空或格式错误"}), 400

        rel_path = data.get('path', '')
        folder_name = data.get('name', '').strip()

        if not folder_name:
            return jsonify({"success": False, "error": "文件夹名称不能为空"}), 400

        if '/' in folder_name or '\\' in folder_name or '..' in folder_name:
            return jsonify({"success": False, "error": "文件夹名称包含非法字符"}), 400

        full_parent = resolve_safe_path(rel_path)
        if full_parent is None:
            return jsonify({"success": False, "error": "路径越权访问被拒绝"}), 403

        full_new_path = os.path.join(full_parent, folder_name)
        if not os.path.abspath(full_new_path).startswith(os.path.abspath(ROOT_DIR)):
            return jsonify({"success": False, "error": "路径越权访问被拒绝"}), 403

        if os.path.exists(full_new_path):
            return jsonify({"success": False, "error": "文件夹已存在"}), 409

        os.makedirs(full_new_path, exist_ok=False)
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"创建文件夹失败: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@traditional_mode_bp.route('/rename', methods=['POST'])
def rename_item():
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({"success": False, "error": "请求体为空或格式错误"}), 400

        rel_path = data.get('path', '')
        old_name = (data.get('old_name') or '').strip()
        new_name = (data.get('new_name') or '').strip()

        if not old_name or not new_name:
            return jsonify({"success": False, "error": "旧名称和新名称不能为空"}), 400

        if '/' in new_name or '\\' in new_name or '..' in new_name:
            return jsonify({"success": False, "error": "新名称包含非法字符"}), 400

        full_parent = resolve_safe_path(rel_path)
        if full_parent is None:
            return jsonify({"success": False, "error": "路径越权访问被拒绝"}), 403

        full_old_path = os.path.join(full_parent, old_name)
        full_new_path = os.path.join(full_parent, new_name)

        if not os.path.abspath(full_old_path).startswith(os.path.abspath(ROOT_DIR)):
            return jsonify({"success": False, "error": "路径越权访问被拒绝"}), 403
        if not os.path.abspath(full_new_path).startswith(os.path.abspath(ROOT_DIR)):
            return jsonify({"success": False, "error": "路径越权访问被拒绝"}), 403

        if not os.path.exists(full_old_path):
            return jsonify({"success": False, "error": "源文件或文件夹不存在"}), 404

        if os.path.exists(full_new_path):
            return jsonify({"success": False, "error": "目标名称已存在"}), 409

        os.rename(full_old_path, full_new_path)
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"重命名失败: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@traditional_mode_bp.route('/delete', methods=['POST'])
def delete_item():
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({"success": False, "error": "请求体为空或格式错误"}), 400

        rel_path = data.get('path', '')
        name = (data.get('name') or '').strip()

        if not name:
            return jsonify({"success": False, "error": "名称不能为空"}), 400

        full_parent = resolve_safe_path(rel_path)
        if full_parent is None:
            return jsonify({"success": False, "error": "路径越权访问被拒绝"}), 403

        full_target = os.path.join(full_parent, name)

        if not os.path.abspath(full_target).startswith(os.path.abspath(ROOT_DIR)):
            return jsonify({"success": False, "error": "路径越权访问被拒绝"}), 403

        if not os.path.exists(full_target):
            return jsonify({"success": False, "error": "文件或文件夹不存在"}), 404

        if os.path.isdir(full_target):
            shutil.rmtree(full_target)
        else:
            os.remove(full_target)

        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"删除失败: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@traditional_mode_bp.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({"success": False, "error": "没有上传文件"}), 400

        uploaded = request.files['file']
        if not uploaded.filename:
            return jsonify({"success": False, "error": "文件名为空"}), 400

        rel_path = request.form.get('path', '')
        full_target_dir = resolve_safe_path(rel_path)
        if full_target_dir is None:
            return jsonify({"success": False, "error": "路径越权访问被拒绝"}), 403

        filename = secure_filename(uploaded.filename)
        if not filename:
            filename = "uploaded_file"

        if not os.path.exists(full_target_dir):
            os.makedirs(full_target_dir, exist_ok=True)

        save_path = os.path.join(full_target_dir, filename)
        if os.path.exists(save_path):
            base, ext = os.path.splitext(filename)
            counter = 1
            while os.path.exists(save_path):
                filename = f"{base}_{counter}{ext}"
                save_path = os.path.join(full_target_dir, filename)
                counter += 1

        uploaded.save(save_path)
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"上传文件失败: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


def _build_snapshot():
    snap: dict[str, float] = {}
    if os.path.exists(ROOT_DIR) and os.path.isdir(ROOT_DIR):
        try:
            for dirpath, _dirnames, filenames in os.walk(ROOT_DIR):
                try:
                    rel = os.path.relpath(dirpath, ROOT_DIR)
                    prefix = '' if rel == '.' else rel.replace(os.sep, '/') + '/'
                    for fname in filenames:
                        fpath = os.path.join(dirpath, fname)
                        try:
                            snap[prefix + fname] = os.stat(fpath).st_mtime
                        except (OSError, IOError):
                            continue
                except OSError:
                    continue
        except OSError:
            pass
    return snap

def _format_snapshot(mtime_map: dict[str, float]):
    return {name: _format_time(ts) for name, ts in mtime_map.items()}

_watchdog_prev: dict[str, float] = {}
_watchdog_changed_callbacks: list[queue.Queue[str]] = []

def _watchdog_loop():
    global _last_snapshot, _watchdog_running, _watchdog_prev
    _watchdog_running = True
    try:
        _watchdog_prev = _build_snapshot()
        while _watchdog_running:
            threading.Event().wait(3.0)
            if not _watchdog_running:
                break
            curr = _build_snapshot()
            changed = []
            for fname, mtime in curr.items():
                if _watchdog_prev.get(fname) != mtime:
                    changed.append(fname)
            for fname in list(_watchdog_prev.keys()):
                if fname not in curr:
                    changed.append(f"__deleted__:{fname}")
            _watchdog_prev = curr
            if changed:
                snap_str = _format_snapshot(curr)
                with _snapshot_lock:
                    _last_snapshot = snap_str
                payload = json.dumps({"changed": changed}, ensure_ascii=False)
                dead = []
                with _watch_clients_lock:
                    for q in _watch_clients:
                        try:
                            q.put_nowait(payload)
                        except queue.Full:
                            dead.append(q)
                for q in dead:
                    with _watch_clients_lock:
                        if q in _watch_clients:
                            _watch_clients.remove(q)
    except Exception as e:
        logger.error(f"Watchdog 线程异常: {e}", exc_info=True)
    finally:
        _watchdog_running = False

def _ensure_watchdog():
    global _watchdog_thread, _watchdog_running
    if not _watchdog_running or (_watchdog_thread and not _watchdog_thread.is_alive()):
        _watchdog_thread = threading.Thread(target=_watchdog_loop, daemon=True, name="fs-watchdog")
        _watchdog_thread.start()

@traditional_mode_bp.route('/watch', methods=['GET'])
def sse_watch():
    rel_path = request.args.get('path', '')
    full_path = resolve_safe_path(rel_path)
    if full_path is None:
        return jsonify({"success": False, "error": "路径越权访问被拒绝"}), 403

    _ensure_watchdog()

    client_q: queue.Queue[str] = queue.Queue(maxsize=20)
    with _watch_clients_lock:
        _watch_clients.append(client_q)

    def generate():
        try:
            with _snapshot_lock:
                initial = json.dumps({"changed": [], "snapshot": dict(_last_snapshot)}, ensure_ascii=False)
            yield f"data: {initial}\n\n"
            while True:
                try:
                    msg = client_q.get(timeout=55)
                    yield f"data: {msg}\n\n"
                except queue.Empty:
                    yield ": heartbeat\n\n"
        except GeneratorExit:
            pass
        finally:
            with _watch_clients_lock:
                if client_q in _watch_clients:
                    _watch_clients.remove(client_q)

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


def register_traditional_mode_routes(app):
    app.register_blueprint(traditional_mode_bp)
