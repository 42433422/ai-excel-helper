"""检查当前输出文件"""
from openpyxl import load_workbook
import os

output_file = r'e:\FHD\424\考勤转换输出.xlsx'

if os.path.exists(output_file):
    wb = load_workbook(output_file)
    print(f'工作表：{wb.sheetnames}')

    if "明细" not in wb.sheetnames:
        print('\n问题确认：前端没有传递 template_relpath 参数！')
        print('输出文件只有 "月度统计" 和 "钉钉解析" 两个工作表')
        print('这说明前端没有使用模板功能')
    else:
        print('\n成功使用模板')
else:
    print('文件不存在')
