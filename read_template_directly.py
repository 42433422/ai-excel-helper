"""直接读取模板文件"""
import sys
from pathlib import Path
from openpyxl import load_workbook
import shutil

# 由于可能存在文件锁定，先复制文件再读取
src = Path(r'e:\FHD\424\考勤 -2026-3 月份考勤统计表.xlsx')
dst = Path(r'e:\FHD\424\temp_copy.xlsx')

if not src.exists():
    print(f"源文件不存在：{src}")
    # 尝试列出目录
    print(f"\n424 目录内容:")
    for f in Path(r'e:\FHD\424').iterdir():
        if f.is_file() and '考勤' in f.name:
            print(f"  {f.name} (size: {f.stat().st_size})")
    sys.exit(1)

print(f"源文件存在：{src}")
print(f"文件大小：{src.stat().st_size}")

# 复制文件
try:
    shutil.copy2(str(src), str(dst))
    print(f"已复制到：{dst}")
    
    # 读取复制的文件
    wb = load_workbook(str(dst))
    print(f"\n工作表：{wb.sheetnames}")
    
    for i, ws_name in enumerate(wb.sheetnames):
        ws = wb[ws_name]
        merged_count = len(list(ws.merged_cells.ranges)) if ws.merged_cells.ranges else 0
        print(f"\n[{i+1}] {ws_name}")
        print(f"  最大行：{ws.max_row}, 最大列：{ws.max_column}")
        print(f"  合并单元格：{merged_count} 个")
        if merged_count > 0:
            for m in ws.merged_cells.ranges:
                print(f"    {m}")
        
        # 显示前 3 行
        print(f"  前 3 行:")
        for row_idx in range(1, min(4, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(8, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(val)
            print(f"    行{row_idx}: {row_data}")
    
    wb.close()
    
except Exception as e:
    print(f"错误：{e}")
    import traceback
    traceback.print_exc()
finally:
    # 清理临时文件
    if dst.exists():
        dst.unlink()
        print(f"\n已清理临时文件：{dst}")
