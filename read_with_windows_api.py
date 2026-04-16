"""使用 Windows API 读取文件"""
import os
import sys
from pathlib import Path, PurePath

# 使用 Windows 的 UNC 路径格式
file_name = "考勤 -2026-3 月份考勤统计表.xlsx"
dir_path = r"e:\FHD\424"

# 列出目录中所有包含"考勤"的文件
print("目录中的文件:")
for f in os.listdir(dir_path):
    if '考勤' in f:
        full_path = os.path.join(dir_path, f)
        print(f"  {f}")
        print(f"    完整路径：{full_path}")
        print(f"    是文件：{os.path.isfile(full_path)}")
        print(f"    大小：{os.path.getsize(full_path)}")
        
        # 尝试用 openpyxl 读取
        if f.endswith('.xlsx') and os.path.isfile(full_path):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(full_path)
                print(f"    工作表：{wb.sheetnames}")
                for ws_name in wb.sheetnames:
                    ws = wb[ws_name]
                    merged = list(ws.merged_cells.ranges) if ws.merged_cells.ranges else []
                    print(f"      {ws_name}: {len(merged)} 个合并单元格")
                    if merged:
                        for m in merged[:3]:
                            print(f"        {m}")
                wb.close()
                print(f"    ✓ 读取成功")
            except Exception as e:
                print(f"    ✗ 读取失败：{e}")
            print()
