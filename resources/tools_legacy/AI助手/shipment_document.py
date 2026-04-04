#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
发货单文档生成器
使用尹玉华1.xlsx模板生成发货单

DEPRECATED:
- 该模块仅作为文档生成底层实现（由 XCAGI 基础设施适配层调用）。
- 编号模式解析/编排已迁移至 XCAGI 内部服务，不应再从路由直接动态加载本模块。
"""

import sys
import os
import re
import sqlite3
from datetime import datetime
from typing import Dict, List
from dataclasses import dataclass
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import logging
import tempfile
import zipfile
import xml.etree.ElementTree as ET

# 配置日志编码
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', encoding='utf-8')
logger = logging.getLogger(__name__)


def _load_workbook_with_repair(template_path: str):
    """
    尝试修复受损 xlsx 后再加载。
    - 清洗所有 xl/*.xml 中的非法控制字符
    - 修复裸 '&' 导致的 XML 非法实体
    - 若仍失败，降级为空工作簿，避免流程直接中断
    """
    def _sanitize_xml_bytes(raw: bytes) -> bytes:
        text = raw.decode("utf-8", errors="ignore")
        # 去掉 XML 1.0 非法控制字符（保留 \t \n \r）
        text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)
        # 修复未转义的 '&'
        text = re.sub(r"&(?!amp;|lt;|gt;|apos;|quot;|#\d+;|#x[0-9A-Fa-f]+;)", "&amp;", text)
        return text.encode("utf-8")

    def _is_numeric_token(value: str) -> bool:
        return bool(re.fullmatch(r"[+-]?\d+(?:\.\d+)?", (value or "").strip()))

    def _repair_worksheet_filters(xml_bytes: bytes, filename: str) -> tuple[bytes, int]:
        """移除 openpyxl 无法解析的 customFilter 条件。"""
        try:
            root = ET.fromstring(xml_bytes)
        except Exception:
            return xml_bytes, 0

        ns_uri = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        ns = {"m": ns_uri}
        removed = 0

        # 查找所有 customFilters/customFilter
        for custom_filters in root.findall(".//m:autoFilter/m:filterColumn/m:customFilters", ns):
            to_remove = []
            for cf in list(custom_filters):
                if not cf.tag.endswith("customFilter"):
                    continue
                val = (cf.attrib.get("val") or "").strip()
                if not val:
                    continue
                operator = (cf.attrib.get("operator") or "equal").strip().lower()
                # openpyxl 对 equal 的字符串要求包含通配符；否则抛 ValueError
                if operator == "equal" and not _is_numeric_token(val) and "*" not in val and "?" not in val:
                    to_remove.append(cf)
            for cf in to_remove:
                custom_filters.remove(cf)
                removed += 1

            # 如果 customFilters 被清空，则连同父节点 filterColumn 一并移除，避免空筛选结构
            if len(list(custom_filters)) == 0:
                for filter_column in root.findall(".//m:autoFilter/m:filterColumn", ns):
                    if custom_filters in list(filter_column):
                        filter_column.remove(custom_filters)
                        break

        if removed == 0:
            return xml_bytes, 0

        repaired = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        logger.warning(f"模板筛选修复: {filename} 移除非法 customFilter {removed} 条")
        return repaired, removed

    try:
        # 先直接尝试一次，避免不必要的修复开销
        return openpyxl.load_workbook(template_path, data_only=True, keep_vba=False)
    except Exception as e:
        logger.warning(f"直接加载模板失败，进入修复流程: {e}")

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        total_removed_filters = 0
        with zipfile.ZipFile(template_path, "r") as zin, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith("xl/") and item.filename.endswith(".xml"):
                    if item.filename.startswith("xl/worksheets/"):
                        data, removed = _repair_worksheet_filters(data, item.filename)
                        total_removed_filters += removed

        if total_removed_filters > 0:
            logger.warning(f"模板筛选修复完成: 共移除非法 customFilter {total_removed_filters} 条")

        try:
            return openpyxl.load_workbook(tmp_path, data_only=True, keep_vba=False)
        except Exception as e:
            logger.error(f"修复后仍无法加载模板，降级为默认工作簿: {e}")
            return openpyxl.Workbook()
    except Exception as e:
        logger.error(f"模板修复流程失败，降级为默认工作簿: {e}")
        return openpyxl.Workbook()
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


@dataclass
class PurchaseUnitInfo:
    """购买单位信息"""
    name: str = ""
    contact_person: str = ""
    contact_phone: str = ""
    address: str = ""
    id: int = None


@dataclass
class ShipmentDocument:
    """发货单文档"""
    filename: str = ""
    filepath: str = ""
    order_number: str = ""
    purchase_unit: PurchaseUnitInfo = None
    products: List[Dict] = None
    total_amount: float = 0.0
    total_quantity: float = 0.0
    
    def to_dict(self) -> dict:
        return {
            "filename": self.filename,
            "filepath": self.filepath,
            "order_number": self.order_number,
            "purchase_unit": self.purchase_unit.__dict__ if self.purchase_unit else None,
            "products": self.products,
            "total_amount": self.total_amount,
            "total_quantity": self.total_quantity
        }


class ShipmentDocumentGenerator:
    """发货单文档生成器"""
    
    # 使用绝对路径，确保无论从哪个目录启动都能正确找到文件
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    TEMPLATE_FOLDER = SCRIPT_DIR
    OUTPUT_FOLDER = os.path.join(SCRIPT_DIR, 'outputs')
    DEFAULT_TEMPLATE = "尹玉华1.xlsx"
    
    def __init__(self, db_path: str = None):
        """初始化"""
        # 获取脚本所在目录
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        logger.info(f"发货单生成器运行目录: {self.base_dir}")
        
        # 使用当前目录中的数据库文件
        if db_path:
            self.db_path = db_path
        else:
            # 使用当前目录的products.db
            self.db_path = os.path.join(self.base_dir, 'products.db')
        
        logger.info(f"数据库路径: {self.db_path}")
        os.makedirs(self.OUTPUT_FOLDER, exist_ok=True)
        
    def _find_template(self, template_name):
        """查找模板文件"""
        import os
        
        # 尝试多个可能的模板目录（使用绝对路径）
        possible_dirs = [
            # 相对于AI助手目录的templates目录
            os.path.join(self.base_dir, 'templates'),
            # 当前目录
            self.base_dir,
            # 相对于根目录的路径
            os.path.join(self.base_dir, '..', 'AI助手', 'templates')
        ]
        
        for directory in possible_dirs:
            template_path = os.path.join(directory, template_name)
            if os.path.exists(template_path):
                logger.info(f"找到模板文件: {template_path}")
                return template_path
        
        # 如果都找不到，尝试在多个位置查找
        root_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(self.base_dir))))
        all_dirs = [
            os.path.join(root_dir, 'AI助手', 'templates'),
            os.path.join(root_dir, 'AI助手'),
            os.path.join(root_dir, 'uploads'),
            self.base_dir
        ]
        
        for directory in all_dirs:
            template_path = os.path.join(directory, template_name)
            if os.path.exists(template_path):
                logger.info(f"在根目录查找找到模板文件: {template_path}")
                return template_path
        
        logger.warning(f"模板文件未找到: {template_name}")
        return None
    
    def generate_document(
        self,
        order_text: str,
        parsed_data: Dict,
        purchase_unit: PurchaseUnitInfo = None,
        template_name: str = None,
        custom_order_number: str = None,
    ) -> ShipmentDocument:
        """
        生成发货单文档
        """
        try:
            # 1. 如果没有指定购买单位信息，从数据库获取
            if purchase_unit is None and parsed_data.get("purchase_unit"):
                purchase_unit = self._get_purchase_unit_info(parsed_data["purchase_unit"])
            
            # 2. 生成订单编号
            order_number = (custom_order_number or "").strip() or self._generate_order_number()
            
            # 3. 准备产品数据
            products = self._prepare_products(parsed_data, order_number)
            
            # 4. 计算总计
            total_amount = sum(p.get("amount", 0) for p in products)
            total_quantity_kg = sum(p.get("quantity_kg", 0) for p in products)
            total_quantity_tins = sum(p.get("quantity_tins", 0) for p in products)
            
            # 5. 使用模板文件
            template_file = template_name or self.DEFAULT_TEMPLATE
            template_path = self._find_template(template_file)
            
            if template_path:
                logger.info(f"找到模板文件: {template_path}")
                try:
                    workbook = openpyxl.load_workbook(template_path)
                except Exception as e:
                    logger.warning(f"正常加载模板失败，尝试容错模式: {e}")
                    try:
                        workbook = openpyxl.load_workbook(template_path, data_only=True, keep_vba=False)
                    except Exception as e2:
                        logger.warning(f"容错模式也失败，尝试修复XML后加载: {e2}")
                        workbook = _load_workbook_with_repair(template_path)
                worksheet = workbook.active
                
                saved_formulas = self._clear_template_data(worksheet)
                
                self._fill_from_template(
                    worksheet,
                    order_number=order_number,
                    purchase_unit=purchase_unit,
                    products=products,
                    total_quantity_kg=total_quantity_kg,
                    total_quantity_tins=total_quantity_tins,
                    total_amount=total_amount,
                    order_text=order_text
                )
            else:
                logger.warning(f"模板文件不存在: {template_file}，创建新文档")
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = "发货单"
                self._create_default_document(
                    worksheet, order_number, purchase_unit, products,
                    total_quantity_kg, total_quantity_tins, total_amount, order_text
                )
            
            # 6. 保存文档
            output_filename = self._generate_output_filename(order_number)
            output_path = os.path.join(self.OUTPUT_FOLDER, output_filename)
            workbook.save(output_path)
            workbook.close()
            
            # 7. 创建文档对象
            doc = ShipmentDocument(
                filename=output_filename,
                filepath=output_path,
                order_number=order_number,
                purchase_unit=purchase_unit,
                products=products,
                total_amount=total_amount,
                total_quantity=total_quantity_kg
            )
            
            logger.info(f"发货单生成成功: {output_filename}")
            return doc
            
        except Exception as e:
            logger.error(f"生成发货单失败: {e}")
            raise
    
    def _clear_template_data(self, worksheet):
        """清空模板中的示例数据"""
        for row in range(4, 16):
            for col in range(1, 11):
                try:
                    cell = worksheet.cell(row=row, column=col)
                    cell.value = None
                except:
                    pass
    
    def _fill_from_template(
        self,
        worksheet,
        order_number: str,
        purchase_unit: PurchaseUnitInfo,
        products: List[Dict],
        total_quantity_kg: float,
        total_quantity_tins: int,
        total_amount: float,
        order_text: str
    ):
        """使用模板填充数据 - 匹配尹玉华1.xlsx原模板格式"""
        # 手动格式化日期以避免locale编码问题
        now = datetime.now()
        year = now.year
        month = now.month
        day = now.day
        today = f"{year}年{month}月{day}日"
        
        # 第2行：购货单位、联系人、日期、订单编号（合并单元格A2:J2）
        header_cell = worksheet.cell(row=2, column=1)
        contact = purchase_unit.contact_person if purchase_unit and purchase_unit.contact_person else ""
        unit_name = purchase_unit.name if purchase_unit and purchase_unit.name else "未指定单位"
        header_info = f"购货单位：{unit_name}       联系人：{contact}              {today}      订单编号：{order_number}"
        header_cell.value = header_info
        logger.info(f"填写购买单位信息: {unit_name}, 联系人: {contact}")
        
        # 填充产品数据（从第4行开始）
        start_row = 4
        for i, product in enumerate(products):
            row = start_row + i
            
            # 产品型号 (A列, A-C合并)
            worksheet.cell(row=row, column=1, value=product.get("model_number", ""))
            
            # 产品名称 (D列)
            worksheet.cell(row=row, column=4, value=product.get("name", ""))
            
            # 数量/件 (E列) - 桶数
            worksheet.cell(row=row, column=5, value=product.get("quantity_tins", 0))
            
            # 规格/KG (F列) - 每桶的KG数
            worksheet.cell(row=row, column=6, value=product.get("tin_spec", 0))
            
            # 数量/KG (G列) - 直接填入数值
            kg_value = float(product.get("quantity_kg", 0))
            if kg_value > 0:
                worksheet.cell(row=row, column=7, value=kg_value)
            
            # 单价/元 (H列)
            unit_price = float(product.get("unit_price", 0))
            if unit_price > 0:
                worksheet.cell(row=row, column=8, value=unit_price)
            
            # 金额/元 (I列) - 直接填入数值
            amount = float(product.get("amount", 0))
            if amount > 0:
                worksheet.cell(row=row, column=9, value=amount)
            elif kg_value > 0 and unit_price > 0:
                worksheet.cell(row=row, column=9, value=round(kg_value * unit_price, 2))
            
            # 备注 (J列) - 保持空白
        
        # 第15行：合计行的数量/件 (E15公式保持SUM(E4:E14))
        if total_quantity_tins > 0:
            worksheet.cell(row=15, column=5, value=total_quantity_tins)
        
        # 第16行：金额区域
        # 大写人民币在A16，金额在I16（保持公式SUM(I4:I14)）
        if total_amount > 0:
            worksheet.cell(row=16, column=4, value=total_amount)
    
    def _create_default_document(
        self,
        worksheet,
        order_number: str,
        purchase_unit: PurchaseUnitInfo,
        products: List[Dict],
        total_quantity_kg: float,
        total_quantity_tins: int,
        total_amount: float,
        order_text: str
    ):
        """创建默认文档格式"""
        # 手动格式化日期以避免locale编码问题
        now = datetime.now()
        year = now.year
        month = now.month
        day = now.day
        today = f"{year}年{month}月{day}日"
        
        # 设置列宽
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 25
        worksheet.column_dimensions['E'].width = 12
        worksheet.column_dimensions['F'].width = 12
        worksheet.column_dimensions['G'].width = 12
        
        # 标题
        worksheet['A1'] = "成都国圣工业有限公司 送货单"
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A1:G1')
        
        # 第2行
        unit_name = purchase_unit.name if purchase_unit and purchase_unit.name else "未指定单位"
        contact = purchase_unit.contact_person if purchase_unit and purchase_unit.contact_person else ""
        header_info = f"购货单位：{unit_name}       联系人：{contact}              {today}      订单编号：{order_number}"
        worksheet['A2'] = header_info
        worksheet.merge_cells('A2:J2')
        
        # 表头
        headers = ["产品型号", "", "", "产品名称", "数量/件", "规格/KG", "数量/KG"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 合并产品型号列
        for i in range(10):
            worksheet.merge_cells(f'A{4+i}:C{4+i}')
        
        # 填充产品
        for i, product in enumerate(products):
            row = 5 + i
            worksheet.cell(row=row, column=1, value=product.get("model_number", ""))
            worksheet.merge_cells(f'A{row}:C{row}')
            worksheet.cell(row=row, column=4, value=product.get("name", ""))
            worksheet.cell(row=row, column=5, value=product.get("quantity_tins", 0))
            worksheet.cell(row=row, column=6, value=product.get("tin_spec", 0))
            worksheet.cell(row=row, column=7, value=product.get("quantity_kg", 0))
        
        # 合计
        total_row = 5 + len(products)
        worksheet.cell(row=total_row, column=1, value="合计")
        worksheet.merge_cells(f'A{total_row}:C{total_row}')
        worksheet.cell(row=total_row, column=5, value=total_quantity_tins)
        worksheet.cell(row=total_row, column=7, value=total_quantity_kg)
        
        # 金额
        worksheet.cell(row=total_row + 1, column=1, value="大写人民币：")
        worksheet.cell(row=total_row + 1, column=4, value=total_amount)
        
        # 备注
        worksheet.cell(row=total_row + 3, column=1, value=f"备注：{order_text}")
        
        # 添加边框
        self._add_borders(worksheet, total_row + 5)
    
    def _add_borders(self, worksheet, max_row: int):
        """添加边框"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, max_row + 1):
            for col in range(1, 8):
                try:
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                except:
                    pass
    
    def _get_purchase_unit_info(self, unit_name: str) -> PurchaseUnitInfo:
        """获取购买单位信息"""
        unit = PurchaseUnitInfo(name=unit_name)
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # 直接查询purchase_units表（正确的表结构）
            cursor.execute('''
                SELECT id, unit_name, contact_person, contact_phone, address
                FROM purchase_units
                WHERE unit_name = ?
                LIMIT 1
            ''', (unit_name,))
            
            row = cursor.fetchone()
            
            # 如果精确匹配失败，尝试模糊匹配
            if not row:
                cursor.execute('''
                    SELECT id, unit_name, contact_person, contact_phone, address
                    FROM purchase_units
                    WHERE unit_name LIKE ?
                    LIMIT 1
                ''', (f'%{unit_name}%',))
                row = cursor.fetchone()
            
            if row:
                unit.id = row[0]
                unit.name = row[1]
                unit.contact_person = row[2] or ""
                unit.contact_phone = row[3] or ""
                unit.address = row[4] or ""
                logger.info(f"成功获取购买单位信息: {unit.name}, 联系人: {unit.contact_person}")
            else:
                logger.warning(f"未找到购买单位信息: {unit_name}")
            
            conn.close()
            
        except Exception as e:
            logger.error(f"获取购买单位信息失败: {e}")
        
        return unit
    
    def _generate_order_number(self) -> str:
        """生成订单编号 - 格式: 26-010003A (年-月-日(00)-序号)"""
        now = datetime.now()
        
        # 格式: 26-月日序号+数量编号
        # 例如: 26-010003A = 2026年1月第3个订单数量A
        year = now.strftime("%y")  # 26
        month = now.strftime("%m")  # 01
        day = "00"  # 日固定为00
        
        # 获取当天订单数量作为序号基数
        sequence = self._get_daily_sequence(now)
        
        # 根据订单数量确定数量编号
        quantity_code = self._get_quantity_code()
        
        return f"{year}-{month}{day}{sequence:03d}{quantity_code}"
    
    def _get_daily_sequence(self, date: datetime) -> int:
        """获取当天订单序号"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            date_str = date.strftime("%Y-%m-%d")
            cursor.execute('''
                SELECT COUNT(*) FROM shipment_records
                WHERE DATE(created_at) = ?
            ''', (date_str,))
            
            count = cursor.fetchone()[0]
            conn.close()
            
            return count + 1
            
        except Exception as e:
            logger.error(f"获取订单序号失败: {e}")
            return 1
    
    def _get_quantity_code(self) -> str:
        """获取数量编号 (A=小, B=中, C=大, D=特大)"""
        return "A"  # 默认A，后续可根据实际数量调整
    
    def _prepare_products(self, parsed_data: Dict, order_number: str) -> List[Dict]:
        """准备产品数据 - 支持多产品"""
        products = []
        
        # 从parsed_data中获取产品列表
        product_list = parsed_data.get("products", [])
        
        # 如果有products列表，遍历处理
        if product_list:
            for p in product_list:
                product = {
                    "order_number": order_number,
                    "model_number": p.get("model_number", ""),
                    "name": p.get("name", ""),
                    "quantity_kg": p.get("quantity_kg", 0),
                    "quantity_tins": p.get("quantity_tins", 0),
                    "tin_spec": p.get("tin_spec", 10.0),
                    "unit_price": p.get("unit_price", 0),
                    "amount": p.get("amount", 0)
                }
                if product["name"]:
                    products.append(product)
        else:
            # 兼容单产品旧逻辑
            product = {
                "order_number": order_number,
                "model_number": parsed_data.get("model_number", ""),
                "name": parsed_data.get("product_name", ""),
                "quantity_kg": parsed_data.get("quantity_kg", 0),
                "quantity_tins": parsed_data.get("quantity_tins", 0),
                "tin_spec": parsed_data.get("tin_spec", 10.0),
                "unit_price": parsed_data.get("unit_price", 0),
                "amount": parsed_data.get("amount", 0)
            }
            
            if product["name"]:
                products.append(product)
        
        return products
    
    def _generate_output_filename(self, order_number: str) -> str:
        """生成输出文件名"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"发货单_{order_number}_{timestamp}.xlsx"
    
    def get_template_list(self) -> List[str]:
        """获取可用模板列表"""
        templates = []
        
        for filename in os.listdir('.'):
            if filename.endswith('.xlsx') and not filename.startswith('~'):
                templates.append(filename)
        
        return templates


class DocumentAPIGenerator:
    """文档API生成器"""

    def __init__(self, db_path: str = None):
        self.generator = ShipmentDocumentGenerator(db_path)
        self.excel_sync_manager = None
        self.excel_sync_enabled = False

    def enable_excel_sync(self, excel_path: str, worksheet_name: str = "25出货"):
        """启用Excel出货记录同步"""
        try:
            from shipment_excel_sync import FixedTemplateSyncManager
            self.excel_sync_manager = FixedTemplateSyncManager(excel_path, worksheet_name)
            self.excel_sync_enabled = True
            logger.info(f"已启用Excel同步: {excel_path}, 工作表: {worksheet_name}")
            return True
        except Exception as e:
            logger.error(f"启用Excel同步失败: {e}")
            return False

    def disable_excel_sync(self):
        """禁用Excel出货记录同步"""
        self.excel_sync_manager = None
        self.excel_sync_enabled = False
        logger.info("已禁用Excel同步")

    def sync_to_excel(self, doc, parsed_data: Dict) -> bool:
        """同步发货数据到Excel出货记录"""
        if not self.excel_sync_enabled or not self.excel_sync_manager:
            logger.warning("Excel同步未启用或管理器未初始化")
            return False

        try:
            # 确保doc是字典
            if hasattr(doc, 'to_dict'):
                doc = doc.to_dict()
            
            # 提取订单信息
            order_no = doc.get('order_number', '')
            order_date = doc.get('order_date', '')
            if not order_date:
                order_date = datetime.now().strftime('%Y-%m-%d')
            
            # 提取产品列表
            products = parsed_data.get('products', [])
            
            # 兼容单产品结构
            if not products:
                # 尝试从旧格式提取产品
                product_name = parsed_data.get('product_name', '')
                if product_name:
                    products = [{
                        'name': product_name,
                        'model_number': parsed_data.get('model_number', ''),
                        'quantity_tins': parsed_data.get('quantity_tins', 0),
                        'tin_spec': parsed_data.get('tin_spec', 10),
                        'unit_price': parsed_data.get('unit_price', 0)
                    }]
                else:
                    logger.warning("未找到产品数据")
                    return False
            
            logger.info(f"准备同步 {len(products)} 个产品到Excel")
            
            # 提取单价映射
            unit_price_map = {}
            for product in products:
                name = product.get('name', '')
                price = product.get('unit_price', 0)
                if name and price:
                    unit_price_map[name] = price
            
            # 调用同步
            success = self.excel_sync_manager.insert_records(order_no, order_date, products, unit_price_map)
            logger.info(f"Excel同步结果: {success}")
            return success
            
        except Exception as e:
            logger.error(f"同步到Excel失败: {e}", exc_info=True)
            return False

    def parse_and_generate(
        self,
        order_text: str,
        custom_mode: bool = False,
        number_mode: bool = False,
        generate_labels: bool = False,
        enable_excel_sync: bool = None,
        custom_order_number: str = None,
    ) -> Dict:
        """
        解析订单并生成文档
        :param order_text: 订单文本
        :param custom_mode: 是否使用自定义模式
        :param number_mode: 是否使用编号模式
        :param generate_labels: 是否生成标签
        :param enable_excel_sync: 是否启用Excel同步（None则使用默认设置）
        :return: 生成结果
        """
        try:
            # 优先使用同目录 AI 增强解析器
            from ai_augmented_parser import AIAugmentedShipmentParser
            parser = AIAugmentedShipmentParser(self.generator.db_path)
        except ImportError:
            try:
                # 回退到同目录传统解析器
                from shipment_parser import ShipmentParser
                parser = ShipmentParser(self.generator.db_path)
            except ImportError:
                # 最终兜底：使用 XCAGI 内部解析函数，避免依赖外部 98k 目录
                logger.warning("未找到本地解析器模块，回退到 XCAGI 内部解析器")
                from app.routes.tools import _parse_order_text

                class _FallbackParsedOrder:
                    def __init__(self, raw_text: str, parsed: Dict):
                        self.raw_text = raw_text
                        self.purchase_unit = str((parsed or {}).get("unit_name") or "")
                        self.products = list((parsed or {}).get("products") or [])
                        self.parsed_data = dict(parsed or {})

                    def is_valid(self) -> bool:
                        return bool(self.products)

                    def to_dict(self) -> Dict:
                        return {
                            "purchase_unit": self.purchase_unit,
                            "products": self.products,
                            "raw_text": self.raw_text,
                            "parsed_data": self.parsed_data,
                            "product_name": (self.products[0].get("name", "") if self.products else ""),
                            "model_number": (self.products[0].get("model_number", "") if self.products else ""),
                            "quantity_kg": sum(float(p.get("quantity_kg", 0) or 0) for p in self.products),
                            "quantity_tins": sum(int(p.get("quantity_tins", 0) or 0) for p in self.products),
                            "tin_spec": (float(self.products[0].get("tin_spec", 0) or 0) if self.products else 0.0),
                            "unit_price": (float(self.products[0].get("unit_price", 0) or 0) if self.products else 0.0),
                            "amount": sum(float(p.get("amount", 0) or 0) for p in self.products),
                        }

                class _FallbackParser:
                    def parse(self, text: str, custom_mode: bool = False, number_mode: bool = False):
                        parsed = _parse_order_text(text) or {}
                        if not parsed.get("success"):
                            return _FallbackParsedOrder(text, {"unit_name": "", "products": []})
                        return _FallbackParsedOrder(text, parsed)

                parser = _FallbackParser()

        # 1. 解析订单
        parsed_order = parser.parse(order_text, custom_mode, number_mode)

        if not parsed_order.is_valid():
            return {
                "success": False,
                "message": "无法解析订单信息",
                "parsed_data": parsed_order.to_dict()
            }

        # 2. 获取购买单位信息
        purchase_unit = None
        if parsed_order.purchase_unit:
            purchase_unit = self.generator._get_purchase_unit_info(parsed_order.purchase_unit)

        # 3. 生成文档（使用尹玉华1.xlsx模板）
        try:
            doc = self.generator.generate_document(
                order_text=order_text,
                parsed_data=parsed_order.to_dict(),
                purchase_unit=purchase_unit,
                template_name="尹玉华1.xlsx",
                custom_order_number=custom_order_number,
            )

            # 4. 创建发货记录
            from shipment_parser import ShipmentRecordManager
            manager = ShipmentRecordManager(self.generator.db_path)
            record_id = manager.create_record(parsed_order, purchase_unit.id if purchase_unit else None)

            # 5. 保存订单到数据库（新增）
            order_info = None
            try:
                from order_manager import OrderManager
                order_manager = OrderManager(self.generator.db_path)
                
                # 准备订单数据
                order_data = {
                    'order_number': doc.order_number,
                    'purchase_unit': parsed_order.purchase_unit,
                    'unit_id': purchase_unit.id if purchase_unit else None,
                    'products': parsed_order.products,
                    'raw_text': order_text,
                    'parsed_data': parsed_order.to_dict()
                }
                
                logger.info(f"准备保存订单数据: {order_data}")
                
                # 保存订单
                order_info = order_manager.create_order(order_data)
                logger.info(f"✅ 订单保存成功: {doc.order_number}, 返回数据: {order_info}")
                
            except Exception as e:
                logger.error(f"保存订单到数据库失败: {e}", exc_info=True)
                order_info = None

            # 6. 同步到Excel出货记录（如果已启用）
            excel_sync_success = False
            should_sync = self.excel_sync_enabled if enable_excel_sync is None else enable_excel_sync
            logger.info(f"Excel同步检查: enable_excel_sync参数={enable_excel_sync}, self.excel_sync_enabled={self.excel_sync_enabled}, should_sync={should_sync}, manager存在={self.excel_sync_manager is not None}")
            if should_sync and self.excel_sync_manager:
                logger.info("🚀 开始同步到Excel出货记录...")
                excel_sync_success = self.sync_to_excel(doc, parsed_order.to_dict())
                logger.info(f"Excel同步结果: {excel_sync_success}")

            result = {
                "success": True,
                "message": "发货单生成成功",
                "record_id": record_id,
                "order_id": order_info.get('order_id') if order_info else None,
                "parsed_data": parsed_order.to_dict(),
                "document": doc.to_dict(),
                "purchase_unit": purchase_unit.__dict__ if purchase_unit else None
            }

            # 添加Excel同步信息
            if self.excel_sync_enabled:
                result["excel_sync"] = {
                    "enabled": True,
                    "success": excel_sync_success
                }

            return result

        except Exception as e:
            logger.error(f"生成发货单失败: {e}")
            return {
                "success": False,
                "message": f"生成发货单失败: {str(e)}",
                "parsed_data": parsed_order.to_dict()
            }


if __name__ == '__main__':
    # 测试文档生成
    api_gen = DocumentAPIGenerator()
    
    test_orders = [
        "蕊芯PU哑光黑面漆20公斤",
        "蕊芯 PU净味面漆稀释剂 180KG",
    ]
    
    print("=== 发货单生成测试（使用尹玉华1.xlsx模板）===\n")
    
    for order_text in test_orders:
        print(f"订单: {order_text}")
        result = api_gen.parse_and_generate(order_text)
        
        if result["success"]:
            print(f"生成成功!")
            print(f"   文档: {result['document']['filename']}")
            print(f"   购买单位: {result['purchase_unit']['name'] if result['purchase_unit'] else '未知'}")
            print(f"   产品: {result['parsed_data']['product_name']}")
            print(f"   数量: {result['parsed_data']['quantity_kg']}kg = {result['parsed_data']['quantity_tins']}桶")
        else:
            print(f"失败: {result['message']}")
        print("-" * 60)