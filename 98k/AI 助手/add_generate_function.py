# 添加 generate_from_template 函数到 excel_template_analyzer.py

function_to_add = '''

def generate_from_template(template_file: str, output_file: str, data: dict, editable_config: dict = None):
    """
    根据模板和数据生成新的 Excel 文件
    
    Args:
        template_file: 原始模板文件路径
        output_file: 输出文件路径
        data: 要填充的数据字典
        editable_config: 可编辑区域配置
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    # 加载模板
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    
    # 根据配置填充数据
    if editable_config:
        # 如果有配置，按配置填充
        for cell_range, cell_data in editable_config.get('cells', {}).items():
            if cell_data.get('field') and cell_data['field'] in data:
                # 解析单元格地址
                col_letter = cell_range[0]
                row_num = int(cell_range[1:])
                
                # 填充数据
                ws[cell_range] = data[cell_data['field']]
    else:
        # 没有配置时，尝试智能填充
        # 遍历所有单元格，查找可编辑区域
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_col + 1):
                cell = ws.cell(row, col)
                cell_addr = cell.coordinate
                
                # 检查是否在可编辑区域
                if data and cell_addr in data:
                    cell.value = data[cell_addr]
    
    # 保存文件
    wb.save(output_file)
'''

# 读取原文件
with open(r'e:\FHD\AI 助手\excel_template_analyzer.py', 'r', encoding='utf-8') as f:
    content = f.read()

# 添加函数
content += function_to_add

# 写回文件
with open(r'e:\FHD\AI 助手\excel_template_analyzer.py', 'w', encoding='utf-8') as f:
    f.write(content)

print('✓ generate_from_template 函数添加成功')
