"""
Excel 数据提取引擎
从 Excel 文件中自动提取结构化数据，识别数据类型（产品、客户、订单等）
并支持数据验证、清洗和导入数据库
"""

import re
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, asdict
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter


@dataclass
class ExtractedData:
    """提取的数据"""
    data_type: str  # products, customers, orders, materials, other
    confidence: float  # 置信度 0-1
    field_mapping: Dict[str, str]  # 列->字段映射
    data: List[Dict[str, Any]]  # 提取的数据列表
    validation: Dict[str, Any]  # 验证结果


@dataclass
class ValidationError:
    """验证错误"""
    row: int
    field: str
    value: Any
    message: str


class ExcelDataExtractor:
    """Excel 数据提取器"""
    
    # 数据类型识别关键词
    PRODUCT_KEYWORDS = ['产品型号', '产品名称', '规格', '型号', '单价', '数量', '单位']
    CUSTOMER_KEYWORDS = ['客户', '联系人', '电话', '地址', '公司', '手机', '传真']
    ORDER_KEYWORDS = ['订单', '送货单', '发货单', '订单编号', '订单号', '日期', '金额']
    MATERIAL_KEYWORDS = ['材料', '原料', '用量', '库存', '批次']
    
    # 字段映射规则
    FIELD_MAPPINGS = {
        # 产品字段
        '产品型号': 'product_code',
        '型号': 'product_code',
        '产品名称': 'product_name',
        '名称': 'product_name',
        '规格': 'specification',
        '规格/KG': 'specification',
        '规格/型号': 'specification',
        '单价': 'unit_price',
        '单价/元': 'unit_price',
        '数量': 'quantity',
        '数量/件': 'quantity',
        '数量/KG': 'quantity',
        '单位': 'unit',
        '金额': 'amount',
        '备注': 'remark',
        
        # 客户字段
        '客户名称': 'customer_name',
        '客户': 'customer_name',
        '购货单位': 'customer_name',
        '联系人': 'contact_person',
        '电话': 'phone',
        '手机': 'mobile',
        '地址': 'address',
        '公司名称': 'company_name',
        '公司': 'company_name',
        
        # 订单字段
        '订单编号': 'order_number',
        '订单号': 'order_number',
        '编号': 'order_number',
        '日期': 'date',
        '送货日期': 'delivery_date',
        '发货日期': 'ship_date',
        '总金额': 'total_amount',
        '合计': 'total_amount',
    }
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.max_row = 0
        self.max_col = 0
        self.headers = []
        
    def extract(self, sheet_name: Optional[str] = None) -> ExtractedData:
        """提取 Excel 数据"""
        self._load_workbook(sheet_name)
        self._extract_headers()
        
        # 识别数据类型
        data_type, confidence = self._identify_data_type()
        
        # 字段映射
        field_mapping = self._map_fields()
        
        # 提取数据
        data = self._extract_data(field_mapping)
        
        # 验证数据
        validation = self._validate_data(data, field_mapping)
        
        return ExtractedData(
            data_type=data_type,
            confidence=confidence,
            field_mapping=field_mapping,
            data=data,
            validation=validation
        )
    
    def _load_workbook(self, sheet_name: Optional[str] = None):
        """加载工作簿"""
        self.workbook = openpyxl.load_workbook(self.file_path)
        
        if sheet_name:
            self.worksheet = self.workbook[sheet_name]
        else:
            self.worksheet = self.workbook.active
            
        self.max_row = self.worksheet.max_row
        self.max_col = self.worksheet.max_column
    
    def _extract_headers(self):
        """提取表头"""
        # 假设表头在第 1-3 行
        headers = []
        for row in range(1, min(4, self.max_row + 1)):
            row_headers = []
            for col in range(1, self.max_col + 1):
                cell = self.worksheet.cell(row, col)
                value = cell.value
                if value:
                    row_headers.append(str(value).strip())
                else:
                    row_headers.append('')
            headers.append(row_headers)
        
        # 合并多行表头
        self.headers = ['_'.join([h[i] for h in headers if h[i]]).strip() 
                       for i in range(self.max_col)]
    
    def _identify_data_type(self) -> Tuple[str, float]:
        """识别数据类型"""
        header_text = ' '.join(self.headers)
        
        scores = {
            'products': 0,
            'customers': 0,
            'orders': 0,
            'materials': 0,
            'other': 0
        }
        
        # 检查产品关键词
        for keyword in self.PRODUCT_KEYWORDS:
            if keyword in header_text:
                scores['products'] += 1
        
        # 检查客户关键词
        for keyword in self.CUSTOMER_KEYWORDS:
            if keyword in header_text:
                scores['customers'] += 1
        
        # 检查订单关键词
        for keyword in self.ORDER_KEYWORDS:
            if keyword in header_text:
                scores['orders'] += 1
        
        # 检查原材料关键词
        for keyword in self.MATERIAL_KEYWORDS:
            if keyword in header_text:
                scores['materials'] += 1
        
        # 找出得分最高的类型
        max_score = max(scores.values())
        if max_score == 0:
            return 'other', 0.5
        
        data_type = [k for k, v in scores.items() if v == max_score][0]
        confidence = min(0.5 + max_score * 0.15, 1.0)  # 置信度计算
        
        return data_type, confidence
    
    def _map_fields(self) -> Dict[str, str]:
        """字段映射"""
        mapping = {}
        
        for col_idx, header in enumerate(self.headers):
            col_letter = get_column_letter(col_idx + 1)
            
            # 精确匹配
            if header in self.FIELD_MAPPINGS:
                mapping[col_letter] = self.FIELD_MAPPINGS[header]
            else:
                # 模糊匹配
                for key, value in self.FIELD_MAPPINGS.items():
                    if key in header or header in key:
                        mapping[col_letter] = value
                        break
        
        return mapping
    
    def _extract_data(self, field_mapping: Dict[str, str]) -> List[Dict[str, Any]]:
        """提取数据"""
        data = []
        
        # 从第 4 行开始（假设前 3 行是表头）
        start_row = 4
        if self.max_row < 4:
            start_row = 1
        
        for row in range(start_row, self.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_letter, field_name in field_mapping.items():
                col_idx = column_index_from_string(col_letter)
                cell = self.worksheet.cell(row, col_idx)
                value = cell.value
                
                if value is not None:
                    row_data[field_name] = value
                    has_data = True
            
            if has_data:
                data.append(row_data)
        
        return data
    
    def _validate_data(self, data: List[Dict], field_mapping: Dict[str, str]) -> Dict[str, Any]:
        """验证数据"""
        errors = []
        valid_count = 0
        invalid_count = 0
        
        for row_idx, row_data in enumerate(data, start=1):
            row_errors = []
            
            # 验证必填字段
            required_fields = ['product_code', 'product_name', 'customer_name', 'order_number']
            for field in required_fields:
                if field in row_data and (row_data[field] is None or str(row_data[field]).strip() == ''):
                    row_errors.append(ValidationError(
                        row=row_idx,
                        field=field,
                        value=row_data.get(field),
                        message=f"必填字段 '{field}' 为空"
                    ))
            
            # 验证数字字段
            numeric_fields = ['quantity', 'unit_price', 'amount', 'total_amount']
            for field in numeric_fields:
                if field in row_data:
                    value = row_data[field]
                    if value is not None and not self._is_numeric(value):
                        row_errors.append(ValidationError(
                            row=row_idx,
                            field=field,
                            value=value,
                            message=f"字段 '{field}' 不是有效数字"
                        ))
            
            # 验证电话号码
            if 'phone' in row_data or 'mobile' in row_data:
                phone = row_data.get('phone') or row_data.get('mobile')
                if phone and not self._is_phone_valid(str(phone)):
                    row_errors.append(ValidationError(
                        row=row_idx,
                        field='phone',
                        value=phone,
                        message="电话号码格式不正确"
                    ))
            
            if row_errors:
                errors.extend(row_errors)
                invalid_count += 1
            else:
                valid_count += 1
        
        return {
            'total': len(data),
            'valid': valid_count,
            'invalid': invalid_count,
            'errors': [asdict(e) for e in errors]
        }
    
    def _is_numeric(self, value: Any) -> bool:
        """检查是否为数字"""
        if isinstance(value, (int, float)):
            return True
        
        try:
            # 尝试转换为数字（去除千分位等）
            cleaned = str(value).replace(',', '').replace(' ', '')
            float(cleaned)
            return True
        except (ValueError, TypeError):
            return False
    
    def _is_phone_valid(self, phone: str) -> bool:
        """验证电话号码"""
        # 简单的电话号码验证
        patterns = [
            r'^1[3-9]\d{9}$',  # 手机号
            r'^0\d{2,3}-?\d{7,8}$',  # 固话
            r'^\d{7,11}$'  # 简单号码
        ]
        
        return any(re.match(pattern, phone) for pattern in patterns)
    
    def clean_data(self, data: List[Dict]) -> List[Dict]:
        """清洗数据"""
        cleaned = []
        
        for row_data in data:
            cleaned_row = {}
            
            for field, value in row_data.items():
                if value is None:
                    continue
                
                # 字符串处理
                if isinstance(value, str):
                    value = value.strip()
                    
                    # 去除单位（如"5 个"→5）
                    if field in ['quantity', 'unit_price', 'amount']:
                        match = re.match(r'^[\d,.]+', value)
                        if match:
                            value = match.group()
                
                # 数字处理
                if isinstance(value, (int, float)) or self._is_numeric(value):
                    try:
                        value = float(str(value).replace(',', '').replace(' ', ''))
                        # 如果是整数，转换为 int
                        if value == int(value):
                            value = int(value)
                    except (ValueError, TypeError):
                        pass
                
                cleaned_row[field] = value
            
            cleaned.append(cleaned_row)
        
        return cleaned
    
    def deduplicate(self, data: List[Dict], key_field: str) -> List[Dict]:
        """去重"""
        seen = set()
        unique = []
        
        for row_data in data:
            key_value = row_data.get(key_field)
            if key_value is not None and key_value not in seen:
                seen.add(key_value)
                unique.append(row_data)
        
        return unique


def extract_data(file_path: str, sheet_name: Optional[str] = None) -> ExtractedData:
    """提取 Excel 数据的便捷函数"""
    extractor = ExcelDataExtractor(file_path)
    return extractor.extract(sheet_name)


def clean_and_validate(file_path: str, data: List[Dict]) -> Tuple[List[Dict], Dict]:
    """清洗并验证数据"""
    extractor = ExcelDataExtractor(file_path)
    cleaned = extractor.clean_data(data)
    validation = extractor._validate_data(cleaned, {})
    return cleaned, validation
